import csv
from collections import defaultdict
from datetime import date, datetime, time
from io import BytesIO, StringIO
import unicodedata

from openpyxl import Workbook, load_workbook
from django.db import transaction

from apps.audit.services import AuditService
from apps.imports.models import ImportBatch, ImportPreviewRow
from apps.scheduling.models import ScheduleAssignment
from apps.workers.models import Area, Shift, SpecialState, Worker


class ImportSampleService:
    WORKERS_HEADERS = ["DNI", "Nombre", "Apellido", "Area", "Sede"]
    SHIFTS_HEADERS = [
        "Area",
        "Turno",
        "Codigo BUK",
        "Hora Inicio",
        "Hora Fin",
        "Inicio Break",
        "Fin Break",
        "Nocturno",
        "Activo",
        "Sede",
    ]

    @staticmethod
    def _yes_no(value):
        return "1" if value else "0"

    @staticmethod
    def _fmt_time(value):
        return value.strftime("%H:%M") if value else ""

    @staticmethod
    def build_sample_payload(*, tenant, property_obj, max_workers=12, max_shifts=20):
        workers = list(
            Worker.objects.filter(tenant=tenant, property=property_obj, active=True)
            .select_related("area")
            .order_by("last_name", "first_name")[: max(1, int(max_workers))]
        )
        shifts = list(
            Shift.objects.filter(tenant=tenant, property=property_obj, active=True)
            .select_related("area")
            .order_by("area__name", "name")[: max(1, int(max_shifts))]
        )

        workers_rows = [
            [
                worker.document_number,
                worker.first_name,
                worker.last_name,
                worker.area.name if worker.area_id else "",
                property_obj.name,
            ]
            for worker in workers
        ]
        shifts_rows = [
            [
                shift.area.name if shift.area_id else "",
                shift.name,
                shift.buk_code,
                ImportSampleService._fmt_time(shift.start_time),
                ImportSampleService._fmt_time(shift.end_time),
                ImportSampleService._fmt_time(shift.break_start),
                ImportSampleService._fmt_time(shift.break_end),
                ImportSampleService._yes_no(shift.is_night_shift),
                ImportSampleService._yes_no(shift.active),
                property_obj.name,
            ]
            for shift in shifts
        ]
        return {
            "workers_headers": list(ImportSampleService.WORKERS_HEADERS),
            "workers_rows": workers_rows,
            "shifts_headers": list(ImportSampleService.SHIFTS_HEADERS),
            "shifts_rows": shifts_rows,
        }

    @staticmethod
    def generate_csv_bytes(*, headers, rows):
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")

    @staticmethod
    def generate_xlsx_bytes(*, sheet_name, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        for row in rows:
            ws.append(row)
        out = BytesIO()
        wb.save(out)
        return out.getvalue()


class ExcelImportService:
    MONTH_SHEETS = {
        "Enero 2026",
        "Febrero 2026",
        "Marzo 2026",
        "Abril 2026",
        "Mayo 2026",
        "Junio 2026",
        "Julio 2026",
        "Agosto 2026",
        "Septiembre 2026",
        "Octubre 2026",
        "Noviembre 2026",
        "Diciembre 2026",
    }

    @staticmethod
    def _normalize_text(value):
        return "" if value is None else str(value).strip()

    @staticmethod
    def _parse_time(value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.time()
        text = str(value).strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_time_range(value):
        text = ExcelImportService._normalize_text(value).replace("–", "-")
        if "-" not in text:
            return None, None
        left, right = text.split("-", 1)
        return ExcelImportService._parse_time(left), ExcelImportService._parse_time(right)

    @staticmethod
    def _normalize_schedule_key(value):
        start_time, end_time = ExcelImportService._parse_time_range(value)
        if start_time and end_time:
            return f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"
        return ExcelImportService._normalize_text(value).replace(" ", "").lower()

    @staticmethod
    def _serialize_time(value):
        if value is None:
            return None
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        return str(value)

    @staticmethod
    def create_preview(*, tenant, property_obj, file_name, file_bytes, user):
        workbook = load_workbook(BytesIO(file_bytes), data_only=False)
        workbook_values = load_workbook(BytesIO(file_bytes), data_only=True)

        sheets_summary = []
        for ws in workbook.worksheets:
            sheets_summary.append(
                {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_col": ws.max_column,
                    "merged_ranges": len(ws.merged_cells.ranges),
                    "has_formulas": any(
                        isinstance(cell.value, str) and cell.value.startswith("=")
                        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30))
                        for cell in row
                    ),
                }
            )

        batch = ImportBatch.objects.create(
            tenant=tenant,
            property=property_obj,
            source_type="excel_original",
            file_name=file_name,
            created_by=user,
            summary={
                "sheets": sheets_summary,
                "detected": {
                    "areas": 0,
                    "workers": 0,
                    "shifts": 0,
                    "special_states": 0,
                    "assignments": 0,
                },
                "new": {"areas": 0, "workers": 0, "shifts": 0, "special_states": 0, "assignments": 0},
                "updates": {"workers": 0, "shifts": 0, "special_states": 0},
                "warnings": 0,
                "errors": 0,
                "skipped": 0,
                "uninterpreted_sheets": [],
            },
        )

        for item in sheets_summary:
            ImportPreviewRow.objects.create(
                batch=batch,
                sheet_name=item["name"],
                row_number=0,
                action="inspect",
                status="ok",
                payload=item,
            )

        areas_by_name = {
            area.name.lower(): area
            for area in Area.objects.filter(tenant=tenant, property=property_obj)
        }
        special_by_name = {
            state.name.lower(): state
            for state in SpecialState.objects.filter(tenant=tenant, property=property_obj)
        }

        # Datos sheet: areas + special states.
        if "Datos" in workbook_values.sheetnames:
            ws = workbook_values["Datos"]
            for row_num in range(6, 19):
                area_name = ExcelImportService._normalize_text(ws.cell(row_num, 3).value)
                if not area_name:
                    continue
                batch.summary["detected"]["areas"] += 1
                exists = area_name.lower() in areas_by_name
                if not exists:
                    batch.summary["new"]["areas"] += 1
                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name="Datos",
                    row_number=row_num * 10 + 1,
                    action="update" if exists else "create",
                    status="ok",
                    payload={"entity": "area", "name": area_name, "property_id": property_obj.id},
                )

            for row_num in range(6, 20):
                state_name = ExcelImportService._normalize_text(ws.cell(row_num, 8).value)
                if not state_name:
                    continue
                batch.summary["detected"]["special_states"] += 1
                exists = state_name.lower() in special_by_name
                if exists:
                    batch.summary["updates"]["special_states"] += 1
                else:
                    batch.summary["new"]["special_states"] += 1
                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name="Datos",
                    row_number=row_num * 10 + 2,
                    action="update" if exists else "create",
                    status="ok",
                    payload={
                        "entity": "special_state",
                        "name": state_name,
                        "buk_code": "D",
                        "property_id": property_obj.id,
                    },
                )

        # Worker registry.
        if "Reg. de trabajadores x area" in workbook_values.sheetnames:
            ws = workbook_values["Reg. de trabajadores x area"]
            for row_num in range(4, ws.max_row + 1):
                doc = WorkerImportService.normalize_document(ws.cell(row_num, 1).value)
                if not doc:
                    continue
                first_name = ExcelImportService._normalize_text(ws.cell(row_num, 2).value)
                last_name = ExcelImportService._normalize_text(ws.cell(row_num, 3).value)
                area_name = ExcelImportService._normalize_text(ws.cell(row_num, 4).value)
                batch.summary["detected"]["workers"] += 1
                if not area_name:
                    batch.summary["errors"] += 1
                    ImportPreviewRow.objects.create(
                        batch=batch,
                        sheet_name="Reg. de trabajadores x area",
                        row_number=row_num,
                        action="skip",
                        status="error",
                        message="Worker row without area.",
                        payload={"entity": "worker", "document_number": doc},
                    )
                    continue
                exists = Worker.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    document_number=doc,
                ).exists()
                if exists:
                    batch.summary["updates"]["workers"] += 1
                else:
                    batch.summary["new"]["workers"] += 1
                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name="Reg. de trabajadores x area",
                    row_number=row_num,
                    action="update" if exists else "create",
                    status="ok",
                    payload={
                        "entity": "worker",
                        "document_number": doc,
                        "first_name": first_name,
                        "last_name": last_name,
                        "area_name": area_name,
                        "property_id": property_obj.id,
                    },
                )

        # Shift registry.
        shift_sheet_name = next((name for name in workbook_values.sheetnames if name.startswith("Reg. Horarios x")), None)
        shift_by_area_schedule = {}
        shift_by_area_schedule_norm = {}
        shift_by_area_code = {}
        shift_by_area_name = {}
        shift_names_by_area = defaultdict(set)
        if shift_sheet_name:
            ws = workbook_values[shift_sheet_name]
            for row_num in range(4, ws.max_row + 1):
                shift_name = ExcelImportService._normalize_text(ws.cell(row_num, 1).value)
                if not shift_name:
                    continue
                buk_code = ExcelImportService._normalize_text(ws.cell(row_num, 2).value)
                area_name = ExcelImportService._normalize_text(ws.cell(row_num, 3).value)
                schedule_text = ExcelImportService._normalize_text(ws.cell(row_num, 4).value)
                break_text = ExcelImportService._normalize_text(ws.cell(row_num, 5).value)

                batch.summary["detected"]["shifts"] += 1
                exists = Shift.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    area__name__iexact=area_name,
                    name=shift_name,
                ).exists()
                if exists:
                    batch.summary["updates"]["shifts"] += 1
                else:
                    batch.summary["new"]["shifts"] += 1

                start_time, end_time = ExcelImportService._parse_time_range(schedule_text)
                break_start, break_end = ExcelImportService._parse_time_range(break_text)
                area_key = area_name.lower()
                shift_by_area_schedule[(area_key, schedule_text.lower())] = shift_name
                shift_by_area_schedule_norm[(area_key, ExcelImportService._normalize_schedule_key(schedule_text))] = shift_name
                if buk_code:
                    shift_by_area_code[(area_key, buk_code.upper())] = shift_name
                shift_by_area_name[(area_key, shift_name.lower())] = shift_name
                shift_names_by_area[area_key].add(shift_name)

                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name=shift_sheet_name,
                    row_number=row_num,
                    action="update" if exists else "create",
                    status="ok",
                    payload={
                        "entity": "shift",
                        "name": shift_name,
                        "buk_code": buk_code,
                        "area_name": area_name,
                        "start_time": ExcelImportService._serialize_time(start_time),
                        "end_time": ExcelImportService._serialize_time(end_time),
                        "break_start": ExcelImportService._serialize_time(break_start),
                        "break_end": ExcelImportService._serialize_time(break_end),
                        "schedule_text": schedule_text,
                        "property_id": property_obj.id,
                    },
                )

        # Monthly assignments.
        interpreted_sheets = {"Datos", "Reg. de trabajadores x area", shift_sheet_name, "Reporte carga BUK"}
        for sheet_name in workbook_values.sheetnames:
            if sheet_name not in ExcelImportService.MONTH_SHEETS:
                if sheet_name not in interpreted_sheets:
                    batch.summary["uninterpreted_sheets"].append(sheet_name)
                continue

            ws = workbook_values[sheet_name]
            date_headers = {}
            for col in range(4, ws.max_column + 1):
                cell_value = ws.cell(5, col).value
                if isinstance(cell_value, datetime):
                    date_headers[col] = cell_value.date()
            current_area = ""
            for row_num in range(6, ws.max_row + 1):
                col_a = ExcelImportService._normalize_text(ws.cell(row_num, 1).value)
                col_b = ExcelImportService._normalize_text(ws.cell(row_num, 2).value)
                col_c = ExcelImportService._normalize_text(ws.cell(row_num, 3).value)
                if col_a and not col_b and not col_c:
                    current_area = col_a
                    continue
                if col_c.upper() != "TURNO":
                    continue
                worker_document = WorkerImportService.normalize_document(ws.cell(row_num, 1).value)
                if not worker_document:
                    continue

                for col, assignment_date in date_headers.items():
                    assignment_value = ExcelImportService._normalize_text(ws.cell(row_num, col).value)
                    if not assignment_value:
                        continue

                    batch.summary["detected"]["assignments"] += 1
                    payload = {
                        "entity": "assignment",
                        "document_number": worker_document,
                        "date": assignment_date.isoformat(),
                        "property_id": property_obj.id,
                    }
                    status = "ok"
                    if assignment_value.upper() in {"OFF", "VACACIONES", "LICENCIA"}:
                        payload["special_state_name"] = assignment_value.upper()
                    else:
                        area_key = current_area.lower()
                        assignment_key_raw = assignment_value.lower()
                        assignment_key_norm = ExcelImportService._normalize_schedule_key(assignment_value)
                        payload["area_name"] = current_area
                        payload["schedule_text"] = assignment_value
                        payload["shift_name"] = shift_by_area_schedule.get(
                            (area_key, assignment_key_raw)
                        )
                        if payload["shift_name"] is None:
                            payload["shift_name"] = shift_by_area_schedule_norm.get((area_key, assignment_key_norm))
                        if payload["shift_name"] is None:
                            payload["shift_name"] = shift_by_area_code.get((area_key, assignment_value.upper()))
                        if payload["shift_name"] is None:
                            payload["shift_name"] = shift_by_area_name.get((area_key, assignment_value.lower()))
                        if payload["shift_name"] is None and len(shift_names_by_area.get(area_key, set())) == 1:
                            payload["shift_name"] = next(iter(shift_names_by_area[area_key]))
                        if payload["shift_name"] is None:
                            status = "warning"
                            batch.summary["warnings"] += 1

                    ImportPreviewRow.objects.create(
                        batch=batch,
                        sheet_name=sheet_name,
                        row_number=row_num * 1000 + col,
                        action="upsert",
                        status=status,
                        payload=payload,
                    )

        batch.save(update_fields=["summary", "updated_at"])
        return batch


class WorkerImportService:
    REQUIRED_COLUMNS = {"dni", "nombre", "apellido", "area"}

    @staticmethod
    def normalize_document(value):
        if value is None:
            return ""
        text = str(value).strip()
        return text.replace(".0", "") if text.endswith(".0") else text

    @staticmethod
    def _normalize_header(header):
        text = str(header or "").strip().lower()
        mapping = {
            "dni": "dni",
            "documento": "dni",
            "nro documento": "dni",
            "numero de documento": "dni",
            "número de documento": "dni",
            "nombre": "nombre",
            "nombres": "nombre",
            "apellido": "apellido",
            "apellidos": "apellido",
            "area": "area",
            "área": "area",
            "sede": "sede",
            "property": "sede",
        }
        return mapping.get(text, text)

    @staticmethod
    def _read_worker_rows(file_name, file_bytes):
        lower = file_name.lower()
        if lower.endswith(".csv"):
            text = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            headers = [WorkerImportService._normalize_header(h) for h in (reader.fieldnames or [])]
            rows = []
            for raw_row in reader:
                row = {}
                for key, value in raw_row.items():
                    row[WorkerImportService._normalize_header(key)] = value
                rows.append(row)
            return headers, rows

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        raw_headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        headers = [WorkerImportService._normalize_header(h) for h in raw_headers]
        rows = []
        for row_num in range(2, ws.max_row + 1):
            payload = {}
            has_any_value = False
            for col_num, key in enumerate(headers, start=1):
                value = ws.cell(row_num, col_num).value
                payload[key] = value
                if value not in (None, ""):
                    has_any_value = True
            if has_any_value:
                rows.append(payload)
        return headers, rows

    @staticmethod
    def create_worker_preview(
        *,
        tenant,
        fallback_property,
        file_name,
        file_bytes,
        user,
        create_missing_areas=False,
        sync_mode=False,
    ):
        headers, rows = WorkerImportService._read_worker_rows(file_name, file_bytes)
        missing = WorkerImportService.REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

        batch = ImportBatch.objects.create(
            tenant=tenant,
            property=fallback_property,
            source_type="workers",
            file_name=file_name,
            created_by=user,
            summary={},
        )

        properties_by_name = {prop.name.strip().lower(): prop for prop in tenant.properties.all()}
        multiple_properties = len(properties_by_name) > 1
        summary = {
            "detected_rows": len(rows),
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "to_deactivate": 0,
            "skipped": 0,
            "errors": 0,
            "new_areas": 0,
            "unknown_properties": 0,
            "mismatched_properties": 0,
            "warnings": 0,
            "sync_mode": bool(sync_mode),
        }
        incoming_documents_by_property = defaultdict(set)

        for row_index, row in enumerate(rows, start=2):
            document = WorkerImportService.normalize_document(row.get("dni"))
            first_name = str(row.get("nombre") or "").strip()
            last_name = str(row.get("apellido") or "").strip()
            area_name = str(row.get("area") or "").strip()
            property_name = str(row.get("sede") or "").strip()

            status = "ok"
            action = "skip"
            message = ""

            if not document or not first_name or not last_name or not area_name:
                status = "error"
                message = "document, first name, last name and area are required"

            if status == "ok" and multiple_properties and not property_name:
                status = "error"
                message = "sede is required when file has mixed properties"

            property_obj = fallback_property
            if status == "ok" and property_name:
                property_obj = properties_by_name.get(property_name.lower())
                if property_obj is None:
                    status = "error"
                    message = f"unknown property: {property_name}"
                    summary["unknown_properties"] += 1
                elif property_obj.id != fallback_property.id:
                    status = "error"
                    message = f"sede no coincide con la sede seleccionada: {property_name}"
                    summary["mismatched_properties"] += 1

            if status == "ok":
                area_obj = Area.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    name__iexact=area_name,
                ).first()
                if area_obj is None and not create_missing_areas:
                    status = "error"
                    message = f"area does not exist: {area_name}"
                elif area_obj is None:
                    action = "create_area"
                    summary["new_areas"] += 1

            if status == "ok":
                incoming_documents_by_property[property_obj.id].add(document)
                exists = Worker.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    document_number=document,
                ).select_related("area").first()
                if exists:
                    changes = {}
                    if exists.first_name != first_name:
                        changes["first_name"] = {"from": exists.first_name, "to": first_name}
                    if exists.last_name != last_name:
                        changes["last_name"] = {"from": exists.last_name, "to": last_name}
                    if exists.area_id != (area_obj.id if area_obj else None):
                        changes["area"] = {
                            "from": exists.area.name if exists.area_id else "",
                            "to": area_name,
                        }
                    if not exists.active:
                        changes["active"] = {"from": False, "to": True}
                    if changes:
                        action = "update"
                        status = "warning"
                        message = "Cambios detectados; revisa antes de confirmar."
                        summary["updated"] += 1
                        summary["warnings"] += 1
                    else:
                        action = "keep"
                        summary["unchanged"] += 1
                else:
                    action = "create"
                    summary["created"] += 1
            else:
                summary["errors"] += 1

            ImportPreviewRow.objects.create(
                batch=batch,
                sheet_name="workers_import",
                row_number=row_index,
                action=action,
                status=status,
                message=message,
                payload={
                    "document_number": document,
                    "first_name": first_name,
                    "last_name": last_name,
                    "area_name": area_name,
                    "property_name": property_name,
                    "property_id": property_obj.id if property_obj else None,
                    "changes": changes if status == "warning" and action == "update" else {},
                },
            )

        if sync_mode:
            incoming_documents = incoming_documents_by_property.get(fallback_property.id, set())
            active_missing_workers = Worker.objects.filter(
                tenant=tenant,
                property=fallback_property,
                active=True,
            ).exclude(document_number__in=incoming_documents)
            next_row_number = 100000
            for worker in active_missing_workers.select_related("area").order_by("last_name", "first_name", "id"):
                summary["to_deactivate"] += 1
                summary["warnings"] += 1
                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name="workers_import",
                    row_number=next_row_number,
                    action="deactivate",
                    status="warning",
                    message="No aparece en el archivo completo de la sede; se marcara como inactivo.",
                    payload={
                        "worker_id": worker.id,
                        "document_number": worker.document_number,
                        "first_name": worker.first_name,
                        "last_name": worker.last_name,
                        "area_name": worker.area.name if worker.area_id else "",
                        "property_name": fallback_property.name,
                        "property_id": fallback_property.id,
                    },
                )
                next_row_number += 1

        batch.summary = summary
        batch.save(update_fields=["summary", "updated_at"])
        return batch

    @staticmethod
    def confirm_worker_import(*, batch):
        created = 0
        updated = 0
        unchanged = 0
        deactivated = 0
        new_areas = 0
        actor = batch.created_by
        rows = batch.preview_rows.filter(status__in=["ok", "warning"]).order_by("row_number")
        with transaction.atomic():
            for row in rows:
                payload = row.payload
                if row.action == "deactivate":
                    worker_id = payload.get("worker_id")
                    worker = Worker.objects.filter(
                        tenant=batch.tenant,
                        property=batch.property,
                        id=worker_id,
                        active=True,
                    ).first()
                    if worker is None:
                        continue
                    before = {
                        "document_number": worker.document_number,
                        "active": worker.active,
                    }
                    worker.active = False
                    worker.save(update_fields=["active", "updated_at"])
                    deactivated += 1
                    if actor:
                        AuditService.log(
                            tenant=batch.tenant,
                            property_obj=worker.property,
                            user=actor,
                            action="worker_deactivate_from_import_sync",
                            entity_type="Worker",
                            entity_id=worker.id,
                            before=before,
                            after={"document_number": worker.document_number, "active": worker.active},
                        )
                    continue
                if row.action == "keep":
                    unchanged += 1
                    continue
                property_id = payload.get("property_id")
                if not property_id:
                    continue
                area, area_created = Area.objects.get_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["area_name"],
                    defaults={"type": "", "active": True},
                )
                if area_created:
                    new_areas += 1
                worker, worker_created = Worker.objects.update_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    document_number=payload["document_number"],
                    defaults={
                        "first_name": payload["first_name"],
                        "last_name": payload["last_name"],
                        "area": area,
                        "active": True,
                    },
                )
                if worker_created:
                    created += 1
                    if actor:
                        AuditService.log(
                            tenant=batch.tenant,
                            property_obj=worker.property,
                            user=actor,
                            action="worker_create_from_import",
                            entity_type="Worker",
                            entity_id=worker.id,
                            before={},
                            after={"document_number": worker.document_number},
                        )
                else:
                    updated += 1
                    if actor:
                        AuditService.log(
                            tenant=batch.tenant,
                            property_obj=worker.property,
                            user=actor,
                            action="worker_update_from_import",
                            entity_type="Worker",
                            entity_id=worker.id,
                            before={},
                            after={"document_number": worker.document_number},
                        )
        batch.summary = {
            **batch.summary,
            "applied_created": created,
            "applied_updated": updated,
            "applied_unchanged": unchanged,
            "applied_deactivated": deactivated,
            "applied_new_areas": new_areas,
        }
        batch.status = "confirmed"
        batch.save(update_fields=["summary", "status", "updated_at"])
        return batch


class ShiftAreaImportService:
    REQUIRED_COLUMNS = {"area", "turno", "buk_code", "start_time", "end_time"}

    @staticmethod
    def _normalize_header(header):
        text = str(header or "").strip().lower()
        mapping = {
            "area": "area",
            "area_name": "area",
            "turno": "turno",
            "nombre turno": "turno",
            "nombre_turno": "turno",
            "shift": "turno",
            "name": "turno",
            "codigo": "buk_code",
            "codigo buk": "buk_code",
            "codigo_buk": "buk_code",
            "buk_code": "buk_code",
            "code": "buk_code",
            "hora inicio": "start_time",
            "hora_inicio": "start_time",
            "inicio": "start_time",
            "start_time": "start_time",
            "hora fin": "end_time",
            "hora_fin": "end_time",
            "fin": "end_time",
            "end_time": "end_time",
            "inicio break": "break_start",
            "break inicio": "break_start",
            "break_start": "break_start",
            "fin break": "break_end",
            "break fin": "break_end",
            "break_end": "break_end",
            "nocturno": "is_night_shift",
            "turno nocturno": "is_night_shift",
            "is_night_shift": "is_night_shift",
            "activo": "active",
            "active": "active",
            "sede": "sede",
        }
        return mapping.get(text, text)

    @staticmethod
    def _read_rows(file_name, file_bytes):
        lower = file_name.lower()
        if lower.endswith(".csv"):
            text = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            headers = [ShiftAreaImportService._normalize_header(h) for h in (reader.fieldnames or [])]
            rows = []
            for raw_row in reader:
                row = {}
                for key, value in raw_row.items():
                    row[ShiftAreaImportService._normalize_header(key)] = value
                rows.append(row)
            return headers, rows

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        raw_headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        headers = [ShiftAreaImportService._normalize_header(h) for h in raw_headers]
        rows = []
        for row_num in range(2, ws.max_row + 1):
            payload = {}
            has_any_value = False
            for col_num, key in enumerate(headers, start=1):
                value = ws.cell(row_num, col_num).value
                payload[key] = value
                if value not in (None, ""):
                    has_any_value = True
            if has_any_value:
                rows.append(payload)
        return headers, rows

    @staticmethod
    def _parse_bool(value, default=False):
        if value in (None, ""):
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in {"1", "true", "si", "yes", "y", "x"}

    @staticmethod
    def create_shift_preview(
        *,
        tenant,
        fallback_property,
        file_name,
        file_bytes,
        user,
        create_missing_areas=False,
        sync_mode=False,
    ):
        headers, rows = ShiftAreaImportService._read_rows(file_name, file_bytes)
        missing = ShiftAreaImportService.REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

        batch = ImportBatch.objects.create(
            tenant=tenant,
            property=fallback_property,
            source_type="shifts_area",
            file_name=file_name,
            created_by=user,
            summary={},
        )

        properties_by_name = {prop.name.strip().lower(): prop for prop in tenant.properties.all()}
        multiple_properties = len(properties_by_name) > 1
        summary = {
            "detected_rows": len(rows),
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "to_deactivate": 0,
            "skipped": 0,
            "errors": 0,
            "new_areas": 0,
            "unknown_properties": 0,
            "mismatched_properties": 0,
            "warnings": 0,
            "sync_mode": bool(sync_mode),
        }
        matched_shift_ids_by_property = defaultdict(set)

        for row_index, row in enumerate(rows, start=2):
            area_name = str(row.get("area") or "").strip()
            shift_name = str(row.get("turno") or "").strip()
            buk_code = str(row.get("buk_code") or "").strip()
            property_name = str(row.get("sede") or "").strip()
            start_time = ExcelImportService._parse_time(row.get("start_time"))
            end_time = ExcelImportService._parse_time(row.get("end_time"))
            break_start = ExcelImportService._parse_time(row.get("break_start"))
            break_end = ExcelImportService._parse_time(row.get("break_end"))
            night_flag_raw = row.get("is_night_shift")
            active_raw = row.get("active")

            status = "ok"
            action = "skip"
            message = ""
            property_obj = fallback_property

            if not area_name or not shift_name or not buk_code or not start_time or not end_time:
                status = "error"
                message = "area, turno, buk_code, start_time y end_time son obligatorios"

            if status == "ok" and multiple_properties and not property_name:
                status = "error"
                message = "sede is required when file has mixed properties"

            if status == "ok" and property_name:
                property_obj = properties_by_name.get(property_name.lower())
                if property_obj is None:
                    status = "error"
                    message = f"unknown property: {property_name}"
                    summary["unknown_properties"] += 1
                elif property_obj.id != fallback_property.id:
                    status = "error"
                    message = f"sede no coincide con la sede seleccionada: {property_name}"
                    summary["mismatched_properties"] += 1

            if status == "ok":
                area_obj = Area.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    name__iexact=area_name,
                ).first()
                if area_obj is None and not create_missing_areas:
                    status = "error"
                    message = f"area does not exist: {area_name}"
                elif area_obj is None:
                    action = "create_area"
                    summary["new_areas"] += 1

            if status == "ok":
                existing = Shift.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    buk_code__iexact=buk_code,
                ).select_related("area").first()
                if existing is None:
                    existing = Shift.objects.filter(
                        tenant=tenant,
                        property=property_obj,
                        area__name__iexact=area_name,
                        name__iexact=shift_name,
                    ).select_related("area").first()
                if existing:
                    matched_shift_ids_by_property[property_obj.id].add(existing.id)
                    parsed_night = ShiftAreaImportService._parse_bool(night_flag_raw, default=None)
                    effective_night = parsed_night if parsed_night is not None else end_time <= start_time
                    effective_active = ShiftAreaImportService._parse_bool(active_raw, default=True)
                    changes = {}
                    if existing.area_id != (area_obj.id if area_obj else None):
                        changes["area"] = {
                            "from": existing.area.name if existing.area_id else "",
                            "to": area_name,
                        }
                    if existing.name != shift_name:
                        changes["name"] = {"from": existing.name, "to": shift_name}
                    if existing.buk_code != buk_code:
                        changes["buk_code"] = {"from": existing.buk_code, "to": buk_code}
                    if existing.start_time != start_time:
                        changes["start_time"] = {
                            "from": ExcelImportService._serialize_time(existing.start_time),
                            "to": ExcelImportService._serialize_time(start_time),
                        }
                    if existing.end_time != end_time:
                        changes["end_time"] = {
                            "from": ExcelImportService._serialize_time(existing.end_time),
                            "to": ExcelImportService._serialize_time(end_time),
                        }
                    if existing.break_start != break_start:
                        changes["break_start"] = {
                            "from": ExcelImportService._serialize_time(existing.break_start),
                            "to": ExcelImportService._serialize_time(break_start),
                        }
                    if existing.break_end != break_end:
                        changes["break_end"] = {
                            "from": ExcelImportService._serialize_time(existing.break_end),
                            "to": ExcelImportService._serialize_time(break_end),
                        }
                    if bool(existing.is_night_shift) != bool(effective_night):
                        changes["is_night_shift"] = {"from": bool(existing.is_night_shift), "to": bool(effective_night)}
                    if bool(existing.active) != bool(effective_active):
                        changes["active"] = {"from": bool(existing.active), "to": bool(effective_active)}
                    if changes:
                        action = "update"
                        status = "warning"
                        message = "Cambios detectados; revisa antes de confirmar."
                        summary["updated"] += 1
                        summary["warnings"] += 1
                    else:
                        action = "keep"
                        summary["unchanged"] += 1
                else:
                    action = "create"
                    summary["created"] += 1
            else:
                summary["errors"] += 1

            parsed_night_payload = ShiftAreaImportService._parse_bool(night_flag_raw, default=None)
            ImportPreviewRow.objects.create(
                batch=batch,
                sheet_name="shifts_import",
                row_number=row_index,
                action=action,
                status=status,
                message=message,
                payload={
                    "area_name": area_name,
                    "shift_name": shift_name,
                    "buk_code": buk_code,
                    "property_name": property_name,
                    "property_id": property_obj.id if property_obj else None,
                    "start_time": ExcelImportService._serialize_time(start_time),
                    "end_time": ExcelImportService._serialize_time(end_time),
                    "break_start": ExcelImportService._serialize_time(break_start),
                    "break_end": ExcelImportService._serialize_time(break_end),
                    "is_night_shift": parsed_night_payload,
                    "active": ShiftAreaImportService._parse_bool(active_raw, default=True),
                    "changes": changes if status == "warning" and action == "update" else {},
                },
            )

        if sync_mode:
            matched_shift_ids = matched_shift_ids_by_property.get(fallback_property.id, set())
            active_missing_shifts = Shift.objects.filter(
                tenant=tenant,
                property=fallback_property,
                active=True,
            ).exclude(id__in=matched_shift_ids)
            next_row_number = 100000
            for shift in active_missing_shifts.select_related("area").order_by("area__name", "name", "id"):
                summary["to_deactivate"] += 1
                summary["warnings"] += 1
                ImportPreviewRow.objects.create(
                    batch=batch,
                    sheet_name="shifts_import",
                    row_number=next_row_number,
                    action="deactivate",
                    status="warning",
                    message="No aparece en el archivo completo de la sede; se marcara como inactivo.",
                    payload={
                        "shift_id": shift.id,
                        "area_name": shift.area.name if shift.area_id else "",
                        "shift_name": shift.name,
                        "buk_code": shift.buk_code,
                        "property_name": fallback_property.name,
                        "property_id": fallback_property.id,
                        "start_time": ExcelImportService._serialize_time(shift.start_time),
                        "end_time": ExcelImportService._serialize_time(shift.end_time),
                    },
                )
                next_row_number += 1

        batch.summary = summary
        batch.save(update_fields=["summary", "updated_at"])
        return batch

    @staticmethod
    def confirm_shift_import(*, batch):
        created = 0
        updated = 0
        unchanged = 0
        deactivated = 0
        new_areas = 0
        actor = batch.created_by
        rows = batch.preview_rows.filter(status__in=["ok", "warning"]).order_by("row_number")
        with transaction.atomic():
            for row in rows:
                payload = row.payload
                if row.action == "deactivate":
                    shift_id = payload.get("shift_id")
                    shift = Shift.objects.filter(
                        tenant=batch.tenant,
                        property=batch.property,
                        id=shift_id,
                        active=True,
                    ).first()
                    if shift is None:
                        continue
                    before = {
                        "name": shift.name,
                        "buk_code": shift.buk_code,
                        "active": shift.active,
                    }
                    shift.active = False
                    shift.save(update_fields=["active", "updated_at"])
                    deactivated += 1
                    if actor:
                        AuditService.log(
                            tenant=batch.tenant,
                            property_obj=shift.property,
                            user=actor,
                            action="shift_deactivate_from_import_sync",
                            entity_type="Shift",
                            entity_id=shift.id,
                            before=before,
                            after={"name": shift.name, "buk_code": shift.buk_code, "active": shift.active},
                        )
                    continue
                if row.action == "keep":
                    unchanged += 1
                    continue
                property_id = payload.get("property_id")
                if not property_id:
                    continue

                area, area_created = Area.objects.get_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["area_name"],
                    defaults={"type": "", "active": True},
                )
                if area_created:
                    new_areas += 1

                start_time = ExcelImportService._parse_time(payload.get("start_time"))
                end_time = ExcelImportService._parse_time(payload.get("end_time"))
                break_start = ExcelImportService._parse_time(payload.get("break_start"))
                break_end = ExcelImportService._parse_time(payload.get("break_end"))
                if not start_time or not end_time:
                    continue

                is_night_shift = payload.get("is_night_shift")
                if is_night_shift is None:
                    is_night_shift = end_time <= start_time

                shift = Shift.objects.filter(
                    tenant=batch.tenant,
                    property_id=property_id,
                    buk_code__iexact=payload["buk_code"],
                ).first()
                if shift is None:
                    shift = Shift.objects.filter(
                        tenant=batch.tenant,
                        property_id=property_id,
                        area=area,
                        name__iexact=payload["shift_name"],
                    ).first()

                if shift is None:
                    shift = Shift.objects.create(
                        tenant=batch.tenant,
                        property_id=property_id,
                        area=area,
                        name=payload["shift_name"],
                        buk_code=payload["buk_code"],
                        start_time=start_time,
                        end_time=end_time,
                        break_start=break_start,
                        break_end=break_end,
                        is_night_shift=bool(is_night_shift),
                        active=bool(payload.get("active", True)),
                    )
                    created += 1
                    if actor:
                        AuditService.log(
                            tenant=batch.tenant,
                            property_obj=shift.property,
                            user=actor,
                            action="shift_create_from_import",
                            entity_type="Shift",
                            entity_id=shift.id,
                            before={},
                            after={"name": shift.name, "buk_code": shift.buk_code},
                        )
                    continue

                shift.area = area
                shift.name = payload["shift_name"]
                shift.buk_code = payload["buk_code"]
                shift.start_time = start_time
                shift.end_time = end_time
                shift.break_start = break_start
                shift.break_end = break_end
                shift.is_night_shift = bool(is_night_shift)
                shift.active = bool(payload.get("active", True))
                shift.save()
                updated += 1
                if actor:
                    AuditService.log(
                        tenant=batch.tenant,
                        property_obj=shift.property,
                        user=actor,
                        action="shift_update_from_import",
                        entity_type="Shift",
                        entity_id=shift.id,
                        before={},
                        after={"name": shift.name, "buk_code": shift.buk_code},
                    )

        batch.summary = {
            **batch.summary,
            "applied_created": created,
            "applied_updated": updated,
            "applied_unchanged": unchanged,
            "applied_deactivated": deactivated,
            "applied_new_areas": new_areas,
        }
        batch.status = "confirmed"
        batch.save(update_fields=["summary", "status", "updated_at"])
        return batch


class ExcelImportApplyService:
    @staticmethod
    def _as_date(value):
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value))

    @staticmethod
    def _normalize_text_key(value):
        text = str(value or "").strip().lower()
        normalized = unicodedata.normalize("NFKD", text)
        return "".join(ch for ch in normalized if not unicodedata.combining(ch))

    @staticmethod
    def _resolve_area_for_assignment(*, tenant, property_id, worker, area_name):
        if area_name:
            area = Area.objects.filter(
                tenant=tenant,
                property_id=property_id,
                name__iexact=area_name,
            ).first()
            if area is not None:
                return area

            normalized_target = ExcelImportApplyService._normalize_text_key(area_name)
            for area_obj in Area.objects.filter(tenant=tenant, property_id=property_id):
                if ExcelImportApplyService._normalize_text_key(area_obj.name) == normalized_target:
                    return area_obj

        return worker.area

    @staticmethod
    def _ensure_auto_shift(*, tenant, property_id, area, schedule_text):
        start_time, end_time = ExcelImportService._parse_time_range(schedule_text)
        if not start_time or not end_time:
            return None, False

        existing = Shift.objects.filter(
            tenant=tenant,
            property_id=property_id,
            area=area,
            start_time=start_time,
            end_time=end_time,
        ).first()
        if existing is not None:
            return existing, False

        name = f"AUTO_{start_time.strftime('%H%M')}-{end_time.strftime('%H%M')}"
        base_code = f"AUTO-{area.id}-{start_time.strftime('%H%M')}-{end_time.strftime('%H%M')}"
        buk_code = base_code
        suffix = 1
        while Shift.objects.filter(
            tenant=tenant,
            property_id=property_id,
            buk_code=buk_code,
        ).exists():
            suffix += 1
            buk_code = f"{base_code}-{suffix}"

        shift, created = Shift.objects.get_or_create(
            tenant=tenant,
            property_id=property_id,
            area=area,
            name=name,
            defaults={
                "buk_code": buk_code,
                "start_time": start_time,
                "end_time": end_time,
                "break_start": None,
                "break_end": None,
                "is_night_shift": end_time <= start_time,
                "active": True,
            },
        )
        return shift, created

    @staticmethod
    def apply_preview_batch(*, batch):
        created = {"areas": 0, "workers": 0, "shifts": 0, "special_states": 0, "assignments": 0}
        updated = {"workers": 0, "shifts": 0, "special_states": 0, "assignments": 0}
        warnings = 0

        rows = list(
            batch.preview_rows.filter(status__in=["ok", "warning"])
            .exclude(action="inspect")
        )
        stage_order = {"area": 1, "special_state": 2, "worker": 3, "shift": 4, "assignment": 5}
        rows.sort(
            key=lambda row: (
                stage_order.get((row.payload or {}).get("entity"), 99),
                row.sheet_name,
                row.row_number,
                row.id,
            )
        )
        for row in rows:
            payload = row.payload
            entity = payload.get("entity")
            property_id = payload.get("property_id") or batch.property_id

            if entity == "area":
                _, was_created = Area.objects.get_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["name"],
                    defaults={"type": "", "active": True},
                )
                if was_created:
                    created["areas"] += 1

            elif entity == "special_state":
                _, was_created = SpecialState.objects.update_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["name"],
                    defaults={"buk_code": payload.get("buk_code", "D"), "active": True},
                )
                if was_created:
                    created["special_states"] += 1
                else:
                    updated["special_states"] += 1

            elif entity == "worker":
                area, _ = Area.objects.get_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["area_name"],
                    defaults={"type": "", "active": True},
                )
                _, was_created = Worker.objects.update_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    document_number=payload["document_number"],
                    defaults={
                        "first_name": payload["first_name"],
                        "last_name": payload["last_name"],
                        "area": area,
                        "active": True,
                    },
                )
                if was_created:
                    created["workers"] += 1
                else:
                    updated["workers"] += 1

            elif entity == "shift":
                area, _ = Area.objects.get_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    name=payload["area_name"],
                    defaults={"type": "", "active": True},
                )
                defaults = {
                    "buk_code": payload.get("buk_code", ""),
                    "start_time": ExcelImportService._parse_time(payload.get("start_time")) or time(0, 0),
                    "end_time": ExcelImportService._parse_time(payload.get("end_time")) or time(0, 0),
                    "break_start": ExcelImportService._parse_time(payload.get("break_start")),
                    "break_end": ExcelImportService._parse_time(payload.get("break_end")),
                    "is_night_shift": False,
                    "active": True,
                }
                _, was_created = Shift.objects.update_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    area=area,
                    name=payload["name"],
                    defaults=defaults,
                )
                if was_created:
                    created["shifts"] += 1
                else:
                    updated["shifts"] += 1

            elif entity == "assignment":
                worker = Worker.objects.filter(
                    tenant=batch.tenant,
                    property_id=property_id,
                    document_number=payload.get("document_number"),
                ).first()
                if not worker:
                    warnings += 1
                    continue

                shift = None
                special_state = None
                if payload.get("special_state_name"):
                    special_state = SpecialState.objects.filter(
                        tenant=batch.tenant,
                        property_id=property_id,
                        name__iexact=payload["special_state_name"],
                    ).first()
                    if special_state is None:
                        special_state = SpecialState.objects.create(
                            tenant=batch.tenant,
                            property_id=property_id,
                            name=payload["special_state_name"],
                            buk_code="D",
                            active=True,
                        )
                else:
                    area_name = payload.get("area_name", worker.area.name)
                    area = ExcelImportApplyService._resolve_area_for_assignment(
                        tenant=batch.tenant,
                        property_id=property_id,
                        worker=worker,
                        area_name=area_name,
                    )
                    shift_name = payload.get("shift_name")
                    if shift_name:
                        shift = Shift.objects.filter(
                            tenant=batch.tenant,
                            property_id=property_id,
                            area=area,
                            name__iexact=shift_name,
                        ).first()
                    if not shift:
                        shift, auto_created = ExcelImportApplyService._ensure_auto_shift(
                            tenant=batch.tenant,
                            property_id=property_id,
                            area=area,
                            schedule_text=payload.get("schedule_text"),
                        )
                        if auto_created:
                            created["shifts"] += 1
                    if not shift:
                        warnings += 1
                        continue

                _, was_created = ScheduleAssignment.objects.update_or_create(
                    tenant=batch.tenant,
                    property_id=property_id,
                    worker=worker,
                    date=ExcelImportApplyService._as_date(payload["date"]),
                    defaults={"shift": shift, "special_state": special_state, "updated_by": batch.created_by},
                )
                if was_created:
                    created["assignments"] += 1
                else:
                    updated["assignments"] += 1

        batch.status = "confirmed"
        batch.summary = {
            **batch.summary,
            "applied_created": created,
            "applied_updated": updated,
            "applied_warnings": warnings,
        }
        batch.save(update_fields=["status", "summary", "updated_at"])
        return batch
