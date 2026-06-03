from django.test import TestCase
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook

from apps.imports.services import ExcelImportApplyService, ExcelImportService, ShiftAreaImportService, WorkerImportService
from apps.scheduling.models import ScheduleAssignment
from apps.tenants.models import Property, Tenant
from apps.users.models import User
from apps.workers.models import Area, Shift, SpecialState, Worker


class WorkerImportServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="ops@pariwana.test", password="StrongPass123")
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepción")
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="11111111",
            first_name="Nombre",
            last_name="Original",
            area=self.area,
        )

    def test_preview_and_confirm_csv_worker_import(self):
        csv_content = (
            "DNI,Nombre,Apellido,Area,Sede\n"
            "11111111,Nombre,Nuevo,Recepción,Pariwana Cusco\n"
            "22222222,Ana,Perez,Housekeeping,Pariwana Cusco\n"
        ).encode("utf-8")
        batch = WorkerImportService.create_worker_preview(
            tenant=self.tenant,
            fallback_property=self.property,
            file_name="workers.csv",
            file_bytes=csv_content,
            user=self.user,
            create_missing_areas=True,
        )
        self.assertEqual(batch.summary["detected_rows"], 2)
        self.assertEqual(batch.summary["errors"], 0)

        WorkerImportService.confirm_worker_import(batch=batch)
        self.assertTrue(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="22222222",
            ).exists()
        )

    def test_sync_worker_import_warns_changes_and_deactivates_missing(self):
        missing = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="33333333",
            first_name="Fuera",
            last_name="Archivo",
            area=self.area,
            active=True,
        )
        csv_content = (
            "DNI,Nombre,Apellido,Area,Sede\n"
            f"11111111,Nombre,Nuevo,{self.area.name},Pariwana Cusco\n"
        ).encode("utf-8")

        batch = WorkerImportService.create_worker_preview(
            tenant=self.tenant,
            fallback_property=self.property,
            file_name="workers.csv",
            file_bytes=csv_content,
            user=self.user,
            create_missing_areas=True,
            sync_mode=True,
        )
        self.assertEqual(batch.summary["updated"], 1)
        self.assertEqual(batch.summary["to_deactivate"], 1)
        self.assertEqual(batch.preview_rows.filter(action="deactivate", status="warning").count(), 1)

        WorkerImportService.confirm_worker_import(batch=batch)
        missing.refresh_from_db()
        self.assertFalse(missing.active)
        updated = Worker.objects.get(tenant=self.tenant, property=self.property, document_number="11111111")
        self.assertEqual(updated.last_name, "Nuevo")


class ExcelImportServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="ops2@pariwana.test", password="StrongPass123")
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")

    def _build_excel_bytes(self):
        wb = Workbook()
        ws_datos = wb.active
        ws_datos.title = "Datos"
        ws_datos.cell(6, 3).value = "Recepcion"
        ws_datos.cell(6, 8).value = "OFF"

        ws_workers = wb.create_sheet("Reg. de trabajadores x area")
        ws_workers.cell(4, 1).value = "12345678"
        ws_workers.cell(4, 2).value = "Ana"
        ws_workers.cell(4, 3).value = "Quispe"
        ws_workers.cell(4, 4).value = "Recepcion"

        ws_shifts = wb.create_sheet("Reg. Horarios x area")
        ws_shifts.cell(4, 1).value = "REC-M"
        ws_shifts.cell(4, 2).value = "REC-M"
        ws_shifts.cell(4, 3).value = "Recepcion"
        ws_shifts.cell(4, 4).value = "06:00 - 14:45"

        ws_month = wb.create_sheet("Abril 2026")
        ws_month.cell(5, 4).value = datetime(2026, 4, 1)
        ws_month.cell(5, 4).number_format = "yyyy-mm-dd"
        ws_month.cell(6, 1).value = "Recepcion"
        ws_month.cell(7, 1).value = "12345678"
        ws_month.cell(7, 2).value = "Ana Quispe"
        ws_month.cell(7, 3).value = "TURNO"
        ws_month.cell(7, 4).value = "06:00 - 14:45"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def test_excel_preview_and_apply(self):
        file_bytes = self._build_excel_bytes()
        batch = ExcelImportService.create_preview(
            tenant=self.tenant,
            property_obj=self.property,
            file_name="source.xlsx",
            file_bytes=file_bytes,
            user=self.user,
        )
        self.assertGreater(batch.preview_rows.count(), 0)

        batch = ExcelImportApplyService.apply_preview_batch(batch=batch)
        self.assertEqual(batch.status, "confirmed")
        self.assertTrue(Area.objects.filter(tenant=self.tenant, property=self.property, name="Recepcion").exists())
        self.assertTrue(Worker.objects.filter(tenant=self.tenant, property=self.property, document_number="12345678").exists())
        self.assertTrue(Shift.objects.filter(tenant=self.tenant, property=self.property, name="REC-M").exists())
        self.assertTrue(SpecialState.objects.filter(tenant=self.tenant, property=self.property, name="OFF").exists())
        self.assertTrue(ScheduleAssignment.objects.filter(tenant=self.tenant, property=self.property).exists())


    def test_excel_preview_maps_schedule_without_spaces(self):
        file_bytes = self._build_excel_bytes()
        wb = Workbook()
        ws_datos = wb.active
        ws_datos.title = "Datos"
        ws_datos.cell(6, 3).value = "Recepcion"
        ws_workers = wb.create_sheet("Reg. de trabajadores x area")
        ws_workers.cell(4, 1).value = "12345678"
        ws_workers.cell(4, 2).value = "Ana"
        ws_workers.cell(4, 3).value = "Quispe"
        ws_workers.cell(4, 4).value = "Recepcion"
        ws_shifts = wb.create_sheet("Reg. Horarios x area")
        ws_shifts.cell(4, 1).value = "REC-M"
        ws_shifts.cell(4, 2).value = "REC-M"
        ws_shifts.cell(4, 3).value = "Recepcion"
        ws_shifts.cell(4, 4).value = "06:00 - 14:45"
        ws_month = wb.create_sheet("Abril 2026")
        ws_month.cell(5, 4).value = datetime(2026, 4, 1)
        ws_month.cell(6, 1).value = "Recepcion"
        ws_month.cell(7, 1).value = "12345678"
        ws_month.cell(7, 2).value = "Ana Quispe"
        ws_month.cell(7, 3).value = "TURNO"
        ws_month.cell(7, 4).value = "06:00-14:45"
        buffer = BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()

        batch = ExcelImportService.create_preview(
            tenant=self.tenant,
            property_obj=self.property,
            file_name="source.xlsx",
            file_bytes=file_bytes,
            user=self.user,
        )
        warning_rows = batch.preview_rows.filter(status="warning", payload__entity="assignment").count()
        self.assertEqual(warning_rows, 0)

    def test_excel_preview_maps_single_shift_area_fallback(self):
        wb = Workbook()
        ws_datos = wb.active
        ws_datos.title = "Datos"
        ws_datos.cell(6, 3).value = "Recursos Humanos"
        ws_workers = wb.create_sheet("Reg. de trabajadores x area")
        ws_workers.cell(4, 1).value = "87654321"
        ws_workers.cell(4, 2).value = "Luz"
        ws_workers.cell(4, 3).value = "Ramos"
        ws_workers.cell(4, 4).value = "Recursos Humanos"
        ws_shifts = wb.create_sheet("Reg. Horarios x area")
        ws_shifts.cell(4, 1).value = "ASISTENTE ADMINISTRATIVO"
        ws_shifts.cell(4, 2).value = "ASIST-ADMIN"
        ws_shifts.cell(4, 3).value = "Recursos Humanos"
        ws_shifts.cell(4, 4).value = "09:00 - 18:00"
        ws_month = wb.create_sheet("Abril 2026")
        ws_month.cell(5, 4).value = datetime(2026, 4, 1)
        ws_month.cell(6, 1).value = "Recursos Humanos"
        ws_month.cell(7, 1).value = "87654321"
        ws_month.cell(7, 2).value = "Luz Ramos"
        ws_month.cell(7, 3).value = "TURNO"
        ws_month.cell(7, 4).value = "13:30-17:30"
        buffer = BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()

        batch = ExcelImportService.create_preview(
            tenant=self.tenant,
            property_obj=self.property,
            file_name="source.xlsx",
            file_bytes=file_bytes,
            user=self.user,
        )
        batch = ExcelImportApplyService.apply_preview_batch(batch=batch)
        self.assertEqual(batch.summary["applied_warnings"], 0)
        self.assertTrue(ScheduleAssignment.objects.filter(tenant=self.tenant, property=self.property).exists())

    def test_excel_apply_creates_auto_shift_for_unknown_schedule(self):
        wb = Workbook()
        ws_datos = wb.active
        ws_datos.title = "Datos"
        ws_datos.cell(6, 3).value = "Bar"
        ws_workers = wb.create_sheet("Reg. de trabajadores x area")
        ws_workers.cell(4, 1).value = "55554444"
        ws_workers.cell(4, 2).value = "Ana"
        ws_workers.cell(4, 3).value = "Lopez"
        ws_workers.cell(4, 4).value = "Bar"
        ws_shifts = wb.create_sheet("Reg. Horarios x area")
        ws_shifts.cell(4, 1).value = "Bar_Regular"
        ws_shifts.cell(4, 2).value = "BAR-REG"
        ws_shifts.cell(4, 3).value = "Bar"
        ws_shifts.cell(4, 4).value = "14:00 - 23:00"
        ws_shifts.cell(5, 1).value = "Bar_Noche"
        ws_shifts.cell(5, 2).value = "BAR-NCH"
        ws_shifts.cell(5, 3).value = "Bar"
        ws_shifts.cell(5, 4).value = "16:00 - 01:00"
        ws_month = wb.create_sheet("Abril 2026")
        ws_month.cell(5, 4).value = datetime(2026, 4, 1)
        ws_month.cell(6, 1).value = "Bar"
        ws_month.cell(7, 1).value = "55554444"
        ws_month.cell(7, 2).value = "Ana Lopez"
        ws_month.cell(7, 3).value = "TURNO"
        ws_month.cell(7, 4).value = "15:15 - 00:00"
        buffer = BytesIO()
        wb.save(buffer)

        batch = ExcelImportService.create_preview(
            tenant=self.tenant,
            property_obj=self.property,
            file_name="source.xlsx",
            file_bytes=buffer.getvalue(),
            user=self.user,
        )
        self.assertGreater(batch.summary["warnings"], 0)
        batch = ExcelImportApplyService.apply_preview_batch(batch=batch)
        self.assertEqual(batch.summary["applied_warnings"], 0)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                area__name__iexact="Bar",
                name="AUTO_1515-0000",
            ).exists()
        )
        self.assertTrue(ScheduleAssignment.objects.filter(tenant=self.tenant, property=self.property).exists())


class ShiftAreaImportServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="shift-import@pariwana.test", password="StrongPass123")
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="REC-M",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
            active=True,
        )

    def test_preview_and_confirm_shifts_csv(self):
        csv_content = (
            "Area,Turno,Codigo BUK,Hora Inicio,Hora Fin,Inicio Break,Fin Break,Nocturno,Activo,Sede\n"
            "Recepcion,REC-M,REC-M,06:00,14:45,10:00,10:30,0,1,Pariwana Cusco\n"
            "Housekeeping,HK-N,HK-N,22:00,06:00,,,1,1,Pariwana Cusco\n"
        ).encode("utf-8")

        batch = ShiftAreaImportService.create_shift_preview(
            tenant=self.tenant,
            fallback_property=self.property,
            file_name="turnos.csv",
            file_bytes=csv_content,
            user=self.user,
            create_missing_areas=True,
        )
        self.assertEqual(batch.summary["detected_rows"], 2)
        self.assertEqual(batch.summary["errors"], 0)

        ShiftAreaImportService.confirm_shift_import(batch=batch)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="HK-N",
            ).exists()
        )

    def test_sync_shift_import_warns_changes_and_deactivates_missing(self):
        missing = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="REC-T",
            buk_code="REC-T",
            start_time="14:45",
            end_time="23:00",
            active=True,
        )
        csv_content = (
            "Area,Turno,Codigo BUK,Hora Inicio,Hora Fin,Inicio Break,Fin Break,Nocturno,Activo,Sede\n"
            "Recepcion,REC-M,REC-M,07:00,15:00,10:00,10:30,0,1,Pariwana Cusco\n"
        ).encode("utf-8")

        batch = ShiftAreaImportService.create_shift_preview(
            tenant=self.tenant,
            fallback_property=self.property,
            file_name="turnos.csv",
            file_bytes=csv_content,
            user=self.user,
            create_missing_areas=True,
            sync_mode=True,
        )
        self.assertEqual(batch.summary["updated"], 1)
        self.assertEqual(batch.summary["to_deactivate"], 1)
        self.assertEqual(batch.preview_rows.filter(action="deactivate", status="warning").count(), 1)

        ShiftAreaImportService.confirm_shift_import(batch=batch)
        missing.refresh_from_db()
        self.assertFalse(missing.active)
        updated = Shift.objects.get(tenant=self.tenant, property=self.property, buk_code="REC-M")
        self.assertEqual(updated.start_time.strftime("%H:%M"), "07:00")

    def test_preview_and_confirm_shifts_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Turnos"
        ws.cell(1, 1).value = "Area"
        ws.cell(1, 2).value = "Turno"
        ws.cell(1, 3).value = "Codigo BUK"
        ws.cell(1, 4).value = "Hora Inicio"
        ws.cell(1, 5).value = "Hora Fin"
        ws.cell(1, 6).value = "Nocturno"
        ws.cell(1, 7).value = "Activo"
        ws.cell(1, 8).value = "Sede"
        ws.cell(2, 1).value = "Recepcion"
        ws.cell(2, 2).value = "REC-T"
        ws.cell(2, 3).value = "REC-T"
        ws.cell(2, 4).value = "14:45"
        ws.cell(2, 5).value = "23:00"
        ws.cell(2, 6).value = 0
        ws.cell(2, 7).value = 1
        ws.cell(2, 8).value = "Pariwana Cusco"
        ws.cell(3, 1).value = "Housekeeping"
        ws.cell(3, 2).value = "HK-N"
        ws.cell(3, 3).value = "HK-N"
        ws.cell(3, 4).value = "22:00"
        ws.cell(3, 5).value = "06:00"
        ws.cell(3, 6).value = 1
        ws.cell(3, 7).value = 1
        ws.cell(3, 8).value = "Pariwana Cusco"
        buffer = BytesIO()
        wb.save(buffer)

        batch = ShiftAreaImportService.create_shift_preview(
            tenant=self.tenant,
            fallback_property=self.property,
            file_name="turnos.xlsx",
            file_bytes=buffer.getvalue(),
            user=self.user,
            create_missing_areas=True,
        )
        self.assertEqual(batch.summary["detected_rows"], 2)
        self.assertEqual(batch.summary["errors"], 0)

        ShiftAreaImportService.confirm_shift_import(batch=batch)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="HK-N",
            ).exists()
        )
