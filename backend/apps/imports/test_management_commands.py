from pathlib import Path
import tempfile
from datetime import time

from django.core.management import call_command
from django.test import TestCase
from openpyxl import load_workbook

from apps.tenants.models import Property, Tenant
from apps.workers.models import Area, Shift, Worker


class GeneratePhase4ImportSamplesCommandTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="12345678",
            first_name="Ana",
            last_name="Quispe",
            area=self.area,
            active=True,
        )
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 0),
            active=True,
        )

    def test_command_generates_csv_and_xlsx_samples(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_dir = Path(tmpdir)
            call_command(
                "generate_phase4_import_samples",
                tenant_slug="pariwana-hostels",
                property_slug="pariwana-cusco",
                output_dir=str(output_dir),
            )

            workers_csv = output_dir / "workers_import_sample.csv"
            shifts_csv = output_dir / "shifts_area_import_sample.csv"
            workers_xlsx = output_dir / "workers_import_sample.xlsx"
            shifts_xlsx = output_dir / "shifts_area_import_sample.xlsx"

            self.assertTrue(workers_csv.exists())
            self.assertTrue(shifts_csv.exists())
            self.assertTrue(workers_xlsx.exists())
            self.assertTrue(shifts_xlsx.exists())

            workers_csv_text = workers_csv.read_text(encoding="utf-8-sig")
            self.assertIn("DNI,Nombre,Apellido,Area,Sede", workers_csv_text)
            self.assertIn("12345678,Ana,Quispe,Recepcion,Pariwana Cusco", workers_csv_text)

            wb = load_workbook(workers_xlsx, data_only=True)
            ws = wb["Trabajadores"]
            self.assertEqual(ws.cell(1, 1).value, "DNI")
            self.assertEqual(ws.cell(2, 1).value, "12345678")

