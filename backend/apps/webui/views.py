import calendar
import ast
import json
import re
from collections import defaultdict
from datetime import date, timedelta, datetime, time
from io import BytesIO, StringIO
from pathlib import Path
import csv
from types import SimpleNamespace
from uuid import uuid4
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.views import redirect_to_login
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Count, Max, Min, Q
from django.db.models.deletion import ProtectedError
from django.db.utils import OperationalError, ProgrammingError
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.audit.models import AuditLog
from apps.audit.services import AuditService
from apps.buk_exports.models import BukExportLog, BukTemplateCompareLog
from apps.buk_exports.services import BukExportService, BukValidationService
from apps.imports.models import ImportBatch
from apps.imports.backup_services import BackupRestoreService
from apps.imports.services import (
    ExcelImportApplyService,
    ExcelImportService,
    ImportSampleService,
    ShiftAreaImportService,
    WorkerImportService,
)
from apps.month_closure.models import MonthClosure
from apps.modules.models import ModuleActivation
from apps.modules.services import ModuleActivationService
from apps.tenants.models import Property, Tenant, TenantStatus, TenantSupportAccessSession
from apps.tenants.services import TenantSupportService
from apps.users.models import RoleChoices, RoleProfile, User, UserAreaPermission, UserPropertyPermission, UserTenantRole
from apps.users.services import PROPERTY_PERMISSION_KEYS, PermissionService, RoleProfileService
from apps.workers.models import Area, Shift, Worker
from apps.webui.forms import AreaForm, PropertyForm, ShiftForm, SpecialStateForm, TenantForm, WorkerForm
from apps.month_closure.services import MonthClosureService
from apps.scheduling.models import ScheduleAssignment, SchedulePatternTemplate, ScheduleRangeTemplate
from apps.scheduling.services import ScheduleAssignmentService
from apps.workers.models import SpecialState


WEEK_PATTERN_KEYS = [
    ("monday", 0),
    ("tuesday", 1),
    ("wednesday", 2),
    ("thursday", 3),
    ("friday", 4),
    ("saturday", 5),
    ("sunday", 6),
]

AUTO_SHIFT_MERGE_CONFIRM_THRESHOLD = 20
WEEKDAY_SHORT_LABELS = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]

WEBUI_MODULE_CATALOG = [
    ("tenants", "Tenants", "Gestion de tenants para crecimiento multi-empresa."),
    ("properties", "Sedes", "Creacion y edicion de sedes por tenant."),
    ("users_permissions", "Roles y permisos", "Roles, permisos por sede y areas autorizadas."),
    ("module_activation", "Modulos activables", "Activacion o pausa de modulos por tenant."),
    ("workers", "Trabajadores", "Gestion de trabajadores activos e inactivos."),
    ("areas", "Areas", "Gestion de areas operativas por sede."),
    ("shifts", "Turnos", "Gestion de turnos y codigos BUK."),
    ("special_states", "Estados especiales", "Estados como descanso, vacaciones o licencia."),
    ("scheduling", "Asignacion de horarios", "Grilla mensual y acciones masivas de horarios."),
    ("control", "Control proximos 15 dias", "Alertas de trabajadores sin turno o estado especial."),
    ("buk_validator", "Validador BUK", "Validaciones bloqueantes y advertencias antes de exportar."),
    ("buk_preview", "Vista previa BUK", "Previsualizacion del reporte antes de descarga."),
    ("buk_export", "Exportador BUK", "Generacion de XLSX/CSV para carga en BUK."),
    ("excel_import", "Importacion / exportacion", "Importaciones desde Excel original y plantillas XLSX/CSV."),
    ("backup", "Backup JSON", "Respaldo/restauracion operativa JSON."),
    ("audit", "Auditoria", "Registro de acciones criticas."),
    ("month_closure", "Cierre de mes", "Cierre y reapertura de periodos por sede."),
]


def _normalize_import_preview_value(value):
    if isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if len(stripped) < 2:
        return value

    likely_structured = (
        (stripped.startswith("{") and stripped.endswith("}"))
        or (stripped.startswith("[") and stripped.endswith("]"))
    )
    if not likely_structured:
        return value

    try:
        parsed = json.loads(stripped)
        if isinstance(parsed, (dict, list)):
            return parsed
    except json.JSONDecodeError:
        pass

    try:
        parsed = ast.literal_eval(stripped)
        if isinstance(parsed, (dict, list)):
            return parsed
    except (ValueError, SyntaxError):
        return value

    return value


def _format_import_preview_value(value):
    normalized = _normalize_import_preview_value(value)
    if isinstance(normalized, (dict, list)):
        return json.dumps(normalized, ensure_ascii=False, indent=2)
    if normalized is None:
        return "-"
    return str(normalized)


def _import_batch_has_errors(batch):
    summary = batch.summary if isinstance(batch.summary, dict) else {}
    try:
        if int(summary.get("errors", 0) or 0) > 0:
            return True
    except (TypeError, ValueError):
        pass
    return batch.preview_rows.filter(status="error").exists()


def _build_import_batch_preview_context(batch, *, row_limit=50):
    if not batch:
        return {
            "selected_batch_summary_rows": [],
            "preview_rows": [],
            "preview_has_more_rows": False,
            "preview_total_rows": 0,
            "preview_blocking_errors": False,
        }

    summary_rows = []
    summary_data = batch.summary if isinstance(batch.summary, dict) else {}
    for raw_key, raw_value in summary_data.items():
        key_label = str(raw_key).replace("_", " ").capitalize()
        normalized_value = _normalize_import_preview_value(raw_value)
        if isinstance(normalized_value, dict):
            summary_rows.append(
                {
                    "key": key_label,
                    "is_block": False,
                    "items": [
                        {
                            "subkey": str(subkey).replace("_", " "),
                            "subvalue": _format_import_preview_value(subvalue),
                        }
                        for subkey, subvalue in normalized_value.items()
                    ],
                    "value": "",
                }
            )
        elif isinstance(normalized_value, list):
            summary_rows.append(
                {
                    "key": key_label,
                    "is_block": True,
                    "items": [],
                    "value": _format_import_preview_value(normalized_value),
                }
            )
        else:
            summary_rows.append(
                {
                    "key": key_label,
                    "is_block": False,
                    "items": [],
                    "value": _format_import_preview_value(normalized_value),
                }
            )

    total_rows = batch.preview_rows.count()
    preview_rows = [
        {
            "sheet_name": item.sheet_name,
            "row_number": item.row_number,
            "status": item.status,
            "action": item.action,
            "message": item.message,
            "payload_pretty": _format_import_preview_value(item.payload),
        }
        for item in batch.preview_rows.order_by("sheet_name", "row_number")[:row_limit]
    ]
    issue_rows = [
        {
            "row_number": item.row_number,
            "status": item.status,
            "action": item.action,
            "message": item.message or "Revisa el detalle de la fila.",
        }
        for item in batch.preview_rows.filter(status__in=["error", "warning"]).order_by("status", "row_number")[:8]
    ]
    return {
        "selected_batch_summary_rows": summary_rows,
        "preview_rows": preview_rows,
        "preview_issue_rows": issue_rows,
        "preview_has_more_rows": total_rows > row_limit,
        "preview_total_rows": total_rows,
        "preview_blocking_errors": _import_batch_has_errors(batch),
    }


def _worker_active_on_date(worker, target_date):
    if not worker.active:
        return False
    if worker.start_date and target_date < worker.start_date:
        return False
    if worker.end_date and target_date > worker.end_date:
        return False
    return True


def _audit_snapshot(model_obj, fields):
    snapshot = {}
    for field_name in fields:
        value = getattr(model_obj, field_name)
        if isinstance(value, (date, datetime, time)):
            snapshot[field_name] = value.isoformat()
        else:
            snapshot[field_name] = value
    return snapshot


def _property_permission_payload_from_request(request, prefix=""):
    return {key: bool(request.POST.get(f"{prefix}{key}")) for key in PROPERTY_PERMISSION_KEYS}


def _property_permission_object_from_payload(payload):
    normalized = RoleProfileService.normalize_permissions(payload)
    return SimpleNamespace(**normalized)


def _apply_role_profile_defaults(request, role_profile):
    permissions = _property_permission_payload_from_request(request)
    if bool(request.POST.get("apply_role_profile_defaults")):
        permissions.update(RoleProfileService.permission_defaults_for_profile(role_profile))
    return permissions


def _selected_properties_from_request(request, tenant, current_property, role):
    all_properties_access = bool(request.POST.get("all_properties_access"))
    tenant_properties = Property.objects.filter(tenant=tenant).order_by("name")
    if role not in {RoleChoices.ADMIN, RoleChoices.OPERATOR}:
        return False, [current_property]
    if all_properties_access:
        return True, list(tenant_properties)
    selected_ids = [int(item) for item in request.POST.getlist("property_ids") if str(item).isdigit()]
    properties = list(tenant_properties.filter(id__in=selected_ids))
    if not properties:
        properties = [current_property]
    return False, properties


def _sync_user_property_permissions(*, user, tenant, properties, permission_payload, all_properties_access):
    selected_ids = [item.id for item in properties]
    for item in properties:
        UserPropertyPermission.objects.update_or_create(
            user=user,
            tenant=tenant,
            property=item,
            defaults=permission_payload,
        )
    if not all_properties_access:
        UserPropertyPermission.objects.filter(user=user, tenant=tenant).exclude(property_id__in=selected_ids).delete()
        UserAreaPermission.objects.filter(user=user, tenant=tenant).exclude(property_id__in=selected_ids).delete()


def _get_exclusive_tenant_user(*, requester, tenant, user_id):
    tenant_role = (
        UserTenantRole.objects.select_related("user")
        .filter(tenant=tenant, user_id=user_id)
        .first()
    )
    if tenant_role is None:
        return None, "Usuario no encontrado en este tenant."
    if (
        not PermissionService.is_super_admin(requester)
        and UserTenantRole.objects.filter(user=tenant_role.user).exclude(tenant=tenant).exists()
    ):
        return None, "No se puede modificar globalmente una cuenta compartida entre tenants."
    return tenant_role.user, None


def _get_role_profile_from_request(request, tenant):
    role_profile_id = str(request.POST.get("role_profile_id", "")).strip()
    if not role_profile_id.isdigit():
        return None
    return RoleProfile.objects.filter(tenant=tenant, id=int(role_profile_id), active=True).first()


def _get_support_session(user, support_session_id):
    if not user.is_super_admin or not support_session_id:
        return None
    return (
        TenantSupportAccessSession.objects.select_related("tenant", "property")
        .filter(
            id=support_session_id,
            started_by=user,
            ended_at__isnull=True,
        )
        .first()
    )


def _get_tenant_options(user, support_session):
    if support_session is not None:
        return Tenant.objects.filter(id=support_session.tenant_id)
    if user.is_super_admin:
        return Tenant.objects.all().order_by("name")
    tenant_ids = user.tenant_roles.values_list("tenant_id", flat=True)
    return Tenant.objects.filter(id__in=tenant_ids).order_by("name")


def _get_property_options(user, tenant, support_session):
    if tenant is None:
        return Property.objects.none()
    if support_session is not None and support_session.property_id:
        return Property.objects.filter(id=support_session.property_id, tenant=tenant)
    if user.is_super_admin:
        return Property.objects.filter(tenant=tenant).order_by("name")
    allowed = PermissionService.get_accessible_property_ids(user, tenant, action="can_access")
    return Property.objects.filter(tenant=tenant, id__in=allowed).order_by("name")


def _can_nav_module(user, tenant, property_obj, module_key, action=None, roles=None):
    if tenant is None:
        return False
    if not PermissionService.user_can_module(user, tenant, module_key):
        return False
    if roles and not PermissionService.user_can_tenant_role(user, tenant, roles):
        return False
    if action and property_obj is not None:
        return PermissionService.user_can_property_action(user, tenant, property_obj, action)
    return action is None


def _build_nav_items(user, tenant, property_obj, current_path="/app/"):
    def nav_item(label, url):
        is_active = current_path == url or (url != "/app/" and current_path.startswith(url))
        return {"label": label, "url": url, "active": is_active}

    items = [nav_item("Dashboard", "/app/")]
    other_items = []
    if PermissionService.is_super_admin(user):
        other_items.append(nav_item("Tenants", "/app/tenants/"))
        other_items.append(nav_item("Modulos", "/app/modules/"))
        other_items.append(nav_item("Soporte", "/app/support/"))
        other_items.append(nav_item("Auditoria global", "/app/audit-global/"))
    if tenant is None:
        if other_items:
            items.append({"label": "Otras funciones", "children": other_items, "active": any(i["active"] for i in other_items)})
        return items

    if _can_nav_module(user, tenant, property_obj, "scheduling", "can_schedule"):
        items.append(nav_item("Asignacion", "/app/scheduling/"))
    if _can_nav_module(user, tenant, property_obj, "properties", roles=["admin"]):
        other_items.append(nav_item("Sedes", "/app/properties/"))
    if _can_nav_module(user, tenant, property_obj, "workers", "can_manage_workers"):
        items.append(nav_item("Trabajadores", "/app/workers/"))
    if _can_nav_module(user, tenant, property_obj, "areas", "can_manage_areas"):
        items.append(nav_item("Areas", "/app/areas/"))
    if _can_nav_module(user, tenant, property_obj, "shifts", "can_manage_shifts"):
        items.append(nav_item("Turnos", "/app/shifts/"))
    if _can_nav_module(user, tenant, property_obj, "special_states", roles=["admin"]):
        items.append(nav_item("Estados especiales", "/app/special-states/"))
    if _can_nav_module(user, tenant, property_obj, "users_permissions", "can_manage_users"):
        items.append(nav_item("Roles y permisos", "/app/users-permissions/"))
    if _can_nav_module(user, tenant, property_obj, "excel_import"):
        can_import_workers = property_obj is not None and PermissionService.user_can_property_action(
            user, tenant, property_obj, "can_manage_workers"
        )
        can_import_shifts = property_obj is not None and PermissionService.user_can_property_action(
            user, tenant, property_obj, "can_manage_shifts"
        )
        if can_import_workers or can_import_shifts:
            other_items.append(nav_item("Importaciones", "/app/imports/"))
    if _can_nav_module(user, tenant, property_obj, "backup", roles=["admin"]):
        other_items.append(nav_item("Backup JSON", "/app/backup/"))
    if _can_nav_module(user, tenant, property_obj, "control", "can_use_control"):
        items.append(nav_item("Control 15 dias", "/app/control/"))
    if _can_nav_module(user, tenant, property_obj, "month_closure", roles=["admin"]):
        other_items.append(nav_item("Cierre de mes", "/app/month-closure/"))
    if _can_nav_module(user, tenant, property_obj, "audit", roles=["admin"]):
        other_items.append(nav_item("Auditoria", "/app/audit/"))
    if (
        _can_nav_module(user, tenant, property_obj, "buk_preview", "can_view_reports")
        or _can_nav_module(user, tenant, property_obj, "buk_preview", "can_export_buk")
    ):
        items.append(nav_item("Reporte BUK", "/app/buk-report/"))
    if other_items:
        items.append({"label": "Otras funciones", "children": other_items, "active": any(i["active"] for i in other_items)})
    return items


def _build_context(request, require_property=False):
    support_session_id = request.session.get("support_session_id")
    support_session = _get_support_session(request.user, support_session_id)
    if support_session is None and "support_session_id" in request.session:
        del request.session["support_session_id"]

    tenant_options = list(_get_tenant_options(request.user, support_session))
    selected_tenant = None
    if support_session is not None:
        selected_tenant = support_session.tenant
    else:
        tenant_id = request.session.get("ui_tenant_id")
        if tenant_id:
            selected_tenant = next((t for t in tenant_options if t.id == tenant_id), None)
        if selected_tenant is None and tenant_options:
            selected_tenant = tenant_options[0]
            request.session["ui_tenant_id"] = selected_tenant.id

    property_options = list(_get_property_options(request.user, selected_tenant, support_session))
    selected_property = None
    if support_session is not None and support_session.property_id:
        selected_property = support_session.property
    else:
        property_id = request.session.get("ui_property_id")
        if property_id:
            selected_property = next((p for p in property_options if p.id == property_id), None)
        if selected_property is None and property_options:
            selected_property = property_options[0]
            request.session["ui_property_id"] = selected_property.id

    if require_property and selected_property is None:
        return {
            "tenant_options": tenant_options,
            "property_options": property_options,
            "selected_tenant": selected_tenant,
            "selected_property": None,
            "support_session": support_session,
            "support_sessions": [],
            "context_error": "No hay sede disponible para este usuario en el tenant seleccionado.",
            "nav_items": _build_nav_items(request.user, selected_tenant, None, request.path),
        }

    if selected_tenant and not request.user.is_super_admin:
        if not PermissionService.user_can_tenant_role(
            request.user,
            selected_tenant,
            ["admin", "operator", "supervisor"],
        ):
            return {
                "tenant_options": tenant_options,
                "property_options": property_options,
                "selected_tenant": selected_tenant,
                "selected_property": selected_property,
                "support_session": support_session,
                "support_sessions": [],
                "context_error": "No tienes permisos para operar en este tenant.",
                "nav_items": _build_nav_items(request.user, selected_tenant, selected_property, request.path),
            }

    if selected_tenant and selected_property:
        if not PermissionService.user_can_property_action(
            request.user,
            selected_tenant,
            selected_property,
            "can_access",
        ):
            return {
                "tenant_options": tenant_options,
                "property_options": property_options,
                "selected_tenant": selected_tenant,
                "selected_property": selected_property,
                "support_session": support_session,
                "support_sessions": [],
                "context_error": "No tienes permisos de acceso en esta sede.",
                "nav_items": _build_nav_items(request.user, selected_tenant, selected_property, request.path),
            }

    support_sessions = []
    if request.user.is_super_admin:
        support_sessions = list(
            TenantSupportAccessSession.objects.select_related("tenant", "property")
            .filter(started_by=request.user, ended_at__isnull=True)
            .order_by("-created_at")
        )

    return {
        "tenant_options": tenant_options,
        "property_options": property_options,
        "selected_tenant": selected_tenant,
        "selected_property": selected_property,
        "support_session": support_session,
        "support_sessions": support_sessions,
        "context_error": "",
        "nav_items": _build_nav_items(request.user, selected_tenant, selected_property, request.path),
    }


def favicon_file(request, file_name="favicon.ico"):
    allowed_files = {
        "favicon.ico": "image/vnd.microsoft.icon",
        "favicon-32x32.png": "image/png",
        "apple-touch-icon.png": "image/png",
    }
    if file_name not in allowed_files:
        raise Http404("Icono no encontrado.")
    icon_path = Path(settings.BASE_DIR) / "apps" / "webui" / "static" / "webui" / "icons" / file_name
    if not icon_path.exists():
        raise Http404("Icono no encontrado.")
    response = FileResponse(icon_path.open("rb"), content_type=allowed_files[file_name])
    response["Cache-Control"] = "public, max-age=3600"
    return response


def login_page(request):
    if request.user.is_authenticated:
        return redirect("webui-dashboard")
    if request.method == "POST":
        email = str(request.POST.get("email", "")).strip()
        password = request.POST.get("password", "")
        user = authenticate(request=request, username=email, password=password)
        if user is None:
            messages.error(request, "Credenciales invalidas.")
        else:
            login(request, user)
            return redirect("webui-dashboard")
    return render(request, "webui/login.html")


@require_http_methods(["POST"])
def logout_page(request):
    logout(request)
    return redirect("webui-login")


@login_required
@require_http_methods(["POST"])
def switch_context(request):
    if request.user.is_super_admin:
        support_session_id = str(request.POST.get("support_session_id", "")).strip()
        if support_session_id:
            support_session = _get_support_session(request.user, support_session_id)
            if support_session is None:
                messages.error(request, "Sesion de soporte invalida o inactiva.")
            else:
                request.session["support_session_id"] = support_session.id
                request.session["ui_tenant_id"] = support_session.tenant_id
                if support_session.property_id:
                    request.session["ui_property_id"] = support_session.property_id
                messages.success(request, "Sesion de soporte activada.")
                return redirect(request.POST.get("next") or "webui-dashboard")
        else:
            if "support_session_id" in request.session:
                del request.session["support_session_id"]
                messages.success(request, "Sesion de soporte desactivada.")

    tenant_id = request.POST.get("tenant_id")
    property_id = request.POST.get("property_id")
    if tenant_id:
        request.session["ui_tenant_id"] = int(tenant_id)
    if property_id:
        request.session["ui_property_id"] = int(property_id)
    return redirect(request.POST.get("next") or "webui-dashboard")


@login_required
@require_GET
def dashboard(request):
    ctx = _build_context(request, require_property=False)
    if ctx.get("selected_tenant") and not ctx.get("context_error"):
        tenant = ctx["selected_tenant"]
        property_obj = ctx["selected_property"]
        stats = {
            "workers": Worker.objects.filter(tenant=tenant, property=property_obj).count() if property_obj else 0,
            "areas": Area.objects.filter(tenant=tenant, property=property_obj).count() if property_obj else 0,
        }
    else:
        stats = {"workers": 0, "areas": 0}
    ctx["stats"] = stats
    if PermissionService.is_super_admin(request.user):
        ctx["system_stats"] = {
            "tenants": Tenant.objects.count(),
            "active_tenants": Tenant.objects.filter(status=TenantStatus.ACTIVE).count(),
            "properties": Property.objects.count(),
            "active_properties": Property.objects.filter(status=TenantStatus.ACTIVE).count(),
            "workers": Worker.objects.filter(active=True).count(),
            "disabled_modules": ModuleActivation.objects.filter(is_enabled=False).count(),
            "active_support_sessions": TenantSupportAccessSession.objects.filter(ended_at__isnull=True).count(),
            "recent_audits": AuditLog.objects.count(),
        }
        ctx["system_recent_audits"] = list(
            AuditLog.objects.select_related("tenant", "property", "user").order_by("-created_at")[:8]
        )
        ctx["system_recent_closures"] = list(
            MonthClosure.objects.select_related("tenant", "property", "closed_by", "reopened_by").order_by("-updated_at")[
                :8
            ]
        )
    return render(request, "webui/dashboard.html", ctx)


@login_required
@require_http_methods(["GET", "POST"])
def areas_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/areas.html", {**ctx, "rows": [], "form": AreaForm()})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "areas"):
        return render(
            request,
            "webui/areas.html",
            {
                **ctx,
                "rows": [],
                "form": AreaForm(),
                "can_manage_areas": False,
                "context_error": "Modulo desactivado: areas.",
            },
        )

    can_manage = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_manage_areas",
    )

    if request.method == "POST":
        if not can_manage:
            return HttpResponseForbidden("No tienes permisos para gestionar areas.")
        action = str(request.POST.get("action", "")).strip()
        if action == "create_area":
            form = AreaForm(request.POST)
            if form.is_valid():
                area = form.save(commit=False)
                area.tenant = tenant
                area.property = property_obj
                try:
                    area.save()
                except IntegrityError:
                    messages.error(request, "Ya existe un area con ese nombre en esta sede.")
                else:
                    AuditService.log(
                        tenant=tenant,
                        property_obj=property_obj,
                        user=request.user,
                        action="create",
                        entity_type="Area",
                        entity_id=area.id,
                        before={},
                        after=_audit_snapshot(area, ["name", "type", "active"]),
                    )
                    messages.success(request, "Area guardada.")
                    return redirect("webui-areas")
            else:
                messages.error(request, "Datos invalidos para crear area.")
        elif action == "update_area":
            area_id = str(request.POST.get("area_id", "")).strip()
            target = Area.objects.filter(tenant=tenant, property=property_obj, id=area_id).first()
            if target is None:
                messages.error(request, "Area no encontrada.")
                return redirect("webui-areas")
            before = _audit_snapshot(target, ["name", "type", "active"])
            target.name = str(request.POST.get("name", "")).strip()
            target.active = bool(request.POST.get("active"))
            if not target.name:
                messages.error(request, "El nombre del area es obligatorio.")
                return redirect("webui-areas")
            try:
                target.save()
            except IntegrityError:
                messages.error(request, "Ya existe un area con ese nombre en esta sede.")
            else:
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="update",
                    entity_type="Area",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(target, ["name", "type", "active"]),
                )
                messages.success(request, "Area actualizada.")
                return redirect("webui-areas")
        elif action == "deactivate_area":
            area_id = str(request.POST.get("area_id", "")).strip()
            replacement_area_id = str(request.POST.get("replacement_area_id", "")).strip()
            target = Area.objects.filter(tenant=tenant, property=property_obj, id=area_id).first()
            if target is None:
                messages.error(request, "Area no encontrada.")
                return redirect("webui-areas")

            worker_count = Worker.objects.filter(tenant=tenant, property=property_obj, area=target).count()
            shift_count = Shift.objects.filter(tenant=tenant, property=property_obj, area=target).count()
            requires_reassignment = (worker_count + shift_count) > 0
            if requires_reassignment:
                if not replacement_area_id.isdigit():
                    messages.error(
                        request,
                        "Debes seleccionar un area destino activa para reasignar trabajadores y turnos.",
                    )
                    return redirect("webui-areas")
                replacement = Area.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    id=int(replacement_area_id),
                    active=True,
                ).exclude(id=target.id).first()
                if replacement is None:
                    messages.error(
                        request,
                        "Debes seleccionar un area destino activa para reasignar trabajadores y turnos.",
                    )
                    return redirect("webui-areas")
                before = _audit_snapshot(target, ["name", "type", "active"])
                with transaction.atomic():
                    Worker.objects.filter(tenant=tenant, property=property_obj, area=target).update(area=replacement)
                    Shift.objects.filter(tenant=tenant, property=property_obj, area=target).update(area=replacement)
                    target.active = False
                    target.save(update_fields=["active", "updated_at"])
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="update",
                    entity_type="Area",
                    entity_id=target.id,
                    before=before,
                    after={
                        **_audit_snapshot(target, ["name", "type", "active"]),
                        "reassigned_to_area_id": replacement.id,
                        "reassigned_to_area_name": replacement.name,
                        "reassigned_workers": worker_count,
                        "reassigned_shifts": shift_count,
                    },
                )
                messages.success(request, f"Area desactivada y reasignada a {replacement.name}.")
                return redirect("webui-areas")

            before = _audit_snapshot(target, ["name", "type", "active"])
            target.active = False
            target.save(update_fields=["active", "updated_at"])
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="update",
                entity_type="Area",
                entity_id=target.id,
                before=before,
                after=_audit_snapshot(target, ["name", "type", "active"]),
            )
            messages.success(request, "Area desactivada.")
            return redirect("webui-areas")
        else:
            messages.error(request, "Accion invalida.")

    areas = list(
        Area.objects.filter(tenant=tenant, property=property_obj)
        .annotate(
            active_workers=Count("workers", filter=Q(workers__active=True)),
            active_shifts=Count("shifts", filter=Q(shifts__active=True)),
        )
        .order_by("name")
    )
    rows = [
        {
            "area": area,
            "active_workers": area.active_workers,
            "active_shifts": area.active_shifts,
            "has_dependencies": (area.active_workers + area.active_shifts) > 0,
        }
        for area in areas
    ]
    area_options = [area for area in areas if area.active]
    return render(
        request,
        "webui/areas.html",
        {
            **ctx,
            "rows": rows,
            "form": AreaForm(),
            "can_manage_areas": can_manage,
            "area_options": area_options,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def workers_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/workers.html", {**ctx, "form": WorkerForm(), "workers": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "workers"):
        return render(
            request,
            "webui/workers.html",
            {
                **ctx,
                "form": WorkerForm(),
                "workers": [],
                "area_options": [],
                "status_filter": "active",
                "can_manage_workers": False,
                "context_error": "Modulo desactivado: workers.",
            },
        )

    can_manage = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_manage_workers",
    )
    can_import_workers = can_manage and PermissionService.user_can_module(request.user, tenant, "excel_import")
    area_options = list(Area.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name"))

    if request.method == "POST":
        if not can_manage:
            return HttpResponseForbidden("No tienes permisos para gestionar trabajadores.")
        action = str(request.POST.get("action", "")).strip()
        if action == "preview_workers_inline":
            if not can_import_workers:
                messages.error(request, "No tienes permisos para importar trabajadores.")
                return redirect("webui-workers")
            if not request.POST.get("confirm_full_sync"):
                messages.error(request, "Confirma que el archivo contiene la lista completa de trabajadores activos de la sede.")
                return redirect("webui-workers")
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo CSV o XLSX.")
                return redirect("webui-workers")
            try:
                batch = WorkerImportService.create_worker_preview(
                    tenant=tenant,
                    fallback_property=property_obj,
                    file_name=uploaded_file.name,
                    file_bytes=uploaded_file.read(),
                    user=request.user,
                    create_missing_areas=bool(request.POST.get("create_missing_areas")),
                    sync_mode=True,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("webui-workers")
            messages.success(request, f"Vista previa creada. Revisa el lote #{batch.id}.")
            return redirect(f"{reverse('webui-workers')}?import_batch_id={batch.id}&import_modal=1")

        if action == "confirm_workers_import_inline":
            if not can_import_workers:
                messages.error(request, "No tienes permisos para confirmar importaciones de trabajadores.")
                return redirect("webui-workers")
            batch_id_raw = str(request.POST.get("batch_id", "")).strip()
            batch = ImportBatch.objects.filter(
                id=int(batch_id_raw) if batch_id_raw.isdigit() else 0,
                tenant=tenant,
                property=property_obj,
                source_type="workers",
                status="preview",
            ).first()
            if batch is None:
                messages.error(request, "Lote no encontrado o ya aplicado.")
                return redirect("webui-workers")
            if _import_batch_has_errors(batch):
                messages.error(request, "No se puede confirmar una importacion con errores bloqueantes.")
                return redirect(f"{reverse('webui-workers')}?import_batch_id={batch.id}&import_modal=1")
            if not request.POST.get("confirm_apply_sync"):
                messages.error(request, "Confirma la aplicacion de la sincronizacion completa.")
                return redirect(f"{reverse('webui-workers')}?import_batch_id={batch.id}&import_modal=1")
            WorkerImportService.confirm_worker_import(batch=batch)
            messages.success(request, f"Lote #{batch.id} confirmado. Trabajadores sincronizados.")
            return redirect("webui-workers")

        if action in {"", "create_worker"}:
            form = WorkerForm(request.POST)
            form.fields["area"].queryset = Area.objects.filter(tenant=tenant, property=property_obj).order_by("name")
            if form.is_valid():
                worker = form.save(commit=False)
                worker.tenant = tenant
                worker.property = property_obj
                worker.save()
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="create",
                    entity_type="Worker",
                    entity_id=worker.id,
                    before={},
                    after=_audit_snapshot(
                        worker,
                        ["document_number", "first_name", "last_name", "area_id", "active", "start_date", "end_date"],
                    ),
                )
                messages.success(request, "Trabajador guardado.")
                return redirect("webui-workers")
        elif action in {"update_worker", "deactivate_worker"}:
            worker_id = str(request.POST.get("worker_id", "")).strip()
            target = Worker.objects.filter(tenant=tenant, property=property_obj, id=worker_id).first()
            if target is None:
                messages.error(request, "Trabajador no encontrado.")
                return redirect("webui-workers")
            before = _audit_snapshot(
                target,
                ["document_number", "first_name", "last_name", "area_id", "active", "start_date", "end_date"],
            )
            if action == "deactivate_worker":
                target.active = False
                if target.end_date is None:
                    target.end_date = timezone.localdate()
                target.save(update_fields=["active", "end_date", "updated_at"])
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="delete",
                    entity_type="Worker",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(
                        target,
                        ["document_number", "first_name", "last_name", "area_id", "active", "start_date", "end_date"],
                    ),
                )
                messages.success(request, "Trabajador desactivado.")
                return redirect("webui-workers")

            area_id = str(request.POST.get("area", "")).strip()
            area = Area.objects.filter(tenant=tenant, property=property_obj, id=area_id, active=True).first()
            if area is None:
                messages.error(request, "Area invalida.")
                return redirect("webui-workers")
            target.document_number = str(request.POST.get("document_number", "")).strip()
            target.first_name = str(request.POST.get("first_name", "")).strip()
            target.last_name = str(request.POST.get("last_name", "")).strip()
            target.area = area
            target.active = bool(request.POST.get("active"))
            if not target.document_number or not target.first_name or not target.last_name:
                messages.error(request, "Documento, nombre y apellido son obligatorios.")
                return redirect("webui-workers")
            try:
                target.save()
            except IntegrityError:
                messages.error(request, "Ya existe un trabajador con ese documento en esta sede.")
            else:
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="update",
                    entity_type="Worker",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(
                        target,
                        ["document_number", "first_name", "last_name", "area_id", "active", "start_date", "end_date"],
                    ),
                )
                messages.success(request, "Trabajador actualizado.")
                return redirect("webui-workers")
        else:
            messages.error(request, "Accion invalida.")

    status_filter = str(request.GET.get("status", "active")).strip().lower()
    if status_filter not in {"active", "inactive", "all"}:
        status_filter = "active"
    queryset = Worker.objects.select_related("area").filter(tenant=tenant, property=property_obj)
    if status_filter == "active":
        queryset = queryset.filter(active=True)
    elif status_filter == "inactive":
        queryset = queryset.filter(active=False)
    queryset = queryset.order_by("last_name", "first_name")

    form = WorkerForm()
    form.fields["area"].queryset = Area.objects.filter(tenant=tenant, property=property_obj).order_by("name")
    import_batch = None
    import_batch_id = str(request.GET.get("import_batch_id", "")).strip()
    if import_batch_id.isdigit():
        import_batch = ImportBatch.objects.filter(
            id=int(import_batch_id),
            tenant=tenant,
            property=property_obj,
            source_type="workers",
        ).first()
    import_context = _build_import_batch_preview_context(import_batch, row_limit=50)

    return render(
        request,
        "webui/workers.html",
        {
            **ctx,
            "workers": queryset,
            "form": form,
            "area_options": area_options,
            "status_filter": status_filter,
            "can_manage_workers": can_manage,
            "can_import_workers": can_import_workers,
            "import_batch": import_batch,
            "import_modal_open": bool(import_batch and request.GET.get("import_modal")),
            **import_context,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def shifts_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/shifts.html", {**ctx, "form": ShiftForm(), "shifts": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "shifts"):
        return render(
            request,
            "webui/shifts.html",
            {
                **ctx,
                "form": ShiftForm(),
                "shifts": [],
                "area_options": [],
                "status_filter": "active",
                "can_manage_shifts": False,
                "context_error": "Modulo desactivado: shifts.",
            },
        )

    can_manage = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_manage_shifts",
    )
    can_import_shifts = can_manage and PermissionService.user_can_module(request.user, tenant, "excel_import")
    area_options = list(Area.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name"))

    if request.method == "POST":
        if not can_manage:
            return HttpResponseForbidden("No tienes permisos para gestionar turnos.")
        action = str(request.POST.get("action", "")).strip()
        if action == "preview_shifts_inline":
            if not can_import_shifts:
                messages.error(request, "No tienes permisos para importar turnos.")
                return redirect("webui-shifts")
            if not request.POST.get("confirm_full_sync"):
                messages.error(request, "Confirma que el archivo contiene la lista completa de turnos activos de la sede.")
                return redirect("webui-shifts")
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo CSV o XLSX.")
                return redirect("webui-shifts")
            try:
                batch = ShiftAreaImportService.create_shift_preview(
                    tenant=tenant,
                    fallback_property=property_obj,
                    file_name=uploaded_file.name,
                    file_bytes=uploaded_file.read(),
                    user=request.user,
                    create_missing_areas=bool(request.POST.get("create_missing_areas")),
                    sync_mode=True,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("webui-shifts")
            messages.success(request, f"Vista previa creada. Revisa el lote #{batch.id}.")
            return redirect(f"{reverse('webui-shifts')}?import_batch_id={batch.id}&import_modal=1")

        if action == "confirm_shifts_import_inline":
            if not can_import_shifts:
                messages.error(request, "No tienes permisos para confirmar importaciones de turnos.")
                return redirect("webui-shifts")
            batch_id_raw = str(request.POST.get("batch_id", "")).strip()
            batch = ImportBatch.objects.filter(
                id=int(batch_id_raw) if batch_id_raw.isdigit() else 0,
                tenant=tenant,
                property=property_obj,
                source_type="shifts_area",
                status="preview",
            ).first()
            if batch is None:
                messages.error(request, "Lote no encontrado o ya aplicado.")
                return redirect("webui-shifts")
            if _import_batch_has_errors(batch):
                messages.error(request, "No se puede confirmar una importacion con errores bloqueantes.")
                return redirect(f"{reverse('webui-shifts')}?import_batch_id={batch.id}&import_modal=1")
            if not request.POST.get("confirm_apply_sync"):
                messages.error(request, "Confirma la aplicacion de la sincronizacion completa.")
                return redirect(f"{reverse('webui-shifts')}?import_batch_id={batch.id}&import_modal=1")
            ShiftAreaImportService.confirm_shift_import(batch=batch)
            messages.success(request, f"Lote #{batch.id} confirmado. Turnos sincronizados.")
            return redirect("webui-shifts")

        if action in {"", "create_shift"}:
            form = ShiftForm(request.POST)
            form.fields["area"].queryset = Area.objects.filter(tenant=tenant, property=property_obj).order_by("name")
            if form.is_valid():
                shift = form.save(commit=False)
                shift.tenant = tenant
                shift.property = property_obj
                shift.save()
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="create",
                    entity_type="Shift",
                    entity_id=shift.id,
                    before={},
                    after=_audit_snapshot(
                        shift,
                        ["area_id", "name", "buk_code", "start_time", "end_time", "is_night_shift", "active"],
                    ),
                )
                messages.success(request, "Turno guardado.")
                return redirect("webui-shifts")
        elif action in {"update_shift", "deactivate_shift"}:
            shift_id = str(request.POST.get("shift_id", "")).strip()
            target = Shift.objects.filter(tenant=tenant, property=property_obj, id=shift_id).first()
            if target is None:
                messages.error(request, "Turno no encontrado.")
                return redirect("webui-shifts")
            before = _audit_snapshot(
                target,
                ["area_id", "name", "buk_code", "start_time", "end_time", "break_start", "break_end", "is_night_shift", "active"],
            )
            if action == "deactivate_shift":
                target.active = False
                target.save(update_fields=["active", "updated_at"])
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="delete",
                    entity_type="Shift",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(
                        target,
                        ["area_id", "name", "buk_code", "start_time", "end_time", "break_start", "break_end", "is_night_shift", "active"],
                    ),
                )
                messages.success(request, "Turno desactivado.")
                return redirect("webui-shifts")

            form = ShiftForm(request.POST, instance=target)
            form.fields["area"].queryset = Area.objects.filter(tenant=tenant, property=property_obj).order_by("name")
            if form.is_valid():
                updated = form.save()
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="update",
                    entity_type="Shift",
                    entity_id=updated.id,
                    before=before,
                    after=_audit_snapshot(
                        updated,
                        ["area_id", "name", "buk_code", "start_time", "end_time", "break_start", "break_end", "is_night_shift", "active"],
                    ),
                )
                messages.success(request, "Turno actualizado.")
                return redirect("webui-shifts")
        else:
            messages.error(request, "Accion invalida.")

    status_filter = str(request.GET.get("status", "active")).strip().lower()
    if status_filter not in {"active", "inactive", "all"}:
        status_filter = "active"
    queryset = Shift.objects.select_related("area").filter(tenant=tenant, property=property_obj)
    if status_filter == "active":
        queryset = queryset.filter(active=True)
    elif status_filter == "inactive":
        queryset = queryset.filter(active=False)
    queryset = queryset.order_by("area__name", "name")

    form = ShiftForm()
    form.fields["area"].queryset = Area.objects.filter(tenant=tenant, property=property_obj).order_by("name")
    import_batch = None
    import_batch_id = str(request.GET.get("import_batch_id", "")).strip()
    if import_batch_id.isdigit():
        import_batch = ImportBatch.objects.filter(
            id=int(import_batch_id),
            tenant=tenant,
            property=property_obj,
            source_type="shifts_area",
        ).first()
    import_context = _build_import_batch_preview_context(import_batch, row_limit=50)

    return render(
        request,
        "webui/shifts.html",
        {
            **ctx,
            "shifts": queryset,
            "form": form,
            "area_options": area_options,
            "status_filter": status_filter,
            "can_manage_shifts": can_manage,
            "can_import_shifts": can_import_shifts,
            "import_batch": import_batch,
            "import_modal_open": bool(import_batch and request.GET.get("import_modal")),
            **import_context,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def auto_shifts_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/auto_shifts.html", {**ctx, "rows": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "shifts"):
        return render(
            request,
            "webui/auto_shifts.html",
            {**ctx, "rows": [], "can_manage_shifts": False, "context_error": "Modulo desactivado: shifts."},
        )

    can_manage = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_manage_shifts",
    )

    def _apply_auto_shift_filters(queryset, *, q, area_id_raw, active_filter):
        filtered = queryset
        if q:
            filtered = filtered.filter(
                Q(name__icontains=q)
                | Q(buk_code__icontains=q)
                | Q(area__name__icontains=q)
            )
        if area_id_raw.isdigit():
            filtered = filtered.filter(area_id=int(area_id_raw))
        if active_filter == "active":
            filtered = filtered.filter(active=True)
        elif active_filter == "inactive":
            filtered = filtered.filter(active=False)
        return filtered

    if request.method == "POST":
        filter_q = str(request.POST.get("q", "")).strip()
        filter_area_id = str(request.POST.get("filter_area_id", "")).strip()
        filter_active = str(request.POST.get("filter_active", "all")).strip()
    else:
        filter_q = str(request.GET.get("q", "")).strip()
        filter_area_id = str(request.GET.get("area_id", "")).strip()
        filter_active = str(request.GET.get("active", "all")).strip()
    if filter_active not in {"all", "active", "inactive"}:
        filter_active = "all"

    base_queryset = Shift.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        name__startswith="AUTO_",
    )
    filtered_queryset = _apply_auto_shift_filters(
        base_queryset,
        q=filter_q,
        area_id_raw=filter_area_id,
        active_filter=filter_active,
    )

    query_params = {}
    if filter_q:
        query_params["q"] = filter_q
    if filter_area_id:
        query_params["area_id"] = filter_area_id
    if filter_active != "all":
        query_params["active"] = filter_active
    redirect_url = "/app/shifts-auto/"
    if query_params:
        redirect_url += "?" + urlencode(query_params)

    if request.method == "POST":
        if not can_manage:
            return HttpResponseForbidden("No tienes permisos para gestionar turnos.")
        action = str(request.POST.get("action", "")).strip()
        if action == "normalize_visible_auto_shifts":
            updated = 0
            skipped = 0
            for target in filtered_queryset:
                area_token = (slugify(target.area.name) or "area").replace("-", "").upper()[:4]
                start_part = target.start_time.strftime("%H%M")
                end_part = target.end_time.strftime("%H%M")
                target.name = f"{target.area.name}_AUTO_{start_part}-{end_part}"
                base_code = f"{area_token}-{start_part}{end_part}"
                candidate = base_code
                suffix = 1
                while Shift.objects.filter(
                    tenant=tenant,
                    property=property_obj,
                    buk_code=candidate,
                ).exclude(id=target.id).exists():
                    suffix += 1
                    candidate = f"{base_code}-{suffix}"
                target.buk_code = candidate
                target.is_night_shift = target.end_time <= target.start_time
                try:
                    target.full_clean()
                    target.save()
                    updated += 1
                except (IntegrityError, ValueError):
                    skipped += 1
            if updated == 0:
                messages.error(request, "No se actualizaron turnos AUTO con el filtro actual.")
            elif skipped:
                messages.success(request, f"Normalizacion aplicada a {updated} turnos. Se omitieron {skipped}.")
            else:
                messages.success(request, f"Normalizacion aplicada a {updated} turnos.")
            return redirect(redirect_url)

        shift_id = str(request.POST.get("shift_id", "")).strip()
        target = Shift.objects.filter(id=shift_id, tenant=tenant, property=property_obj).first()
        if target is None:
            messages.error(request, "Turno no encontrado.")
            return redirect(redirect_url)
        if not target.name.startswith("AUTO_"):
            messages.error(request, "Solo se permite esta accion sobre turnos AUTO.")
            return redirect(redirect_url)

        if action == "merge_auto_shift":
            destination_id = str(request.POST.get("destination_shift_id", "")).strip()
            source_mode = str(request.POST.get("source_mode", "deactivate")).strip()
            merge_confirm_text = str(request.POST.get("merge_confirm_text", "")).strip().upper()
            if source_mode not in {"keep", "deactivate", "delete"}:
                source_mode = "deactivate"
            destination = Shift.objects.filter(
                id=destination_id,
                tenant=tenant,
                property=property_obj,
                area=target.area,
            ).exclude(id=target.id).first()
            if destination is None:
                messages.error(request, "Selecciona un turno destino valido de la misma area.")
                return redirect(redirect_url)

            assignment_source_qs = ScheduleAssignment.objects.filter(
                tenant=tenant,
                property=property_obj,
                shift=target,
            ).order_by("date", "id")
            impact_count = assignment_source_qs.count()
            if impact_count >= AUTO_SHIFT_MERGE_CONFIRM_THRESHOLD and merge_confirm_text != "CONFIRMAR":
                messages.error(
                    request,
                    f"Fusion de alto impacto ({impact_count} asignaciones). Escribe CONFIRMAR para continuar.",
                )
                return redirect(redirect_url)

            impact_dates = assignment_source_qs.aggregate(first_date=Min("date"), last_date=Max("date"))
            moved_assignment_ids_sample = list(assignment_source_qs.values_list("id", flat=True)[:50])
            moved = assignment_source_qs.update(
                shift=destination,
                updated_by=request.user,
            )
            before = {
                "source_shift_id": target.id,
                "source_name": target.name,
                "source_buk_code": target.buk_code,
                "destination_shift_id": destination.id,
                "destination_name": destination.name,
                "destination_buk_code": destination.buk_code,
                "moved_assignments": moved,
                "moved_assignment_ids_sample": moved_assignment_ids_sample,
                "moved_assignment_ids_sample_truncated": impact_count > len(moved_assignment_ids_sample),
                "moved_assignments_date_from": (
                    impact_dates["first_date"].isoformat() if impact_dates["first_date"] else None
                ),
                "moved_assignments_date_to": (
                    impact_dates["last_date"].isoformat() if impact_dates["last_date"] else None
                ),
            }
            if source_mode == "deactivate":
                target.active = False
                target.save(update_fields=["active", "updated_at"])
            elif source_mode == "delete":
                try:
                    target.delete()
                except ProtectedError:
                    target.active = False
                    target.save(update_fields=["active", "updated_at"])
                    source_mode = "deactivate"
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="auto_shift_merge",
                entity_type="Shift",
                entity_id=destination.id,
                before=before,
                after={
                    "source_mode": source_mode,
                    "source_shift_active_after_merge": target.active if source_mode != "delete" else False,
                },
            )
            messages.success(request, f"Fusion completada. Asignaciones movidas: {moved}.")
            return redirect(redirect_url)

        if action == "update_auto_shift":
            target.name = str(request.POST.get("name", "")).strip() or target.name
            target.buk_code = str(request.POST.get("buk_code", "")).strip() or target.buk_code
            target.active = bool(request.POST.get("active"))
            target.is_night_shift = bool(request.POST.get("is_night_shift"))
            try:
                target.full_clean()
                target.save()
            except (IntegrityError, ValueError):
                messages.error(request, "No se pudo guardar el turno (nombre/codigo duplicado o invalido).")
            else:
                messages.success(request, "Turno AUTO actualizado.")
            return redirect(redirect_url)

        if action == "deactivate_auto_shift":
            target.active = False
            target.save(update_fields=["active", "updated_at"])
            messages.success(request, "Turno AUTO desactivado.")
            return redirect(redirect_url)

        messages.error(request, "Accion invalida.")
        return redirect(redirect_url)

    area_shift_options = {}
    for item in Shift.objects.filter(tenant=tenant, property=property_obj).order_by("area__name", "name"):
        if item.name.startswith("AUTO_"):
            continue
        area_shift_options.setdefault(item.area_id, []).append(item)

    impact_map = {}
    impact_rows = (
        ScheduleAssignment.objects.filter(tenant=tenant, property=property_obj, shift__name__startswith="AUTO_")
        .values("shift_id")
        .annotate(assignments=Count("id"), first_date=Min("date"), last_date=Max("date"))
    )
    for item in impact_rows:
        impact_map[item["shift_id"]] = {
            "assignments": item["assignments"],
            "first_date": item["first_date"],
            "last_date": item["last_date"],
        }

    rows = []
    for shift in filtered_queryset.order_by("area__name", "name"):
        impact = impact_map.get(shift.id, {"assignments": 0, "first_date": None, "last_date": None})
        rows.append(
            {
                "shift": shift,
                "merge_options": area_shift_options.get(shift.area_id, []),
                "impact": impact,
            }
        )
    area_options = list(Area.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name"))
    return render(
        request,
        "webui/auto_shifts.html",
        {
            **ctx,
            "rows": rows,
            "can_manage_shifts": can_manage,
            "area_options": area_options,
            "filter_q": filter_q,
            "filter_area_id": int(filter_area_id) if filter_area_id.isdigit() else None,
            "filter_active": filter_active,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def special_states_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/special_states.html", {**ctx, "form": SpecialStateForm()})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "special_states"):
        return render(
            request,
            "webui/special_states.html",
            {
                **ctx,
                "states": [],
                "form": SpecialStateForm(),
                "can_manage_special_states": False,
                "context_error": "Modulo desactivado: special_states.",
            },
        )

    can_manage = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_manage_shifts",
    )
    form = SpecialStateForm()
    if request.method == "POST":
        if not can_manage:
            return HttpResponseForbidden("No tienes permisos para gestionar estados especiales.")
        action = str(request.POST.get("action", "create_state")).strip()
        if action == "create_state":
            form = SpecialStateForm(request.POST)
            if form.is_valid():
                state = form.save(commit=False)
                state.tenant = tenant
                state.property = property_obj
                try:
                    state.save()
                except IntegrityError:
                    messages.error(request, "Ya existe un estado especial con ese nombre en esta sede.")
                else:
                    AuditService.log(
                        tenant=tenant,
                        property_obj=property_obj,
                        user=request.user,
                        action="create",
                        entity_type="SpecialState",
                        entity_id=state.id,
                        before={},
                        after=_audit_snapshot(state, ["name", "buk_code", "active"]),
                    )
                    messages.success(request, "Estado especial guardado.")
                    return redirect("webui-special-states")
            else:
                messages.error(request, "Datos invalidos para crear estado especial.")
        elif action in {"update_state", "deactivate_state"}:
            state_id = str(request.POST.get("state_id", "")).strip()
            target = SpecialState.objects.filter(tenant=tenant, property=property_obj, id=state_id).first()
            if target is None:
                messages.error(request, "Estado especial no encontrado.")
                return redirect("webui-special-states")
            before = _audit_snapshot(target, ["name", "buk_code", "active"])
            if action == "deactivate_state":
                target.active = False
                target.save(update_fields=["active", "updated_at"])
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="delete",
                    entity_type="SpecialState",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(target, ["name", "buk_code", "active"]),
                )
                messages.success(request, "Estado especial desactivado.")
                return redirect("webui-special-states")

            target.name = str(request.POST.get("name", "")).strip()
            target.buk_code = str(request.POST.get("buk_code", "")).strip()
            target.active = bool(request.POST.get("active"))
            if not target.name:
                messages.error(request, "El nombre del estado especial es obligatorio.")
                return redirect("webui-special-states")
            try:
                target.save()
            except IntegrityError:
                messages.error(request, "Ya existe un estado especial con ese nombre en esta sede.")
            else:
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="update",
                    entity_type="SpecialState",
                    entity_id=target.id,
                    before=before,
                    after=_audit_snapshot(target, ["name", "buk_code", "active"]),
                )
                messages.success(request, "Estado especial actualizado.")
                return redirect("webui-special-states")
        else:
            messages.error(request, "Accion invalida.")
    queryset = SpecialState.objects.filter(tenant=tenant, property=property_obj).order_by("-active", "name")

    return render(
        request,
        "webui/special_states.html",
        {
            **ctx,
            "states": queryset,
            "form": form,
            "can_manage_special_states": can_manage,
        },
    )


@login_required
@require_GET
def scheduling_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/scheduling.html", {**ctx})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return render(
            request,
            "webui/scheduling.html",
            {
                **ctx,
                "rows": [],
                "days": [],
                "context_error": "Modulo desactivado: scheduling.",
            },
        )

    month_raw = str(request.GET.get("month", "")).strip()
    worker_query = str(request.GET.get("worker_q", "")).strip()
    focus_date_raw = str(request.GET.get("focus_date", "")).strip()
    edited_worker_id_raw = str(request.GET.get("edited_worker_id", "")).strip()
    edited_date_raw = str(request.GET.get("edited_date", "")).strip()
    today = date.today()
    if not month_raw:
        month_raw = f"{today.year:04d}-{today.month:02d}"
    try:
        year, month = [int(x) for x in month_raw.split("-", 1)]
    except ValueError:
        year, month = today.year, today.month
        month_raw = f"{year:04d}-{month:02d}"
    focus_date = None
    focus_date_value = ""
    if focus_date_raw:
        try:
            parsed_focus_date = date.fromisoformat(focus_date_raw)
        except ValueError:
            parsed_focus_date = None
        if parsed_focus_date and parsed_focus_date.year == year and parsed_focus_date.month == month:
            focus_date = parsed_focus_date
            focus_date_value = parsed_focus_date.isoformat()
    edited_worker_id = int(edited_worker_id_raw) if edited_worker_id_raw.isdigit() else None
    edited_date = None
    if edited_date_raw:
        try:
            parsed_edited_date = date.fromisoformat(edited_date_raw)
        except ValueError:
            parsed_edited_date = None
        if parsed_edited_date and parsed_edited_date.year == year and parsed_edited_date.month == month:
            edited_date = parsed_edited_date

    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    month_prev_value = f"{prev_year:04d}-{prev_month:02d}"
    month_next_value = f"{next_year:04d}-{next_month:02d}"

    area_id = request.GET.get("area_id")
    allowed_area_ids = PermissionService.get_accessible_area_ids(
        request.user,
        tenant,
        property_obj,
        action="can_view",
    )
    area_queryset = Area.objects.filter(
        tenant=tenant,
        property=property_obj,
        active=True,
        id__in=allowed_area_ids,
    ).order_by("name")
    selected_area = area_queryset.filter(id=area_id).first() if area_id else None

    workers = Worker.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        active=True,
        area_id__in=allowed_area_ids,
    )
    if selected_area:
        workers = workers.filter(area=selected_area)
    if worker_query:
        workers = workers.filter(
            Q(first_name__icontains=worker_query)
            | Q(last_name__icontains=worker_query)
            | Q(document_number__icontains=worker_query)
        )
    workers = workers.order_by("area__name", "last_name", "first_name")

    _, total_days = calendar.monthrange(year, month)
    days = []
    for day in range(1, total_days + 1):
        work_date = date(year, month, day)
        days.append(
            {
                "date": work_date,
                "day": day,
                "weekday_label": WEEKDAY_SHORT_LABELS[work_date.weekday()],
                "is_weekend": work_date.weekday() >= 5,
                "is_focus": bool(focus_date and work_date == focus_date),
            }
        )

    start_date = date(year, month, 1)
    end_date = date(year, month, total_days)
    assignments = ScheduleAssignment.objects.select_related("shift", "special_state").filter(
        tenant=tenant,
        property=property_obj,
        date__gte=start_date,
        date__lte=end_date,
        worker__area_id__in=allowed_area_ids,
    )
    assignment_index = {(item.worker_id, item.date): item for item in assignments}

    states = list(
        SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name")
    )
    shifts = list(
        Shift.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
            area_id__in=allowed_area_ids,
        )
        .select_related("area")
        .order_by("area__name", "name", "buk_code")
    )
    shifts_by_area = {}
    for shift in shifts:
        shifts_by_area.setdefault(shift.area_id, []).append(shift)
    for values in shifts_by_area.values():
        values.sort(key=lambda x: (x.name, x.buk_code))

    can_schedule_property = PermissionService.user_can_property_action(
        request.user,
        tenant,
        property_obj,
        "can_schedule",
    )
    is_tenant_admin = PermissionService.user_can_tenant_role(request.user, tenant, ["admin"])
    is_month_closed = MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=year,
        month=month,
    )
    can_schedule = can_schedule_property and not is_month_closed

    rows = []
    scheduled_cells_count = 0
    missing_cells_count = 0
    editable_cells_count = 0
    night_cells_count = 0
    for worker in workers:
        if not PermissionService.user_can_area_view(request.user, tenant, property_obj, worker.area):
            continue
        area_shifts = shifts_by_area.get(worker.area_id, [])
        row_cells = []
        area_allowed = PermissionService.user_can_area_schedule(
            request.user,
            tenant,
            property_obj,
            worker.area,
        )
        for day_meta in days:
            assignment = assignment_index.get((worker.id, day_meta["date"]))
            selected_value = ""
            code = ""
            is_night = False
            selected_shift_id = None
            selected_state_id = None
            selected_value = ""
            display_label = "Sin asignar"
            if assignment:
                if assignment.shift_id:
                    selected_value = f"shift:{assignment.shift_id}"
                    code = assignment.shift.buk_code
                    display_label = (
                        f"{assignment.shift.start_time:%H:%M}\u2013{assignment.shift.end_time:%H:%M}"
                    )
                    is_night = bool(assignment.shift.is_night_shift)
                    selected_shift_id = assignment.shift_id
                elif assignment.special_state_id:
                    selected_value = f"state:{assignment.special_state_id}"
                    code = assignment.special_state.buk_code or assignment.special_state.name
                    display_label = assignment.special_state.name
                    selected_state_id = assignment.special_state_id
            is_editable = can_schedule and area_allowed
            if code:
                scheduled_cells_count += 1
            else:
                missing_cells_count += 1
            if is_editable:
                editable_cells_count += 1
            if is_night:
                night_cells_count += 1
            row_cells.append(
                {
                    "date": day_meta["date"],
                    "day": day_meta["day"],
                    "weekday_label": day_meta["weekday_label"],
                    "is_weekend": day_meta["is_weekend"],
                    "selected_value": selected_value,
                    "selected_shift_id": selected_shift_id,
                    "selected_state_id": selected_state_id,
                    "display_code": code,
                    "display_label": display_label,
                    "is_night": is_night,
                    "is_focus": bool(focus_date and day_meta["date"] == focus_date),
                    "is_recent_edit": bool(
                        edited_worker_id
                        and edited_date
                        and worker.id == edited_worker_id
                        and day_meta["date"] == edited_date
                    ),
                    "editable": is_editable,
                    "shift_options": area_shifts,
                    "state_options": states,
                }
            )
        mobile_weeks = []
        for index in range(0, len(row_cells), 7):
            week_cells = row_cells[index : index + 7]
            mobile_weeks.append(
                {
                    "label": week_cells[0]["date"].strftime("%d/%m") if week_cells else "",
                    "has_missing_assignments": any(not cell["display_code"] for cell in week_cells),
                    "cells": week_cells,
                }
            )
        rows.append(
            {
                "worker": worker,
                "area_allowed": area_allowed,
                "cells": row_cells,
                "mobile_weeks": mobile_weeks,
                "has_missing_assignments": any(not cell["display_code"] for cell in row_cells),
            }
        )
    bulk_worker_options = [row["worker"] for row in rows if row["area_allowed"]]
    if selected_area:
        bulk_shift_options = list(shifts_by_area.get(selected_area.id, []))
    else:
        bulk_shift_options = list(shifts)
    week_pattern_templates = []
    week_pattern_templates_admin = []
    range_templates = []
    range_templates_admin = []
    templates_load_error = ""
    try:
        template_queryset = SchedulePatternTemplate.objects.filter(
            tenant=tenant,
            property=property_obj,
        )
        if selected_area:
            template_queryset = template_queryset.filter(area=selected_area)
        else:
            template_queryset = template_queryset.filter(area__isnull=True)
        week_pattern_templates = list(template_queryset.filter(active=True).order_by("name", "id"))
        week_pattern_templates_admin = list(template_queryset.order_by("-active", "name", "id"))
        range_template_queryset = ScheduleRangeTemplate.objects.filter(
            tenant=tenant,
            property=property_obj,
        )
        if selected_area:
            range_template_queryset = range_template_queryset.filter(area=selected_area)
        else:
            range_template_queryset = range_template_queryset.filter(area__isnull=True)
        range_templates = list(range_template_queryset.filter(active=True).order_by("name", "id"))
        range_templates_admin = list(range_template_queryset.order_by("-active", "name", "id"))
    except (OperationalError, ProgrammingError):
        templates_load_error = "No se cargaron plantillas de asignacion. Ejecuta migraciones pendientes."
    range_template_state_ids = []
    for template_item in range_templates_admin:
        for range_item in list(template_item.ranges or []):
            state_id = range_item.get("special_state_id")
            if isinstance(state_id, int):
                range_template_state_ids.append(state_id)
    range_template_states_map = {
        item.id: item
        for item in SpecialState.objects.filter(
            id__in=range_template_state_ids,
            tenant=tenant,
            property=property_obj,
        )
    }
    range_template_health_map = {
        item.id: _range_template_health(
            template=item,
            states_map=range_template_states_map,
            year=year,
            month=month,
        )
        for item in range_templates_admin
    }
    range_template_risk_filter = _normalize_range_template_risk_filter(request.GET.get("range_template_risk", "all"))
    range_templates_admin_rows = [
        {"template": item, "health": range_template_health_map.get(item.id)}
        for item in range_templates_admin
    ]
    if range_template_risk_filter != "all":
        range_templates_admin_rows = [
            row
            for row in range_templates_admin_rows
            if row.get("health") and row["health"]["level"] == range_template_risk_filter
        ]
    range_editor_rows = []
    range_editor_template_id = None
    range_template_edit_id_raw = str(request.GET.get("range_template_edit_id", "")).strip()
    if range_templates_admin and range_template_edit_id_raw.isdigit():
        edit_template = ScheduleRangeTemplate.objects.filter(
            id=int(range_template_edit_id_raw),
            tenant=tenant,
            property=property_obj,
        ).first()
        if edit_template is not None and (not edit_template.area_id or (selected_area and edit_template.area_id == selected_area.id)):
            _, month_days_total = calendar.monthrange(year, month)
            for item in list(edit_template.ranges or []):
                start_day = item.get("start_day")
                end_day = item.get("end_day")
                state_id = item.get("special_state_id")
                if not isinstance(start_day, int) or not isinstance(end_day, int) or not isinstance(state_id, int):
                    continue
                if start_day < 1 or start_day > month_days_total or end_day < start_day:
                    continue
                from_date = date(year, month, start_day)
                to_date = date(year, month, min(end_day, month_days_total))
                range_editor_rows.append(
                    {
                        "date_from": from_date.isoformat(),
                        "date_to": to_date.isoformat(),
                        "state_id": state_id,
                    }
                )
            range_editor_template_id = edit_template.id
            if range_editor_rows:
                messages.info(
                    request,
                    f"Editando contenido de plantilla: {edit_template.name}.",
                )
    range_template_preview = None
    preview_template_id_raw = str(request.GET.get("range_template_preview_id", "")).strip()
    if range_templates_admin and preview_template_id_raw.isdigit():
        preview_template = ScheduleRangeTemplate.objects.filter(
            id=int(preview_template_id_raw),
            tenant=tenant,
            property=property_obj,
        ).first()
        if preview_template is not None and (not preview_template.area_id or (selected_area and preview_template.area_id == selected_area.id)):
            state_ids = []
            raw_ranges = list(preview_template.ranges or [])
            for item in raw_ranges:
                state_id = item.get("special_state_id")
                if isinstance(state_id, int):
                    state_ids.append(state_id)
            states_map = {
                item.id: item
                for item in SpecialState.objects.filter(
                    id__in=state_ids,
                    tenant=tenant,
                    property=property_obj,
                )
            }
            _, month_days_total = calendar.monthrange(year, month)
            detail_rows = []
            day_hits = {}
            valid_ranges_for_impact = []
            for index, item in enumerate(raw_ranges):
                start_day = item.get("start_day")
                end_day = item.get("end_day")
                state_id = item.get("special_state_id")
                state_obj = states_map.get(state_id)
                if not isinstance(start_day, int) or not isinstance(end_day, int):
                    detail_rows.append(
                        {
                            "index": index + 1,
                            "date_from": "-",
                            "date_to": "-",
                            "state_name": "Invalido",
                            "state_code": "",
                            "status": "Rango invalido",
                        }
                    )
                    continue
                if start_day < 1 or start_day > month_days_total or end_day < start_day:
                    detail_rows.append(
                        {
                            "index": index + 1,
                            "date_from": "-",
                            "date_to": "-",
                            "state_name": "Invalido",
                            "state_code": "",
                            "status": "Fuera de mes",
                        }
                    )
                    continue
                clamped_end = min(end_day, month_days_total)
                date_from = date(year, month, start_day)
                date_to = date(year, month, clamped_end)
                if state_obj is None:
                    state_name = "Estado no disponible"
                    state_code = ""
                    status = "Invalido"
                else:
                    state_name = state_obj.name
                    state_code = state_obj.buk_code or ""
                    status = "OK" if state_obj.active else "Inactivo"
                for day in range(start_day, clamped_end + 1):
                    day_hits[day] = day_hits.get(day, 0) + 1
                if state_obj is not None and state_obj.active:
                    valid_ranges_for_impact.append(
                        {
                            "date_from": date_from,
                            "date_to": date_to,
                            "state": state_obj,
                        }
                    )
                detail_rows.append(
                    {
                        "index": index + 1,
                        "date_from": date_from,
                        "date_to": date_to,
                        "state_name": state_name,
                        "state_code": state_code,
                        "status": status,
                    }
                )
            overlap_days = sum(1 for count in day_hits.values() if count > 1)
            covered_days = len(day_hits)
            plans = {}
            for range_item in valid_ranges_for_impact:
                total_range_days = (range_item["date_to"] - range_item["date_from"]).days + 1
                for worker in bulk_worker_options:
                    for offset in range(total_range_days):
                        target_date = range_item["date_from"] + timedelta(days=offset)
                        plans[(worker.id, target_date)] = {
                            "worker": worker,
                            "date": target_date,
                            "shift": None,
                            "special_state": range_item["state"],
                        }
            impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
            range_template_preview = {
                "template": preview_template,
                "rows": detail_rows,
                "month_days_total": month_days_total,
                "covered_days": covered_days,
                "overlap_days": overlap_days,
                "impact": impact,
                "impact_workers": len(bulk_worker_options),
            }
    pending_previews = list(request.session.get("scheduling_pending_previews", []))

    return render(
        request,
        "webui/scheduling.html",
        {
            **ctx,
            "days": days,
            "rows": rows,
            "month_value": month_raw,
            "month_prev_value": month_prev_value,
            "month_next_value": month_next_value,
            "worker_query": worker_query,
            "focus_date_value": focus_date_value,
            "areas": area_queryset,
            "selected_area_id": selected_area.id if selected_area else None,
            "show_area_selector": is_tenant_admin or area_queryset.count() > 1,
            "area_filter_label": "Todas" if is_tenant_admin else "Todas mis áreas",
            "can_schedule": can_schedule,
            "is_tenant_admin": is_tenant_admin,
            "is_month_closed": is_month_closed,
            "scheduled_cells_count": scheduled_cells_count,
            "missing_cells_count": missing_cells_count,
            "editable_cells_count": editable_cells_count,
            "night_cells_count": night_cells_count,
            "bulk_state_options": states,
            "bulk_shift_options": bulk_shift_options,
            "bulk_worker_options": bulk_worker_options,
            "week_pattern_templates": week_pattern_templates,
            "week_pattern_templates_admin": week_pattern_templates_admin,
            "range_templates": range_templates,
            "range_templates_admin": range_templates_admin,
            "range_templates_admin_rows": range_templates_admin_rows,
            "range_template_risk_filter": range_template_risk_filter,
            "range_editor_rows": range_editor_rows,
            "range_editor_template_id": range_editor_template_id,
            "range_template_preview": range_template_preview,
            "pending_previews": pending_previews,
            "templates_load_error": templates_load_error,
            "is_scheduling_page": True,
            "has_single_property": len(ctx["property_options"]) == 1,
        },
    )


@login_required
@require_GET
def scheduling_team_report_pdf(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_access"):
        return HttpResponseForbidden("No tienes acceso a esta sede.")

    area_id = str(request.GET.get("area_id", "")).strip()
    if not area_id.isdigit():
        return HttpResponseForbidden("Debe seleccionar un area.")
    area = Area.objects.filter(tenant=tenant, property=property_obj, id=int(area_id), active=True).first()
    if area is None:
        return HttpResponseForbidden("Area no encontrada.")
    if not PermissionService.user_can_area_view(request.user, tenant, property_obj, area):
        return HttpResponseForbidden("No tienes permisos para ver esta area.")

    today = timezone.localdate()
    date_from = _parse_date_or_default(request.GET.get("date_from", ""), today)
    date_to = _parse_date_or_default(request.GET.get("date_to", ""), today + timedelta(days=6))
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    if (date_to - date_from).days > 45:
        return HttpResponseForbidden("El reporte PDF permite un rango maximo de 45 dias.")

    date_columns = []
    cursor = date_from
    while cursor <= date_to:
        date_columns.append(cursor)
        cursor += timedelta(days=1)

    assignments = (
        ScheduleAssignment.objects.select_related("worker", "shift", "special_state")
        .filter(
            tenant=tenant,
            property=property_obj,
            worker__area=area,
            worker__active=True,
            date__gte=date_from,
            date__lte=date_to,
        )
        .order_by("date", "shift__start_time", "shift__name", "worker__last_name", "worker__first_name")
    )
    grid = defaultdict(list)
    shift_rows = {}
    state_rows = {}
    for assignment in assignments:
        worker_name = f"{assignment.worker.first_name} {assignment.worker.last_name}".strip()
        if assignment.shift_id:
            key = ("shift", assignment.shift_id)
            shift_rows[key] = assignment.shift
        elif assignment.special_state_id:
            key = ("state", assignment.special_state_id)
            state_rows[key] = assignment.special_state
        else:
            continue
        grid[(key, assignment.date)].append(worker_name)

    row_keys = []
    for key, shift in sorted(
        shift_rows.items(),
        key=lambda item: (item[1].start_time, item[1].end_time, item[1].name),
    ):
        row_keys.append((key, f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}"))
    for key, state in sorted(state_rows.items(), key=lambda item: item[1].name):
        row_keys.append((key, state.name))

    date_chunks = [date_columns[index : index + 7] for index in range(0, len(date_columns), 7)]

    buffer = BytesIO()
    response = HttpResponse(content_type="application/pdf")
    file_name = f"reporte_equipo_{property_obj.slug}_{area.name.lower().replace(' ', '_')}_{date_from}_{date_to}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    small = styles["BodyText"]
    small.fontSize = 6.5
    small.leading = 7.5
    header_style = styles["Heading2"]
    story = [
        Paragraph(f"{property_obj.name} - {area.name}", header_style),
        Paragraph(f"{date_from.strftime('%d-%m-%Y')} a {date_to.strftime('%d-%m-%Y')}", styles["BodyText"]),
        Spacer(1, 8),
    ]

    first_col_width = 76
    usable_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    remaining_width = usable_width - first_col_width
    for chunk_index, chunk_dates in enumerate(date_chunks):
        if chunk_index:
            story.append(Spacer(1, 10))
        story.append(
            Paragraph(
                f"Bloque {chunk_index + 1}: {chunk_dates[0].strftime('%d-%m-%Y')} a {chunk_dates[-1].strftime('%d-%m-%Y')}",
                styles["BodyText"],
            )
        )
        story.append(Spacer(1, 4))
        header = [Paragraph("Turno", small)]
        for day in chunk_dates:
            header.append(Paragraph(f"{day.strftime('%d/%m')}<br/>{WEEKDAY_SHORT_LABELS[day.weekday()]}", small))
        table_data = [header]
        if row_keys:
            for key, label in row_keys:
                row = [Paragraph(label, small)]
                for day in chunk_dates:
                    names = sorted(grid.get((key, day), []))
                    row.append(Paragraph("<br/>".join(names) if names else "-", small))
                table_data.append(row)
        else:
            table_data.append([Paragraph("Sin asignaciones", small)] + [Paragraph("-", small) for _ in chunk_dates])

        date_col_width = remaining_width / max(1, len(chunk_dates))
        table = Table(table_data, repeatRows=1, colWidths=[first_col_width] + [date_col_width] * len(chunk_dates))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f8fafc")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
    doc.build(story)
    response.write(buffer.getvalue())
    return response


@login_required
@require_http_methods(["POST"])
def scheduling_assign(request):
    wants_json = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.headers.get("accept", "").lower().find("application/json") >= 0
    )

    def json_error(message, status=400):
        return JsonResponse({"ok": False, "message": message}, status=status)

    def json_success(message, *, display_code="", is_empty=False, assignment_value=""):
        return JsonResponse(
            {
                "ok": True,
                "message": message,
                "display_code": display_code or "",
                "is_empty": bool(is_empty),
                "assignment_value": assignment_value,
            }
        )

    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        if wants_json:
            return json_error(ctx["context_error"], status=400)
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        if wants_json:
            return json_error("Modulo desactivado: scheduling.", status=403)
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        if wants_json:
            return json_error("No tienes permisos para asignar en esta sede.", status=403)
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    worker_query = str(request.POST.get("worker_q", "")).strip()
    focus_date = str(request.POST.get("focus_date", "")).strip()
    worker_id = request.POST.get("worker_id")
    work_date_raw = request.POST.get("work_date")
    assignment_value = str(request.POST.get("assignment_value", "")).strip()

    redirect_url = _build_scheduling_redirect_url(month_value, area_value, worker_query, focus_date)

    if not worker_id or not work_date_raw:
        if wants_json:
            return json_error("Debe seleccionar trabajador y fecha.")
        messages.error(request, "Debe seleccionar trabajador y fecha.")
        return redirect(redirect_url)

    try:
        work_date = date.fromisoformat(work_date_raw)
    except ValueError:
        if wants_json:
            return json_error("Fecha invalida.")
        messages.error(request, "Fecha invalida.")
        return redirect(redirect_url)

    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=work_date.year,
        month=work_date.month,
    ):
        if wants_json:
            return json_error("El mes esta cerrado para esta sede.", status=409)
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    worker = Worker.objects.select_related("area").filter(
        id=worker_id,
        tenant=tenant,
        property=property_obj,
    ).first()
    if worker is None:
        if wants_json:
            return json_error("Trabajador no encontrado.", status=404)
        messages.error(request, "Trabajador no encontrado.")
        return redirect(redirect_url)

    if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
        if wants_json:
            return json_error("No tienes permisos para asignar en esta area.", status=403)
        return HttpResponseForbidden("No tienes permisos para asignar en esta area.")

    if assignment_value == "":
        existing_assignment = ScheduleAssignment.objects.filter(
            tenant=tenant,
            property=property_obj,
            worker=worker,
            date=work_date,
        ).select_related("shift", "special_state").first()
        if existing_assignment is None:
            if wants_json:
                return json_success("Celda sin asignacion.", display_code="", is_empty=True, assignment_value="")
            messages.info(request, "No habia asignacion para eliminar en esa fecha.")
            return redirect(redirect_url)

        before = {
            "shift_id": existing_assignment.shift_id,
            "special_state_id": existing_assignment.special_state_id,
            "worker_id": existing_assignment.worker_id,
            "date": existing_assignment.date.isoformat(),
        }
        assignment_id = existing_assignment.id
        existing_assignment.delete()
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_assignment_delete",
            entity_type="ScheduleAssignment",
            entity_id=assignment_id,
            before=before,
            after={},
        )
        if wants_json:
            return json_success("Asignacion eliminada.", display_code="", is_empty=True, assignment_value="")
        messages.success(request, "Asignacion eliminada.")
        return redirect(
            _build_scheduling_redirect_url(
                month_value,
                area_value,
                worker_query,
                work_date.isoformat(),
                edited_worker_id=str(worker.id),
                edited_date=work_date.isoformat(),
            )
        )

    shift = None
    special_state = None
    if assignment_value.startswith("shift:"):
        shift_id = assignment_value.split(":", 1)[1]
        shift = Shift.objects.filter(id=shift_id, tenant=tenant, property=property_obj, area=worker.area).first()
        if shift is None:
            if wants_json:
                return json_error("Turno invalido para esta area.")
            messages.error(request, "Turno invalido para esta area.")
            return redirect(redirect_url)
    elif assignment_value.startswith("state:"):
        state_id = assignment_value.split(":", 1)[1]
        special_state = SpecialState.objects.filter(id=state_id, tenant=tenant, property=property_obj).first()
        if special_state is None:
            if wants_json:
                return json_error("Estado especial invalido.")
            messages.error(request, "Estado especial invalido.")
            return redirect(redirect_url)
    else:
        if wants_json:
            return json_error("Seleccion invalida.")
        messages.error(request, "Seleccion invalida.")
        return redirect(redirect_url)

    ScheduleAssignmentService.upsert_assignment(
        tenant=tenant,
        property_obj=property_obj,
        worker=worker,
        date=work_date,
        shift=shift,
        special_state=special_state,
        user=request.user,
    )
    if shift is not None:
        if wants_json:
            return json_success(
                "Asignacion guardada.",
                display_code=shift.buk_code,
                is_empty=False,
                assignment_value=f"shift:{shift.id}",
            )
    elif special_state is not None and wants_json:
        return json_success(
            "Asignacion guardada.",
            display_code=special_state.buk_code or special_state.name,
            is_empty=False,
            assignment_value=f"state:{special_state.id}",
        )
    messages.success(request, "Asignacion guardada.")
    return redirect(
        _build_scheduling_redirect_url(
            month_value,
            area_value,
            worker_query,
            work_date.isoformat(),
            edited_worker_id=str(worker.id),
            edited_date=work_date.isoformat(),
        )
    )


@login_required
@require_http_methods(["POST"])
def scheduling_copy_cell_range(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    worker_query = str(request.POST.get("worker_q", "")).strip()
    source_date_raw = str(request.POST.get("source_date", "")).strip()
    date_from_raw = str(request.POST.get("date_from", "")).strip()
    date_to_raw = str(request.POST.get("date_to", "")).strip()
    source_worker_id = str(request.POST.get("source_worker_id", "")).strip()
    assignment_value = str(request.POST.get("assignment_value", "")).strip()
    redirect_url = _build_scheduling_redirect_url(month_value, area_value, worker_query, source_date_raw)

    if not source_worker_id or not assignment_value:
        messages.error(request, "Debe copiar desde una celda con turno o estado especial.")
        return redirect(redirect_url)
    try:
        source_date = date.fromisoformat(source_date_raw)
        date_from = date.fromisoformat(date_from_raw)
        date_to = date.fromisoformat(date_to_raw)
    except ValueError:
        messages.error(request, "Rango de fechas invalido.")
        return redirect(redirect_url)

    if date_to < date_from:
        messages.error(request, "La fecha final no puede ser menor que la fecha inicial.")
        return redirect(redirect_url)
    total_days = (date_to - date_from).days + 1
    if total_days > 62:
        messages.error(request, "El rango maximo para copiar es de 62 dias.")
        return redirect(redirect_url)
    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=date_from,
        end_date=date_to,
    ):
        messages.error(request, "El rango destino incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    worker = Worker.objects.select_related("area").filter(
        id=source_worker_id,
        tenant=tenant,
        property=property_obj,
        active=True,
    ).first()
    if worker is None:
        messages.error(request, "Trabajador no encontrado o inactivo.")
        return redirect(redirect_url)
    if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
        return HttpResponseForbidden("No tienes permisos para asignar en esta area.")

    shifts_map = {
        shift.id: shift
        for shift in Shift.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        ).select_related("area")
    }
    states_map = {
        state.id: state
        for state in SpecialState.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        )
    }
    shift, special_state, error = _resolve_assignment_value_for_worker(
        assignment_value=assignment_value,
        worker=worker,
        shifts_map=shifts_map,
        states_map=states_map,
    )
    if error:
        messages.error(request, "La asignacion origen ya no es valida para este trabajador.")
        return redirect(redirect_url)

    plans = {}
    for offset in range(total_days):
        target_date = date_from + timedelta(days=offset)
        plans[(worker.id, target_date)] = {
            "worker": worker,
            "date": target_date,
            "shift": shift,
            "special_state": special_state,
        }
    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)

    with transaction.atomic():
        copied = 0
        for plan in plans.values():
            ScheduleAssignmentService.upsert_assignment(
                tenant=tenant,
                property_obj=property_obj,
                worker=plan["worker"],
                date=plan["date"],
                shift=plan["shift"],
                special_state=plan["special_state"],
                user=request.user,
            )
            copied += 1
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_copy_cell_range_apply",
            entity_type="ScheduleAssignment",
            entity_id=f"{worker.id}:{date_from.isoformat()}:{date_to.isoformat()}",
            before={},
            after={
                "worker_id": worker.id,
                "source_date": source_date.isoformat(),
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
                "assignment_value": assignment_value,
                "shift_id": shift.id if shift else None,
                "special_state_id": special_state.id if special_state else None,
                "copied": copied,
                "impact": impact,
            },
        )

    messages.success(
        request,
        (
            f"Asignacion copiada a {copied} dias "
            f"({impact['to_create']} nuevas, {impact['to_update']} actualizadas, {impact['unchanged']} sin cambios)."
        ),
    )
    return redirect(
        _build_scheduling_redirect_url(
            month_value,
            area_value,
            worker_query,
            date_from.isoformat(),
            edited_worker_id=str(worker.id),
            edited_date=date_from.isoformat(),
        )
    )


@login_required
@require_http_methods(["POST"])
def scheduling_preview_cancel(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url(
        month_value,
        area_value,
        str(request.POST.get("worker_q", "")).strip(),
        str(request.POST.get("focus_date", "")).strip(),
    )
    preview_id = str(request.POST.get("preview_id", "")).strip()
    if not preview_id:
        messages.error(request, "Preview invalido.")
        return redirect(redirect_url)

    previews = list(request.session.get("scheduling_pending_previews", []))
    updated = []
    removed_item = None
    for item in previews:
        if removed_item is None and str(item.get("id", "")) == preview_id:
            removed_item = item
            continue
        updated.append(item)
    if removed_item is None:
        messages.error(request, "Preview no encontrado.")
        return redirect(redirect_url)

    request.session["scheduling_pending_previews"] = updated
    request.session.modified = True
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_preview_canceled",
        entity_type="SchedulingPreview",
        entity_id=preview_id,
        before=removed_item,
        after={},
    )
    messages.success(request, "Vista previa cancelada.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_state(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    worker_query = str(request.POST.get("worker_q", "")).strip()
    focus_date = str(request.POST.get("focus_date", "")).strip()
    work_date_raw = str(request.POST.get("work_date", "")).strip()
    state_id = str(request.POST.get("special_state_id", "")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))

    redirect_url = _build_scheduling_redirect_url(month_value, area_value, worker_query, focus_date)

    if not work_date_raw or not state_id:
        messages.error(request, "Debe seleccionar fecha y estado especial.")
        return redirect(redirect_url)
    try:
        work_date = date.fromisoformat(work_date_raw)
    except ValueError:
        messages.error(request, "Fecha invalida.")
        return redirect(redirect_url)

    special_state = SpecialState.objects.filter(id=state_id, tenant=tenant, property=property_obj, active=True).first()
    if special_state is None:
        messages.error(request, "Estado especial invalido.")
        return redirect(redirect_url)

    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=work_date.year,
        month=work_date.month,
    ):
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    workers = Worker.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        active=True,
    )
    if area_value.isdigit():
        workers = workers.filter(area_id=int(area_value))
    if worker_query:
        workers = workers.filter(
            Q(first_name__icontains=worker_query)
            | Q(last_name__icontains=worker_query)
            | Q(document_number__icontains=worker_query)
        )

    plans = {}
    for worker in workers:
        if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            continue
        plans[(worker.id, work_date)] = {
            "worker": worker,
            "date": work_date,
            "shift": None,
            "special_state": special_state,
        }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_fields = {
            "month": month_value,
            "area_id": area_value,
            "worker_q": worker_query,
            "work_date": work_date.isoformat(),
            "special_state_id": str(special_state.id),
        }
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-bulk-state"),
            label="Asignacion masiva por dia",
            summary=f"Total: {impact['total']} | Nuevas: {impact['to_create']} | Actualizaciones: {impact['to_update']} | Sin cambios: {impact['unchanged']}",
            fields=preview_fields,
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "scheduling_bulk_state_apply",
                "redirect_url": redirect_url,
                "preview_fields": preview_fields,
                "impact": impact,
            },
        )
        if impact["total"] == 0:
            messages.warning(request, "Vista previa generada sin trabajadores visibles con permiso.")
        else:
            messages.info(
                request,
                f"Vista previa: {impact['total']} filas ({impact['to_create']} nuevas, {impact['to_update']} actualizaciones, {impact['unchanged']} sin cambios).",
            )
        return redirect(redirect_url)

    preview_fields = {
        "month": month_value,
        "area_id": area_value,
        "worker_q": worker_query,
        "work_date": work_date.isoformat(),
        "special_state_id": str(special_state.id),
    }
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-bulk-state"),
        fields=preview_fields,
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=str(consumed_preview.get("id", "")),
            before=consumed_preview,
            after={
                "operation": "scheduling_bulk_state_apply",
                "redirect_url": redirect_url,
            },
        )

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_state_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{work_date.isoformat()}",
        before={},
        after={
            "work_date": work_date.isoformat(),
            "special_state_id": special_state.id,
            "special_state_code": special_state.buk_code or "",
            "applied": applied,
            "area_filter": int(area_value) if area_value.isdigit() else None,
            "impact": impact,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios (sin trabajadores visibles con permiso).")
    else:
        messages.success(request, f"Estado especial aplicado a {applied} trabajadores.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_shift(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    worker_query = str(request.POST.get("worker_q", "")).strip()
    focus_date = str(request.POST.get("focus_date", "")).strip()
    work_date_raw = str(request.POST.get("work_date", "")).strip()
    shift_id = str(request.POST.get("shift_id", "")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url(month_value, area_value, worker_query, focus_date)

    if not work_date_raw or not shift_id:
        messages.error(request, "Debe seleccionar fecha y turno.")
        return redirect(redirect_url)
    try:
        work_date = date.fromisoformat(work_date_raw)
    except ValueError:
        messages.error(request, "Fecha invalida.")
        return redirect(redirect_url)

    shift = Shift.objects.select_related("area").filter(
        id=shift_id,
        tenant=tenant,
        property=property_obj,
        active=True,
    ).first()
    if shift is None:
        messages.error(request, "Turno invalido.")
        return redirect(redirect_url)

    if area_value.isdigit() and shift.area_id != int(area_value):
        messages.error(request, "El turno seleccionado no pertenece al area filtrada.")
        return redirect(redirect_url)

    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=work_date.year,
        month=work_date.month,
    ):
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    workers = Worker.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        active=True,
    )
    if area_value.isdigit():
        workers = workers.filter(area_id=int(area_value))
    if worker_query:
        workers = workers.filter(
            Q(first_name__icontains=worker_query)
            | Q(last_name__icontains=worker_query)
            | Q(document_number__icontains=worker_query)
        )

    plans = {}
    for worker in workers:
        if worker.area_id != shift.area_id:
            continue
        if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            continue
        plans[(worker.id, work_date)] = {
            "worker": worker,
            "date": work_date,
            "shift": shift,
            "special_state": None,
        }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    preview_fields = {
        "month": month_value,
        "area_id": area_value,
        "worker_q": worker_query,
        "work_date": work_date.isoformat(),
        "shift_id": str(shift.id),
    }
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-bulk-shift"),
            label="Asignacion masiva por turno (dia)",
            summary=f"Total: {impact['total']} | Nuevas: {impact['to_create']} | Actualizaciones: {impact['to_update']} | Sin cambios: {impact['unchanged']}",
            fields=preview_fields,
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "scheduling_bulk_shift_apply",
                "redirect_url": redirect_url,
                "preview_fields": preview_fields,
                "impact": impact,
            },
        )
        if impact["total"] == 0:
            messages.warning(request, "Vista previa generada sin trabajadores visibles con permiso.")
        else:
            messages.info(
                request,
                f"Vista previa: {impact['total']} filas ({impact['to_create']} nuevas, {impact['to_update']} actualizaciones, {impact['unchanged']} sin cambios).",
            )
        return redirect(redirect_url)

    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-bulk-shift"),
        fields=preview_fields,
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=str(consumed_preview.get("id", "")),
            before=consumed_preview,
            after={
                "operation": "scheduling_bulk_shift_apply",
                "redirect_url": redirect_url,
            },
        )

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1

    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_shift_apply",
        entity_type="ScheduleAssignment",
        entity_id=None,
        before={},
        after={
            "month": month_value,
            "area_id": area_value or None,
            "work_date": work_date.isoformat(),
            "shift_id": shift.id,
            "applied": applied,
            "dry_run": False,
            "impact": impact,
        },
    )
    if applied == 0:
        messages.error(request, "No se aplicaron cambios (sin trabajadores visibles con permiso).")
    else:
        messages.success(request, f"Turno aplicado a {applied} trabajadores.")
    return redirect(redirect_url)


def _build_scheduling_redirect_url(
    month_value,
    area_value,
    worker_query="",
    focus_date="",
    edited_worker_id="",
    edited_date="",
):
    params = {"month": month_value}
    if area_value:
        params["area_id"] = area_value
    worker_query_text = str(worker_query or "").strip()
    if worker_query_text:
        params["worker_q"] = worker_query_text
    focus_date_text = str(focus_date or "").strip()
    if focus_date_text:
        try:
            date.fromisoformat(focus_date_text)
            params["focus_date"] = focus_date_text
        except ValueError:
            pass
    edited_worker_text = str(edited_worker_id or "").strip()
    if edited_worker_text.isdigit():
        params["edited_worker_id"] = edited_worker_text
    edited_date_text = str(edited_date or "").strip()
    if edited_date_text:
        try:
            date.fromisoformat(edited_date_text)
            params["edited_date"] = edited_date_text
        except ValueError:
            pass
    return f"/app/scheduling/?{urlencode(params)}"


def _build_scheduling_redirect_url_from_post(request, month_value, area_value):
    return _build_scheduling_redirect_url(
        month_value,
        area_value,
        str(request.POST.get("worker_q", "")).strip(),
        str(request.POST.get("focus_date", "")).strip(),
    )


def _parse_month_or_none(month_value):
    raw_value = str(month_value or "").strip()
    if not raw_value:
        return None, None
    parts = raw_value.split("-", 1)
    if len(parts) != 2:
        return None, None
    try:
        year = int(parts[0])
        month = int(parts[1])
    except ValueError:
        return None, None
    if month < 1 or month > 12:
        return None, None
    return year, month


def _is_any_month_closed(*, tenant, property_obj, start_date, end_date):
    cursor = start_date
    while cursor <= end_date:
        if MonthClosureService.is_closed(
            tenant=tenant,
            property_obj=property_obj,
            year=cursor.year,
            month=cursor.month,
        ):
            return True
        next_month = cursor.replace(day=28) + timedelta(days=4)
        cursor = next_month.replace(day=1)
    return False


def _get_visible_workers_for_scheduling(*, request, tenant, property_obj, area_value):
    workers = Worker.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        active=True,
    )
    if str(area_value or "").isdigit():
        workers = workers.filter(area_id=int(area_value))
    worker_query = str(request.POST.get("worker_q") or request.GET.get("worker_q") or "").strip()
    if worker_query:
        workers = workers.filter(
            Q(first_name__icontains=worker_query)
            | Q(last_name__icontains=worker_query)
            | Q(document_number__icontains=worker_query)
        )
    visible_workers = []
    for worker in workers:
        if PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            visible_workers.append(worker)
    return visible_workers


def _get_target_workers_for_scheduling(*, request, tenant, property_obj, area_value, selected_worker_ids):
    visible_workers = _get_visible_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
    )
    if not selected_worker_ids:
        return visible_workers
    selected_set = set(selected_worker_ids)
    return [worker for worker in visible_workers if worker.id in selected_set]


def _parse_selected_worker_ids(raw_values):
    worker_ids = []
    for raw_value in raw_values:
        raw_text = str(raw_value or "").strip()
        if raw_text.isdigit():
            worker_ids.append(int(raw_text))
    return worker_ids


def _parse_multi_range_rows(post_data):
    raw_from_values = post_data.getlist("range_date_from")
    raw_to_values = post_data.getlist("range_date_to")
    raw_state_values = post_data.getlist("range_state_id")
    rows_count = max(len(raw_from_values), len(raw_to_values), len(raw_state_values))
    rows = []
    for index in range(rows_count):
        from_raw = str(raw_from_values[index]).strip() if index < len(raw_from_values) else ""
        to_raw = str(raw_to_values[index]).strip() if index < len(raw_to_values) else ""
        state_raw = str(raw_state_values[index]).strip() if index < len(raw_state_values) else ""
        rows.append(
            {
                "index": index + 1,
                "from_raw": from_raw,
                "to_raw": to_raw,
                "state_raw": state_raw,
            }
        )
    return rows


def _range_template_health(*, template, states_map, year, month):
    _, month_days_total = calendar.monthrange(year, month)
    invalid_ranges = 0
    overlap_days = 0
    covered_days = 0
    day_hits = {}
    for item in list(template.ranges or []):
        start_day = item.get("start_day")
        end_day = item.get("end_day")
        state_id = item.get("special_state_id")
        state_obj = states_map.get(state_id)
        if (
            not isinstance(start_day, int)
            or not isinstance(end_day, int)
            or start_day < 1
            or end_day < start_day
            or start_day > month_days_total
            or state_obj is None
            or not state_obj.active
        ):
            invalid_ranges += 1
            continue
        clamped_end = min(end_day, month_days_total)
        for day in range(start_day, clamped_end + 1):
            day_hits[day] = day_hits.get(day, 0) + 1
    covered_days = len(day_hits)
    overlap_days = sum(1 for count in day_hits.values() if count > 1)
    if invalid_ranges > 0:
        level = "error"
        label = "Rojo"
    elif overlap_days > 0:
        level = "warning"
        label = "Amarillo"
    elif covered_days == 0 or covered_days < month_days_total:
        level = "warning"
        label = "Amarillo"
    else:
        level = "ok"
        label = "Verde"
    return {
        "level": level,
        "label": label,
        "invalid_ranges": invalid_ranges,
        "overlap_days": overlap_days,
        "covered_days": covered_days,
        "month_days_total": month_days_total,
    }


def _normalize_range_template_risk_filter(raw_value):
    value = str(raw_value or "").strip().lower()
    if value in {"error", "warning", "ok"}:
        return value
    return "all"


def _next_range_template_version_name(*, tenant, property_obj, area_obj, source_name):
    base_name = str(source_name or "").strip()
    match = re.match(r"^(.*)\s+v(\d+)$", base_name, flags=re.IGNORECASE)
    if match:
        root_name = match.group(1).strip()
    else:
        root_name = base_name
    if not root_name:
        root_name = "Plantilla subrangos"

    names = list(
        ScheduleRangeTemplate.objects.filter(
            tenant=tenant,
            property=property_obj,
            area=area_obj,
            name__istartswith=root_name,
        ).values_list("name", flat=True)
    )
    max_version = 1
    version_pattern = re.compile(rf"^{re.escape(root_name)}\s+v(\d+)$", re.IGNORECASE)
    for name in names:
        text = str(name or "").strip()
        if text.lower() == root_name.lower():
            max_version = max(max_version, 1)
            continue
        version_match = version_pattern.match(text)
        if version_match:
            try:
                number = int(version_match.group(1))
            except ValueError:
                continue
            max_version = max(max_version, number)
    return f"{root_name} v{max_version + 1}"


def _extract_week_pattern_from_source(source):
    pattern = {}
    has_any_value = False
    for key, weekday in WEEK_PATTERN_KEYS:
        value = str(source.get(f"{key}_value", "")).strip()
        pattern[weekday] = value
        if value:
            has_any_value = True
    return pattern, has_any_value


def _serialize_week_pattern(pattern):
    serialized = {}
    for key, weekday in WEEK_PATTERN_KEYS:
        serialized[key] = str(pattern.get(weekday, "")).strip()
    return serialized


def _deserialize_week_pattern(pattern_dict):
    source = pattern_dict or {}
    parsed = {}
    for key, weekday in WEEK_PATTERN_KEYS:
        parsed[weekday] = str(source.get(key, "")).strip()
    return parsed


def _copy_assignment(*, tenant, property_obj, worker, target_date, source_assignment, user, copy_kind="all"):
    if source_assignment is None:
        return False
    shift, special_state = _resolve_copy_payload(source_assignment=source_assignment, copy_kind=copy_kind)
    if shift is None and special_state is None:
        return False
    ScheduleAssignmentService.upsert_assignment(
        tenant=tenant,
        property_obj=property_obj,
        worker=worker,
        date=target_date,
        shift=shift,
        special_state=special_state,
        user=user,
    )
    return True


def _resolve_assignment_value_for_worker(*, assignment_value, worker, shifts_map, states_map):
    raw_value = str(assignment_value or "").strip()
    if not raw_value:
        return None, None, "empty"
    if raw_value.startswith("shift:"):
        shift_id = raw_value.split(":", 1)[1].strip()
        if not shift_id.isdigit():
            return None, None, "invalid"
        shift = shifts_map.get(int(shift_id))
        if shift is None or shift.area_id != worker.area_id:
            return None, None, "invalid"
        return shift, None, None
    if raw_value.startswith("state:"):
        state_id = raw_value.split(":", 1)[1].strip()
        if not state_id.isdigit():
            return None, None, "invalid"
        state = states_map.get(int(state_id))
        if state is None:
            return None, None, "invalid"
        return None, state, None
    return None, None, "invalid"


def _resolve_copy_payload(*, source_assignment, copy_kind):
    if source_assignment is None:
        return None, None
    shift = None
    special_state = None
    if copy_kind == "shift":
        if source_assignment.shift_id is None:
            return None, None
        shift = source_assignment.shift
    elif copy_kind == "state":
        if source_assignment.special_state_id is None:
            return None, None
        special_state = source_assignment.special_state
    else:
        shift = source_assignment.shift
        special_state = source_assignment.special_state
    return shift, special_state


def _summarize_assignment_plans(*, tenant, property_obj, plans):
    if not plans:
        return {"total": 0, "to_create": 0, "to_update": 0, "unchanged": 0}
    worker_ids = sorted({item["worker"].id for item in plans.values()})
    min_date = min(item["date"] for item in plans.values())
    max_date = max(item["date"] for item in plans.values())
    existing_map = {
        (item.worker_id, item.date): item
        for item in ScheduleAssignment.objects.filter(
            tenant=tenant,
            property=property_obj,
            worker_id__in=worker_ids,
            date__gte=min_date,
            date__lte=max_date,
        )
    }
    to_create = 0
    to_update = 0
    unchanged = 0
    for key, plan in plans.items():
        current = existing_map.get(key)
        target_shift_id = plan["shift"].id if plan["shift"] is not None else None
        target_state_id = plan["special_state"].id if plan["special_state"] is not None else None
        if current is None:
            to_create += 1
            continue
        if current.shift_id == target_shift_id and current.special_state_id == target_state_id:
            unchanged += 1
        else:
            to_update += 1
    return {
        "total": len(plans),
        "to_create": to_create,
        "to_update": to_update,
        "unchanged": unchanged,
    }


def _is_truthy(raw_value):
    return str(raw_value or "").strip().lower() in {"1", "true", "on", "yes"}


def _normalize_preview_fields(fields):
    normalized = {}
    for key, value in (fields or {}).items():
        if isinstance(value, (list, tuple)):
            normalized[str(key)] = [str(item) for item in value if str(item).strip() != ""]
        else:
            text = str(value or "").strip()
            if text != "":
                normalized[str(key)] = text
    return normalized


def _preview_fields_to_template(normalized_fields):
    items = []
    for key, value in normalized_fields.items():
        if isinstance(value, list):
            values = value
        else:
            values = [value]
        items.append({"name": key, "values": values})
    return items


def _queue_scheduling_preview(*, request, action_url, label, summary, fields):
    normalized_fields = _normalize_preview_fields(fields)
    previews = list(request.session.get("scheduling_pending_previews", []))
    preview_id = str(uuid4())
    previews.insert(
        0,
        {
            "id": preview_id,
            "action_url": action_url,
            "label": label,
            "summary": summary,
            "fields": _preview_fields_to_template(normalized_fields),
            "_match_key": json.dumps({"action_url": action_url, "fields": normalized_fields}, ensure_ascii=False, sort_keys=True),
        },
    )
    request.session["scheduling_pending_previews"] = previews[:8]
    request.session.modified = True
    return preview_id


def _consume_scheduling_preview(*, request, action_url, fields):
    previews = list(request.session.get("scheduling_pending_previews", []))
    normalized_fields = _normalize_preview_fields(fields)
    target_key = json.dumps({"action_url": action_url, "fields": normalized_fields}, ensure_ascii=False, sort_keys=True)
    updated = []
    removed_item = None
    for item in previews:
        if removed_item is None and item.get("_match_key") == target_key:
            removed_item = item
            continue
        updated.append(item)
    if removed_item is not None:
        request.session["scheduling_pending_previews"] = updated
        request.session.modified = True
    return removed_item


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_range_state(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    date_from_raw = str(request.POST.get("date_from", "")).strip()
    date_to_raw = str(request.POST.get("date_to", "")).strip()
    state_id = str(request.POST.get("special_state_id", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    dry_run = _is_truthy(request.POST.get("dry_run"))

    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not date_from_raw or not date_to_raw or not state_id:
        messages.error(request, "Debe seleccionar fecha inicial, fecha final y estado especial.")
        return redirect(redirect_url)
    try:
        date_from = date.fromisoformat(date_from_raw)
        date_to = date.fromisoformat(date_to_raw)
    except ValueError:
        messages.error(request, "Rango de fechas invalido.")
        return redirect(redirect_url)
    if date_from > date_to:
        messages.error(request, "La fecha inicial no puede ser mayor a la fecha final.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year and selected_month:
        if (
            date_from.year != selected_year
            or date_from.month != selected_month
            or date_to.year != selected_year
            or date_to.month != selected_month
        ):
            messages.error(request, "El rango debe pertenecer al mes seleccionado.")
            return redirect(redirect_url)

    special_state = SpecialState.objects.filter(id=state_id, tenant=tenant, property=property_obj, active=True).first()
    if special_state is None:
        messages.error(request, "Estado especial invalido.")
        return redirect(redirect_url)

    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=date_from,
        end_date=date_to,
    ):
        messages.error(request, "El rango incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )

    plans = {}
    total_days = (date_to - date_from).days + 1
    for worker in workers:
        if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            continue
        for offset in range(total_days):
            target_date = date_from + timedelta(days=offset)
            plans[(worker.id, target_date)] = {
                "worker": worker,
                "date": target_date,
                "shift": None,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/bulk-range-state/",
            label="Asignacion masiva por rango",
            summary=(
                f"Rango {date_from} a {date_to}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "special_state_id": state_id,
                "worker_ids": request.POST.getlist("worker_ids"),
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "bulk-range-state",
                "summary": (
                    f"Rango {date_from} a {date_to}: total={impact['total']}, crear={impact['to_create']}, "
                    f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa rango {date_from} a {date_to}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/bulk-range-state/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "special_state_id": state_id,
            "worker_ids": request.POST.getlist("worker_ids"),
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "bulk-range-state"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_range_state_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{date_from.isoformat()}:{date_to.isoformat()}",
        before={},
        after={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "special_state_id": special_state.id,
            "special_state_code": special_state.buk_code or "",
            "applied": applied,
            "impact": impact,
            "selected_workers": len(selected_worker_ids),
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios (sin trabajadores visibles con permiso).")
    else:
        messages.success(request, f"Estado especial aplicado en rango ({date_from} a {date_to}) con {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_multi_range_state(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    raw_from_values = request.POST.getlist("range_date_from")
    raw_to_values = request.POST.getlist("range_date_to")
    raw_state_values = request.POST.getlist("range_state_id")
    rows_count = max(len(raw_from_values), len(raw_to_values), len(raw_state_values))
    if rows_count == 0:
        messages.error(request, "Debe ingresar al menos un subrango.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    parsed_ranges = []
    for index in range(rows_count):
        from_raw = str(raw_from_values[index]).strip() if index < len(raw_from_values) else ""
        to_raw = str(raw_to_values[index]).strip() if index < len(raw_to_values) else ""
        state_raw = str(raw_state_values[index]).strip() if index < len(raw_state_values) else ""
        if not from_raw and not to_raw and not state_raw:
            continue
        if not from_raw or not to_raw or not state_raw:
            messages.error(request, f"Subrango #{index + 1} incompleto.")
            return redirect(redirect_url)
        try:
            date_from = date.fromisoformat(from_raw)
            date_to = date.fromisoformat(to_raw)
        except ValueError:
            messages.error(request, f"Subrango #{index + 1} con fechas invalidas.")
            return redirect(redirect_url)
        if date_from > date_to:
            messages.error(request, f"Subrango #{index + 1} con rango invalido (desde > hasta).")
            return redirect(redirect_url)
        if selected_year and selected_month:
            if (
                date_from.year != selected_year
                or date_from.month != selected_month
                or date_to.year != selected_year
                or date_to.month != selected_month
            ):
                messages.error(request, f"Subrango #{index + 1} fuera del mes seleccionado.")
                return redirect(redirect_url)
        if _is_any_month_closed(
            tenant=tenant,
            property_obj=property_obj,
            start_date=date_from,
            end_date=date_to,
        ):
            messages.error(request, f"Subrango #{index + 1} incluye un mes cerrado para esta sede.")
            return redirect(redirect_url)
        parsed_ranges.append(
            {
                "date_from": date_from,
                "date_to": date_to,
                "state_id": state_raw,
                "from_raw": from_raw,
                "to_raw": to_raw,
            }
        )

    if not parsed_ranges:
        messages.error(request, "Debe ingresar al menos un subrango valido.")
        return redirect(redirect_url)

    state_ids = [int(item["state_id"]) for item in parsed_ranges if str(item["state_id"]).isdigit()]
    if len(state_ids) != len(parsed_ranges):
        messages.error(request, "Uno o mas estados especiales son invalidos.")
        return redirect(redirect_url)

    states_map = {
        item.id: item
        for item in SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
            active=True,
        )
    }
    for index, item in enumerate(parsed_ranges):
        state_obj = states_map.get(int(item["state_id"]))
        if state_obj is None:
            messages.error(request, f"Subrango #{index + 1} con estado especial invalido.")
            return redirect(redirect_url)
        item["state"] = state_obj

    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para aplicar subrangos.")
        return redirect(redirect_url)

    plans = {}
    for item in parsed_ranges:
        total_days = (item["date_to"] - item["date_from"]).days + 1
        for worker in workers:
            for offset in range(total_days):
                target_date = item["date_from"] + timedelta(days=offset)
                plans[(worker.id, target_date)] = {
                    "worker": worker,
                    "date": target_date,
                    "shift": None,
                    "special_state": item["state"],
                }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    preview_fields = {
        "month": month_value,
        "area_id": area_value,
        "worker_ids": request.POST.getlist("worker_ids"),
        "range_date_from": [item["from_raw"] for item in parsed_ranges],
        "range_date_to": [item["to_raw"] for item in parsed_ranges],
        "range_state_id": [str(item["state"].id) for item in parsed_ranges],
    }

    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-bulk-multi-range-state"),
            label="Subrangos por estado",
            summary=(
                f"Subrangos={len(parsed_ranges)} total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields=preview_fields,
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "bulk-multi-range-state",
                "ranges_count": len(parsed_ranges),
                "impact": impact,
            },
        )
        messages.info(
            request,
            (
                f"Vista previa subrangos: {len(parsed_ranges)} rangos, total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1

    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-bulk-multi-range-state"),
        fields=preview_fields,
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "bulk-multi-range-state"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_multi_range_state_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{month_value or 'no-month'}",
        before={},
        after={
            "ranges_count": len(parsed_ranges),
            "ranges": [
                {
                    "date_from": item["date_from"].isoformat(),
                    "date_to": item["date_to"].isoformat(),
                    "special_state_id": item["state"].id,
                    "special_state_code": item["state"].buk_code or "",
                }
                for item in parsed_ranges
            ],
            "applied": applied,
            "impact": impact,
            "selected_workers": len(selected_worker_ids),
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios con los subrangos seleccionados.")
    else:
        messages.success(request, f"Subrangos aplicados con {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_sundays_state(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    state_id = str(request.POST.get("special_state_id", "")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not state_id:
        messages.error(request, "Debe seleccionar estado especial.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year is None or selected_month is None:
        messages.error(request, "Mes invalido.")
        return redirect(redirect_url)

    special_state = SpecialState.objects.filter(id=state_id, tenant=tenant, property=property_obj, active=True).first()
    if special_state is None:
        messages.error(request, "Estado especial invalido.")
        return redirect(redirect_url)

    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=selected_year,
        month=selected_month,
    ):
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    _, total_days = calendar.monthrange(selected_year, selected_month)
    sunday_dates = [
        date(selected_year, selected_month, day)
        for day in range(1, total_days + 1)
        if date(selected_year, selected_month, day).weekday() == 6
    ]

    workers = Worker.objects.select_related("area").filter(
        tenant=tenant,
        property=property_obj,
        active=True,
    )
    if area_value.isdigit():
        workers = workers.filter(area_id=int(area_value))

    plans = {}
    for worker in workers:
        if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            continue
        for sunday_date in sunday_dates:
            plans[(worker.id, sunday_date)] = {
                "worker": worker,
                "date": sunday_date,
                "shift": None,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/bulk-sundays-state/",
            label="OFF domingos",
            summary=(
                f"Domingos {selected_year:04d}-{selected_month:02d}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "special_state_id": state_id,
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "bulk-sundays-state",
                "summary": (
                    f"Domingos {selected_year:04d}-{selected_month:02d}: total={impact['total']}, "
                    f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa domingos {selected_year:04d}-{selected_month:02d}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/bulk-sundays-state/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "special_state_id": state_id,
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "bulk-sundays-state"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_sundays_state_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{selected_year:04d}-{selected_month:02d}",
        before={},
        after={
            "year": selected_year,
            "month": selected_month,
            "special_state_id": special_state.id,
            "special_state_code": special_state.buk_code or "",
            "sundays": len(sunday_dates),
            "applied": applied,
            "impact": impact,
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios (sin trabajadores visibles con permiso).")
    else:
        messages.success(request, f"Estado especial aplicado en {len(sunday_dates)} domingos ({applied} asignaciones).")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_bulk_week_pattern(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    date_from_raw = str(request.POST.get("date_from", "")).strip()
    date_to_raw = str(request.POST.get("date_to", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not date_from_raw or not date_to_raw:
        messages.error(request, "Debe seleccionar fecha inicial y fecha final.")
        return redirect(redirect_url)
    try:
        date_from = date.fromisoformat(date_from_raw)
        date_to = date.fromisoformat(date_to_raw)
    except ValueError:
        messages.error(request, "Rango de fechas invalido.")
        return redirect(redirect_url)
    if date_from > date_to:
        messages.error(request, "La fecha inicial no puede ser mayor a la fecha final.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year and selected_month:
        if (
            date_from.year != selected_year
            or date_from.month != selected_month
            or date_to.year != selected_year
            or date_to.month != selected_month
        ):
            messages.error(request, "El rango debe pertenecer al mes seleccionado.")
            return redirect(redirect_url)

    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=date_from,
        end_date=date_to,
    ):
        messages.error(request, "El rango incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    day_pattern, has_any_value = _extract_week_pattern_from_source(request.POST)
    if not has_any_value:
        messages.error(request, "Debe seleccionar al menos un turno/estado en el patron semanal.")
        return redirect(redirect_url)

    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para aplicar el patron.")
        return redirect(redirect_url)

    shifts_map = {
        item.id: item
        for item in Shift.objects.filter(tenant=tenant, property=property_obj, active=True).select_related("area")
    }
    states_map = {
        item.id: item
        for item in SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True)
    }

    plans = {}
    skipped_invalid = 0
    total_days = (date_to - date_from).days + 1
    for worker in workers:
        for offset in range(total_days):
            target_date = date_from + timedelta(days=offset)
            assignment_value = day_pattern.get(target_date.weekday(), "")
            shift, special_state, error = _resolve_assignment_value_for_worker(
                assignment_value=assignment_value,
                worker=worker,
                shifts_map=shifts_map,
                states_map=states_map,
            )
            if error == "empty":
                continue
            if error is not None:
                skipped_invalid += 1
                continue
            plans[(worker.id, target_date)] = {
                "worker": worker,
                "date": target_date,
                "shift": shift,
                "special_state": special_state,
            }
    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/bulk-week-pattern/",
            label="Patron semanal",
            summary=(
                f"Patron {date_from} a {date_to}: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "worker_ids": request.POST.getlist("worker_ids"),
                "monday_value": request.POST.get("monday_value", ""),
                "tuesday_value": request.POST.get("tuesday_value", ""),
                "wednesday_value": request.POST.get("wednesday_value", ""),
                "thursday_value": request.POST.get("thursday_value", ""),
                "friday_value": request.POST.get("friday_value", ""),
                "saturday_value": request.POST.get("saturday_value", ""),
                "sunday_value": request.POST.get("sunday_value", ""),
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "bulk-week-pattern",
                "summary": (
                    f"Patron {date_from} a {date_to}: total={impact['total']}, crear={impact['to_create']}, "
                    f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa patron: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/bulk-week-pattern/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "worker_ids": request.POST.getlist("worker_ids"),
            "monday_value": request.POST.get("monday_value", ""),
            "tuesday_value": request.POST.get("tuesday_value", ""),
            "wednesday_value": request.POST.get("wednesday_value", ""),
            "thursday_value": request.POST.get("thursday_value", ""),
            "friday_value": request.POST.get("friday_value", ""),
            "saturday_value": request.POST.get("saturday_value", ""),
            "sunday_value": request.POST.get("sunday_value", ""),
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "bulk-week-pattern"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_bulk_week_pattern_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{date_from.isoformat()}:{date_to.isoformat()}",
        before={},
        after={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "applied": applied,
            "skipped_invalid": skipped_invalid,
            "impact": impact,
            "selected_workers": len(selected_worker_ids),
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios con el patron seleccionado.")
    elif skipped_invalid:
        messages.success(
            request,
            f"Patron aplicado: {applied} asignaciones. Se omitieron {skipped_invalid} celdas por incompatibilidad de turno/area.",
        )
    else:
        messages.success(request, f"Patron aplicado: {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_operational_rule(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    date_from_raw = str(request.POST.get("date_from", "")).strip()
    date_to_raw = str(request.POST.get("date_to", "")).strip()
    weekday_value = str(request.POST.get("weekday_value", "")).strip()
    sunday_state_id = str(request.POST.get("sunday_state_id", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not date_from_raw or not date_to_raw:
        messages.error(request, "Debe seleccionar fecha inicial y fecha final.")
        return redirect(redirect_url)
    try:
        date_from = date.fromisoformat(date_from_raw)
        date_to = date.fromisoformat(date_to_raw)
    except ValueError:
        messages.error(request, "Rango de fechas invalido.")
        return redirect(redirect_url)
    if date_from > date_to:
        messages.error(request, "La fecha inicial no puede ser mayor a la fecha final.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year and selected_month:
        if (
            date_from.year != selected_year
            or date_from.month != selected_month
            or date_to.year != selected_year
            or date_to.month != selected_month
        ):
            messages.error(request, "El rango debe pertenecer al mes seleccionado.")
            return redirect(redirect_url)

    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=date_from,
        end_date=date_to,
    ):
        messages.error(request, "El rango incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    if not weekday_value:
        messages.error(request, "Debe seleccionar turno o estado para lunes-sabado.")
        return redirect(redirect_url)

    sunday_state = SpecialState.objects.filter(
        id=sunday_state_id,
        tenant=tenant,
        property=property_obj,
        active=True,
    ).first()
    if sunday_state is None:
        messages.error(request, "Debe seleccionar un estado especial valido para domingos.")
        return redirect(redirect_url)

    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para aplicar la regla.")
        return redirect(redirect_url)

    shifts_map = {
        item.id: item
        for item in Shift.objects.filter(tenant=tenant, property=property_obj, active=True).select_related("area")
    }
    states_map = {
        item.id: item
        for item in SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True)
    }

    plans = {}
    skipped_invalid = 0
    total_days = (date_to - date_from).days + 1
    for worker in workers:
        for offset in range(total_days):
            target_date = date_from + timedelta(days=offset)
            if target_date.weekday() == 6:
                shift = None
                special_state = sunday_state
            else:
                shift, special_state, error = _resolve_assignment_value_for_worker(
                    assignment_value=weekday_value,
                    worker=worker,
                    shifts_map=shifts_map,
                    states_map=states_map,
                )
                if error is not None:
                    skipped_invalid += 1
                    continue
            plans[(worker.id, target_date)] = {
                "worker": worker,
                "date": target_date,
                "shift": shift,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-operational-rule"),
            label="Regla operativa",
            summary=(
                f"Regla {date_from} a {date_to}: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "weekday_value": weekday_value,
                "sunday_state_id": sunday_state_id,
                "worker_ids": request.POST.getlist("worker_ids"),
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "operational-rule",
                "summary": (
                    f"Regla {date_from} a {date_to}: total={impact['total']}, crear={impact['to_create']}, "
                    f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa regla operativa: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-operational-rule"),
        fields={
            "month": month_value,
            "area_id": area_value,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "weekday_value": weekday_value,
            "sunday_state_id": sunday_state_id,
            "worker_ids": request.POST.getlist("worker_ids"),
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "operational-rule"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_operational_rule_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{date_from.isoformat()}:{date_to.isoformat()}",
        before={},
        after={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "weekday_value": weekday_value,
            "sunday_state_id": sunday_state.id,
            "applied": applied,
            "skipped_invalid": skipped_invalid,
            "impact": impact,
            "selected_workers": len(selected_worker_ids),
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if applied == 0:
        messages.error(request, "No se aplicaron cambios con la regla operativa.")
    elif skipped_invalid:
        messages.success(
            request,
            f"Regla operativa aplicada: {applied} asignaciones. Se omitieron {skipped_invalid} celdas por incompatibilidad.",
        )
    else:
        messages.success(request, f"Regla operativa aplicada: {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_save_range_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_name = str(request.POST.get("template_name", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_name:
        messages.error(request, "Debe ingresar nombre de plantilla.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year is None or selected_month is None:
        messages.error(request, "Selecciona un mes valido para guardar la plantilla.")
        return redirect(redirect_url)

    area_obj = None
    if area_value.isdigit():
        area_obj = Area.objects.filter(
            id=int(area_value),
            tenant=tenant,
            property=property_obj,
            active=True,
        ).first()
        if area_obj is None:
            messages.error(request, "Area invalida para guardar plantilla.")
            return redirect(redirect_url)

    range_rows = _parse_multi_range_rows(request.POST)
    parsed_ranges = []
    for row in range_rows:
        if not row["from_raw"] and not row["to_raw"] and not row["state_raw"]:
            continue
        if not row["from_raw"] or not row["to_raw"] or not row["state_raw"]:
            messages.error(request, f"Subrango #{row['index']} incompleto.")
            return redirect(redirect_url)
        try:
            from_date = date.fromisoformat(row["from_raw"])
            to_date = date.fromisoformat(row["to_raw"])
        except ValueError:
            messages.error(request, f"Subrango #{row['index']} con fechas invalidas.")
            return redirect(redirect_url)
        if from_date > to_date:
            messages.error(request, f"Subrango #{row['index']} invalido (desde > hasta).")
            return redirect(redirect_url)
        if (
            from_date.year != selected_year
            or from_date.month != selected_month
            or to_date.year != selected_year
            or to_date.month != selected_month
        ):
            messages.error(request, f"Subrango #{row['index']} fuera del mes seleccionado.")
            return redirect(redirect_url)
        if not row["state_raw"].isdigit():
            messages.error(request, f"Subrango #{row['index']} con estado invalido.")
            return redirect(redirect_url)
        parsed_ranges.append(
            {
                "start_day": from_date.day,
                "end_day": to_date.day,
                "special_state_id": int(row["state_raw"]),
            }
        )

    if not parsed_ranges:
        messages.error(request, "Debes completar al menos un subrango para guardar plantilla.")
        return redirect(redirect_url)

    state_ids = [item["special_state_id"] for item in parsed_ranges]
    valid_states = set(
        SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
            active=True,
        ).values_list("id", flat=True)
    )
    for index, item in enumerate(parsed_ranges):
        if item["special_state_id"] not in valid_states:
            messages.error(request, f"Subrango #{index + 1} con estado invalido.")
            return redirect(redirect_url)

    template, created = ScheduleRangeTemplate.objects.get_or_create(
        tenant=tenant,
        property=property_obj,
        area=area_obj,
        name=template_name,
        defaults={
            "ranges": parsed_ranges,
            "created_by": request.user,
            "updated_by": request.user,
        },
    )
    before = list(template.ranges or [])
    template.ranges = parsed_ranges
    template.active = True
    template.updated_by = request.user
    template.full_clean()
    template.save()

    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_save",
        entity_type="ScheduleRangeTemplate",
        entity_id=template.id,
        before=before,
        after=template.ranges,
    )
    if created:
        messages.success(request, f"Plantilla de subrangos guardada: {template.name}.")
    else:
        messages.success(request, f"Plantilla de subrangos actualizada: {template.name}.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_update_range_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    template_name = str(request.POST.get("template_name", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit() or not template_name:
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    before = {
        "name": template.name,
        "active": template.active,
        "ranges": template.ranges,
    }
    template.name = template_name
    template.active = bool(request.POST.get("active"))
    template.updated_by = request.user
    try:
        template.full_clean()
        template.save()
    except (IntegrityError, ValueError):
        messages.error(request, "No se pudo actualizar la plantilla (nombre duplicado o invalido).")
        return redirect(redirect_url)

    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_update",
        entity_type="ScheduleRangeTemplate",
        entity_id=template.id,
        before=before,
        after={
            "name": template.name,
            "active": template.active,
            "ranges": template.ranges,
        },
    )
    messages.success(request, "Plantilla de subrangos actualizada.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_update_range_template_ranges(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year is None or selected_month is None:
        messages.error(request, "Selecciona un mes valido para actualizar contenido.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    range_rows = _parse_multi_range_rows(request.POST)
    parsed_ranges = []
    for row in range_rows:
        if not row["from_raw"] and not row["to_raw"] and not row["state_raw"]:
            continue
        if not row["from_raw"] or not row["to_raw"] or not row["state_raw"]:
            messages.error(request, f"Subrango #{row['index']} incompleto.")
            return redirect(redirect_url)
        try:
            from_date = date.fromisoformat(row["from_raw"])
            to_date = date.fromisoformat(row["to_raw"])
        except ValueError:
            messages.error(request, f"Subrango #{row['index']} con fechas invalidas.")
            return redirect(redirect_url)
        if from_date > to_date:
            messages.error(request, f"Subrango #{row['index']} invalido (desde > hasta).")
            return redirect(redirect_url)
        if (
            from_date.year != selected_year
            or from_date.month != selected_month
            or to_date.year != selected_year
            or to_date.month != selected_month
        ):
            messages.error(request, f"Subrango #{row['index']} fuera del mes seleccionado.")
            return redirect(redirect_url)
        if not row["state_raw"].isdigit():
            messages.error(request, f"Subrango #{row['index']} con estado invalido.")
            return redirect(redirect_url)
        parsed_ranges.append(
            {
                "start_day": from_date.day,
                "end_day": to_date.day,
                "special_state_id": int(row["state_raw"]),
            }
        )
    if not parsed_ranges:
        messages.error(request, "Debes completar al menos un subrango para actualizar contenido.")
        return redirect(redirect_url)

    state_ids = [item["special_state_id"] for item in parsed_ranges]
    valid_states = set(
        SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
            active=True,
        ).values_list("id", flat=True)
    )
    for index, item in enumerate(parsed_ranges):
        if item["special_state_id"] not in valid_states:
            messages.error(request, f"Subrango #{index + 1} con estado invalido.")
            return redirect(redirect_url)

    before = list(template.ranges or [])
    template.ranges = parsed_ranges
    template.updated_by = request.user
    template.full_clean()
    template.save()
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_ranges_update",
        entity_type="ScheduleRangeTemplate",
        entity_id=template.id,
        before=before,
        after=template.ranges,
    )
    messages.success(request, "Contenido de plantilla de subrangos actualizado.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_create_range_template_version(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year is None or selected_month is None:
        messages.error(request, "Selecciona un mes valido para crear version.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    range_rows = _parse_multi_range_rows(request.POST)
    parsed_ranges = []
    for row in range_rows:
        if not row["from_raw"] and not row["to_raw"] and not row["state_raw"]:
            continue
        if not row["from_raw"] or not row["to_raw"] or not row["state_raw"]:
            messages.error(request, f"Subrango #{row['index']} incompleto.")
            return redirect(redirect_url)
        try:
            from_date = date.fromisoformat(row["from_raw"])
            to_date = date.fromisoformat(row["to_raw"])
        except ValueError:
            messages.error(request, f"Subrango #{row['index']} con fechas invalidas.")
            return redirect(redirect_url)
        if from_date > to_date:
            messages.error(request, f"Subrango #{row['index']} invalido (desde > hasta).")
            return redirect(redirect_url)
        if (
            from_date.year != selected_year
            or from_date.month != selected_month
            or to_date.year != selected_year
            or to_date.month != selected_month
        ):
            messages.error(request, f"Subrango #{row['index']} fuera del mes seleccionado.")
            return redirect(redirect_url)
        if not row["state_raw"].isdigit():
            messages.error(request, f"Subrango #{row['index']} con estado invalido.")
            return redirect(redirect_url)
        parsed_ranges.append(
            {
                "start_day": from_date.day,
                "end_day": to_date.day,
                "special_state_id": int(row["state_raw"]),
            }
        )
    if not parsed_ranges:
        messages.error(request, "Debes completar al menos un subrango para crear una version.")
        return redirect(redirect_url)

    state_ids = [item["special_state_id"] for item in parsed_ranges]
    valid_states = set(
        SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
            active=True,
        ).values_list("id", flat=True)
    )
    for index, item in enumerate(parsed_ranges):
        if item["special_state_id"] not in valid_states:
            messages.error(request, f"Subrango #{index + 1} con estado invalido.")
            return redirect(redirect_url)

    version_name = _next_range_template_version_name(
        tenant=tenant,
        property_obj=property_obj,
        area_obj=template.area,
        source_name=template.name,
    )
    new_template = ScheduleRangeTemplate.objects.create(
        tenant=tenant,
        property=property_obj,
        area=template.area,
        name=version_name,
        ranges=parsed_ranges,
        active=True,
        created_by=request.user,
        updated_by=request.user,
    )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_version_create",
        entity_type="ScheduleRangeTemplate",
        entity_id=new_template.id,
        before={
            "source_template_id": template.id,
            "source_name": template.name,
            "source_ranges": template.ranges,
        },
        after={
            "new_template_id": new_template.id,
            "new_name": new_template.name,
            "new_ranges": new_template.ranges,
        },
    )
    messages.success(request, f"Nueva version creada: {new_template.name}.")
    return redirect(f"{redirect_url}&range_template_edit_id={new_template.id}")


@login_required
@require_http_methods(["POST"])
def scheduling_clone_range_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    base_name = f"{template.name} (copia {timestamp})"
    clone_name = base_name
    suffix = 1
    while ScheduleRangeTemplate.objects.filter(
        tenant=tenant,
        property=property_obj,
        area=template.area,
        name=clone_name,
    ).exists():
        suffix += 1
        clone_name = f"{base_name}-{suffix}"

    clone = ScheduleRangeTemplate.objects.create(
        tenant=tenant,
        property=property_obj,
        area=template.area,
        name=clone_name,
        ranges=list(template.ranges or []),
        active=True,
        created_by=request.user,
        updated_by=request.user,
    )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_clone",
        entity_type="ScheduleRangeTemplate",
        entity_id=clone.id,
        before={
            "source_template_id": template.id,
            "source_name": template.name,
            "source_ranges": template.ranges,
        },
        after={
            "clone_template_id": clone.id,
            "clone_name": clone.name,
            "clone_ranges": clone.ranges,
        },
    )
    messages.success(request, f"Plantilla clonada para correccion: {clone.name}.")
    return redirect(f"{redirect_url}&range_template_edit_id={clone.id}")


@login_required
@require_http_methods(["POST"])
def scheduling_delete_range_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    before = {
        "name": template.name,
        "active": template.active,
        "ranges": template.ranges,
    }
    template_id_value = template.id
    template.delete()
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_delete",
        entity_type="ScheduleRangeTemplate",
        entity_id=template_id_value,
        before=before,
        after={},
    )
    messages.success(request, "Plantilla de subrangos eliminada.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_range_template_preview(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Selecciona una plantilla valida para ver detalle.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    preview_url = f"{redirect_url}&range_template_preview_id={template.id}"
    return redirect(preview_url)


@login_required
@require_http_methods(["POST"])
def scheduling_apply_range_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    allow_risky_template = _is_truthy(request.POST.get("allow_risky_template"))
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not template_id.isdigit():
        messages.error(request, "Debe seleccionar plantilla.")
        return redirect(redirect_url)
    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year is None or selected_month is None:
        messages.error(request, "Mes invalido.")
        return redirect(redirect_url)

    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=selected_year,
        month=selected_month,
    ):
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    template = ScheduleRangeTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
        active=True,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)
    if template.area_id and (not area_value.isdigit() or int(area_value) != template.area_id):
        messages.error(request, "Esta plantilla requiere filtrar por su area antes de aplicarla.")
        return redirect(redirect_url)

    template_ranges = list(template.ranges or [])
    if not template_ranges:
        messages.error(request, "La plantilla no tiene subrangos configurados.")
        return redirect(redirect_url)

    state_ids = []
    for item in template_ranges:
        state_id = item.get("special_state_id")
        if not isinstance(state_id, int):
            messages.error(request, "La plantilla contiene estados invalidos.")
            return redirect(redirect_url)
        state_ids.append(state_id)
    states_map_all = {
        item.id: item
        for item in SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
        )
    }
    template_health = _range_template_health(
        template=template,
        states_map=states_map_all,
        year=selected_year,
        month=selected_month,
    )
    is_admin = PermissionService.user_can_tenant_role(request.user, tenant, ["admin"])
    if template_health["level"] == "error" and not is_admin:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="schedule_range_template_override_denied" if allow_risky_template else "schedule_range_template_blocked",
            entity_type="ScheduleRangeTemplate",
            entity_id=template.id,
            before={"health": template_health},
            after={"allow_risky_template": bool(allow_risky_template)},
        )
    if template_health["level"] == "error" and not (is_admin and allow_risky_template):
        messages.error(
            request,
            "La plantilla tiene semaforo rojo. Corrigela o usa override de administrador.",
        )
        return redirect(redirect_url)
    if template_health["level"] == "error" and is_admin and allow_risky_template:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="schedule_range_template_override_requested",
            entity_type="ScheduleRangeTemplate",
            entity_id=template.id,
            before={"health": template_health},
            after={"allow_risky_template": True},
        )
    states_map = {
        item.id: item
        for item in SpecialState.objects.filter(
            id__in=state_ids,
            tenant=tenant,
            property=property_obj,
            active=True,
        )
    }

    _, total_days = calendar.monthrange(selected_year, selected_month)
    parsed_ranges = []
    skipped_invalid_ranges = 0
    for item in template_ranges:
        start_day = item.get("start_day")
        end_day = item.get("end_day")
        state_id = item.get("special_state_id")
        if not isinstance(start_day, int) or not isinstance(end_day, int) or start_day < 1 or end_day < 1 or start_day > end_day:
            skipped_invalid_ranges += 1
            continue
        if start_day > total_days:
            skipped_invalid_ranges += 1
            continue
        state = states_map.get(state_id)
        if state is None:
            skipped_invalid_ranges += 1
            continue
        parsed_ranges.append(
            {
                "date_from": date(selected_year, selected_month, start_day),
                "date_to": date(selected_year, selected_month, min(end_day, total_days)),
                "state": state,
            }
        )
    if not parsed_ranges:
        messages.error(request, "La plantilla no contiene subrangos aplicables al mes seleccionado.")
        return redirect(redirect_url)

    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para aplicar la plantilla.")
        return redirect(redirect_url)

    plans = {}
    for item in parsed_ranges:
        total_range_days = (item["date_to"] - item["date_from"]).days + 1
        for worker in workers:
            for offset in range(total_range_days):
                target_date = item["date_from"] + timedelta(days=offset)
                plans[(worker.id, target_date)] = {
                    "worker": worker,
                    "date": target_date,
                    "shift": None,
                    "special_state": item["state"],
                }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    preview_fields = {
        "month": month_value,
        "area_id": area_value,
        "template_id": template_id,
        "worker_ids": request.POST.getlist("worker_ids"),
        "allow_risky_template": "1" if allow_risky_template else "",
    }
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-apply-range-template"),
            label="Aplicar plantilla subrangos",
            summary=(
                f"Plantilla {template.name}: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid_ranges}"
            ),
            fields=preview_fields,
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "apply-range-template",
                "template_id": template.id,
                "impact": impact,
                "skipped_invalid_ranges": skipped_invalid_ranges,
            },
        )
        messages.info(
            request,
            (
                f"Vista previa plantilla subrangos: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid_ranges}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-apply-range-template"),
        fields=preview_fields,
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "apply-range-template"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_range_template_apply",
        entity_type="ScheduleRangeTemplate",
        entity_id=template.id,
        before={},
        after={
            "month": month_value,
            "applied": applied,
            "skipped_invalid_ranges": skipped_invalid_ranges,
            "impact": impact,
            "selected_workers": len(selected_worker_ids),
            "area_filter": int(area_value) if area_value.isdigit() else None,
            "allow_risky_template": bool(allow_risky_template),
        },
    )
    if applied == 0:
        messages.error(request, "No se aplicaron cambios con la plantilla de subrangos.")
    else:
        messages.success(request, f"Plantilla de subrangos aplicada: {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_save_week_pattern_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_name = str(request.POST.get("template_name", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_name:
        messages.error(request, "Debe ingresar nombre de plantilla.")
        return redirect(redirect_url)

    area_obj = None
    if area_value.isdigit():
        area_obj = Area.objects.filter(
            id=int(area_value),
            tenant=tenant,
            property=property_obj,
            active=True,
        ).first()
        if area_obj is None:
            messages.error(request, "Area invalida para guardar plantilla.")
            return redirect(redirect_url)

    day_pattern, has_any_value = _extract_week_pattern_from_source(request.POST)
    if not has_any_value:
        messages.error(request, "Debe seleccionar al menos un turno/estado para guardar plantilla.")
        return redirect(redirect_url)

    template, created = SchedulePatternTemplate.objects.get_or_create(
        tenant=tenant,
        property=property_obj,
        area=area_obj,
        name=template_name,
        defaults={
            "pattern": _serialize_week_pattern(day_pattern),
            "created_by": request.user,
            "updated_by": request.user,
        },
    )
    before = _serialize_week_pattern(template.pattern)
    template.pattern = _serialize_week_pattern(day_pattern)
    template.updated_by = request.user
    template.active = True
    template.full_clean()
    template.save()
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_pattern_template_save",
        entity_type="SchedulePatternTemplate",
        entity_id=template.id,
        before=before,
        after=template.pattern,
    )
    if created:
        messages.success(request, f"Plantilla guardada: {template.name}.")
    else:
        messages.success(request, f"Plantilla actualizada: {template.name}.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_update_week_pattern_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    template_name = str(request.POST.get("template_name", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not template_id.isdigit() or not template_name:
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)
    template = SchedulePatternTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    before = {
        "name": template.name,
        "active": template.active,
        "pattern": template.pattern,
    }
    template.name = template_name
    template.active = bool(request.POST.get("active"))
    template.updated_by = request.user
    try:
        template.full_clean()
        template.save()
    except (IntegrityError, ValueError):
        messages.error(request, "No se pudo actualizar la plantilla (nombre duplicado o invalido).")
        return redirect(redirect_url)

    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_pattern_template_update",
        entity_type="SchedulePatternTemplate",
        entity_id=template.id,
        before=before,
        after={
            "name": template.name,
            "active": template.active,
            "pattern": template.pattern,
        },
    )
    messages.success(request, "Plantilla actualizada.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_delete_week_pattern_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if not template_id.isdigit():
        messages.error(request, "Plantilla invalida.")
        return redirect(redirect_url)

    template = SchedulePatternTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    before = {
        "name": template.name,
        "active": template.active,
        "pattern": template.pattern,
    }
    template_id_value = template.id
    template.delete()
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_pattern_template_delete",
        entity_type="SchedulePatternTemplate",
        entity_id=template_id_value,
        before=before,
        after={},
    )
    messages.success(request, "Plantilla eliminada.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_apply_week_pattern_template(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    date_from_raw = str(request.POST.get("date_from", "")).strip()
    date_to_raw = str(request.POST.get("date_to", "")).strip()
    template_id = str(request.POST.get("template_id", "")).strip()
    selected_worker_ids = _parse_selected_worker_ids(request.POST.getlist("worker_ids"))
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)

    if not template_id.isdigit():
        messages.error(request, "Debe seleccionar plantilla.")
        return redirect(redirect_url)
    if not date_from_raw or not date_to_raw:
        messages.error(request, "Debe seleccionar fecha inicial y fecha final.")
        return redirect(redirect_url)
    try:
        date_from = date.fromisoformat(date_from_raw)
        date_to = date.fromisoformat(date_to_raw)
    except ValueError:
        messages.error(request, "Rango de fechas invalido.")
        return redirect(redirect_url)
    if date_from > date_to:
        messages.error(request, "La fecha inicial no puede ser mayor a la fecha final.")
        return redirect(redirect_url)

    selected_year, selected_month = _parse_month_or_none(month_value)
    if selected_year and selected_month:
        if (
            date_from.year != selected_year
            or date_from.month != selected_month
            or date_to.year != selected_year
            or date_to.month != selected_month
        ):
            messages.error(request, "El rango debe pertenecer al mes seleccionado.")
            return redirect(redirect_url)

    template = SchedulePatternTemplate.objects.filter(
        id=int(template_id),
        tenant=tenant,
        property=property_obj,
        active=True,
    ).first()
    if template is None:
        messages.error(request, "Plantilla no encontrada.")
        return redirect(redirect_url)

    if template.area_id and (not area_value.isdigit() or int(area_value) != template.area_id):
        messages.error(request, "Esta plantilla requiere filtrar por su area antes de aplicarla.")
        return redirect(redirect_url)

    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=date_from,
        end_date=date_to,
    ):
        messages.error(request, "El rango incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    day_pattern = _deserialize_week_pattern(template.pattern)
    workers = _get_target_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
        selected_worker_ids=selected_worker_ids,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para aplicar la plantilla.")
        return redirect(redirect_url)

    shifts_map = {
        item.id: item
        for item in Shift.objects.filter(tenant=tenant, property=property_obj, active=True).select_related("area")
    }
    states_map = {
        item.id: item
        for item in SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True)
    }

    plans = {}
    skipped_invalid = 0
    total_days = (date_to - date_from).days + 1
    for worker in workers:
        for offset in range(total_days):
            target_date = date_from + timedelta(days=offset)
            assignment_value = day_pattern.get(target_date.weekday(), "")
            shift, special_state, error = _resolve_assignment_value_for_worker(
                assignment_value=assignment_value,
                worker=worker,
                shifts_map=shifts_map,
                states_map=states_map,
            )
            if error == "empty":
                continue
            if error is not None:
                skipped_invalid += 1
                continue
            plans[(worker.id, target_date)] = {
                "worker": worker,
                "date": target_date,
                "shift": shift,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/apply-week-pattern-template/",
            label="Aplicar plantilla semanal",
            summary=(
                f"Plantilla {template.name}: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "template_id": template_id,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "worker_ids": request.POST.getlist("worker_ids"),
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "apply-week-pattern-template",
                "summary": (
                    f"Plantilla {template.name}: total={impact['total']}, crear={impact['to_create']}, "
                    f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa plantilla: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}, omitidas={skipped_invalid}."
            ),
        )
        return redirect(redirect_url)

    applied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        applied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/apply-week-pattern-template/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "template_id": template_id,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "worker_ids": request.POST.getlist("worker_ids"),
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "apply-week-pattern-template"},
        )

    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="schedule_pattern_template_apply",
        entity_type="SchedulePatternTemplate",
        entity_id=template.id,
        before={},
        after={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "applied": applied,
            "skipped_invalid": skipped_invalid,
            "impact": impact,
        },
    )
    if applied == 0:
        messages.error(request, "No se aplicaron cambios con la plantilla seleccionada.")
    elif skipped_invalid:
        messages.success(
            request,
            f"Plantilla aplicada: {applied} asignaciones. Se omitieron {skipped_invalid} celdas por incompatibilidad de turno/area.",
        )
    else:
        messages.success(request, f"Plantilla aplicada: {applied} asignaciones.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_copy_week(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    source_start_raw = str(request.POST.get("source_week_start", "")).strip()
    target_start_raw = str(request.POST.get("target_week_start", "")).strip()
    copy_kind = str(request.POST.get("copy_kind", "all")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if copy_kind not in {"all", "shift", "state"}:
        copy_kind = "all"

    if not source_start_raw or not target_start_raw:
        messages.error(request, "Debe seleccionar semana origen y semana destino.")
        return redirect(redirect_url)
    try:
        source_start = date.fromisoformat(source_start_raw)
        target_start = date.fromisoformat(target_start_raw)
    except ValueError:
        messages.error(request, "Semana origen/destino invalida.")
        return redirect(redirect_url)

    source_end = source_start + timedelta(days=6)
    target_end = target_start + timedelta(days=6)
    if _is_any_month_closed(
        tenant=tenant,
        property_obj=property_obj,
        start_date=target_start,
        end_date=target_end,
    ):
        messages.error(request, "La semana destino incluye un mes cerrado para esta sede.")
        return redirect(redirect_url)

    workers = _get_visible_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para copiar.")
        return redirect(redirect_url)

    worker_ids = [worker.id for worker in workers]
    source_assignments = ScheduleAssignment.objects.select_related("shift", "special_state").filter(
        tenant=tenant,
        property=property_obj,
        worker_id__in=worker_ids,
        date__gte=source_start,
        date__lte=source_end,
    )
    source_index = {(item.worker_id, item.date): item for item in source_assignments}

    plans = {}
    for worker in workers:
        for offset in range(7):
            from_date = source_start + timedelta(days=offset)
            to_date = target_start + timedelta(days=offset)
            shift, special_state = _resolve_copy_payload(
                source_assignment=source_index.get((worker.id, from_date)),
                copy_kind=copy_kind,
            )
            if shift is None and special_state is None:
                continue
            plans[(worker.id, to_date)] = {
                "worker": worker,
                "date": to_date,
                "shift": shift,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/copy-week/",
            label="Copiar semana",
            summary=(
                f"Semana {source_start} -> {target_start}: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "source_week_start": source_start_raw,
                "target_week_start": target_start_raw,
                "copy_kind": copy_kind,
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "copy-week",
                "summary": (
                    f"Semana {source_start} -> {target_start}: total={impact['total']}, crear={impact['to_create']}, "
                    f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa copia semana: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    copied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        copied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/copy-week/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "source_week_start": source_start_raw,
            "target_week_start": target_start_raw,
            "copy_kind": copy_kind,
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "copy-week"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_copy_week_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{source_start.isoformat()}:{target_start.isoformat()}",
        before={},
        after={
            "source_week_start": source_start.isoformat(),
            "target_week_start": target_start.isoformat(),
            "copy_kind": copy_kind,
            "copied": copied,
            "impact": impact,
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if copied == 0:
        messages.error(request, "No se encontraron asignaciones para copiar en la semana origen.")
    else:
        messages.success(request, f"Semana copiada: {copied} asignaciones actualizadas.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_copy_month(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    source_month_value = str(request.POST.get("source_month", "")).strip()
    copy_kind = str(request.POST.get("copy_kind", "all")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if copy_kind not in {"all", "shift", "state"}:
        copy_kind = "all"

    target_year, target_month = _parse_month_or_none(month_value)
    if target_year is None or target_month is None:
        messages.error(request, "Mes destino invalido.")
        return redirect(redirect_url)
    source_year, source_month = _parse_month_or_none(source_month_value)
    if source_year is None or source_month is None:
        messages.error(request, "Mes origen invalido.")
        return redirect(redirect_url)
    if source_year == target_year and source_month == target_month:
        messages.error(request, "El mes origen debe ser distinto al mes destino.")
        return redirect(redirect_url)
    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=target_year,
        month=target_month,
    ):
        messages.error(request, "El mes destino esta cerrado para esta sede.")
        return redirect(redirect_url)

    _, target_days = calendar.monthrange(target_year, target_month)
    _, source_days = calendar.monthrange(source_year, source_month)

    workers = _get_visible_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para copiar.")
        return redirect(redirect_url)

    worker_ids = [worker.id for worker in workers]
    source_start = date(source_year, source_month, 1)
    source_end = date(source_year, source_month, source_days)
    source_assignments = ScheduleAssignment.objects.select_related("shift", "special_state").filter(
        tenant=tenant,
        property=property_obj,
        worker_id__in=worker_ids,
        date__gte=source_start,
        date__lte=source_end,
    )
    source_index = {(item.worker_id, item.date): item for item in source_assignments}

    plans = {}
    for worker in workers:
        max_day = min(target_days, source_days)
        for day in range(1, max_day + 1):
            from_date = date(source_year, source_month, day)
            to_date = date(target_year, target_month, day)
            shift, special_state = _resolve_copy_payload(
                source_assignment=source_index.get((worker.id, from_date)),
                copy_kind=copy_kind,
            )
            if shift is None and special_state is None:
                continue
            plans[(worker.id, to_date)] = {
                "worker": worker,
                "date": to_date,
                "shift": shift,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url=reverse("webui-scheduling-copy-month"),
            label="Copiar desde mes origen",
            summary=(
                f"Mes {source_year:04d}-{source_month:02d} -> {target_year:04d}-{target_month:02d}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "source_month": source_month_value,
                "copy_kind": copy_kind,
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "copy-month",
                "summary": (
                    f"Mes {source_year:04d}-{source_month:02d} -> {target_year:04d}-{target_month:02d}: total={impact['total']}, "
                    f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa copia mes: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    copied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        copied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url=reverse("webui-scheduling-copy-month"),
        fields={
            "month": month_value,
            "area_id": area_value,
            "source_month": source_month_value,
            "copy_kind": copy_kind,
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "copy-month"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_copy_month_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{target_year:04d}-{target_month:02d}",
        before={},
        after={
            "target_year": target_year,
            "target_month": target_month,
            "source_year": source_year,
            "source_month": source_month,
            "copy_kind": copy_kind,
            "copied": copied,
            "impact": impact,
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if copied == 0:
        messages.error(request, "No se encontraron asignaciones para copiar en el mes origen.")
    else:
        messages.success(request, f"Mes copiado: {copied} asignaciones actualizadas.")
    return redirect(redirect_url)


@login_required
@require_http_methods(["POST"])
def scheduling_copy_previous_month(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        messages.error(request, ctx["context_error"])
        return redirect("webui-scheduling")

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "scheduling"):
        return HttpResponseForbidden("Modulo desactivado: scheduling.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
        return HttpResponseForbidden("No tienes permisos para asignar en esta sede.")

    month_value = str(request.POST.get("month", "")).strip()
    area_value = str(request.POST.get("area_id", "")).strip()
    copy_kind = str(request.POST.get("copy_kind", "all")).strip()
    dry_run = _is_truthy(request.POST.get("dry_run"))
    redirect_url = _build_scheduling_redirect_url_from_post(request, month_value, area_value)
    if copy_kind not in {"all", "shift", "state"}:
        copy_kind = "all"

    year, month = _parse_month_or_none(month_value)
    if year is None or month is None:
        messages.error(request, "Mes invalido.")
        return redirect(redirect_url)
    if MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=year,
        month=month,
    ):
        messages.error(request, "El mes esta cerrado para esta sede.")
        return redirect(redirect_url)

    first_day = date(year, month, 1)
    prev_month_last_day = first_day - timedelta(days=1)
    prev_year = prev_month_last_day.year
    prev_month = prev_month_last_day.month
    _, target_days = calendar.monthrange(year, month)
    _, source_days = calendar.monthrange(prev_year, prev_month)

    workers = _get_visible_workers_for_scheduling(
        request=request,
        tenant=tenant,
        property_obj=property_obj,
        area_value=area_value,
    )
    if not workers:
        messages.error(request, "No hay trabajadores visibles con permiso para copiar.")
        return redirect(redirect_url)

    worker_ids = [worker.id for worker in workers]
    source_start = date(prev_year, prev_month, 1)
    source_end = date(prev_year, prev_month, source_days)
    source_assignments = ScheduleAssignment.objects.select_related("shift", "special_state").filter(
        tenant=tenant,
        property=property_obj,
        worker_id__in=worker_ids,
        date__gte=source_start,
        date__lte=source_end,
    )
    source_index = {(item.worker_id, item.date): item for item in source_assignments}

    plans = {}
    for worker in workers:
        max_day = min(target_days, source_days)
        for day in range(1, max_day + 1):
            from_date = date(prev_year, prev_month, day)
            to_date = date(year, month, day)
            shift, special_state = _resolve_copy_payload(
                source_assignment=source_index.get((worker.id, from_date)),
                copy_kind=copy_kind,
            )
            if shift is None and special_state is None:
                continue
            plans[(worker.id, to_date)] = {
                "worker": worker,
                "date": to_date,
                "shift": shift,
                "special_state": special_state,
            }

    impact = _summarize_assignment_plans(tenant=tenant, property_obj=property_obj, plans=plans)
    if dry_run:
        preview_id = _queue_scheduling_preview(
            request=request,
            action_url="/app/scheduling/copy-previous-month/",
            label="Copiar mes anterior",
            summary=(
                f"Mes {prev_year:04d}-{prev_month:02d} -> {year:04d}-{month:02d}: total={impact['total']}, "
                f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
            ),
            fields={
                "month": month_value,
                "area_id": area_value,
                "copy_kind": copy_kind,
            },
        )
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_created",
            entity_type="SchedulingPreview",
            entity_id=preview_id,
            before={},
            after={
                "operation": "copy-previous-month",
                "summary": (
                    f"Mes {prev_year:04d}-{prev_month:02d} -> {year:04d}-{month:02d}: total={impact['total']}, "
                    f"crear={impact['to_create']}, actualizar={impact['to_update']}, sin cambios={impact['unchanged']}"
                ),
            },
        )
        messages.info(
            request,
            (
                f"Vista previa copia mes anterior: total={impact['total']}, crear={impact['to_create']}, "
                f"actualizar={impact['to_update']}, sin cambios={impact['unchanged']}."
            ),
        )
        return redirect(redirect_url)

    copied = 0
    for plan in plans.values():
        ScheduleAssignmentService.upsert_assignment(
            tenant=tenant,
            property_obj=property_obj,
            worker=plan["worker"],
            date=plan["date"],
            shift=plan["shift"],
            special_state=plan["special_state"],
            user=request.user,
        )
        copied += 1
    consumed_preview = _consume_scheduling_preview(
        request=request,
        action_url="/app/scheduling/copy-previous-month/",
        fields={
            "month": month_value,
            "area_id": area_value,
            "copy_kind": copy_kind,
        },
    )
    if consumed_preview is not None:
        AuditService.log(
            tenant=tenant,
            property_obj=property_obj,
            user=request.user,
            action="scheduling_preview_confirmed",
            entity_type="SchedulingPreview",
            entity_id=consumed_preview.get("id", "unknown"),
            before=consumed_preview,
            after={"operation": "copy-previous-month"},
        )
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="scheduling_copy_previous_month_apply",
        entity_type="ScheduleAssignment",
        entity_id=f"{property_obj.id}:{year:04d}-{month:02d}",
        before={},
        after={
            "target_year": year,
            "target_month": month,
            "source_year": prev_year,
            "source_month": prev_month,
            "copy_kind": copy_kind,
            "copied": copied,
            "impact": impact,
            "area_filter": int(area_value) if area_value.isdigit() else None,
        },
    )

    if copied == 0:
        messages.error(request, "No se encontraron asignaciones para copiar en el mes anterior.")
    else:
        messages.success(request, f"Mes anterior copiado: {copied} asignaciones actualizadas.")
    return redirect(redirect_url)


def _parse_date_or_default(raw_value, default_value):
    try:
        return date.fromisoformat(str(raw_value).strip())
    except ValueError:
        return default_value


@login_required
@require_http_methods(["GET", "POST"])
def buk_report_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/buk_report.html", {**ctx, "rows": [], "issues": [], "date_columns": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    property_id_raw = request.POST.get("property_id") if request.method == "POST" else request.GET.get("property_id")
    if property_id_raw and str(property_id_raw).isdigit():
        requested_property = next(
            (item for item in ctx.get("property_options", []) if item.id == int(property_id_raw)),
            None,
        )
        if requested_property is not None:
            property_obj = requested_property
            ctx["selected_property"] = requested_property
            request.session["ui_property_id"] = requested_property.id
    can_export_report = PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_export_buk"
    )
    can_view_report = can_export_report or PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_view_reports"
    )
    if not can_view_report:
        return render(
            request,
            "webui/buk_report.html",
            {
                **ctx,
                "rows": [],
                "issues": [],
                "date_columns": [],
                "can_export_buk": False,
                "context_error": "No tienes permisos para ver reporte BUK en esta sede.",
            },
        )

    module_preview = PermissionService.user_can_module(request.user, tenant, "buk_preview")
    module_validator = PermissionService.user_can_module(request.user, tenant, "buk_validator")
    module_export = PermissionService.user_can_module(request.user, tenant, "buk_export")
    is_admin_export = PermissionService.user_can_tenant_role(request.user, tenant, ["admin"])
    if not module_preview:
        return render(
            request,
            "webui/buk_report.html",
            {
                **ctx,
                "rows": [],
                "issues": [],
                "date_columns": [],
                "can_export_buk": can_export_report,
                "context_error": "Modulo desactivado: buk_preview.",
                "module_export_enabled": module_export,
                "is_admin_export": is_admin_export,
            },
        )

    today = timezone.localdate()
    month_start = date(today.year, today.month, 1)
    next_month_anchor = month_start + timedelta(days=32)
    next_month_start = date(next_month_anchor.year, next_month_anchor.month, 1)
    month_end = next_month_start - timedelta(days=1)

    source = request.POST if request.method == "POST" else request.GET
    selected_area_ids = []
    for raw in source.getlist("area_ids"):
        raw_text = str(raw).strip()
        if raw_text.isdigit():
            selected_area_ids.append(int(raw_text))
    selected_worker_ids = []
    for raw in source.getlist("worker_ids"):
        raw_text = str(raw).strip()
        if raw_text.isdigit():
            selected_worker_ids.append(int(raw_text))
    area_options = []
    for area in Area.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name"):
        if PermissionService.user_can_area_view(request.user, tenant, property_obj, area):
            area_options.append(area)
    if not area_options:
        return render(
            request,
            "webui/buk_report.html",
            {
                **ctx,
                "rows": [],
                "issues": [],
                "date_columns": [],
                "can_export_buk": can_export_report,
                "context_error": "No tienes areas autorizadas para visualizar el reporte BUK.",
                "module_export_enabled": module_export,
                "module_validator_enabled": module_validator,
                "is_admin_export": is_admin_export,
            },
        )
    valid_area_ids = {item.id for item in area_options}
    selected_area_ids = [area_id for area_id in selected_area_ids if area_id in valid_area_ids]
    effective_area_ids = selected_area_ids if selected_area_ids else list(valid_area_ids)
    worker_options_qs = Worker.objects.filter(tenant=tenant, property=property_obj, active=True).select_related("area")
    worker_options_qs = worker_options_qs.filter(area_id__in=effective_area_ids)
    worker_options = list(worker_options_qs.order_by("last_name", "first_name"))
    valid_worker_ids = {item.id for item in worker_options}
    selected_worker_ids = [worker_id for worker_id in selected_worker_ids if worker_id in valid_worker_ids]

    date_from = _parse_date_or_default(source.get("date_from", ""), month_start)
    date_to = _parse_date_or_default(source.get("date_to", ""), month_end)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    date_columns = []
    cursor = date_from
    while cursor <= date_to:
        date_columns.append({"iso": cursor.isoformat(), "label": cursor.strftime("%d-%m-%Y")})
        cursor += timedelta(days=1)

    issues = []
    if module_validator:
        issues = BukValidationService.validate_assignments(
            tenant=tenant,
            property_obj=property_obj,
            date_from=date_from,
            date_to=date_to,
            area_ids=effective_area_ids,
            worker_ids=selected_worker_ids,
        )
    error_count = sum(1 for issue in issues if issue.severity == "error")
    warning_count = sum(1 for issue in issues if issue.severity == "warning")
    info_count = sum(1 for issue in issues if issue.severity == "info")
    template_compare_result = None

    if request.method == "POST" and request.POST.get("action") == "export":
        if not can_export_report:
            return HttpResponseForbidden("No tienes permisos para exportar reporte BUK.")
        if not module_export:
            messages.error(request, "Modulo desactivado: buk_export.")
            return redirect(f"/app/buk-report/?date_from={date_from.isoformat()}&date_to={date_to.isoformat()}")

        export_with_observations = str(request.POST.get("export_with_observations", "")).strip().lower() in {
            "1",
            "true",
            "on",
            "yes",
        }
        blocking_errors = [issue for issue in issues if issue.severity == "error"]
        if blocking_errors and not export_with_observations:
            messages.error(request, "No se puede exportar porque existen errores bloqueantes.")
        elif blocking_errors and export_with_observations and not is_admin_export:
            messages.error(request, "Solo administradores pueden exportar con observaciones.")
        else:
            output_format = str(request.POST.get("format", "xlsx")).strip().lower()
            file_base = f"buk_{property_obj.slug}_{date_from.isoformat()}_{date_to.isoformat()}"
            if output_format == "csv":
                content = BukExportService.generate_csv_text(
                    tenant=tenant,
                    property_obj=property_obj,
                    date_from=date_from,
                    date_to=date_to,
                    area_ids=effective_area_ids,
                    worker_ids=selected_worker_ids,
                )
                response = HttpResponse(content, content_type="text/csv; charset=utf-8")
                response["Content-Disposition"] = f'attachment; filename="{file_base}.csv"'
                file_name = f"{file_base}.csv"
            else:
                content = BukExportService.generate_xlsx_bytes(
                    tenant=tenant,
                    property_obj=property_obj,
                    date_from=date_from,
                    date_to=date_to,
                    area_ids=effective_area_ids,
                    worker_ids=selected_worker_ids,
                )
                response = HttpResponse(
                    content,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{file_base}.xlsx"'
                file_name = f"{file_base}.xlsx"

            log = BukExportService.log_export(
                tenant=tenant,
                property_obj=property_obj,
                date_from=date_from,
                date_to=date_to,
                generated_by=request.user if request.user.is_authenticated else None,
                file_name=file_name,
                validation_issues=issues,
                export_with_observations=export_with_observations,
            )
            response["X-Buk-Export-Log-Id"] = str(log.id)
            return response
    elif request.method == "POST" and request.POST.get("action") == "compare_template":
        if not can_export_report:
            return HttpResponseForbidden("No tienes permisos para comparar/exportar reporte BUK.")
        reference_file = request.FILES.get("reference_file")
        if not reference_file:
            messages.error(request, "Debes seleccionar un archivo XLSX de referencia.")
        else:
            download_compare_json = str(request.POST.get("download_compare_json", "")).strip().lower() in {
                "1",
                "true",
                "on",
                "yes",
            }
            reference_file_bytes = reference_file.read()
            candidate_xlsx = BukExportService.generate_xlsx_bytes(
                tenant=tenant,
                property_obj=property_obj,
                date_from=date_from,
                date_to=date_to,
                area_ids=effective_area_ids,
                worker_ids=selected_worker_ids,
            )
            try:
                template_compare_result = BukExportService.compare_template_compatibility(
                    reference_file_bytes=reference_file_bytes,
                    candidate_file_bytes=candidate_xlsx,
                )
            except Exception:
                messages.error(request, "No se pudo comparar la plantilla. Verifica que el archivo sea XLSX valido.")
            else:
                compare_log = BukExportService.log_template_compare(
                    tenant=tenant,
                    property_obj=property_obj,
                    compared_by=request.user if request.user.is_authenticated else None,
                    date_from=date_from,
                    date_to=date_to,
                    sheet_name="Reporte carga BUK",
                    reference_file_name=getattr(reference_file, "name", "") or "",
                    reference_file_bytes=reference_file_bytes,
                    result=template_compare_result,
                )
                template_compare_result["compare_log_id"] = compare_log.id
                if download_compare_json:
                    report_payload = {
                        "tenant": tenant.slug,
                        "property": property_obj.slug,
                        "date_from": date_from.isoformat(),
                        "date_to": date_to.isoformat(),
                        "sheet_name": "Reporte carga BUK",
                        "result": template_compare_result,
                    }
                    file_name = (
                        f"buk_template_compare_{property_obj.slug}_{date_from.isoformat()}_{date_to.isoformat()}.json"
                    )
                    response = HttpResponse(
                        json.dumps(report_payload, ensure_ascii=False, indent=2),
                        content_type="application/json; charset=utf-8",
                    )
                    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
                    return response
                if template_compare_result["is_compatible"]:
                    messages.success(request, "Comparacion completada: estructura compatible con la referencia.")
                else:
                    messages.error(request, "Comparacion completada: se detectaron diferencias estructurales.")

    rows = BukExportService.build_preview_rows(
        tenant=tenant,
        property_obj=property_obj,
        date_from=date_from,
        date_to=date_to,
        area_ids=effective_area_ids,
        worker_ids=selected_worker_ids,
    )
    valid_codes = set(
        Shift.objects.filter(tenant=tenant, property=property_obj, active=True)
        .exclude(buk_code="")
        .values_list("buk_code", flat=True)
    )
    valid_state_codes = set(
        SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True)
        .exclude(buk_code="")
        .values_list("buk_code", flat=True)
    )
    valid_state_names = set(
        SpecialState.objects.filter(tenant=tenant, property=property_obj, active=True)
        .exclude(name="")
        .values_list("name", flat=True)
    )
    valid_codes = valid_codes.union(valid_state_codes).union(valid_state_names)

    for row in rows:
        day_cells = []
        for day in date_columns:
            value = row["days"].get(day["iso"], "")
            is_empty = value == ""
            is_invalid = (not is_empty) and (value not in valid_codes)
            day_cells.append(
                {
                    "value": value,
                    "is_empty": is_empty,
                    "is_invalid": is_invalid,
                }
            )
        row["day_cells"] = day_cells
    has_blocking_errors = any(issue.severity == "error" for issue in issues)

    compare_result_filter = str(request.GET.get("compare_result", "")).strip().lower()
    compare_user_filter = str(request.GET.get("compare_user", "")).strip()
    compare_from_raw = str(request.GET.get("compare_from", "")).strip()
    compare_to_raw = str(request.GET.get("compare_to", "")).strip()
    compare_from = _parse_date_or_default(compare_from_raw, None) if compare_from_raw else None
    compare_to = _parse_date_or_default(compare_to_raw, None) if compare_to_raw else None
    compare_page_size_raw = str(request.GET.get("compare_page_size", "")).strip()
    if compare_page_size_raw.isdigit():
        compare_page_size = int(compare_page_size_raw)
    else:
        compare_limit_raw = str(request.GET.get("compare_limit", "10")).strip()
        compare_page_size = int(compare_limit_raw) if compare_limit_raw.isdigit() else 10
    compare_page_size = max(1, min(compare_page_size, 100))
    compare_page_raw = str(request.GET.get("compare_page", "1")).strip()
    compare_page = int(compare_page_raw) if compare_page_raw.isdigit() else 1
    compare_page = max(1, compare_page)

    compare_logs_qs = BukTemplateCompareLog.objects.filter(
        tenant=tenant,
        property=property_obj,
    ).select_related("compared_by")
    if compare_result_filter == "compatible":
        compare_logs_qs = compare_logs_qs.filter(is_compatible=True)
    elif compare_result_filter == "incompatible":
        compare_logs_qs = compare_logs_qs.filter(is_compatible=False)
    if compare_user_filter:
        compare_logs_qs = compare_logs_qs.filter(compared_by__email__icontains=compare_user_filter)
    if compare_from:
        compare_logs_qs = compare_logs_qs.filter(compared_at__date__gte=compare_from)
    if compare_to:
        compare_logs_qs = compare_logs_qs.filter(compared_at__date__lte=compare_to)

    compare_total = compare_logs_qs.count()
    compare_total_pages = max(1, (compare_total + compare_page_size - 1) // compare_page_size)
    if compare_page > compare_total_pages:
        compare_page = compare_total_pages
    compare_offset = (compare_page - 1) * compare_page_size
    recent_compare_logs = list(
        compare_logs_qs.order_by("-compared_at", "-id")[compare_offset : compare_offset + compare_page_size]
    )

    compare_query_pairs = [
        ("date_from", date_from.isoformat()),
        ("date_to", date_to.isoformat()),
        ("compare_result", compare_result_filter),
        ("compare_user", compare_user_filter),
        ("compare_from", compare_from.isoformat() if compare_from else ""),
        ("compare_to", compare_to.isoformat() if compare_to else ""),
        ("compare_page_size", str(compare_page_size)),
    ]
    for area_id in selected_area_ids:
        compare_query_pairs.append(("area_ids", str(area_id)))
    for worker_id in selected_worker_ids:
        compare_query_pairs.append(("worker_ids", str(worker_id)))

    compare_prev_url = None
    compare_next_url = None
    if compare_page > 1:
        compare_prev_url = f"/app/buk-report/?{urlencode(compare_query_pairs + [('compare_page', compare_page - 1)])}"
    if compare_page < compare_total_pages:
        compare_next_url = f"/app/buk-report/?{urlencode(compare_query_pairs + [('compare_page', compare_page + 1)])}"
    compare_csv_url = f"/app/buk-report/compare-logs/download-csv/?{urlencode(compare_query_pairs)}"

    compare_api_pairs = [
        ("tenant_id", str(tenant.id)),
        ("property_id", str(property_obj.id)),
        ("page", str(compare_page)),
        ("page_size", str(compare_page_size)),
    ]
    if compare_result_filter == "compatible":
        compare_api_pairs.append(("is_compatible", "1"))
    elif compare_result_filter == "incompatible":
        compare_api_pairs.append(("is_compatible", "0"))
    if compare_user_filter:
        compare_api_pairs.append(("user", compare_user_filter))
    if compare_from:
        compare_api_pairs.append(("compared_from", compare_from.isoformat()))
    if compare_to:
        compare_api_pairs.append(("compared_to", compare_to.isoformat()))
    compare_api_url = f"/api/buk/compare-template-logs/?{urlencode(compare_api_pairs)}"

    return render(
        request,
        "webui/buk_report.html",
        {
            **ctx,
            "rows": rows,
            "issues": issues,
            "date_columns": date_columns,
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "has_blocking_errors": has_blocking_errors,
            "can_export_buk": can_export_report,
            "module_export_enabled": module_export,
            "module_validator_enabled": module_validator,
            "is_admin_export": is_admin_export,
            "error_count": error_count,
            "warning_count": warning_count,
            "info_count": info_count,
            "area_options": area_options,
            "selected_area_ids": selected_area_ids,
            "worker_options": worker_options,
            "selected_worker_ids": selected_worker_ids,
            "template_compare_result": template_compare_result,
            "recent_compare_logs": recent_compare_logs,
            "compare_result_filter": compare_result_filter,
            "compare_user_filter": compare_user_filter,
            "compare_from": compare_from.isoformat() if compare_from else "",
            "compare_to": compare_to.isoformat() if compare_to else "",
            "compare_page_size": compare_page_size,
            "compare_page": compare_page,
            "compare_total": compare_total,
            "compare_total_pages": compare_total_pages,
            "compare_prev_url": compare_prev_url,
            "compare_next_url": compare_next_url,
            "compare_csv_url": compare_csv_url,
            "compare_api_url": compare_api_url,
        },
    )


@login_required
@require_GET
def buk_report_compare_log_download(request, log_id):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    can_view_buk_history = PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_export_buk"
    ) or PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_view_reports")
    if not can_view_buk_history:
        return HttpResponseForbidden("No tienes permisos para acceder al historial BUK.")
    if not PermissionService.user_can_module(request.user, tenant, "buk_preview"):
        return HttpResponseForbidden("Modulo desactivado: buk_preview.")

    item = (
        BukTemplateCompareLog.objects.filter(
            id=log_id,
            tenant=tenant,
            property=property_obj,
        )
        .select_related("compared_by")
        .first()
    )
    if item is None:
        return HttpResponseForbidden("Log de comparacion no encontrado para esta sede.")

    payload = {
        "compare_log_id": item.id,
        "tenant": tenant.slug,
        "property": property_obj.slug,
        "date_from": item.date_from.isoformat(),
        "date_to": item.date_to.isoformat(),
        "sheet_name": item.sheet_name,
        "reference_file_name": item.reference_file_name,
        "reference_file_sha256": item.reference_file_sha256,
        "reference_file_size_bytes": item.reference_file_size_bytes,
        "is_compatible": item.is_compatible,
        "errors_count": item.errors_count,
        "warnings_count": item.warnings_count,
        "compared_by": item.compared_by.email if item.compared_by_id else None,
        "result": item.result_payload,
    }
    file_name = f"buk_template_compare_log_{item.id}.json"
    response = HttpResponse(
        json.dumps(payload, ensure_ascii=False, indent=2),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


@login_required
@require_GET
def buk_report_compare_logs_csv_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    can_view_buk_history = PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_export_buk"
    ) or PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_view_reports")
    if not can_view_buk_history:
        return HttpResponseForbidden("No tienes permisos para acceder al historial BUK.")
    if not PermissionService.user_can_module(request.user, tenant, "buk_preview"):
        return HttpResponseForbidden("Modulo desactivado: buk_preview.")

    compare_result_filter = str(request.GET.get("compare_result", "")).strip().lower()
    compare_user_filter = str(request.GET.get("compare_user", "")).strip()
    compare_from_raw = str(request.GET.get("compare_from", "")).strip()
    compare_to_raw = str(request.GET.get("compare_to", "")).strip()
    compare_from = _parse_date_or_default(compare_from_raw, None) if compare_from_raw else None
    compare_to = _parse_date_or_default(compare_to_raw, None) if compare_to_raw else None

    compare_logs_qs = BukTemplateCompareLog.objects.filter(
        tenant=tenant,
        property=property_obj,
    ).select_related("compared_by")
    if compare_result_filter == "compatible":
        compare_logs_qs = compare_logs_qs.filter(is_compatible=True)
    elif compare_result_filter == "incompatible":
        compare_logs_qs = compare_logs_qs.filter(is_compatible=False)
    if compare_user_filter:
        compare_logs_qs = compare_logs_qs.filter(compared_by__email__icontains=compare_user_filter)
    if compare_from:
        compare_logs_qs = compare_logs_qs.filter(compared_at__date__gte=compare_from)
    if compare_to:
        compare_logs_qs = compare_logs_qs.filter(compared_at__date__lte=compare_to)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "id",
            "compared_at",
            "date_from",
            "date_to",
            "sheet_name",
            "reference_file_name",
            "reference_file_sha256",
            "reference_file_size_bytes",
            "is_compatible",
            "errors_count",
            "warnings_count",
            "compared_by",
        ]
    )
    for item in compare_logs_qs.order_by("-compared_at", "-id")[:5000]:
        writer.writerow(
            [
                item.id,
                item.compared_at.isoformat(),
                item.date_from.isoformat(),
                item.date_to.isoformat(),
                item.sheet_name,
                item.reference_file_name,
                item.reference_file_sha256,
                item.reference_file_size_bytes,
                "1" if item.is_compatible else "0",
                item.errors_count,
                item.warnings_count,
                item.compared_by.email if item.compared_by_id else "",
            ]
        )
    file_name = f"buk_template_compare_logs_{property_obj.slug}.csv"
    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


@login_required
@require_http_methods(["GET", "POST"])
def imports_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(
            request,
            "webui/imports.html",
            {**ctx, "batches": [], "selected_batch": None, "preview_rows": []},
        )

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return render(
            request,
            "webui/imports.html",
            {
                **ctx,
                "batches": [],
                "selected_batch": None,
                "preview_rows": [],
                "context_error": "Modulo desactivado: excel_import.",
            },
        )
    can_manage_workers = PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_manage_workers"
    )
    can_manage_shifts = PermissionService.user_can_property_action(
        request.user, tenant, property_obj, "can_manage_shifts"
    )
    if not (can_manage_workers or can_manage_shifts):
        return render(
            request,
            "webui/imports.html",
            {
                **ctx,
                "batches": [],
                "selected_batch": None,
                "preview_rows": [],
                "context_error": "No tienes permisos para gestionar importaciones en esta sede.",
            },
        )

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()

        if action == "preview_workers":
            if not can_manage_workers:
                messages.error(request, "No tienes permisos para importar trabajadores.")
                return redirect("/app/imports/")
            if not request.POST.get("confirm_full_sync"):
                messages.error(request, "Confirma que el archivo contiene la lista completa de trabajadores de la sede.")
                return redirect("/app/imports/")
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo CSV o XLSX.")
            else:
                try:
                    batch = WorkerImportService.create_worker_preview(
                        tenant=tenant,
                        fallback_property=property_obj,
                        file_name=uploaded_file.name,
                        file_bytes=uploaded_file.read(),
                        user=request.user,
                        create_missing_areas=bool(request.POST.get("create_missing_areas")),
                        sync_mode=True,
                    )
                    messages.success(request, f"Vista previa creada. Lote #{batch.id}.")
                    return redirect(f"/app/imports/?batch_id={batch.id}")
                except ValueError as exc:
                    messages.error(request, str(exc))

        elif action == "preview_excel":
            messages.error(request, "La importacion del Excel original ya no esta disponible desde la aplicacion.")
            return redirect("/app/imports/")

        elif action == "preview_shifts_area":
            if not can_manage_shifts:
                messages.error(request, "No tienes permisos para importar turnos.")
                return redirect("/app/imports/")
            if not request.POST.get("confirm_full_sync"):
                messages.error(request, "Confirma que el archivo contiene la lista completa de turnos de la sede.")
                return redirect("/app/imports/")
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo XLSX o CSV.")
            else:
                try:
                    batch = ShiftAreaImportService.create_shift_preview(
                        tenant=tenant,
                        fallback_property=property_obj,
                        file_name=uploaded_file.name,
                        file_bytes=uploaded_file.read(),
                        user=request.user,
                        create_missing_areas=bool(request.POST.get("create_missing_areas")),
                        sync_mode=True,
                    )
                    messages.success(request, f"Vista previa de turnos creada. Lote #{batch.id}.")
                    return redirect(f"/app/imports/?batch_id={batch.id}")
                except ValueError as exc:
                    messages.error(request, str(exc))

        elif action in {"confirm_batch", "cancel_batch"}:
            batch_id_raw = str(request.POST.get("batch_id", "")).strip()
            if not batch_id_raw.isdigit():
                messages.error(request, "Lote invalido.")
                return redirect("/app/imports/")
            batch = (
                ImportBatch.objects.filter(
                    id=int(batch_id_raw),
                    tenant=tenant,
                    property=property_obj,
                )
                .first()
            )
            if batch is None:
                messages.error(request, "Lote no encontrado.")
                return redirect("/app/imports/")
            if action == "confirm_batch":
                if batch.status != "preview":
                    messages.error(request, "El lote ya no esta en estado preview.")
                else:
                    if _import_batch_has_errors(batch):
                        messages.error(request, "No se puede confirmar una importacion con errores bloqueantes.")
                        return redirect(f"/app/imports/?batch_id={batch.id}")
                    requires_sync_confirmation = batch.source_type in {"workers", "shifts_area"} and bool(
                        batch.summary.get("sync_mode")
                    )
                    if requires_sync_confirmation and not request.POST.get("confirm_apply_sync"):
                        messages.error(
                            request,
                            "Confirma la aplicacion de la sincronizacion completa antes de continuar.",
                        )
                        return redirect(f"/app/imports/?batch_id={batch.id}")
                    if batch.source_type == "workers":
                        if not can_manage_workers:
                            messages.error(request, "No tienes permisos para confirmar importaciones de trabajadores.")
                            return redirect(f"/app/imports/?batch_id={batch.id}")
                        batch = WorkerImportService.confirm_worker_import(batch=batch)
                    elif batch.source_type == "excel_original":
                        messages.error(request, "La importacion del Excel original ya no esta disponible.")
                        return redirect(f"/app/imports/?batch_id={batch.id}")
                    elif batch.source_type == "shifts_area":
                        if not can_manage_shifts:
                            messages.error(request, "No tienes permisos para confirmar importaciones de turnos.")
                            return redirect(f"/app/imports/?batch_id={batch.id}")
                        batch = ShiftAreaImportService.confirm_shift_import(batch=batch)
                    else:
                        batch.status = "confirmed"
                        batch.save(update_fields=["status", "updated_at"])
                    messages.success(request, f"Lote #{batch.id} confirmado.")
                return redirect(f"/app/imports/?batch_id={batch.id}")

            batch.status = "cancelled"
            batch.save(update_fields=["status", "updated_at"])
            messages.success(request, f"Lote #{batch.id} cancelado.")
            return redirect(f"/app/imports/?batch_id={batch.id}")

    batches = list(
        ImportBatch.objects.filter(tenant=tenant, property=property_obj)
        .select_related("created_by")
        .order_by("-created_at")[:20]
    )
    source_labels = {
        "workers": "Trabajadores",
        "shifts_area": "Turnos por area",
        "excel_original": "Excel original",
    }
    status_labels = {
        "preview": "Vista previa",
        "confirmed": "Confirmado",
        "cancelled": "Cancelado",
        "failed": "Fallido",
    }
    batches_rows = [
        {
            "batch": item,
            "source_label": source_labels.get(item.source_type, item.source_type),
            "status_label": status_labels.get(item.status, item.status),
        }
        for item in batches
    ]
    selected_batch = None
    batch_id = request.GET.get("batch_id")
    if batch_id and str(batch_id).isdigit():
        selected_batch = next((item for item in batches if item.id == int(batch_id)), None)

    selected_batch_source_label = ""
    selected_batch_status_label = ""
    if selected_batch:
        selected_batch_source_label = source_labels.get(selected_batch.source_type, selected_batch.source_type)
        selected_batch_status_label = status_labels.get(selected_batch.status, selected_batch.status)
    import_context = _build_import_batch_preview_context(selected_batch, row_limit=50)

    return render(
        request,
        "webui/imports.html",
        {
            **ctx,
            "batches": batches,
            "batches_rows": batches_rows,
            "selected_batch": selected_batch,
            "selected_batch_source_label": selected_batch_source_label,
            "selected_batch_status_label": selected_batch_status_label,
            "import_batch": selected_batch,
            "import_modal_open": bool(selected_batch and batch_id),
            **import_context,
            "can_import_workers": can_manage_workers,
            "can_import_shifts": can_manage_shifts,
        },
    )


@login_required
@require_GET
def workers_template_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return HttpResponseForbidden("Modulo desactivado: excel_import.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_workers"):
        return HttpResponseForbidden("No tienes permisos para gestionar trabajadores en esta sede.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores"
    headers = ["DNI", "Nombre", "Apellido", "Area", "Sede"]
    for col, value in enumerate(headers, start=1):
        ws.cell(1, col, value)

    # Ejemplo de referencia editable por el usuario.
    ws.cell(2, 1, "12345678")
    ws.cell(2, 2, "Nombre")
    ws.cell(2, 3, "Apellido")
    ws.cell(2, 4, "Recepción")
    ws.cell(2, 5, property_obj.name)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 24

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_trabajadores.xlsx"'
    return response


@login_required
@require_GET
def shifts_template_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return HttpResponseForbidden("Modulo desactivado: excel_import.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_shifts"):
        return HttpResponseForbidden("No tienes permisos para gestionar turnos en esta sede.")

    wb = Workbook()
    ws = wb.active
    ws.title = "TurnosPorArea"
    headers = [
        "Area",
        "Turno",
        "Codigo BUK",
        "Hora Inicio",
        "Hora Fin",
        "Inicio Break",
        "Fin Break",
        "Nocturno",
        "Activo",
        "Sede",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(1, col, value)

    ws.cell(2, 1, "Recepción")
    ws.cell(2, 2, "REC-M")
    ws.cell(2, 3, "REC-M")
    ws.cell(2, 4, "06:00")
    ws.cell(2, 5, "14:45")
    ws.cell(2, 6, "10:00")
    ws.cell(2, 7, "10:30")
    ws.cell(2, 8, "0")
    ws.cell(2, 9, "1")
    ws.cell(2, 10, property_obj.name)

    for key, width in {
        "A": 20,
        "B": 18,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 10,
        "I": 10,
        "J": 24,
    }.items():
        ws.column_dimensions[key].width = width

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_turnos_por_area.xlsx"'
    return response


def _sample_file_format_from_request(request):
    requested = str(request.GET.get("format", "xlsx")).strip().lower()
    return requested if requested in {"xlsx", "csv"} else "xlsx"


@login_required
@require_GET
def workers_sample_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return HttpResponseForbidden("Modulo desactivado: excel_import.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_workers"):
        return HttpResponseForbidden("No tienes permisos para gestionar trabajadores en esta sede.")

    payload = ImportSampleService.build_sample_payload(
        tenant=tenant,
        property_obj=property_obj,
        max_workers=20,
        max_shifts=20,
    )
    headers = payload["workers_headers"]
    rows = payload["workers_rows"]
    output_format = _sample_file_format_from_request(request)

    if output_format == "csv":
        csv_bytes = ImportSampleService.generate_csv_bytes(headers=headers, rows=rows)
        response = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="muestra_trabajadores.csv"'
        return response

    xlsx_bytes = ImportSampleService.generate_xlsx_bytes(
        sheet_name="Trabajadores",
        headers=headers,
        rows=rows,
    )
    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="muestra_trabajadores.xlsx"'
    return response


@login_required
@require_GET
def shifts_sample_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return HttpResponseForbidden("Modulo desactivado: excel_import.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_shifts"):
        return HttpResponseForbidden("No tienes permisos para gestionar turnos en esta sede.")

    payload = ImportSampleService.build_sample_payload(
        tenant=tenant,
        property_obj=property_obj,
        max_workers=20,
        max_shifts=40,
    )
    headers = payload["shifts_headers"]
    rows = payload["shifts_rows"]
    output_format = _sample_file_format_from_request(request)

    if output_format == "csv":
        csv_bytes = ImportSampleService.generate_csv_bytes(headers=headers, rows=rows)
        response = HttpResponse(csv_bytes, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="muestra_turnos_por_area.csv"'
        return response

    xlsx_bytes = ImportSampleService.generate_xlsx_bytes(
        sheet_name="Turnos",
        headers=headers,
        rows=rows,
    )
    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="muestra_turnos_por_area.xlsx"'
    return response


@login_required
@require_GET
def backup_download(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return HttpResponseForbidden(ctx["context_error"])

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_tenant_role(request.user, tenant, ["admin"]):
        return HttpResponseForbidden("Solo administradores pueden descargar backups.")
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return HttpResponseForbidden("Modulo desactivado: excel_import.")
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_workers"):
        return HttpResponseForbidden("No tienes permisos para gestionar trabajadores en esta sede.")

    payload = BackupRestoreService.create_backup_payload(tenant=tenant, property_obj=property_obj)
    content = BackupRestoreService.payload_to_json_bytes(payload)
    file_name = f"backup_{tenant.slug}_{property_obj.slug}_{timezone.localdate().isoformat()}.json"
    AuditService.log(
        tenant=tenant,
        property_obj=property_obj,
        user=request.user,
        action="backup_export_json",
        entity_type="Backup",
        entity_id=file_name,
        before={},
        after={
            "areas": len(payload.get("areas", [])),
            "workers": len(payload.get("workers", [])),
            "shifts": len(payload.get("shifts", [])),
            "assignments": len(payload.get("assignments", [])),
        },
    )
    response = HttpResponse(content, content_type="application/json; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    return response


@login_required
@require_http_methods(["GET", "POST"])
def backup_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/backup.html", {**ctx})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_tenant_role(request.user, tenant, ["admin"]):
        return render(
            request,
            "webui/backup.html",
            {**ctx, "context_error": "Solo administradores pueden usar backup/restauracion."},
        )
    if not PermissionService.user_can_module(request.user, tenant, "excel_import"):
        return render(
            request,
            "webui/backup.html",
            {**ctx, "context_error": "Modulo desactivado: excel_import."},
        )
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_workers"):
        return render(
            request,
            "webui/backup.html",
            {**ctx, "context_error": "No tienes permisos para gestionar datos en esta sede."},
        )

    preview_summary = None
    payload_json = ""
    source_name = ""
    preview_sync_mode = False
    if request.method == "POST":
        action = str(request.POST.get("action", "apply_restore")).strip()
        sync_mode = str(request.POST.get("sync_mode", "")).strip().lower() in {"1", "true", "on", "yes"}
        payload = None

        if action == "preview_restore":
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo JSON.")
                return redirect("webui-backup")
            source_name = uploaded_file.name
            try:
                payload = json.loads(uploaded_file.read().decode("utf-8"))
            except (json.JSONDecodeError, UnicodeDecodeError):
                messages.error(request, "Archivo JSON invalido.")
                return redirect("webui-backup")

            preview_summary = BackupRestoreService.preview_restore_from_payload(
                tenant=tenant,
                property_obj=property_obj,
                payload=payload,
                sync_mode=sync_mode,
            )
            payload_json = json.dumps(payload, ensure_ascii=False)
            preview_sync_mode = sync_mode
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="backup_restore_preview_json",
                entity_type="Backup",
                entity_id=source_name,
                before={},
                after={**preview_summary, "sync_mode": sync_mode},
            )
            messages.success(request, "Vista previa generada. Revisa los cambios antes de aplicar.")

        elif action == "apply_restore":
            payload_json = str(request.POST.get("payload_json", "")).strip()
            source_name = str(request.POST.get("source_name", "restore.json")).strip() or "restore.json"
            if payload_json:
                try:
                    payload = json.loads(payload_json)
                except json.JSONDecodeError:
                    messages.error(request, "Payload de restauracion invalido.")
                    return redirect("webui-backup")
            else:
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Debes seleccionar un archivo JSON.")
                    return redirect("webui-backup")
                source_name = uploaded_file.name
                try:
                    payload = json.loads(uploaded_file.read().decode("utf-8"))
                except (json.JSONDecodeError, UnicodeDecodeError):
                    messages.error(request, "Archivo JSON invalido.")
                    return redirect("webui-backup")

            summary = BackupRestoreService.restore_from_payload(
                tenant=tenant,
                property_obj=property_obj,
                payload=payload,
                sync_mode=sync_mode,
            )
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="backup_restore_json",
                entity_type="Backup",
                entity_id=source_name,
                before={},
                after={**summary, "sync_mode": sync_mode},
            )
            messages.success(request, "Restauracion completada correctamente.")
            return redirect("webui-backup")

    return render(
        request,
        "webui/backup.html",
        {
            **ctx,
            "preview_summary": preview_summary,
            "payload_json": payload_json,
            "source_name": source_name,
            "preview_sync_mode": preview_sync_mode,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def month_closure_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/month_closure.html", {**ctx, "closures": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_tenant_role(request.user, tenant, ["admin", "operator", "supervisor"]):
        return render(
            request,
            "webui/month_closure.html",
            {**ctx, "closures": [], "context_error": "No tienes permisos para este modulo."},
        )
    if not PermissionService.user_can_module(request.user, tenant, "month_closure"):
        return render(
            request,
            "webui/month_closure.html",
            {**ctx, "closures": [], "context_error": "Modulo desactivado: month_closure."},
        )
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_access"):
        return render(
            request,
            "webui/month_closure.html",
            {**ctx, "closures": [], "context_error": "No tienes acceso a esta sede."},
        )

    is_admin = PermissionService.user_can_tenant_role(request.user, tenant, ["admin"])
    month_raw = str(request.POST.get("month") if request.method == "POST" else request.GET.get("month", "")).strip()
    today = timezone.localdate()
    if not month_raw:
        month_raw = f"{today.year:04d}-{today.month:02d}"
    try:
        year, month = [int(x) for x in month_raw.split("-", 1)]
    except ValueError:
        year, month = today.year, today.month
        month_raw = f"{year:04d}-{month:02d}"

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()
        if not is_admin:
            return HttpResponseForbidden("Solo administradores pueden cerrar o reabrir meses.")
        if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_schedule"):
            return HttpResponseForbidden("No tienes permisos para gestionar cierre de mes en esta sede.")
        if action == "close":
            MonthClosureService.close_month(
                tenant=tenant,
                property_obj=property_obj,
                year=year,
                month=month,
                user=request.user,
            )
            messages.success(request, f"Mes {month:02d}-{year} cerrado.")
            return redirect(f"/app/month-closure/?month={month_raw}")
        if action == "reopen":
            MonthClosureService.reopen_month(
                tenant=tenant,
                property_obj=property_obj,
                year=year,
                month=month,
                user=request.user,
            )
            messages.success(request, f"Mes {month:02d}-{year} reabierto.")
            return redirect(f"/app/month-closure/?month={month_raw}")

    is_closed = MonthClosureService.is_closed(
        tenant=tenant,
        property_obj=property_obj,
        year=year,
        month=month,
    )
    closures = list(
        MonthClosure.objects.filter(tenant=tenant, property=property_obj)
        .select_related("closed_by", "reopened_by")
        .order_by("-year", "-month", "-updated_at")[:24]
    )

    closure_keys = {(item.year, item.month) for item in closures}
    current_period_key = (year, month)
    closure_keys.add(current_period_key)
    buk_export_log_by_period = {}
    if closure_keys:
        min_year, min_month = min(closure_keys)
        max_year, max_month = max(closure_keys)
        min_period_start = date(min_year, min_month, 1)
        max_period_end = date(max_year, max_month, calendar.monthrange(max_year, max_month)[1])
        export_logs = list(
            BukExportLog.objects.filter(
                tenant=tenant,
                property=property_obj,
                date_from__lte=max_period_end,
                date_to__gte=min_period_start,
            ).order_by("-generated_at", "-id")
        )
        for log in export_logs:
            month_cursor = date(log.date_from.year, log.date_from.month, 1)
            month_limit = date(log.date_to.year, log.date_to.month, 1)
            while month_cursor <= month_limit:
                key = (month_cursor.year, month_cursor.month)
                if key in closure_keys and key not in buk_export_log_by_period:
                    buk_export_log_by_period[key] = log
                next_anchor = month_cursor + timedelta(days=32)
                month_cursor = date(next_anchor.year, next_anchor.month, 1)

    for item in closures:
        item.buk_export_log = buk_export_log_by_period.get((item.year, item.month))
        item.buk_export_generated = item.buk_export_log is not None

    current_period_export_log = buk_export_log_by_period.get(current_period_key)

    return render(
        request,
        "webui/month_closure.html",
        {
            **ctx,
            "closures": closures,
            "month_value": month_raw,
            "is_closed": is_closed,
            "is_admin": is_admin,
            "current_period_export_log": current_period_export_log,
        },
    )


@login_required
@require_GET
def control_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(
            request,
            "webui/control.html",
            {**ctx, "rows": [], "pending_count": 0, "affected_areas": 0, "affected_workers": 0, "area_summary": []},
        )

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_module(request.user, tenant, "control"):
        return render(
            request,
            "webui/control.html",
            {
                **ctx,
                "rows": [],
                "pending_count": 0,
                "affected_areas": 0,
                "affected_workers": 0,
                "area_summary": [],
                "context_error": "Modulo desactivado: control.",
            },
        )
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_use_control"):
        return render(
            request,
            "webui/control.html",
            {
                **ctx,
                "rows": [],
                "pending_count": 0,
                "affected_areas": 0,
                "affected_workers": 0,
                "area_summary": [],
                "context_error": "No tienes permisos para usar Control en esta sede.",
            },
        )

    start_date = timezone.localdate()
    end_date = start_date + timedelta(days=14)
    workers = Worker.objects.filter(
        tenant=tenant,
        property=property_obj,
        active=True,
    ).select_related("area").order_by("area__name", "last_name", "first_name")
    assignments = ScheduleAssignment.objects.filter(
        tenant=tenant,
        property=property_obj,
        date__gte=start_date,
        date__lte=end_date,
    ).select_related("worker")
    assignment_index = {(item.worker_id, item.date): item for item in assignments}

    rows = []
    affected_areas = set()
    affected_workers = set()
    area_missing_days = defaultdict(int)
    area_required_days = defaultdict(int)
    area_covered_days = defaultdict(int)
    area_worker_ids = defaultdict(set)
    for worker in workers:
        if not PermissionService.user_can_area_schedule(request.user, tenant, property_obj, worker.area):
            continue
        for offset in range(15):
            day = start_date + timedelta(days=offset)
            if not _worker_active_on_date(worker, day):
                continue
            area_name = worker.area.name if worker.area_id else "Sin area"
            area_required_days[area_name] += 1
            assignment = assignment_index.get((worker.id, day))
            covered = bool(assignment and (assignment.shift_id or assignment.special_state_id))
            if covered:
                area_covered_days[area_name] += 1
                continue
            rows.append(
                {
                    "date": day,
                    "worker_id": worker.id,
                    "worker_name": f"{worker.last_name}, {worker.first_name}",
                    "document_number": worker.document_number,
                    "area": area_name,
                    "go_url": "/app/scheduling/?"
                    + urlencode(
                        [
                            ("month", day.strftime("%Y-%m")),
                            ("area_id", str(worker.area_id or "")),
                            ("worker_q", worker.document_number),
                            ("focus_date", day.isoformat()),
                        ]
                    ),
                }
            )
            affected_areas.add(area_name)
            affected_workers.add(worker.id)
            area_missing_days[area_name] += 1
            area_worker_ids[area_name].add(worker.id)

    area_summary = []
    for area_name, worker_ids in area_worker_ids.items():
        required_days = area_required_days[area_name]
        covered_days = area_covered_days[area_name]
        missing_days = area_missing_days[area_name]
        coverage_percent = round((covered_days / required_days) * 100) if required_days else 100
        if coverage_percent < 70 or missing_days >= 10:
            risk_level = "Alta"
            risk_class = "high"
        elif coverage_percent < 90 or missing_days >= 3:
            risk_level = "Media"
            risk_class = "medium"
        else:
            risk_level = "Baja"
            risk_class = "low"
        area_summary.append(
            {
                "area": area_name,
                "affected_workers": len(worker_ids),
                "missing_days": missing_days,
                "required_days": required_days,
                "covered_days": covered_days,
                "coverage_percent": coverage_percent,
                "risk_level": risk_level,
                "risk_class": risk_class,
            }
        )
    area_summary.sort(key=lambda item: (-item["missing_days"], item["area"]))

    return render(
        request,
        "webui/control.html",
        {
            **ctx,
            "rows": rows,
            "pending_count": len(rows),
            "affected_areas": len(affected_areas),
            "affected_workers": len(affected_workers),
            "area_summary": area_summary,
            "start_date": start_date,
            "end_date": end_date,
        },
    )


def _apply_audit_filters(request, queryset):
    action_filter = str(request.GET.get("action", "")).strip()
    entity_filter = str(request.GET.get("entity_type", "")).strip()
    user_filter = str(request.GET.get("user_email", "")).strip()
    date_from_raw = str(request.GET.get("date_from", "")).strip()
    date_to_raw = str(request.GET.get("date_to", "")).strip()

    if action_filter:
        queryset = queryset.filter(action__icontains=action_filter)
    if entity_filter:
        queryset = queryset.filter(entity_type__icontains=entity_filter)
    if user_filter:
        queryset = queryset.filter(user__email__icontains=user_filter)
    if date_from_raw:
        date_from = _parse_date_or_default(date_from_raw, None)
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to_raw:
        date_to = _parse_date_or_default(date_to_raw, None)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

    return queryset, {
        "filter_action": action_filter,
        "filter_entity_type": entity_filter,
        "filter_user_email": user_filter,
        "filter_date_from": date_from_raw,
        "filter_date_to": date_to_raw,
    }


@login_required
@require_GET
def audit_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(request, "webui/audit.html", {**ctx, "rows": []})

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_tenant_role(request.user, tenant, ["admin", "operator", "supervisor"]):
        return render(
            request,
            "webui/audit.html",
            {**ctx, "rows": [], "context_error": "No tienes permisos para este modulo."},
        )
    if not PermissionService.user_can_module(request.user, tenant, "audit"):
        return render(
            request,
            "webui/audit.html",
            {**ctx, "rows": [], "context_error": "Modulo desactivado: audit."},
        )
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_access"):
        return render(
            request,
            "webui/audit.html",
            {**ctx, "rows": [], "context_error": "No tienes acceso a esta sede."},
        )

    queryset = AuditLog.objects.select_related("tenant", "user", "property").filter(
        tenant=tenant,
        property=property_obj,
    )
    queryset, filter_context = _apply_audit_filters(request, queryset)

    rows = list(queryset.order_by("-created_at")[:300])
    return render(
        request,
        "webui/audit.html",
        {
            **ctx,
            "rows": rows,
            **filter_context,
            "audit_form_action": "/app/audit/",
            "audit_title": "Auditoria de sede",
        },
    )


@login_required
@require_GET
def audit_global_page(request):
    ctx = _build_context(request, require_property=False)
    if not PermissionService.is_super_admin(request.user):
        return render(
            request,
            "webui/audit.html",
            {
                **ctx,
                "rows": [],
                "is_global_audit": True,
                "audit_form_action": "/app/audit-global/",
                "audit_title": "Auditoria global",
                "context_error": "Solo Super Administrador puede ver auditoria global.",
            },
        )

    tenant_filter = str(request.GET.get("tenant_id", "")).strip()
    property_filter = str(request.GET.get("property_id", "")).strip()
    queryset = AuditLog.objects.select_related("tenant", "property", "user").all()
    if tenant_filter.isdigit():
        queryset = queryset.filter(tenant_id=int(tenant_filter))
    if property_filter.isdigit():
        queryset = queryset.filter(property_id=int(property_filter))
    queryset, filter_context = _apply_audit_filters(request, queryset)

    tenant_filter_options = list(Tenant.objects.all().order_by("name"))
    property_filter_options = list(Property.objects.select_related("tenant").all().order_by("tenant__name", "name"))
    rows = list(queryset.order_by("-created_at")[:300])
    return render(
        request,
        "webui/audit.html",
        {
            **ctx,
            "rows": rows,
            **filter_context,
            "is_global_audit": True,
            "audit_form_action": "/app/audit-global/",
            "audit_title": "Auditoria global",
            "tenant_filter_options": tenant_filter_options,
            "property_filter_options": property_filter_options,
            "filter_tenant_id": tenant_filter,
            "filter_property_id": property_filter,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def support_page(request):
    ctx = _build_context(request, require_property=False)
    if not PermissionService.is_super_admin(request.user):
        return render(
            request,
            "webui/support.html",
            {
                **ctx,
                "rows": [],
                "tenant_rows": [],
                "property_rows": [],
                "context_error": "Solo Super Administrador puede gestionar sesiones de soporte.",
            },
        )

    tenants = list(Tenant.objects.all().order_by("name"))
    selected_tenant_id = request.POST.get("tenant_id") or request.GET.get("tenant_id") or request.session.get("ui_tenant_id")
    selected_tenant = Tenant.objects.filter(id=selected_tenant_id).first() if selected_tenant_id else None
    if selected_tenant is None and tenants:
        selected_tenant = tenants[0]
    property_rows = list(Property.objects.filter(tenant=selected_tenant).order_by("name")) if selected_tenant else []

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()
        if action == "start_support":
            tenant_id = str(request.POST.get("tenant_id", "")).strip()
            tenant = Tenant.objects.filter(id=tenant_id).first()
            if tenant is None:
                messages.error(request, "Tenant no encontrado.")
                return redirect("webui-support")
            property_id = str(request.POST.get("property_id", "")).strip()
            property_obj = None
            if property_id:
                property_obj = Property.objects.filter(id=property_id, tenant=tenant).first()
                if property_obj is None:
                    messages.error(request, "La sede no pertenece al tenant.")
                    return redirect("webui-support")
            reason = str(request.POST.get("reason", "")).strip()
            session = TenantSupportService.start_session(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                reason=reason,
            )
            request.session["support_session_id"] = session.id
            request.session["ui_tenant_id"] = tenant.id
            if property_obj:
                request.session["ui_property_id"] = property_obj.id
            messages.success(request, "Sesion de soporte iniciada y activada.")
            return redirect("webui-support")

        if action == "stop_support":
            session_id = str(request.POST.get("session_id", "")).strip()
            session = TenantSupportAccessSession.objects.filter(
                id=session_id,
                started_by=request.user,
                ended_at__isnull=True,
            ).first()
            if session is None:
                messages.error(request, "Sesion activa no encontrada.")
                return redirect("webui-support")
            TenantSupportService.stop_session(
                session=session,
                user=request.user,
                reason=str(request.POST.get("reason", "")).strip(),
            )
            if str(request.session.get("support_session_id", "")) == str(session.id):
                request.session.pop("support_session_id", None)
            messages.success(request, "Sesion de soporte cerrada.")
            return redirect("webui-support")

        if action == "stop_all_support":
            sessions = TenantSupportAccessSession.objects.filter(
                started_by=request.user,
                ended_at__isnull=True,
            ).select_related("tenant", "property")
            reason = str(request.POST.get("reason", "")).strip()
            closed_count = 0
            for session in sessions:
                TenantSupportService.stop_session(session=session, user=request.user, reason=reason)
                closed_count += 1
            request.session.pop("support_session_id", None)
            messages.success(request, f"Sesiones de soporte cerradas: {closed_count}.")
            return redirect("webui-support")

        messages.error(request, "Accion invalida.")

    active_rows = list(
        TenantSupportAccessSession.objects.filter(started_by=request.user, ended_at__isnull=True)
        .select_related("tenant", "property", "started_by")
        .order_by("-created_at")
    )
    history_rows = list(
        TenantSupportAccessSession.objects.filter(started_by=request.user)
        .select_related("tenant", "property", "started_by", "ended_by")
        .order_by("-created_at")[:20]
    )

    return render(
        request,
        "webui/support.html",
        {
            **ctx,
            "tenant_rows": tenants,
            "property_rows": property_rows,
            "support_rows": active_rows,
            "history_rows": history_rows,
            "support_selected_tenant": selected_tenant,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def modules_page(request):
    ctx = _build_context(request, require_property=False)
    if not PermissionService.is_super_admin(request.user):
        return render(
            request,
            "webui/modules.html",
            {**ctx, "rows": [], "context_error": "Solo Super Administrador puede gestionar modulos."},
        )

    tenant = ctx.get("selected_tenant")
    if tenant is None:
        return render(
            request,
            "webui/modules.html",
            {**ctx, "rows": [], "context_error": "Selecciona un tenant para gestionar modulos."},
        )

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()
        if action == "update_modules":
            enabled_keys = set(request.POST.getlist("enabled_modules"))
            catalog_keys = {module_key for module_key, _, _ in WEBUI_MODULE_CATALOG}
            with transaction.atomic():
                for module_key in catalog_keys:
                    ModuleActivationService.set_state(
                        tenant=tenant,
                        module_key=module_key,
                        is_enabled=module_key in enabled_keys,
                        user=request.user,
                    )
            messages.success(request, "Modulos actualizados.")
            return redirect("webui-modules")
        messages.error(request, "Accion invalida.")

    activations = {
        item.module_key: item
        for item in ModuleActivation.objects.filter(tenant=tenant, module_key__in=[key for key, _, _ in WEBUI_MODULE_CATALOG])
    }
    rows = []
    for module_key, label, description in WEBUI_MODULE_CATALOG:
        activation = activations.get(module_key)
        rows.append(
            {
                "module_key": module_key,
                "label": label,
                "description": description,
                "is_enabled": bool(activation and activation.is_enabled),
                "updated_at": activation.updated_at if activation else None,
                "enabled_by": activation.enabled_by if activation else None,
            }
        )

    return render(
        request,
        "webui/modules.html",
        {
            **ctx,
            "rows": rows,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def tenants_page(request):
    ctx = _build_context(request, require_property=False)
    if not PermissionService.is_super_admin(request.user):
        return render(
            request,
            "webui/tenants.html",
            {**ctx, "rows": [], "form": TenantForm(), "context_error": "Solo Super Administrador puede gestionar tenants."},
        )

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()
        if action == "create_tenant":
            post_data = request.POST.copy()
            if not str(post_data.get("slug", "")).strip():
                post_data["slug"] = slugify(str(post_data.get("name", "")).strip())
            form = TenantForm(post_data)
            if form.is_valid():
                tenant = form.save(commit=False)
                if tenant.settings is None:
                    tenant.settings = {}
                try:
                    tenant.save()
                except IntegrityError:
                    messages.error(request, "El tenant o slug ya existe.")
                else:
                    AuditService.log(
                        tenant=tenant,
                        user=request.user,
                        action="create",
                        entity_type="Tenant",
                        entity_id=tenant.id,
                        after=_audit_snapshot(tenant, ["name", "slug", "status", "settings"]),
                    )
                    messages.success(request, "Tenant creado.")
                    return redirect("webui-tenants")
            else:
                messages.error(request, "Datos invalidos para crear tenant.")
        elif action == "update_tenant":
            tenant_id = str(request.POST.get("tenant_id", "")).strip()
            target = Tenant.objects.filter(id=tenant_id).first()
            if target is None:
                messages.error(request, "Tenant no encontrado.")
                return redirect("webui-tenants")
            post_data = request.POST.copy()
            if not str(post_data.get("slug", "")).strip():
                post_data["slug"] = slugify(str(post_data.get("name", target.name)).strip())
            before = _audit_snapshot(target, ["name", "slug", "status", "settings"])
            form = TenantForm(post_data, instance=target)
            if form.is_valid():
                try:
                    tenant = form.save()
                except IntegrityError:
                    messages.error(request, "El tenant o slug ya existe.")
                else:
                    AuditService.log(
                        tenant=tenant,
                        user=request.user,
                        action="update",
                        entity_type="Tenant",
                        entity_id=tenant.id,
                        before=before,
                        after=_audit_snapshot(tenant, ["name", "slug", "status", "settings"]),
                    )
                    messages.success(request, "Tenant actualizado.")
                    return redirect("webui-tenants")
            else:
                messages.error(request, "Datos invalidos para actualizar tenant.")
        else:
            messages.error(request, "Accion invalida.")

    rows = list(Tenant.objects.all().order_by("name"))
    for item in rows:
        item.settings_json = json.dumps(item.settings or {}, ensure_ascii=True, indent=2)
    return render(
        request,
        "webui/tenants.html",
        {
            **ctx,
            "rows": rows,
            "form": TenantForm(initial={"settings": {}}),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def properties_page(request):
    ctx = _build_context(request, require_property=False)
    if ctx.get("context_error"):
        return render(request, "webui/properties.html", {**ctx, "rows": [], "form": PropertyForm()})

    tenant = ctx["selected_tenant"]
    if tenant is None:
        return render(
            request,
            "webui/properties.html",
            {**ctx, "rows": [], "form": PropertyForm(), "context_error": "Selecciona un tenant para gestionar sedes."},
        )

    if not PermissionService.user_can_tenant_role(request.user, tenant, ["admin"]):
        return render(
            request,
            "webui/properties.html",
            {**ctx, "rows": [], "form": PropertyForm(), "context_error": "Solo administradores pueden gestionar sedes."},
        )
    if not PermissionService.user_can_module(request.user, tenant, "properties"):
        return render(
            request,
            "webui/properties.html",
            {**ctx, "rows": [], "form": PropertyForm(), "context_error": "Modulo desactivado: properties."},
        )

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()
        if action == "create_property":
            post_data = request.POST.copy()
            if not str(post_data.get("slug", "")).strip():
                post_data["slug"] = slugify(str(post_data.get("name", "")).strip())
            form = PropertyForm(post_data)
            if form.is_valid():
                property_obj = form.save(commit=False)
                property_obj.tenant = tenant
                try:
                    with transaction.atomic():
                        property_obj.save()
                        if not PermissionService.is_super_admin(request.user):
                            UserPropertyPermission.objects.update_or_create(
                                user=request.user,
                                tenant=tenant,
                                property=property_obj,
                                defaults={
                                    "can_access": True,
                                    "can_schedule": True,
                                    "can_export_buk": True,
                                    "can_manage_workers": True,
                                    "can_manage_shifts": True,
                                    "can_manage_areas": True,
                                    "can_manage_users": True,
                                    "can_view_reports": True,
                                    "can_use_control": True,
                                },
                            )
                        AuditService.log(
                            tenant=tenant,
                            property_obj=property_obj,
                            user=request.user,
                            action="create",
                            entity_type="Property",
                            entity_id=property_obj.id,
                            after=_audit_snapshot(property_obj, ["name", "slug", "location", "status"]),
                        )
                except IntegrityError:
                    messages.error(request, "La sede o slug ya existe en este tenant.")
                else:
                    messages.success(request, "Sede creada.")
                    return redirect("webui-properties")
            else:
                messages.error(request, "Datos invalidos para crear sede.")
        elif action == "update_property":
            property_id = str(request.POST.get("property_id", "")).strip()
            target = Property.objects.filter(tenant=tenant, id=property_id).first()
            if target is None:
                messages.error(request, "Sede no encontrada.")
                return redirect("webui-properties")
            post_data = request.POST.copy()
            if not str(post_data.get("slug", "")).strip():
                post_data["slug"] = slugify(str(post_data.get("name", target.name)).strip())
            before = _audit_snapshot(target, ["name", "slug", "location", "status"])
            form = PropertyForm(post_data, instance=target)
            if form.is_valid():
                try:
                    property_obj = form.save()
                    AuditService.log(
                        tenant=tenant,
                        property_obj=property_obj,
                        user=request.user,
                        action="update",
                        entity_type="Property",
                        entity_id=property_obj.id,
                        before=before,
                        after=_audit_snapshot(property_obj, ["name", "slug", "location", "status"]),
                    )
                except IntegrityError:
                    messages.error(request, "La sede o slug ya existe en este tenant.")
                else:
                    messages.success(request, "Sede actualizada.")
                    return redirect("webui-properties")
            else:
                messages.error(request, "Datos invalidos para actualizar sede.")
        else:
            messages.error(request, "Accion invalida.")

    rows = list(Property.objects.filter(tenant=tenant).order_by("name"))
    return render(
        request,
        "webui/properties.html",
        {
            **ctx,
            "rows": rows,
            "form": PropertyForm(),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def users_permissions_page(request):
    ctx = _build_context(request, require_property=True)
    if ctx.get("context_error"):
        return render(
            request,
            "webui/users_permissions.html",
            {**ctx, "rows": [], "areas": []},
        )

    tenant = ctx["selected_tenant"]
    property_obj = ctx["selected_property"]
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_manage_users"):
        return render(
            request,
            "webui/users_permissions.html",
            {**ctx, "rows": [], "areas": [], "context_error": "Solo administradores pueden gestionar usuarios."},
        )
    if not PermissionService.user_can_module(request.user, tenant, "users_permissions"):
        return render(
            request,
            "webui/users_permissions.html",
            {**ctx, "rows": [], "areas": [], "context_error": "Modulo desactivado: users_permissions."},
        )
    if not PermissionService.user_can_property_action(request.user, tenant, property_obj, "can_access"):
        return render(
            request,
            "webui/users_permissions.html",
            {**ctx, "rows": [], "areas": [], "context_error": "No tienes acceso a esta sede."},
        )

    role_profiles = list(RoleProfileService.get_active_profiles(tenant))

    if request.method == "POST":
        action = str(request.POST.get("action", "")).strip()

        if action == "create_role_profile":
            name = str(request.POST.get("name", "")).strip()
            code = slugify(str(request.POST.get("code", "")).strip() or name)
            base_role = str(request.POST.get("base_role", "")).strip()
            if not name or not code or base_role not in {"admin", "operator", "supervisor"}:
                messages.error(request, "Nombre, codigo y rol base son obligatorios para crear un rol.")
                return redirect("webui-users-permissions")
            permissions = _property_permission_payload_from_request(request)
            try:
                profile = RoleProfile.objects.create(
                    tenant=tenant,
                    code=code,
                    name=name,
                    base_role=base_role,
                    description=str(request.POST.get("description", "")).strip(),
                    permissions=RoleProfileService.normalize_permissions(permissions),
                    is_system=False,
                    active=True,
                )
            except IntegrityError:
                messages.error(request, "Ya existe un rol con ese codigo en este tenant.")
            else:
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="create",
                    entity_type="RoleProfile",
                    entity_id=profile.id,
                    before={},
                    after=_audit_snapshot(profile, ["code", "name", "base_role", "permissions", "active"]),
                )
                messages.success(request, "Rol creado.")
            return redirect("webui-users-permissions")

        if action in {"update_role_profile", "deactivate_role_profile"}:
            profile_id = str(request.POST.get("role_profile_id", "")).strip()
            profile = RoleProfile.objects.filter(tenant=tenant, id=profile_id).first()
            if profile is None:
                messages.error(request, "Rol no encontrado.")
                return redirect("webui-users-permissions")
            before = _audit_snapshot(profile, ["code", "name", "base_role", "permissions", "active"])
            if action == "deactivate_role_profile":
                if profile.is_system:
                    messages.error(request, "Los roles base del sistema no se pueden desactivar.")
                    return redirect("webui-users-permissions")
                profile.active = False
                profile.save(update_fields=["active", "updated_at"])
                AuditService.log(
                    tenant=tenant,
                    property_obj=property_obj,
                    user=request.user,
                    action="delete",
                    entity_type="RoleProfile",
                    entity_id=profile.id,
                    before=before,
                    after=_audit_snapshot(profile, ["code", "name", "base_role", "permissions", "active"]),
                )
                messages.success(request, "Rol desactivado.")
                return redirect("webui-users-permissions")

            profile.name = str(request.POST.get("name", "")).strip()
            base_role = str(request.POST.get("base_role", "")).strip()
            if base_role in {"admin", "operator", "supervisor"}:
                profile.base_role = base_role
            profile.description = str(request.POST.get("description", "")).strip()
            profile.permissions = RoleProfileService.normalize_permissions(_property_permission_payload_from_request(request))
            if not profile.name:
                messages.error(request, "El nombre del rol es obligatorio.")
                return redirect("webui-users-permissions")
            profile.save()
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="update",
                entity_type="RoleProfile",
                entity_id=profile.id,
                before=before,
                after=_audit_snapshot(profile, ["code", "name", "base_role", "permissions", "active"]),
            )
            messages.success(request, "Rol actualizado.")
            return redirect("webui-users-permissions")

        if action == "create_user":
            email = str(request.POST.get("email", "")).strip().lower()
            password = str(request.POST.get("password", "")).strip()
            first_name = str(request.POST.get("first_name", "")).strip()
            last_name = str(request.POST.get("last_name", "")).strip()
            role = str(request.POST.get("role", "")).strip()
            role_profile = _get_role_profile_from_request(request, tenant)
            if role_profile is not None:
                role = role_profile.base_role

            if not email or not role:
                messages.error(request, "Email y tipo de usuario son obligatorios.")
                return redirect("webui-users-permissions")
            if role not in {"admin", "operator", "supervisor"}:
                messages.error(request, "Tipo de usuario invalido.")
                return redirect("webui-users-permissions")

            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "is_active": True,
                },
            )
            if created:
                if not password:
                    messages.error(request, "La contrasena es obligatoria para nuevos usuarios.")
                    user.delete()
                    return redirect("webui-users-permissions")
                user.set_password(password)
            else:
                if first_name:
                    user.first_name = first_name
                if last_name:
                    user.last_name = last_name
                if password:
                    user.set_password(password)
            user.save()

            permission_payload = _apply_role_profile_defaults(request, role_profile)
            all_properties_access, selected_properties = _selected_properties_from_request(
                request,
                tenant,
                property_obj,
                role,
            )
            UserTenantRole.objects.update_or_create(
                user=user,
                tenant=tenant,
                defaults={
                    "role": role,
                    "role_profile": role_profile,
                    "all_properties_access": all_properties_access,
                    "property_permissions_template": RoleProfileService.normalize_permissions(permission_payload),
                },
            )
            _sync_user_property_permissions(
                user=user,
                tenant=tenant,
                properties=selected_properties,
                permission_payload=permission_payload,
                all_properties_access=all_properties_access,
            )
            selected_area_ids = [int(x) for x in request.POST.getlist("area_ids") if str(x).isdigit()]
            valid_areas = Area.objects.filter(tenant=tenant, property=property_obj, id__in=selected_area_ids)
            for area in valid_areas:
                UserAreaPermission.objects.update_or_create(
                    user=user,
                    tenant=tenant,
                    property=property_obj,
                    area=area,
                    defaults={"can_view": True, "can_schedule": True},
                )

            messages.success(request, "Usuario guardado con permisos.")
            return redirect("webui-users-permissions")

        if action == "update_property_permissions":
            user_id = str(request.POST.get("user_id", "")).strip()
            if not user_id.isdigit():
                messages.error(request, "Usuario invalido.")
                return redirect("webui-users-permissions")
            target_user, target_error = _get_exclusive_tenant_user(
                requester=request.user,
                tenant=tenant,
                user_id=int(user_id),
            )
            if target_user is None:
                messages.error(request, target_error)
                return redirect("webui-users-permissions")

            before_user = _audit_snapshot(target_user, ["email", "first_name", "last_name", "is_active", "is_super_admin"])
            target_user.first_name = str(request.POST.get("first_name", "")).strip()
            target_user.last_name = str(request.POST.get("last_name", "")).strip()
            target_user.save(update_fields=["first_name", "last_name", "updated_at"])

            role = str(request.POST.get("role", "")).strip()
            role_profile = _get_role_profile_from_request(request, tenant)
            if role_profile is not None:
                role = role_profile.base_role
            permission_payload = _apply_role_profile_defaults(request, role_profile)
            if role not in {"admin", "operator", "supervisor"}:
                existing_role = UserTenantRole.objects.filter(user=target_user, tenant=tenant).first()
                role = existing_role.role if existing_role else ""
            all_properties_access, selected_properties = _selected_properties_from_request(
                request,
                tenant=tenant,
                current_property=property_obj,
                role=role,
            )
            if role in {"admin", "operator", "supervisor"}:
                UserTenantRole.objects.update_or_create(
                    user=target_user,
                    tenant=tenant,
                    defaults={
                        "role": role,
                        "role_profile": role_profile,
                        "all_properties_access": all_properties_access,
                        "property_permissions_template": RoleProfileService.normalize_permissions(permission_payload),
                    },
                )
            _sync_user_property_permissions(
                user=target_user,
                tenant=tenant,
                properties=selected_properties,
                permission_payload=permission_payload,
                all_properties_access=all_properties_access,
            )
            selected_area_ids = [int(x) for x in request.POST.getlist("area_ids") if str(x).isdigit()]
            UserAreaPermission.objects.filter(
                user=target_user,
                tenant=tenant,
                property=property_obj,
            ).exclude(area_id__in=selected_area_ids).delete()
            valid_areas = Area.objects.filter(tenant=tenant, property=property_obj, id__in=selected_area_ids)
            for area in valid_areas:
                UserAreaPermission.objects.update_or_create(
                    user=target_user,
                    tenant=tenant,
                    property=property_obj,
                    area=area,
                    defaults={"can_view": True, "can_schedule": True},
                )
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="update",
                entity_type="User",
                entity_id=target_user.id,
                before=before_user,
                after=_audit_snapshot(target_user, ["email", "first_name", "last_name", "is_active", "is_super_admin"]),
            )
            messages.success(request, "Permisos actualizados.")
            return redirect("webui-users-permissions")

        if action == "reset_user_password":
            user_id = str(request.POST.get("user_id", "")).strip()
            new_password = str(request.POST.get("new_password", "")).strip()
            confirm_password = str(request.POST.get("confirm_password", "")).strip()
            if not user_id.isdigit():
                messages.error(request, "Usuario invalido.")
                return redirect("webui-users-permissions")
            target_user, target_error = _get_exclusive_tenant_user(
                requester=request.user,
                tenant=tenant,
                user_id=int(user_id),
            )
            if target_user is None:
                messages.error(request, target_error)
                return redirect("webui-users-permissions")
            if target_user.is_super_admin:
                messages.error(request, "No se puede restablecer la clave de un Super Administrador desde esta pantalla.")
                return redirect("webui-users-permissions")
            if not target_user.is_active:
                messages.error(request, "No se puede restablecer la clave de una cuenta inactiva.")
                return redirect("webui-users-permissions")
            if not new_password or new_password != confirm_password:
                messages.error(request, "La nueva contrasena y la confirmacion deben coincidir.")
                return redirect("webui-users-permissions")
            try:
                validate_password(new_password, target_user)
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))
                return redirect("webui-users-permissions")

            before = {"email": target_user.email, "password_reset": False}
            target_user.set_password(new_password)
            target_user.save(update_fields=["password", "updated_at"])
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="password_reset",
                entity_type="User",
                entity_id=target_user.id,
                before=before,
                after={"email": target_user.email, "password_reset": True, "reset_by_user_id": request.user.id},
            )
            messages.success(
                request,
                "Contrasena restablecida. Entrega la nueva clave al usuario por un canal seguro.",
            )
            return redirect("webui-users-permissions")

        if action == "deactivate_user":
            user_id = str(request.POST.get("user_id", "")).strip()
            if not user_id.isdigit():
                messages.error(request, "Usuario invalido.")
                return redirect("webui-users-permissions")
            target_user, target_error = _get_exclusive_tenant_user(
                requester=request.user,
                tenant=tenant,
                user_id=int(user_id),
            )
            if target_user is None:
                messages.error(request, target_error)
                return redirect("webui-users-permissions")
            if target_user.id == request.user.id:
                messages.error(request, "No puedes desactivar tu propio usuario desde esta pantalla.")
                return redirect("webui-users-permissions")
            if target_user.is_super_admin:
                messages.error(request, "No se puede desactivar un Super Administrador desde esta pantalla.")
                return redirect("webui-users-permissions")
            if not target_user.is_active:
                messages.info(request, "La cuenta ya se encontraba inactiva.")
                return redirect("webui-users-permissions")
            before = _audit_snapshot(target_user, ["email", "is_active", "is_super_admin"])
            target_user.is_active = False
            target_user.save(update_fields=["is_active", "updated_at"])
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="delete",
                entity_type="User",
                entity_id=target_user.id,
                before=before,
                after=_audit_snapshot(target_user, ["email", "is_active", "is_super_admin"]),
            )
            messages.success(request, "Usuario desactivado.")
            return redirect("webui-users-permissions")

        if action == "reactivate_user":
            user_id = str(request.POST.get("user_id", "")).strip()
            if not user_id.isdigit():
                messages.error(request, "Usuario invalido.")
                return redirect("webui-users-permissions")
            target_user, target_error = _get_exclusive_tenant_user(
                requester=request.user,
                tenant=tenant,
                user_id=int(user_id),
            )
            if target_user is None:
                messages.error(request, target_error)
                return redirect("webui-users-permissions")
            if target_user.is_super_admin:
                messages.error(request, "No se puede reactivar un Super Administrador desde esta pantalla.")
                return redirect("webui-users-permissions")
            if target_user.is_active:
                messages.info(request, "La cuenta ya se encontraba activa.")
                return redirect("webui-users-permissions")
            before = _audit_snapshot(target_user, ["email", "is_active", "is_super_admin"])
            target_user.is_active = True
            target_user.save(update_fields=["is_active", "updated_at"])
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="reactivate",
                entity_type="User",
                entity_id=target_user.id,
                before=before,
                after=_audit_snapshot(target_user, ["email", "is_active", "is_super_admin"]),
            )
            messages.success(request, "Usuario reactivado.")
            return redirect("webui-users-permissions")

        if action == "delete_user_permanently":
            user_id = str(request.POST.get("user_id", "")).strip()
            if not user_id.isdigit():
                messages.error(request, "Usuario invalido.")
                return redirect("webui-users-permissions")
            target_user, target_error = _get_exclusive_tenant_user(
                requester=request.user,
                tenant=tenant,
                user_id=int(user_id),
            )
            if target_user is None:
                messages.error(request, target_error)
                return redirect("webui-users-permissions")
            if target_user.id == request.user.id:
                messages.error(request, "No puedes borrar definitivamente tu propio usuario desde esta pantalla.")
                return redirect("webui-users-permissions")
            if target_user.is_super_admin:
                messages.error(request, "No se puede borrar un Super Administrador desde esta pantalla.")
                return redirect("webui-users-permissions")
            if target_user.is_active:
                messages.error(request, "Primero desactiva el usuario antes de borrarlo definitivamente.")
                return redirect("webui-users-permissions")
            before = _audit_snapshot(target_user, ["email", "is_active", "is_super_admin"])
            target_id = target_user.id
            target_email = target_user.email
            try:
                target_user.delete()
            except ProtectedError:
                messages.error(request, "No se pudo borrar el usuario porque tiene registros protegidos asociados.")
                return redirect("webui-users-permissions")
            AuditService.log(
                tenant=tenant,
                property_obj=property_obj,
                user=request.user,
                action="delete_permanent",
                entity_type="User",
                entity_id=target_id,
                before=before,
                after={"email": target_email, "deleted": True},
            )
            messages.success(request, "Usuario borrado definitivamente.")
            return redirect("webui-users-permissions")

    tenant_properties = list(Property.objects.filter(tenant=tenant).order_by("name"))
    areas = list(Area.objects.filter(tenant=tenant, property=property_obj, active=True).order_by("name"))
    role_map = {
        item.user_id: item
        for item in UserTenantRole.objects.filter(tenant=tenant).select_related("user", "role_profile")
    }
    tenant_prop_perms = list(
        UserPropertyPermission.objects.filter(tenant=tenant).select_related("user", "property")
    )
    perm_by_user_property = {(item.user_id, item.property_id): item for item in tenant_prop_perms}
    property_ids_map = {}
    for item in tenant_prop_perms:
        property_ids_map.setdefault(item.user_id, set()).add(item.property_id)
    area_perms = list(
        UserAreaPermission.objects.filter(tenant=tenant, property=property_obj).select_related("user", "area")
    )
    area_map = {}
    for item in area_perms:
        area_map.setdefault(item.user_id, set()).add(item.area_id)

    property_names_by_id = {item.id: item.name for item in tenant_properties}
    rows = []
    for tenant_role in role_map.values():
        role_profile_name = tenant_role.role_profile.name if tenant_role.role_profile_id else ""
        role_label = dict(RoleChoices.choices).get(tenant_role.role, tenant_role.role)
        selected_property_ids = (
            {item.id for item in tenant_properties}
            if tenant_role.all_properties_access
            else property_ids_map.get(tenant_role.user_id, set())
        )
        selected_property_names = [
            property_names_by_id[property_id]
            for property_id in sorted(selected_property_ids, key=lambda item: property_names_by_id.get(item, ""))
            if property_id in property_names_by_id
        ]
        current_permission = perm_by_user_property.get((tenant_role.user_id, property_obj.id))
        if current_permission is None and tenant_role.all_properties_access:
            current_permission = _property_permission_object_from_payload(tenant_role.property_permissions_template)
        elif current_permission is None:
            current_permission = _property_permission_object_from_payload({})
        rows.append(
            {
                "user": tenant_role.user,
                "role": tenant_role.role,
                "role_label": role_label,
                "role_profile_name": role_profile_name,
                "role_profile_id": tenant_role.role_profile_id,
                "all_properties_access": tenant_role.all_properties_access,
                "selected_property_ids": selected_property_ids,
                "selected_property_names": selected_property_names,
                "permission": current_permission,
                "selected_area_ids": area_map.get(tenant_role.user_id, set()),
            }
        )
    rows.sort(key=lambda item: item["user"].email)

    return render(
        request,
        "webui/users_permissions.html",
        {
            **ctx,
            "rows": rows,
            "areas": areas,
            "tenant_properties": tenant_properties,
            "role_profiles": role_profiles,
            "permission_keys": PROPERTY_PERMISSION_KEYS,
        },
    )


def root_redirect(request):
    if request.user.is_authenticated:
        return redirect("webui-dashboard")
    return redirect_to_login(next="/app/", login_url="/app/login/")

