import json
from datetime import date, time, timedelta
from io import BytesIO
from unittest.mock import patch

from django.core.management import call_command
from django.core.management.base import CommandError
from django.db.utils import OperationalError
from django.test import TestCase
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from apps.audit.models import AuditLog
from apps.buk_exports.models import BukExportLog, BukTemplateCompareLog
from apps.imports.models import ImportBatch, ImportPreviewRow
from apps.month_closure.models import MonthClosure, MonthClosureStatus
from apps.modules.models import ModuleActivation
from apps.scheduling.models import ScheduleAssignment, SchedulePatternTemplate, ScheduleRangeTemplate
from apps.tenants.models import Property, Tenant, TenantSupportAccessSession
from apps.users.models import RoleChoices, RoleProfile, User, UserAreaPermission, UserPropertyPermission, UserTenantRole
from apps.users.services import PermissionService
from apps.workers.models import Area, Shift, SpecialState, Worker


class WebUiAuthTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="ui@pariwana.test", password="StrongPass123")

    def test_login_page_loads(self):
        response = self.client.get(reverse("webui-login"))
        self.assertEqual(response.status_code, 200)

    def test_login_redirects_to_dashboard(self):
        response = self.client.post(
            reverse("webui-login"),
            {"email": "ui@pariwana.test", "password": "StrongPass123"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("webui-dashboard"))

    def test_protected_webui_uses_app_login_url(self):
        response = self.client.get(reverse("webui-scheduling"))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.url.startswith("/app/login/?next=/app/scheduling/"))


class WebUiDashboardTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="12345678",
            first_name="Ana",
            last_name="Rojas",
            active=True,
        )
        self.super_admin = User.objects.create_user(
            email="dashboard-super@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
        )
        self.admin = User.objects.create_user(email="dashboard-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="workers", is_enabled=False)
        TenantSupportAccessSession.objects.create(
            tenant=self.tenant,
            property=self.property,
            started_by=self.super_admin,
            reason="dashboard",
        )
        AuditLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            user=self.super_admin,
            action="support_access_start",
            entity_type="TenantSupportAccessSession",
            entity_id="1",
            before={},
            after={},
        )
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status="closed",
            closed_by=self.admin,
        )

    def _set_context(self, user):
        self.client.force_login(user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_super_admin_dashboard_shows_system_status(self):
        self._set_context(self.super_admin)
        response = self.client.get(reverse("webui-dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Estado general del sistema")
        self.assertContains(response, "Tenants activos")
        self.assertContains(response, "Soporte activo")
        self.assertContains(response, "Modulos deshabilitados")
        self.assertContains(response, "Auditoria reciente")
        self.assertContains(response, "support_access_start")
        self.assertContains(response, "Cierres recientes")

    def test_regular_admin_dashboard_hides_system_status(self):
        self._set_context(self.admin)
        response = self.client.get(reverse("webui-dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Estado general del sistema")
        self.assertContains(response, "Trabajadores")


class WebUiTenantsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.super_admin = User.objects.create_user(
            email="super-admin@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
        )
        self.admin = User.objects.create_user(email="tenant-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )

    def test_super_admin_can_open_tenants_page_and_menu(self):
        self.client.force_login(self.super_admin)
        response = self.client.get(reverse("webui-tenants"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Crear tenant")
        self.assertContains(response, "Tenants registrados")
        self.assertContains(response, "/app/tenants/")

    def test_non_super_admin_cannot_manage_tenants(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("webui-tenants"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo Super Administrador puede gestionar tenants.")

    def test_super_admin_can_create_tenant(self):
        self.client.force_login(self.super_admin)
        response = self.client.post(
            reverse("webui-tenants"),
            {
                "action": "create_tenant",
                "name": "Nuevo Hostel",
                "slug": "nuevo-hostel",
                "status": "active",
                "settings": '{"country": "PE"}',
            },
        )
        self.assertEqual(response.status_code, 302)
        tenant = Tenant.objects.get(slug="nuevo-hostel")
        self.assertEqual(tenant.settings["country"], "PE")
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=tenant,
                user=self.super_admin,
                action="create",
                entity_type="Tenant",
            ).exists()
        )

    def test_super_admin_can_update_tenant(self):
        self.client.force_login(self.super_admin)
        response = self.client.post(
            reverse("webui-tenants"),
            {
                "action": "update_tenant",
                "tenant_id": self.tenant.id,
                "name": "Pariwana Hostels Peru",
                "slug": "pariwana-hostels-peru",
                "status": "inactive",
                "settings": '{"support": true}',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.tenant.refresh_from_db()
        self.assertEqual(self.tenant.name, "Pariwana Hostels Peru")
        self.assertEqual(self.tenant.status, "inactive")
        self.assertTrue(self.tenant.settings["support"])
        log = AuditLog.objects.get(
            tenant=self.tenant,
            user=self.super_admin,
            action="update",
            entity_type="Tenant",
        )
        self.assertEqual(log.before["name"], "Pariwana Hostels")
        self.assertEqual(log.after["name"], "Pariwana Hostels Peru")


class WebUiModulesTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.super_admin = User.objects.create_user(
            email="modules-super-admin@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
        )
        self.admin = User.objects.create_user(email="modules-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="workers", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="excel_import", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="areas", is_enabled=False)

    def _activate_super_admin_context(self):
        self.client.force_login(self.super_admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_super_admin_can_open_modules_page_and_menu(self):
        self._activate_super_admin_context()
        response = self.client.get(reverse("webui-modules"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modulos activables")
        self.assertContains(response, "Trabajadores")
        self.assertContains(response, "/app/modules/")

    def test_non_super_admin_cannot_manage_modules(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-modules"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo Super Administrador puede gestionar modulos.")

    def test_super_admin_can_update_module_states(self):
        self._activate_super_admin_context()
        response = self.client.post(
            reverse("webui-modules"),
            {
                "action": "update_modules",
                "enabled_modules": ["areas", "buk_export"],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(ModuleActivation.objects.get(tenant=self.tenant, module_key="workers").is_enabled)
        self.assertTrue(ModuleActivation.objects.get(tenant=self.tenant, module_key="areas").is_enabled)
        self.assertTrue(ModuleActivation.objects.get(tenant=self.tenant, module_key="buk_export").is_enabled)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                user=self.super_admin,
                action="module_activation_change",
                entity_type="ModuleActivation",
            ).exists()
        )


class WebUiSupportTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.super_admin = User.objects.create_user(
            email="support-super-admin@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
        )
        self.admin = User.objects.create_user(email="support-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )

    def _activate_super_admin_context(self):
        self.client.force_login(self.super_admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_super_admin_can_open_support_page_and_menu(self):
        self._activate_super_admin_context()
        response = self.client.get(reverse("webui-support"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Iniciar soporte")
        self.assertContains(response, "Sesiones activas")
        self.assertContains(response, "/app/support/")

    def test_non_super_admin_cannot_manage_support(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-support"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo Super Administrador puede gestionar sesiones de soporte.")

    def test_super_admin_can_start_support_and_activate_context(self):
        self._activate_super_admin_context()
        response = self.client.post(
            reverse("webui-support"),
            {
                "action": "start_support",
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "reason": "revisar horarios",
            },
        )
        self.assertEqual(response.status_code, 302)
        support_session = TenantSupportAccessSession.objects.get(started_by=self.super_admin)
        self.assertEqual(support_session.tenant, self.tenant)
        self.assertEqual(support_session.property, self.property)
        self.assertEqual(self.client.session["support_session_id"], support_session.id)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.super_admin,
                action="support_access_start",
                entity_type="TenantSupportAccessSession",
            ).exists()
        )

    def test_super_admin_can_stop_support(self):
        support_session = TenantSupportAccessSession.objects.create(
            tenant=self.tenant,
            property=self.property,
            started_by=self.super_admin,
            reason="debug",
        )
        self._activate_super_admin_context()
        session = self.client.session
        session["support_session_id"] = support_session.id
        session.save()

        response = self.client.post(
            reverse("webui-support"),
            {
                "action": "stop_support",
                "session_id": support_session.id,
                "reason": "resuelto",
            },
        )
        self.assertEqual(response.status_code, 302)
        support_session.refresh_from_db()
        self.assertIsNotNone(support_session.ended_at)
        self.assertNotIn("support_session_id", self.client.session)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.super_admin,
                action="support_access_stop",
                entity_type="TenantSupportAccessSession",
            ).exists()
        )


class WebUiPropertiesTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.user = User.objects.create_user(email="prop-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="properties", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_properties_page_loads(self):
        self._activate_context()
        response = self.client.get(reverse("webui-properties"))
        self.assertEqual(response.status_code, 200)

    def test_properties_page_can_create_property(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-properties"),
            {
                "action": "create_property",
                "name": "Pariwana Lima",
                "slug": "pariwana-lima",
                "location": "Lima",
                "status": "active",
            },
        )
        self.assertEqual(response.status_code, 302)
        created_property = Property.objects.get(
            tenant=self.tenant,
            name="Pariwana Lima",
            slug="pariwana-lima",
        )
        created_permission = UserPropertyPermission.objects.get(
            user=self.user,
            tenant=self.tenant,
            property=created_property,
        )
        self.assertTrue(created_permission.can_access)
        self.assertTrue(created_permission.can_manage_users)
        self.assertTrue(created_permission.can_export_buk)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=created_property,
                user=self.user,
                action="create",
                entity_type="Property",
            ).exists()
        )

    def test_properties_page_can_update_property(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-properties"),
            {
                "action": "update_property",
                "property_id": self.property.id,
                "name": "Pariwana Cusco Centro",
                "slug": "pariwana-cusco-centro",
                "location": "Cusco Centro",
                "status": "active",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.property.refresh_from_db()
        self.assertEqual(self.property.name, "Pariwana Cusco Centro")
        self.assertEqual(self.property.slug, "pariwana-cusco-centro")
        log = AuditLog.objects.get(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="update",
            entity_type="Property",
        )
        self.assertEqual(log.before["name"], "Pariwana Cusco")
        self.assertEqual(log.after["name"], "Pariwana Cusco Centro")


class WebUiNavigationPermissionTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.user = User.objects.create_user(email="reports-only@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
            can_view_reports=True,
            can_export_buk=False,
            can_use_control=False,
            can_manage_users=False,
        )
        UserAreaPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )
        for module_key in ["scheduling", "buk_preview", "buk_validator", "buk_export", "control", "users_permissions"]:
            ModuleActivation.objects.create(tenant=self.tenant, module_key=module_key, is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_menu_uses_property_permissions(self):
        self._activate_context()
        response = self.client.get(reverse("webui-dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asignacion")
        self.assertContains(response, "Reporte BUK")
        self.assertContains(response, "reports-only@pariwana.test")
        self.assertNotContains(response, 'href="/app/workers/"')
        self.assertNotContains(response, 'href="/app/areas/"')
        self.assertNotContains(response, 'href="/app/shifts/"')
        self.assertNotContains(response, 'href="/app/special-states/"')
        self.assertNotContains(response, 'href="/app/users-permissions/"')
        self.assertNotContains(response, 'href="/app/control/"')
        self.assertNotContains(response, 'href="/app/month-closure/"')
        self.assertNotContains(response, 'href="/app/audit/"')

    def test_topbar_shows_user_full_name_when_available(self):
        self.user.first_name = "Rosa"
        self.user.last_name = "Paredes"
        self.user.save(update_fields=["first_name", "last_name"])
        self._activate_context()

        response = self.client.get(reverse("webui-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rosa Paredes")
        self.assertContains(response, "Salir")

    def test_operator_menu_only_shows_enabled_actions(self):
        operator = User.objects.create_user(email="operator-nav@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
            can_manage_workers=True,
            can_manage_shifts=True,
            can_manage_areas=False,
            can_manage_users=False,
            can_use_control=False,
            can_view_reports=False,
            can_export_buk=False,
        )
        for module_key in ["workers", "areas", "shifts", "special_states", "excel_import"]:
            ModuleActivation.objects.get_or_create(tenant=self.tenant, module_key=module_key, defaults={"is_enabled": True})
        self.client.force_login(operator)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="/app/workers/"')
        self.assertContains(response, 'href="/app/shifts/"')
        self.assertContains(response, 'href="/app/imports/"')
        self.assertNotContains(response, 'href="/app/areas/"')
        self.assertNotContains(response, 'href="/app/special-states/"')
        self.assertNotContains(response, 'href="/app/users-permissions/"')
        self.assertNotContains(response, 'href="/app/control/"')

    def test_buk_report_view_only_hides_export_and_compare(self):
        self._activate_context()
        response = self.client.get(reverse("webui-buk-report"), {"date_from": "2026-06-01", "date_to": "2026-06-01"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vista previa BUK")
        self.assertNotContains(response, 'name="action" value="export"')
        self.assertNotContains(response, 'name="action" value="compare_template"')
        self.assertNotContains(response, "Comparar con Excel base")


class WebUiAreasTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.user = User.objects.create_user(email="areas-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="areas", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_areas_page_can_create_area(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-areas"),
            {
                "action": "create_area",
                "name": "Housekeeping",
                "type": "Operacion",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Area.objects.filter(
                tenant=self.tenant,
                property=self.property,
                name="Housekeeping",
                active=True,
            ).exists()
        )

    def test_areas_forms_hide_type_field(self):
        self._activate_context()
        response = self.client.get(reverse("webui-areas"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="type"', html=False)

    def test_areas_page_deactivate_requires_reassignment_when_dependencies_exist(self):
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="77770000",
            first_name="Ana",
            last_name="Rojas",
            active=True,
        )
        shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="M",
            buk_code="REC-M",
            start_time="07:00",
            end_time="15:00",
        )
        replacement = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        self._activate_context()

        response = self.client.post(
            reverse("webui-areas"),
            {"action": "deactivate_area", "area_id": str(self.area.id)},
        )
        self.assertEqual(response.status_code, 302)
        self.area.refresh_from_db()
        self.assertTrue(self.area.active)

        response = self.client.post(
            reverse("webui-areas"),
            {
                "action": "deactivate_area",
                "area_id": str(self.area.id),
                "replacement_area_id": str(replacement.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.area.refresh_from_db()
        shift.refresh_from_db()
        worker = Worker.objects.get(document_number="77770000", tenant=self.tenant, property=self.property)
        self.assertFalse(self.area.active)
        self.assertEqual(worker.area_id, replacement.id)
        self.assertEqual(shift.area_id, replacement.id)


class WebUiWorkersTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")

        self.user = User.objects.create_user(email="admin-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
            can_manage_shifts=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="workers", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="excel_import", is_enabled=True)

    def test_create_worker_from_webui(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-workers"),
            {
                "document_number": "12312312",
                "first_name": "Ana",
                "last_name": "Quispe",
                "area": self.area.id,
                "active": "on",
                "buk_employee_code": "EMP-001",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="12312312",
            ).exists()
        )

    def test_worker_create_form_hides_buk_and_date_fields(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-workers"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="buk_employee_code"', html=False)
        self.assertNotContains(response, 'name="start_date"', html=False)
        self.assertNotContains(response, 'name="end_date"', html=False)

    def test_supervisor_cannot_create_worker_from_webui(self):
        supervisor = User.objects.create_user(email="worker-supervisor@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=False,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-workers"),
            {
                "document_number": "99998888",
                "first_name": "Luis",
                "last_name": "Paredes",
                "area": self.area.id,
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="99998888",
            ).exists()
        )

    def test_update_and_deactivate_worker_from_webui(self):
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="66667777",
            first_name="Mario",
            last_name="Lopez",
            active=True,
        )
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-workers"),
            {
                "action": "update_worker",
                "worker_id": str(worker.id),
                "document_number": "66667777",
                "first_name": "Mario Updated",
                "last_name": "Lopez Updated",
                "area": str(area_2.id),
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        worker.refresh_from_db()
        self.assertEqual(worker.first_name, "Mario Updated")
        self.assertEqual(worker.area_id, area_2.id)

        response = self.client.post(
            reverse("webui-workers"),
            {
                "action": "deactivate_worker",
                "worker_id": str(worker.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        worker.refresh_from_db()
        self.assertFalse(worker.active)

    def test_workers_page_hides_inactive_by_default(self):
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="11112222",
            first_name="Activo",
            last_name="Uno",
            active=True,
        )
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="33334444",
            first_name="Inactivo",
            last_name="Dos",
            active=False,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-workers"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "11112222")
        self.assertNotContains(response, "33334444")

    def test_workers_inline_import_preview_and_confirm(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()
        csv_bytes = (
            "DNI,Nombre,Apellido,Area,Sede\n"
            "12121212,Ana,Importada,Recepcion,Pariwana Cusco\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("workers.csv", csv_bytes, content_type="text/csv")

        response = self.client.post(
            reverse("webui-workers"),
            {
                "action": "preview_workers_inline",
                "create_missing_areas": "1",
                "confirm_full_sync": "1",
                "file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch = ImportBatch.objects.latest("id")
        self.assertEqual(batch.source_type, "workers")
        response = self.client.get(f"{reverse('webui-workers')}?import_batch_id={batch.id}&import_modal=1")
        self.assertContains(response, "Vista previa de trabajadores")
        self.assertContains(response, "Confirmar importacion")

        response = self.client.post(
            reverse("webui-workers"),
            {
                "action": "confirm_workers_import_inline",
                "batch_id": str(batch.id),
                "confirm_apply_sync": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="12121212",
            ).exists()
        )


class WebUiShiftsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")

        self.user = User.objects.create_user(email="shift-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_shifts=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="shifts", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="excel_import", is_enabled=True)

    def test_create_shift_from_webui(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "area": self.area.id,
                "name": "Manana",
                "buk_code": "REC-M",
                "start_time": "06:00",
                "end_time": "14:45",
                "is_night_shift": "",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="REC-M",
            ).exists()
        )

    def test_supervisor_cannot_create_shift_from_webui(self):
        supervisor = User.objects.create_user(email="shift-supervisor@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_shifts=False,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "area": self.area.id,
                "name": "Nocturno",
                "buk_code": "REC-N",
                "start_time": "22:00",
                "end_time": "06:00",
                "is_night_shift": "on",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="REC-N",
            ).exists()
        )

    def test_update_and_deactivate_shift_from_webui(self):
        shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
            active=True,
        )
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "action": "update_shift",
                "shift_id": str(shift.id),
                "area": str(area_2.id),
                "name": "Tarde",
                "buk_code": "BAR-T",
                "start_time": "15:00",
                "end_time": "23:00",
                "is_night_shift": "on",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        shift.refresh_from_db()
        self.assertEqual(shift.area_id, area_2.id)
        self.assertEqual(shift.name, "Tarde")

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "action": "deactivate_shift",
                "shift_id": str(shift.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        shift.refresh_from_db()
        self.assertFalse(shift.active)

    def test_shifts_page_hides_inactive_by_default(self):
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Act",
            buk_code="REC-A",
            start_time="06:00",
            end_time="14:00",
            active=True,
        )
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Ina",
            buk_code="REC-I",
            start_time="14:00",
            end_time="22:00",
            active=False,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-shifts"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "REC-A")
        self.assertNotContains(response, "REC-I")

    def test_shifts_inline_import_preview_and_confirm(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()
        csv_bytes = (
            "Area,Turno,Codigo BUK,Hora Inicio,Hora Fin,Nocturno,Activo,Sede\n"
            "Recepcion,REC-T,REC-T,14:45,23:00,0,1,Pariwana Cusco\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("turnos.csv", csv_bytes, content_type="text/csv")

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "action": "preview_shifts_inline",
                "create_missing_areas": "1",
                "confirm_full_sync": "1",
                "file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch = ImportBatch.objects.latest("id")
        self.assertEqual(batch.source_type, "shifts_area")
        response = self.client.get(f"{reverse('webui-shifts')}?import_batch_id={batch.id}&import_modal=1")
        self.assertContains(response, "Vista previa de turnos")
        self.assertContains(response, "Confirmar importacion")

        response = self.client.post(
            reverse("webui-shifts"),
            {
                "action": "confirm_shifts_import_inline",
                "batch_id": str(batch.id),
                "confirm_apply_sync": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="REC-T",
            ).exists()
        )

    def test_auto_shifts_page_can_update_and_deactivate(self):
        auto_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="AUTO_1515-0000",
            buk_code="AUTO-1",
            start_time="15:15",
            end_time="00:00",
            is_night_shift=True,
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-auto-shifts"))
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "update_auto_shift",
                "shift_id": auto_shift.id,
                "name": "AUTO_BAR_1515-0000",
                "buk_code": "BAR-T-EXT",
                "is_night_shift": "on",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        auto_shift.refresh_from_db()
        self.assertEqual(auto_shift.name, "AUTO_BAR_1515-0000")
        self.assertEqual(auto_shift.buk_code, "BAR-T-EXT")
        self.assertTrue(auto_shift.active)

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "deactivate_auto_shift",
                "shift_id": auto_shift.id,
            },
        )
        self.assertEqual(response.status_code, 302)
        auto_shift.refresh_from_db()
        self.assertFalse(auto_shift.active)

    def test_auto_shifts_page_filters_and_bulk_normalize_visible(self):
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        shift_a = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="AUTO_1515-0000",
            buk_code="AUTO-A",
            start_time="15:15",
            end_time="00:00",
            is_night_shift=True,
            active=True,
        )
        shift_b = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            name="AUTO_1600-0100",
            buk_code="AUTO-B",
            start_time="16:00",
            end_time="01:00",
            is_night_shift=True,
            active=True,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-auto-shifts"), {"q": "1600"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AUTO_1600-0100")
        self.assertNotContains(response, "AUTO_1515-0000")

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "normalize_visible_auto_shifts",
                "filter_area_id": str(area_2.id),
                "filter_active": "active",
            },
        )
        self.assertEqual(response.status_code, 302)
        shift_a.refresh_from_db()
        shift_b.refresh_from_db()
        self.assertEqual(shift_a.name, "AUTO_1515-0000")
        self.assertEqual(shift_b.name, "Bar_AUTO_1600-0100")
        self.assertTrue(shift_b.buk_code.startswith("BAR-16000100"))

    def test_auto_shifts_page_can_merge_auto_shift_into_official(self):
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="90900111",
            first_name="Lia",
            last_name="Torres",
            active=True,
        )
        official_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Recepcion_Tarde",
            buk_code="REC-TRD",
            start_time="14:00",
            end_time="23:00",
            is_night_shift=False,
            active=True,
        )
        auto_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="AUTO_1400-2300",
            buk_code="AUTO-REC-14002300",
            start_time="14:00",
            end_time="23:00",
            is_night_shift=False,
            active=True,
        )
        assignment = ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date="2026-06-12",
            shift=auto_shift,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "merge_auto_shift",
                "shift_id": auto_shift.id,
                "destination_shift_id": official_shift.id,
                "source_mode": "deactivate",
            },
        )
        self.assertEqual(response.status_code, 302)
        assignment.refresh_from_db()
        auto_shift.refresh_from_db()
        self.assertEqual(assignment.shift_id, official_shift.id)
        self.assertFalse(auto_shift.active)
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            action="auto_shift_merge",
            entity_id=str(official_shift.id),
        ).latest("id")
        self.assertEqual(log.before["source_shift_id"], auto_shift.id)
        self.assertEqual(log.before["moved_assignments"], 1)
        self.assertIn(assignment.id, log.before["moved_assignment_ids_sample"])
        self.assertEqual(log.before["moved_assignments_date_from"], "2026-06-12")
        self.assertEqual(log.before["moved_assignments_date_to"], "2026-06-12")
        self.assertFalse(log.after["source_shift_active_after_merge"])

    def test_auto_shifts_page_shows_merge_impact_summary(self):
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="90900122",
            first_name="Mia",
            last_name="Rojas",
            active=True,
        )
        auto_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="AUTO_1000-1845",
            buk_code="AUTO-REC-10001845",
            start_time="10:00",
            end_time="18:45",
            is_night_shift=False,
            active=True,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date="2026-06-01",
            shift=auto_shift,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date="2026-06-03",
            shift=auto_shift,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-auto-shifts"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "2 asignaciones")
        self.assertContains(response, "2026-06-01 a 2026-06-03")

    def test_auto_shift_merge_high_impact_requires_confirmation_text(self):
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="90900123",
            first_name="Ari",
            last_name="Mena",
            active=True,
        )
        official_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Recepcion_Noche",
            buk_code="REC-NCH",
            start_time="22:00",
            end_time="07:00",
            is_night_shift=True,
            active=True,
        )
        auto_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="AUTO_2200-0700",
            buk_code="AUTO-REC-22000700",
            start_time="22:00",
            end_time="07:00",
            is_night_shift=True,
            active=True,
        )
        for day in range(1, 22):
            ScheduleAssignment.objects.create(
                tenant=self.tenant,
                property=self.property,
                worker=worker,
                date=f"2026-06-{day:02d}",
                shift=auto_shift,
            )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "merge_auto_shift",
                "shift_id": auto_shift.id,
                "destination_shift_id": official_shift.id,
                "source_mode": "deactivate",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                shift=auto_shift,
            ).count(),
            21,
        )

        response = self.client.post(
            reverse("webui-auto-shifts"),
            {
                "action": "merge_auto_shift",
                "shift_id": auto_shift.id,
                "destination_shift_id": official_shift.id,
                "source_mode": "deactivate",
                "merge_confirm_text": "CONFIRMAR",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                shift=auto_shift,
            ).count(),
            0,
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            action="auto_shift_merge",
            entity_id=str(official_shift.id),
        ).latest("id")
        self.assertEqual(log.before["moved_assignments"], 21)
        self.assertFalse(log.before["moved_assignment_ids_sample_truncated"])


class WebUiSpecialStatesTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")

        self.user = User.objects.create_user(email="state-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_shifts=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="special_states", is_enabled=True)

    def test_create_special_state_from_webui(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-special-states"),
            {
                "name": "VAC",
                "buk_code": "VAC",
                "active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            SpecialState.objects.filter(
                tenant=self.tenant,
                property=self.property,
                name="VAC",
                buk_code="VAC",
            ).exists()
        )


class WebUiSchedulingTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="44556677",
            first_name="Luz",
            last_name="Mamani",
            active=True,
        )
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
        )
        self.state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="OFF",
            buk_code="OFF",
        )

        self.user = User.objects.create_user(email="sched-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
            can_use_control=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="scheduling", is_enabled=True)

    def test_scheduling_page_loads(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mes anterior")
        self.assertContains(response, "Mes siguiente")
        self.assertContains(response, "Programadas:")
        self.assertContains(response, "Pendientes:")
        self.assertContains(response, 'aria-label="Vista movil de asignacion"')
        self.assertContains(response, "Lun")

    def test_scheduling_page_handles_template_table_missing_gracefully(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        with patch("apps.webui.views.SchedulePatternTemplate.objects.filter", side_effect=OperationalError("missing")):
            response = self.client.get(reverse("webui-scheduling"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se cargaron plantillas de asignacion")

    def test_scheduling_page_hides_workers_outside_supervisor_area_permissions(self):
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            document_number="11112222",
            first_name="Carlos",
            last_name="Rojas",
            active=True,
        )
        supervisor = User.objects.create_user(email="sched-supervisor@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        UserAreaPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )

        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "44556677")
        self.assertNotContains(response, "11112222")

    def test_operator_with_restricted_area_cannot_assign_worker_outside_area(self):
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        worker_2 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            document_number="33334444",
            first_name="Pedro",
            last_name="Diaz",
            active=True,
        )
        shift_2 = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            name="Bar tarde",
            buk_code="BAR-T",
            start_time="14:00",
            end_time="22:00",
        )
        operator = User.objects.create_user(email="sched-operator-limited@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        UserAreaPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )
        self.client.force_login(operator)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": worker_2.id,
                "work_date": "2026-06-15",
                "assignment_value": f"shift:{shift_2.id}",
            },
        )
        self.assertEqual(response.status_code, 403)

    def test_scheduling_assign_creates_assignment(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-15",
                "assignment_value": f"shift:{self.shift.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-15",
                shift=self.shift,
            ).exists()
        )

    def test_scheduling_assign_ajax_creates_assignment(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-15",
                "assignment_value": f"shift:{self.shift.id}",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["display_code"], "REC-M")
        self.assertEqual(payload["assignment_value"], f"shift:{self.shift.id}")
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-15",
                shift=self.shift,
            ).exists()
        )

    def test_scheduling_assign_can_clear_existing_assignment(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-18",
            shift=self.shift,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-18",
                "assignment_value": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-18",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_assignment_delete",
                entity_type="ScheduleAssignment",
            ).exists()
        )

    def test_scheduling_assign_ajax_can_clear_existing_assignment(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-18",
            shift=self.shift,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-18",
                "assignment_value": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertTrue(payload["is_empty"])
        self.assertEqual(payload["display_code"], "")
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-18",
            ).exists()
        )

    def test_scheduling_page_filters_workers_by_query(self):
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="99887766",
            first_name="Carlos",
            last_name="Lopez",
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06", "worker_q": "44556677"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "44556677")
        self.assertNotContains(response, "99887766")

    def test_scheduling_assign_redirect_keeps_worker_query(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "area_id": str(self.area.id),
                "worker_q": "44556677",
                "worker_id": self.worker.id,
                "work_date": "2026-06-21",
                "assignment_value": f"shift:{self.shift.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("worker_q=44556677", response.url)

    def test_scheduling_assign_redirect_keeps_focus_date(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "area_id": str(self.area.id),
                "worker_q": "44556677",
                "focus_date": "2026-06-21",
                "worker_id": self.worker.id,
                "work_date": "2026-06-21",
                "assignment_value": f"shift:{self.shift.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("focus_date=2026-06-21", response.url)
        self.assertIn(f"edited_worker_id={self.worker.id}", response.url)
        self.assertIn("edited_date=2026-06-21", response.url)

    def test_scheduling_page_highlights_recently_edited_cell(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling"),
            {
                "month": "2026-06",
                "focus_date": "2026-06-21",
                "edited_worker_id": str(self.worker.id),
                "edited_date": "2026-06-21",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-day-date="2026-06-21" class="focus-day-cell recent-edit-cell"')

    def test_scheduling_page_highlights_focus_date(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling"),
            {"month": "2026-06", "focus_date": "2026-06-15"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-day-date="2026-06-15" class="focus-day-header"')

    def test_scheduling_page_shows_copy_button_for_assigned_cells(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-15",
            shift=self.shift,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-open-copy-assignment')
        self.assertContains(response, 'data-assignment-value="shift:')
        self.assertContains(response, 'id="copy-assignment-modal"')

    def test_scheduling_copy_cell_range_copies_assignment_to_worker_dates(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-cell-range"),
            {
                "month": "2026-06",
                "area_id": str(self.area.id),
                "source_worker_id": str(self.worker.id),
                "source_date": "2026-06-10",
                "assignment_value": f"shift:{self.shift.id}",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
            },
        )

        self.assertEqual(response.status_code, 302)
        for work_date in ["2026-06-10", "2026-06-11", "2026-06-12"]:
            self.assertTrue(
                ScheduleAssignment.objects.filter(
                    tenant=self.tenant,
                    property=self.property,
                    worker=self.worker,
                    date=work_date,
                    shift=self.shift,
                ).exists()
            )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_copy_cell_range_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["copied"], 3)
        self.assertEqual(log.after["date_from"], "2026-06-10")
        self.assertEqual(log.after["date_to"], "2026-06-12")

    def test_scheduling_copy_cell_range_blocks_closed_target_month(self):
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
            closed_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-cell-range"),
            {
                "month": "2026-06",
                "source_worker_id": str(self.worker.id),
                "source_date": "2026-06-10",
                "assignment_value": f"shift:{self.shift.id}",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).exists()
        )

    def test_scheduling_bulk_state_creates_assignments(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-state"),
            {
                "month": "2026-06",
                "work_date": "2026-06-16",
                "special_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-16",
                special_state=self.state,
            ).exists()
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_bulk_state_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["work_date"], "2026-06-16")
        self.assertEqual(log.after["special_state_id"], self.state.id)
        self.assertEqual(log.after["applied"], 1)

    def test_scheduling_bulk_shift_creates_assignments_for_shift_area(self):
        other_area = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        other_worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=other_area,
            document_number="66778899",
            first_name="Mario",
            last_name="Rojas",
            active=True,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-shift"),
            {
                "month": "2026-06",
                "work_date": "2026-06-17",
                "shift_id": str(self.shift.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-17",
                shift=self.shift,
            ).exists()
        )
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=other_worker,
                date="2026-06-17",
            ).exists()
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_bulk_shift_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["shift_id"], self.shift.id)
        self.assertEqual(log.after["applied"], 1)

    def test_scheduling_bulk_shift_dry_run_does_not_write(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-shift"),
            {
                "month": "2026-06",
                "work_date": "2026-06-17",
                "shift_id": str(self.shift.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-17",
            ).exists()
        )

    def test_scheduling_bulk_shift_redirect_keeps_focus_date(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-shift"),
            {
                "month": "2026-06",
                "work_date": "2026-06-17",
                "shift_id": str(self.shift.id),
                "focus_date": "2026-06-17",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("focus_date=2026-06-17", response.url)

    def test_scheduling_page_bulk_shift_options_respect_selected_area(self):
        other_area = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=other_area,
            name="Bar tarde",
            buk_code="BAR-T",
            start_time="14:00",
            end_time="22:00",
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling"),
            {"month": "2026-06", "area_id": str(self.area.id)},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "06:00-14:45 (Manana)")
        self.assertNotContains(response, "14:00-22:00 (Bar tarde)")

    def test_scheduling_team_report_pdf_downloads_for_allowed_area(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 6, 1),
            shift=self.shift,
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling-team-report-pdf"),
            {
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "area_id": str(self.area.id),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_scheduling_team_report_pdf_accepts_month_range(self):
        for day in range(1, 31):
            ScheduleAssignment.objects.create(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date=date(2026, 6, day),
                shift=self.shift,
                created_by=self.user,
                updated_by=self.user,
            )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling-team-report-pdf"),
            {
                "date_from": "2026-06-01",
                "date_to": "2026-06-30",
                "area_id": str(self.area.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_scheduling_bulk_state_respects_worker_query_filter(self):
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="99887766",
            first_name="Carlos",
            last_name="Lopez",
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-state"),
            {
                "month": "2026-06",
                "work_date": "2026-06-18",
                "special_state_id": str(self.state.id),
                "worker_q": "44556677",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("worker_q=44556677", response.url)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                date="2026-06-18",
                special_state=self.state,
            ).count(),
            1,
        )

    def test_scheduling_bulk_state_dry_run_does_not_write(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-state"),
            {
                "month": "2026-06",
                "work_date": "2026-06-16",
                "special_state_id": str(self.state.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-16",
            ).exists()
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_bulk_state_apply",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_created",
                entity_type="SchedulingPreview",
            ).exists()
        )

    def test_scheduling_bulk_state_redirect_keeps_focus_date(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-state"),
            {
                "month": "2026-06",
                "work_date": "2026-06-16",
                "special_state_id": str(self.state.id),
                "focus_date": "2026-06-16",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("focus_date=2026-06-16", response.url)

    def test_scheduling_bulk_range_state_creates_assignments(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                special_state=self.state,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).count(),
            3,
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_bulk_range_state_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["date_from"], "2026-06-10")
        self.assertEqual(log.after["date_to"], "2026-06-12")
        self.assertEqual(log.after["special_state_id"], self.state.id)
        self.assertEqual(log.after["applied"], 3)

    def test_scheduling_bulk_range_state_dry_run_does_not_write(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_bulk_range_state_apply",
            ).exists()
        )

    def test_scheduling_dry_run_creates_confirm_preview_and_confirm_applies(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Confirmar vista previa")
        self.assertContains(response, "Asignacion masiva por rango")

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                special_state=self.state,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).count(),
            3,
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_created",
                entity_type="SchedulingPreview",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_confirmed",
                entity_type="SchedulingPreview",
            ).exists()
        )

    def test_scheduling_preview_cancel_removes_pending_preview_without_writing_assignments(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
                "dry_run": "1",
            },
        )

        session = self.client.session
        pending = list(session.get("scheduling_pending_previews", []))
        self.assertTrue(pending)
        preview_id = str(pending[0]["id"])

        response = self.client.post(
            reverse("webui-scheduling-preview-cancel"),
            {
                "month": "2026-06",
                "preview_id": preview_id,
                "focus_date": "2026-06-11",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("focus_date=2026-06-11", response.url)

        session = self.client.session
        self.assertEqual(list(session.get("scheduling_pending_previews", [])), [])
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).count(),
            0,
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_canceled",
                entity_type="SchedulingPreview",
                entity_id=preview_id,
            ).exists()
        )

    def test_scheduling_bulk_sundays_state_creates_assignments(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-sundays-state"),
            {
                "month": "2026-06",
                "special_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        sunday_dates = {"2026-06-07", "2026-06-14", "2026-06-21", "2026-06-28"}
        assignments = ScheduleAssignment.objects.filter(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            special_state=self.state,
        )
        self.assertEqual(assignments.count(), 4)
        self.assertEqual({item.date.isoformat() for item in assignments}, sunday_dates)
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_bulk_sundays_state_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["year"], 2026)
        self.assertEqual(log.after["month"], 6)
        self.assertEqual(log.after["sundays"], 4)
        self.assertEqual(log.after["applied"], 4)

    def test_scheduling_bulk_range_state_blocked_when_month_closed(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-10",
                "date_to": "2026-06-12",
                "special_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).exists()
        )

    def test_scheduling_bulk_multi_range_state_applies_and_last_range_wins(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-multi-range-state"),
            {
                "month": "2026-06",
                "range_date_from": ["2026-06-10", "2026-06-12"],
                "range_date_to": ["2026-06-12", "2026-06-13"],
                "range_state_id": [str(self.state.id), str(vac_state.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-10",
                special_state=self.state,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-11",
                special_state=self.state,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-12",
                special_state=vac_state,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-13",
                special_state=vac_state,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_bulk_multi_range_state_apply",
                entity_type="ScheduleAssignment",
            ).exists()
        )

    def test_scheduling_bulk_multi_range_state_dry_run_does_not_write(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-multi-range-state"),
            {
                "month": "2026-06",
                "range_date_from": ["2026-06-10", "2026-06-12"],
                "range_date_to": ["2026-06-12", "2026-06-13"],
                "range_state_id": [str(self.state.id), str(vac_state.id)],
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-13",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_bulk_multi_range_state_apply",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_created",
                entity_type="SchedulingPreview",
            ).exists()
        )

    def test_scheduling_bulk_multi_range_state_blocked_when_month_closed(self):
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-multi-range-state"),
            {
                "month": "2026-06",
                "range_date_from": ["2026-06-10"],
                "range_date_to": ["2026-06-12"],
                "range_state_id": [str(self.state.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-10",
                date__lte="2026-06-12",
            ).exists()
        )

    def test_scheduling_copy_week_copies_assignments(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-01",
            shift=self.shift,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-03",
            special_state=self.state,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-week"),
            {
                "month": "2026-06",
                "source_week_start": "2026-06-01",
                "target_week_start": "2026-06-08",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-08",
                shift=self.shift,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-10",
                special_state=self.state,
            ).exists()
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_copy_week_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["source_week_start"], "2026-06-01")
        self.assertEqual(log.after["target_week_start"], "2026-06-08")
        self.assertEqual(log.after["copy_kind"], "all")
        self.assertEqual(log.after["copied"], 2)

    def test_scheduling_copy_week_redirect_keeps_focus_date(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-week"),
            {
                "month": "2026-06",
                "source_week_start": "2026-06-01",
                "target_week_start": "2026-06-08",
                "focus_date": "2026-06-09",
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("focus_date=2026-06-09", response.url)

    def test_scheduling_copy_week_dry_run_does_not_write(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-01",
            shift=self.shift,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-week"),
            {
                "month": "2026-06",
                "source_week_start": "2026-06-01",
                "target_week_start": "2026-06-08",
                "copy_kind": "all",
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-08",
            ).exists()
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_copy_week_apply",
            ).exists()
        )

    def test_scheduling_copy_week_blocked_when_target_week_has_closed_month(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-05-28",
            shift=self.shift,
        )
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-week"),
            {
                "month": "2026-06",
                "source_week_start": "2026-05-25",
                "target_week_start": "2026-06-01",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-04",
            ).exists()
        )

    def test_scheduling_copy_previous_month_copies_assignments(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-05-05",
            shift=self.shift,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-previous-month"),
            {
                "month": "2026-06",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-05",
                shift=self.shift,
            ).exists()
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_copy_previous_month_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["target_year"], 2026)
        self.assertEqual(log.after["target_month"], 6)
        self.assertEqual(log.after["source_year"], 2026)
        self.assertEqual(log.after["source_month"], 5)
        self.assertEqual(log.after["copy_kind"], "all")
        self.assertEqual(log.after["copied"], 1)

    def test_scheduling_copy_previous_month_blocked_when_month_closed(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-05-05",
            shift=self.shift,
        )
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-previous-month"),
            {
                "month": "2026-06",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-05",
            ).exists()
        )

    def test_scheduling_copy_month_copies_assignments(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-04-05",
            shift=self.shift,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-month"),
            {
                "month": "2026-06",
                "source_month": "2026-04",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-05",
                shift=self.shift,
            ).exists()
        )
        log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="scheduling_copy_month_apply",
            entity_type="ScheduleAssignment",
        ).latest("id")
        self.assertEqual(log.after["target_year"], 2026)
        self.assertEqual(log.after["target_month"], 6)
        self.assertEqual(log.after["source_year"], 2026)
        self.assertEqual(log.after["source_month"], 4)
        self.assertEqual(log.after["copy_kind"], "all")
        self.assertEqual(log.after["copied"], 1)

    def test_scheduling_copy_month_dry_run_does_not_write(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-04-05",
            shift=self.shift,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-month"),
            {
                "month": "2026-06",
                "source_month": "2026-04",
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-05",
            ).exists()
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_copy_month_apply",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_preview_created",
                entity_type="SchedulingPreview",
            ).exists()
        )

    def test_scheduling_copy_month_blocked_when_target_month_closed(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-04-05",
            shift=self.shift,
        )
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-month"),
            {
                "month": "2026-06",
                "source_month": "2026-04",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-05",
            ).exists()
        )

    def test_scheduling_bulk_range_state_can_target_selected_workers(self):
        worker_2 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="11112222",
            first_name="Jose",
            last_name="Flores",
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-range-state"),
            {
                "month": "2026-06",
                "date_from": "2026-06-20",
                "date_to": "2026-06-21",
                "special_state_id": str(self.state.id),
                "worker_ids": [str(self.worker.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-20",
                date__lte="2026-06-21",
                special_state=self.state,
            ).count(),
            2,
        )
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=worker_2,
                date__gte="2026-06-20",
                date__lte="2026-06-21",
            ).count(),
            0,
        )

    def test_scheduling_copy_week_shift_only_does_not_copy_special_states(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-01",
            shift=self.shift,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-02",
            special_state=self.state,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-week"),
            {
                "month": "2026-06",
                "source_week_start": "2026-06-01",
                "target_week_start": "2026-06-08",
                "copy_kind": "shift",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-08",
                shift=self.shift,
            ).exists()
        )
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-09",
            ).exists()
        )

    def test_scheduling_copy_previous_month_state_only_does_not_copy_shifts(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-05-03",
            special_state=self.state,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-05-04",
            shift=self.shift,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-copy-previous-month"),
            {
                "month": "2026-06",
                "copy_kind": "state",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-03",
                special_state=self.state,
            ).exists()
        )
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-04",
            ).exists()
        )

    def test_scheduling_bulk_week_pattern_assigns_weekday_values(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-week-pattern"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "monday_value": f"shift:{self.shift.id}",
                "sunday_value": f"state:{self.state.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                shift=self.shift,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-07",
                special_state=self.state,
            ).exists()
        )
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-02",
            ).exists()
        )

    def test_scheduling_bulk_week_pattern_can_target_selected_workers(self):
        worker_2 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="22334455",
            first_name="Mia",
            last_name="Ruiz",
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-week-pattern"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-01",
                "monday_value": f"state:{self.state.id}",
                "worker_ids": [str(self.worker.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                special_state=self.state,
            ).exists()
        )
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=worker_2,
                date="2026-06-01",
            ).exists()
        )

    def test_scheduling_bulk_week_pattern_skips_invalid_shift_for_worker_area(self):
        area_2 = Area.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Bar",
        )
        shift_2 = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            name="Bar tarde",
            buk_code="BAR-T",
            start_time="14:00",
            end_time="22:00",
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-bulk-week-pattern"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-01",
                "monday_value": f"shift:{shift_2.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
            ).exists()
        )

    def test_scheduling_operational_rule_applies_weekday_and_sunday(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-operational-rule"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "weekday_value": f"shift:{self.shift.id}",
                "sunday_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                shift=self.shift,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-07",
                special_state=self.state,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_operational_rule_apply",
                entity_type="ScheduleAssignment",
            ).exists()
        )

    def test_scheduling_operational_rule_dry_run_does_not_write(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-operational-rule"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "weekday_value": f"shift:{self.shift.id}",
                "sunday_state_id": str(self.state.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-01",
                date__lte="2026-06-07",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="scheduling_operational_rule_apply",
            ).exists()
        )

    def test_scheduling_operational_rule_blocked_when_month_closed(self):
        MonthClosure.objects.create(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
            status=MonthClosureStatus.CLOSED,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-operational-rule"),
            {
                "month": "2026-06",
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "weekday_value": f"shift:{self.shift.id}",
                "sunday_state_id": str(self.state.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-01",
                date__lte="2026-06-07",
            ).exists()
        )

    def test_scheduling_save_range_template_creates_template(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-save-range-template"),
            {
                "month": "2026-06",
                "template_name": "Vacaciones quincena",
                "range_date_from": ["2026-06-01", "2026-06-15"],
                "range_date_to": ["2026-06-07", "2026-06-16"],
                "range_state_id": [str(self.state.id), str(vac_state.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        template = ScheduleRangeTemplate.objects.get(
            tenant=self.tenant,
            property=self.property,
            name="Vacaciones quincena",
        )
        self.assertEqual(len(template.ranges), 2)
        self.assertEqual(template.ranges[0]["start_day"], 1)
        self.assertEqual(template.ranges[0]["end_day"], 7)

    def test_scheduling_range_template_table_shows_semaphore(self):
        ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla con error",
            ranges=[{"start_day": 1, "end_day": 3, "special_state_id": 999999}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Plantilla con error")
        self.assertContains(response, "Rojo")

    def test_scheduling_range_template_table_filters_by_risk(self):
        ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla roja",
            ranges=[{"start_day": 1, "end_day": 3, "special_state_id": 999999}],
            created_by=self.user,
            updated_by=self.user,
        )
        ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla verde",
            ranges=[{"start_day": 1, "end_day": 30, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-scheduling"),
            {"month": "2026-06", "range_template_risk": "error"},
        )
        self.assertEqual(response.status_code, 200)
        rows = list(response.context["range_templates_admin_rows"])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["template"].name, "Plantilla roja")
        self.assertEqual(rows[0]["health"]["level"], "error")

    def test_scheduling_clone_range_template_creates_copy_and_audit(self):
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla a corregir",
            ranges=[{"start_day": 1, "end_day": 3, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-clone-range-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("range_template_edit_id=", response.url)
        clones = ScheduleRangeTemplate.objects.filter(
            tenant=self.tenant,
            property=self.property,
            area=template.area,
            ranges=template.ranges,
        ).order_by("id")
        self.assertEqual(clones.count(), 2)
        clone = clones.last()
        self.assertNotEqual(clone.id, template.id)
        self.assertIn("Plantilla a corregir (copia", clone.name)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_clone",
                entity_type="ScheduleRangeTemplate",
                entity_id=clone.id,
            ).exists()
        )
        edit_response = self.client.get(response.url)
        self.assertEqual(edit_response.status_code, 200)
        self.assertContains(edit_response, 'name="range_date_from" value="2026-06-01"', html=False)
        self.assertContains(edit_response, f'name="template_id"', html=False)

    def test_scheduling_create_range_template_version_from_editor_rows(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla versionable",
            ranges=[{"start_day": 1, "end_day": 2, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-create-range-template-version"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "range_date_from": ["2026-06-10", "2026-06-15"],
                "range_date_to": ["2026-06-12", "2026-06-16"],
                "range_state_id": [str(vac_state.id), str(self.state.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("range_template_edit_id=", response.url)

        versions = list(
            ScheduleRangeTemplate.objects.filter(
                tenant=self.tenant,
                property=self.property,
                name__istartswith="Plantilla versionable",
            ).order_by("id")
        )
        self.assertEqual(len(versions), 2)
        new_template = versions[-1]
        self.assertNotEqual(new_template.id, template.id)
        self.assertEqual(new_template.name, "Plantilla versionable v2")
        self.assertEqual(
            new_template.ranges,
            [
                {"start_day": 10, "end_day": 12, "special_state_id": vac_state.id},
                {"start_day": 15, "end_day": 16, "special_state_id": self.state.id},
            ],
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_version_create",
                entity_type="ScheduleRangeTemplate",
                entity_id=new_template.id,
            ).exists()
        )

    def test_scheduling_apply_range_template_creates_assignments(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla Rango",
            ranges=[
                {"start_day": 1, "end_day": 2, "special_state_id": self.state.id},
                {"start_day": 3, "end_day": 3, "special_state_id": vac_state.id},
            ],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-range-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "allow_risky_template": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                special_state=self.state,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-03",
                special_state=vac_state,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_apply",
                entity_type="ScheduleRangeTemplate",
            ).exists()
        )

    def test_scheduling_apply_range_template_dry_run_does_not_write(self):
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla dry range",
            ranges=[{"start_day": 1, "end_day": 2, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-range-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-01",
                date__lte="2026-06-02",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_apply",
            ).exists()
        )

    def test_scheduling_apply_range_template_blocks_red_for_non_admin(self):
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla roja",
            ranges=[
                {"start_day": 1, "end_day": 2, "special_state_id": self.state.id},
                {"start_day": 3, "end_day": 4, "special_state_id": 999999},
            ],
            created_by=self.user,
            updated_by=self.user,
        )
        operator = User.objects.create_user(email="operator-range@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        self.client.force_login(operator)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-range-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-01",
                date__lte="2026-06-04",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=operator,
                action="schedule_range_template_apply",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=operator,
                action__in=["schedule_range_template_override_denied", "schedule_range_template_blocked"],
                entity_type="ScheduleRangeTemplate",
                entity_id=template.id,
            ).exists()
        )

    def test_scheduling_apply_range_template_allows_admin_override_on_red(self):
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla roja admin",
            ranges=[
                {"start_day": 1, "end_day": 2, "special_state_id": self.state.id},
                {"start_day": 3, "end_day": 4, "special_state_id": 999999},
            ],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-range-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "allow_risky_template": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                special_state=self.state,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_override_requested",
                entity_type="ScheduleRangeTemplate",
                entity_id=template.id,
            ).exists()
        )
        apply_log = AuditLog.objects.filter(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="schedule_range_template_apply",
            entity_type="ScheduleRangeTemplate",
            entity_id=template.id,
        ).latest("id")
        self.assertTrue(apply_log.after["allow_risky_template"])
        self.assertEqual(apply_log.after["skipped_invalid_ranges"], 1)

    def test_scheduling_page_hides_range_override_for_operator(self):
        operator = User.objects.create_user(email="operator-hide@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        self.client.force_login(operator)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-scheduling"), {"month": "2026-06"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Aplicar con observaciones (solo administrador)")

    def test_scheduling_range_template_preview_shows_detail(self):
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla detalle",
            ranges=[{"start_day": 5, "end_day": 7, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-range-template-preview"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Detalle plantilla: Plantilla detalle")
        self.assertContains(response, "2026-06-05")
        self.assertContains(response, "2026-06-07")
        self.assertContains(response, "OFF")
        self.assertContains(response, "Impacto estimado")
        self.assertContains(response, "Total 3")

    def test_scheduling_update_range_template_ranges_replaces_content(self):
        vac_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
        )
        template = ScheduleRangeTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla editable",
            ranges=[{"start_day": 1, "end_day": 2, "special_state_id": self.state.id}],
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-update-range-template-ranges"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "range_date_from": ["2026-06-10", "2026-06-15"],
                "range_date_to": ["2026-06-12", "2026-06-16"],
                "range_state_id": [str(vac_state.id), str(self.state.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        template.refresh_from_db()
        self.assertEqual(len(template.ranges), 2)
        self.assertEqual(template.ranges[0]["start_day"], 10)
        self.assertEqual(template.ranges[0]["special_state_id"], vac_state.id)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_range_template_ranges_update",
                entity_type="ScheduleRangeTemplate",
                entity_id=template.id,
            ).exists()
        )

    def test_scheduling_save_week_pattern_template_creates_template(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-save-week-pattern-template"),
            {
                "month": "2026-06",
                "template_name": "Patron base",
                "monday_value": f"shift:{self.shift.id}",
                "sunday_value": f"state:{self.state.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        template = SchedulePatternTemplate.objects.get(
            tenant=self.tenant,
            property=self.property,
            name="Patron base",
        )
        self.assertEqual(template.pattern.get("monday"), f"shift:{self.shift.id}")
        self.assertEqual(template.pattern.get("sunday"), f"state:{self.state.id}")

    def test_scheduling_apply_week_pattern_template_creates_assignments(self):
        template = SchedulePatternTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Patron aplicar",
            pattern={
                "monday": f"shift:{self.shift.id}",
                "sunday": f"state:{self.state.id}",
            },
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-week-pattern-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
                shift=self.shift,
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-07",
                special_state=self.state,
            ).exists()
        )

    def test_scheduling_apply_week_pattern_template_dry_run_does_not_write(self):
        template = SchedulePatternTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Patron dry run",
            pattern={
                "monday": f"shift:{self.shift.id}",
                "sunday": f"state:{self.state.id}",
            },
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-week-pattern-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "date_from": "2026-06-01",
                "date_to": "2026-06-07",
                "dry_run": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date__gte="2026-06-01",
                date__lte="2026-06-07",
            ).count(),
            0,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="schedule_pattern_template_apply",
            ).exists()
        )

    def test_scheduling_apply_week_pattern_template_requires_matching_area_filter(self):
        area_2 = Area.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Bar",
        )
        template = SchedulePatternTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            name="Patron bar",
            pattern={"monday": f"state:{self.state.id}"},
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-apply-week-pattern-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "date_from": "2026-06-01",
                "date_to": "2026-06-01",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-01",
            ).exists()
        )

    def test_scheduling_update_week_pattern_template(self):
        template = SchedulePatternTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla Inicial",
            pattern={"monday": f"state:{self.state.id}"},
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-update-week-pattern-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
                "template_name": "Plantilla Actualizada",
            },
        )
        self.assertEqual(response.status_code, 302)
        template.refresh_from_db()
        self.assertEqual(template.name, "Plantilla Actualizada")
        self.assertFalse(template.active)

    def test_scheduling_delete_week_pattern_template(self):
        template = SchedulePatternTemplate.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Plantilla Borrar",
            pattern={"monday": f"state:{self.state.id}"},
            active=True,
        )
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-scheduling-delete-week-pattern-template"),
            {
                "month": "2026-06",
                "template_id": str(template.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(SchedulePatternTemplate.objects.filter(id=template.id).exists())


class WebUiBukReportTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="77889900",
            first_name="Mario",
            last_name="Soto",
            active=True,
        )
        self.worker_2 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area_2,
            document_number="99887766",
            first_name="Lucia",
            last_name="Perez",
            active=True,
        )
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
        )
        self.shift_2 = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area_2,
            name="Tarde",
            buk_code="BAR-T",
            start_time="14:00",
            end_time="22:00",
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-10",
            shift=self.shift,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker_2,
            date="2026-06-10",
            shift=self.shift_2,
        )

        self.user = User.objects.create_user(email="buk-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_validator", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_export", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_buk_report_page_loads(self):
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mario Soto")

    def test_buk_report_csv_export_works(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "export",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "format": "csv",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])

    def test_buk_report_preview_marks_empty_cells(self):
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-11", "date_to": "2026-06-11"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<td class="cell-empty"></td>', html=False)

    def test_buk_report_preview_marks_invalid_cells(self):
        self.shift.active = False
        self.shift.save(update_fields=["active", "updated_at"])

        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<td class="cell-invalid">REC-M</td>', html=False)

    def test_buk_report_preview_can_filter_by_area(self):
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10", "area_ids": [str(self.area.id)]},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mario Soto")
        self.assertNotContains(response, "Lucia Perez")

    def test_buk_report_preview_can_filter_by_worker(self):
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10", "worker_ids": [str(self.worker.id)]},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mario Soto")
        self.assertNotContains(response, "Lucia Perez")

    def test_supervisor_buk_report_is_limited_to_permitted_areas(self):
        supervisor = User.objects.create_user(email="buk-supervisor@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        UserAreaPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mario Soto")
        self.assertNotContains(response, "Lucia Perez")

    def test_supervisor_buk_report_ignores_unpermitted_area_filter(self):
        supervisor = User.objects.create_user(email="buk-supervisor-2@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        UserAreaPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(
            reverse("webui-buk-report"),
            {
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "area_ids": [str(self.area_2.id)],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mario Soto")
        self.assertNotContains(response, "Lucia Perez")

    def test_supervisor_buk_report_csv_export_only_includes_permitted_area(self):
        supervisor = User.objects.create_user(email="buk-supervisor-csv@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        UserAreaPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "export",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "format": "csv",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        content = response.content.decode("utf-8")
        self.assertIn("Mario Soto", content)
        self.assertNotIn("Lucia Perez", content)

    def test_operator_cannot_export_with_observations(self):
        self.worker.document_number = ""
        self.worker.save(update_fields=["document_number", "updated_at"])
        self._activate_context()
        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "export",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "format": "xlsx",
                "export_with_observations": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/html", response["Content-Type"])

    def test_buk_report_can_compare_template_file(self):
        self._activate_context()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte carga BUK"
        ws.cell(1, 1, "Trabajadores")
        ws.cell(2, 1, "RUT")
        ws.cell(2, 2, "Nombre")
        ws.cell(2, 3, "Area")
        ws.cell(1, 4, "06-2026")
        ws.cell(2, 4, "10-06-2026")
        ws.cell(3, 1, "12345678")
        ws.freeze_panes = "D3"
        out = BytesIO()
        wb.save(out)
        uploaded = SimpleUploadedFile(
            "referencia_buk.xlsx",
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "compare_template",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "reference_file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Herramienta avanzada: comparar con Excel base")
        self.assertNotContains(response, "Log ID:")
        self.assertTrue(BukTemplateCompareLog.objects.exists())
        log = BukTemplateCompareLog.objects.latest("id")
        self.assertEqual(len(log.reference_file_sha256), 64)
        self.assertGreater(log.reference_file_size_bytes, 0)

    def test_buk_report_compare_can_download_json_report(self):
        self._activate_context()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte carga BUK"
        ws.cell(1, 1, "Trabajadores")
        ws.cell(2, 1, "RUT")
        ws.cell(2, 2, "Nombre")
        ws.cell(2, 3, "Area")
        ws.cell(1, 4, "06-2026")
        ws.cell(2, 4, "10-06-2026")
        ws.cell(3, 1, "12345678")
        ws.freeze_panes = "D3"
        out = BytesIO()
        wb.save(out)
        uploaded = SimpleUploadedFile(
            "referencia_buk.xlsx",
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "compare_template",
                "download_compare_json": "1",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "reference_file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertEqual(BukTemplateCompareLog.objects.count(), 1)
        log = BukTemplateCompareLog.objects.first()
        self.assertEqual(len(log.reference_file_sha256), 64)
        self.assertGreater(log.reference_file_size_bytes, 0)

    def test_buk_report_page_hides_compare_history(self):
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 6, 1),
            date_to=date(2026, 6, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok.xlsx",
            reference_file_sha256="c" * 64,
            reference_file_size_bytes=111,
            is_compatible=True,
            errors_count=0,
            warnings_count=0,
            result_payload={"is_compatible": True},
        )
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 6, 3),
            date_to=date(2026, 6, 4),
            sheet_name="Reporte carga BUK",
            reference_file_name="incompatible.xlsx",
            reference_file_sha256="d" * 64,
            reference_file_size_bytes=222,
            is_compatible=False,
            errors_count=2,
            warnings_count=1,
            result_payload={"is_compatible": False},
        )
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "compare_result": "incompatible",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Historial de comparaciones tecnicas")
        self.assertNotContains(response, "incompatible.xlsx")
        self.assertNotContains(response, "ok.xlsx")

    def test_buk_report_history_can_download_json_by_log_id(self):
        item = BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 6, 1),
            date_to=date(2026, 6, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok.xlsx",
            reference_file_sha256="c" * 64,
            reference_file_size_bytes=111,
            is_compatible=True,
            errors_count=0,
            warnings_count=0,
            result_payload={"is_compatible": True},
        )
        self._activate_context()
        response = self.client.get(reverse("webui-buk-report-compare-log-download", args=[item.id]))
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])

    def test_buk_report_history_can_download_csv(self):
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 6, 1),
            date_to=date(2026, 6, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok-ui-csv.xlsx",
            reference_file_sha256="c" * 64,
            reference_file_size_bytes=111,
            is_compatible=True,
            errors_count=0,
            warnings_count=0,
            result_payload={"is_compatible": True},
        )
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report-compare-logs-csv-download"),
            {
                "compare_result": "compatible",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertIn("ok-ui-csv.xlsx", response.content.decode("utf-8"))

    def test_buk_report_history_is_not_rendered_on_page(self):
        for idx in range(1, 5):
            BukTemplateCompareLog.objects.create(
                tenant=self.tenant,
                property=self.property,
                compared_by=self.user,
                date_from=date(2026, 6, idx),
                date_to=date(2026, 6, idx),
                sheet_name="Reporte carga BUK",
                reference_file_name=f"hist-{idx}.xlsx",
                reference_file_sha256=str(idx) * 64,
                reference_file_size_bytes=100 + idx,
                is_compatible=True,
                errors_count=0,
                warnings_count=0,
                result_payload={"is_compatible": True},
            )
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "compare_from": "2026-01-01",
                "compare_page_size": "1",
                "compare_page": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Pagina 1 de 4")
        self.assertNotContains(response, "hist-4.xlsx")
        self.assertNotContains(response, "hist-3.xlsx")
        self.assertNotContains(response, "/api/buk/compare-template-logs/?")
        self.assertNotContains(response, "Descargar CSV historial")


class WebUiBukReportAdminOverrideTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="12121212",
            first_name="Mario",
            last_name="Soto",
            active=True,
        )
        self.worker.document_number = ""
        self.worker.save(update_fields=["document_number", "updated_at"])
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-06-10",
            shift=self.shift,
        )

        self.user = User.objects.create_user(email="buk-admin-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_validator", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_export", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_admin_can_export_with_observations(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-buk-report"),
            {
                "action": "export",
                "date_from": "2026-06-10",
                "date_to": "2026-06-10",
                "format": "xlsx",
                "export_with_observations": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        log = BukExportLog.objects.latest("id")
        self.assertEqual(log.validation_status, "observations")

    def test_admin_preview_shows_validation_summary_counts(self):
        self._activate_context()
        response = self.client.get(
            reverse("webui-buk-report"),
            {"date_from": "2026-06-10", "date_to": "2026-06-10"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Errores bloqueantes: 1")


class WebUiControlTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="90909090",
            first_name="Rosa",
            last_name="Perez",
            active=True,
        )
        self.user = User.objects.create_user(email="control-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
            can_use_control=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="control", is_enabled=True)

    def test_control_page_shows_pending_rows(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-control"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rosa")
        self.assertContains(response, "worker_q=90909090")
        self.assertContains(response, f"area_id={self.area.id}")
        self.assertContains(response, "focus_date=")
        self.assertContains(response, "Resumen por area")
        self.assertContains(response, "Recepcion")
        self.assertContains(response, "Alerta Alta")
        self.assertContains(response, "Cobertura 0%")
        self.assertContains(response, "15 dias sin asignacion de 15 dias requeridos")
        self.assertContains(response, "Detalle de pendientes")
        self.assertContains(response, "Actualizar")

    def test_control_page_excludes_days_outside_worker_validity_range(self):
        self.worker.start_date = date.today() + timedelta(days=30)
        self.worker.save(update_fields=["start_date"])
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-control"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "90909090")

    def test_control_page_respects_operator_area_permissions_when_defined(self):
        area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=area_2,
            document_number="56565656",
            first_name="Luis",
            last_name="Mamani",
            active=True,
        )
        UserAreaPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            can_view=True,
            can_schedule=True,
        )

        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-control"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "90909090")
        self.assertNotContains(response, "56565656")

    def test_control_page_denies_user_without_control_permission(self):
        supervisor = User.objects.create_user(email="control-supervisor@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=supervisor, tenant=self.tenant, role=RoleChoices.SUPERVISOR)
        UserPropertyPermission.objects.create(
            user=supervisor,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        self.client.force_login(supervisor)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        response = self.client.get(reverse("webui-control"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No tienes permisos para usar Control en esta sede.")


class WebUiImportsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.user = User.objects.create_user(email="imports-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="excel_import", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_imports_page_loads(self):
        self._activate_context()
        response = self.client.get(reverse("webui-imports"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Importar Excel original")
        self.assertNotContains(response, 'value="preview_excel"')

    def test_imports_page_shows_human_readable_source_label(self):
        ImportBatch.objects.create(
            tenant=self.tenant,
            property=self.property,
            source_type="workers",
            file_name="workers.csv",
            status="preview",
            created_by=self.user,
            summary={"detected_rows": 1},
        )
        self._activate_context()
        response = self.client.get(reverse("webui-imports"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Trabajadores")

    def test_workers_import_preview_and_confirm(self):
        self._activate_context()
        csv_bytes = (
            "DNI,Nombre,Apellido,Area,Sede\n"
            "11111111,Ana,Quispe,Recepcion,Pariwana Cusco\n"
            "22222222,Luis,Rojas,Housekeeping,Pariwana Cusco\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("workers.csv", csv_bytes, content_type="text/csv")

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "preview_workers",
                "create_missing_areas": "1",
                "confirm_full_sync": "1",
                "file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch = ImportBatch.objects.latest("id")
        self.assertEqual(batch.status, "preview")

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "confirm_batch",
                "batch_id": str(batch.id),
                "confirm_apply_sync": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="22222222",
            ).exists()
        )

    def test_workers_template_download(self):
        self._activate_context()
        response = self.client.get(reverse("webui-workers-template-download"))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

    def test_shifts_template_download(self):
        self._activate_context()
        response = self.client.get(reverse("webui-shifts-template-download"))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

    def test_workers_sample_download_xlsx_and_csv(self):
        self._activate_context()
        Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="99887766",
            first_name="Muestra",
            last_name="QA",
            area=self.area,
            active=True,
        )
        response_xlsx = self.client.get(reverse("webui-workers-sample-download"))
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        self.assertTrue(response_xlsx.content.startswith(b"PK"))

        response_csv = self.client.get(f"{reverse('webui-workers-sample-download')}?format=csv")
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        csv_text = response_csv.content.decode("utf-8-sig")
        self.assertIn("DNI,Nombre,Apellido,Area,Sede", csv_text)
        self.assertIn("99887766,Muestra,QA,Recepcion,Pariwana Cusco", csv_text)

    def test_shifts_sample_download_xlsx_and_csv(self):
        self._activate_context()
        Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="REC-N",
            buk_code="REC-N",
            start_time=time(22, 0),
            end_time=time(6, 0),
            is_night_shift=True,
            active=True,
        )

        response_xlsx = self.client.get(reverse("webui-shifts-sample-download"))
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        self.assertTrue(response_xlsx.content.startswith(b"PK"))

        response_csv = self.client.get(f"{reverse('webui-shifts-sample-download')}?format=csv")
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        csv_text = response_csv.content.decode("utf-8-sig")
        self.assertIn("Area,Turno,Codigo BUK,Hora Inicio,Hora Fin", csv_text)
        self.assertIn("Recepcion,REC-N,REC-N,22:00,06:00", csv_text)

    def test_shifts_import_preview_and_confirm(self):
        self._activate_context()
        csv_bytes = (
            "Area,Turno,Codigo BUK,Hora Inicio,Hora Fin,Nocturno,Activo,Sede\n"
            "Recepcion,REC-T,REC-T,14:45,23:00,0,1,Pariwana Cusco\n"
            "Housekeeping,HK-N,HK-N,22:00,06:00,1,1,Pariwana Cusco\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("turnos.csv", csv_bytes, content_type="text/csv")

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "preview_shifts_area",
                "create_missing_areas": "1",
                "confirm_full_sync": "1",
                "file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch = ImportBatch.objects.latest("id")
        self.assertEqual(batch.source_type, "shifts_area")
        self.assertEqual(batch.status, "preview")

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "confirm_batch",
                "batch_id": str(batch.id),
                "confirm_apply_sync": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="HK-N",
            ).exists()
        )

    def test_shifts_import_preview_and_confirm_xlsx(self):
        self._activate_context()
        wb = Workbook()
        ws = wb.active
        ws.title = "Turnos"
        ws.cell(1, 1).value = "Area"
        ws.cell(1, 2).value = "Turno"
        ws.cell(1, 3).value = "Codigo BUK"
        ws.cell(1, 4).value = "Hora Inicio"
        ws.cell(1, 5).value = "Hora Fin"
        ws.cell(1, 6).value = "Nocturno"
        ws.cell(1, 7).value = "Activo"
        ws.cell(1, 8).value = "Sede"
        ws.cell(2, 1).value = "Recepcion"
        ws.cell(2, 2).value = "REC-T"
        ws.cell(2, 3).value = "REC-T"
        ws.cell(2, 4).value = "14:45"
        ws.cell(2, 5).value = "23:00"
        ws.cell(2, 6).value = 0
        ws.cell(2, 7).value = 1
        ws.cell(2, 8).value = "Pariwana Cusco"
        ws.cell(3, 1).value = "Housekeeping"
        ws.cell(3, 2).value = "HK-N"
        ws.cell(3, 3).value = "HK-N"
        ws.cell(3, 4).value = "22:00"
        ws.cell(3, 5).value = "06:00"
        ws.cell(3, 6).value = 1
        ws.cell(3, 7).value = 1
        ws.cell(3, 8).value = "Pariwana Cusco"
        stream = BytesIO()
        wb.save(stream)
        uploaded = SimpleUploadedFile(
            "turnos.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "preview_shifts_area",
                "create_missing_areas": "1",
                "confirm_full_sync": "1",
                "file": uploaded,
            },
        )
        self.assertEqual(response.status_code, 302)
        batch = ImportBatch.objects.latest("id")
        self.assertEqual(batch.source_type, "shifts_area")
        self.assertEqual(batch.status, "preview")

        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "confirm_batch",
                "batch_id": str(batch.id),
                "confirm_apply_sync": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Shift.objects.filter(
                tenant=self.tenant,
                property=self.property,
                buk_code="HK-N",
            ).exists()
        )

    def test_imports_page_formats_preview_payload_and_summary(self):
        batch = ImportBatch.objects.create(
            tenant=self.tenant,
            property=self.property,
            source_type="excel_original",
            file_name="source.xlsx",
            status="preview",
            created_by=self.user,
            summary={
                "detected": "{'workers': 2, 'shifts': 1}",
                "warnings": ["fila 4 incompleta"],
            },
        )
        ImportPreviewRow.objects.create(
            batch=batch,
            sheet_name="Abril 2026",
            row_number=7,
            action="inspect",
            status="ok",
            payload={"entity": "assignment", "worker": "12345678"},
        )

        self._activate_context()
        response = self.client.get(reverse("webui-imports"), {"batch_id": str(batch.id)})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<strong>workers</strong>: 2", html=True)
        self.assertContains(response, "&quot;entity&quot;: &quot;assignment&quot;")

    def test_imports_preview_modal_shows_blocking_errors(self):
        batch = ImportBatch.objects.create(
            tenant=self.tenant,
            property=self.property,
            source_type="shifts_area",
            file_name="turnos.xlsx",
            status="preview",
            created_by=self.user,
            summary={"detected_rows": 1, "errors": 1},
        )
        ImportPreviewRow.objects.create(
            batch=batch,
            sheet_name="shifts_import",
            row_number=2,
            action="skip",
            status="error",
            message="area does not exist: Recepción",
            payload={"area_name": "Recepción"},
        )

        self._activate_context()
        response = self.client.get(reverse("webui-imports"), {"batch_id": str(batch.id)})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Problemas detectados")
        self.assertContains(response, "Fila 2")
        self.assertContains(response, "area does not exist: Recepción")
        self.assertContains(response, "El boton esta deshabilitado porque hay errores bloqueantes.")

    def test_operator_without_shift_permission_cannot_preview_shift_import(self):
        operator = User.objects.create_user(email="imports-operator@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
            can_manage_shifts=False,
        )
        self.client.force_login(operator)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

        csv_bytes = (
            "Area,Turno,Codigo BUK,Hora Inicio,Hora Fin,Sede\n"
            "Recepcion,REC-T,REC-T,14:45,23:00,Pariwana Cusco\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("turnos.csv", csv_bytes, content_type="text/csv")
        response = self.client.post(
            reverse("webui-imports"),
            {
                "action": "preview_shifts_area",
                "create_missing_areas": "1",
                "file": uploaded,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No tienes permisos para importar turnos.")
        self.assertFalse(ImportBatch.objects.filter(tenant=self.tenant, source_type="shifts_area").exists())


class WebUiMonthClosureTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="40404040",
            first_name="Jorge",
            last_name="Puma",
            active=True,
        )
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
        )

        self.user = User.objects.create_user(email="month-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="month_closure", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="scheduling", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_admin_can_close_and_reopen_month(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-month-closure"),
            {
                "action": "close",
                "month": "2026-06",
            },
        )
        self.assertEqual(response.status_code, 302)
        closure = MonthClosure.objects.get(
            tenant=self.tenant,
            property=self.property,
            year=2026,
            month=6,
        )
        self.assertEqual(closure.status, MonthClosureStatus.CLOSED)

        response = self.client.post(
            reverse("webui-month-closure"),
            {
                "action": "reopen",
                "month": "2026-06",
            },
        )
        self.assertEqual(response.status_code, 302)
        closure.refresh_from_db()
        self.assertEqual(closure.status, MonthClosureStatus.OPEN)

    def test_scheduling_assign_blocked_when_month_closed(self):
        self._activate_context()
        self.client.post(
            reverse("webui-month-closure"),
            {"action": "close", "month": "2026-06"},
        )

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-20",
                "assignment_value": f"shift:{self.shift.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-20",
            ).exists()
        )

    def test_scheduling_assign_ajax_blocked_when_month_closed(self):
        self._activate_context()
        self.client.post(
            reverse("webui-month-closure"),
            {"action": "close", "month": "2026-06"},
        )

        response = self.client.post(
            reverse("webui-scheduling-assign"),
            {
                "month": "2026-06",
                "worker_id": self.worker.id,
                "work_date": "2026-06-20",
                "assignment_value": f"shift:{self.shift.id}",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 409)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("mes esta cerrado", payload["message"])
        self.assertFalse(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker=self.worker,
                date="2026-06-20",
            ).exists()
        )

    def test_month_closure_page_shows_buk_export_generated_status(self):
        self._activate_context()
        self.client.post(
            reverse("webui-month-closure"),
            {"action": "close", "month": "2026-06"},
        )
        BukExportLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            date_from=date(2026, 6, 1),
            date_to=date(2026, 6, 30),
            generated_by=self.user,
            file_name="buk_pariwana_cusco_2026-06.xlsx",
            validation_status="ok",
            errors_count=0,
            warnings_count=0,
        )
        response = self.client.get(reverse("webui-month-closure"), {"month": "2026-06"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reporte BUK del periodo:")
        self.assertContains(response, "GENERADO")
        self.assertContains(response, "buk_pariwana_cusco_2026-06.xlsx")


class WebUiAuditTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.user = User.objects.create_user(email="audit-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="audit", is_enabled=True)
        AuditLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="month_close",
            entity_type="MonthClosure",
            entity_id="1",
            before={"status": "open"},
            after={"status": "closed"},
        )
        AuditLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            user=self.user,
            action="worker_update",
            entity_type="Worker",
            entity_id="2",
            before={"last_name": "A"},
            after={"last_name": "B"},
        )

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_audit_page_loads(self):
        self._activate_context()
        response = self.client.get(reverse("webui-audit"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "month_close")

    def test_audit_page_filters_by_action(self):
        self._activate_context()
        response = self.client.get(reverse("webui-audit"), {"action": "month_close"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "month_close")
        self.assertNotContains(response, "worker_update")

    def test_audit_page_shows_module_disabled_message(self):
        ModuleActivation.objects.filter(tenant=self.tenant, module_key="audit").update(is_enabled=False)
        self._activate_context()
        response = self.client.get(reverse("webui-audit"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modulo desactivado: audit.")


class WebUiGlobalAuditTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.other_tenant = Tenant.objects.create(name="Otro Tenant", slug="otro-tenant")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.other_property = Property.objects.create(tenant=self.other_tenant, name="Otra Sede", slug="otra-sede")
        self.super_admin = User.objects.create_user(
            email="global-audit-super@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
        )
        self.admin = User.objects.create_user(email="global-audit-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        AuditLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            user=self.super_admin,
            action="support_access_start",
            entity_type="TenantSupportAccessSession",
            entity_id="11",
            before={},
            after={"reason": "support"},
        )
        AuditLog.objects.create(
            tenant=self.other_tenant,
            property=self.other_property,
            user=self.super_admin,
            action="tenant_update",
            entity_type="Tenant",
            entity_id="22",
            before={"name": "A"},
            after={"name": "B"},
        )

    def _activate_super_admin_context(self):
        self.client.force_login(self.super_admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_super_admin_can_view_global_audit_and_menu(self):
        self._activate_super_admin_context()
        response = self.client.get(reverse("webui-audit-global"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Auditoria global")
        self.assertContains(response, "Pariwana Hostels")
        self.assertContains(response, "Otro Tenant")
        self.assertContains(response, "support_access_start")
        self.assertContains(response, "tenant_update")
        self.assertContains(response, "/app/audit-global/")

    def test_global_audit_filters_by_tenant(self):
        self._activate_super_admin_context()
        response = self.client.get(reverse("webui-audit-global"), {"tenant_id": self.tenant.id})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "support_access_start")
        self.assertNotContains(response, "tenant_update")

    def test_non_super_admin_cannot_view_global_audit(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()
        response = self.client.get(reverse("webui-audit-global"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo Super Administrador puede ver auditoria global.")


class WebUiUsersPermissionsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.property_2 = Property.objects.create(tenant=self.tenant, name="Pariwana Lima", slug="pariwana-lima")
        self.area_1 = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Housekeeping")
        self.user = User.objects.create_user(email="users-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="users_permissions", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def _create_foreign_tenant_user(self, *, email, is_active=True):
        foreign_tenant = Tenant.objects.create(
            name=f"Tenant externo {email}",
            slug=f"tenant-externo-{email.split('@')[0]}",
        )
        foreign_property = Property.objects.create(
            tenant=foreign_tenant,
            name="Sede externa",
            slug="sede-externa",
        )
        target = User.objects.create_user(
            email=email,
            password="ForeignPass123",
            first_name="Sin cambios",
            last_name="Externo",
            is_active=is_active,
        )
        UserTenantRole.objects.create(user=target, tenant=foreign_tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=foreign_tenant,
            property=foreign_property,
            can_access=True,
        )
        return foreign_tenant, foreign_property, target

    def test_users_permissions_page_loads(self):
        self._activate_context()
        response = self.client.get(reverse("webui-users-permissions"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pariwana Cusco")
        self.assertContains(response, "Pariwana Lima")
        self.assertContains(response, "Tipo de usuario")
        self.assertContains(response, "Sedes permitidas")
        self.assertContains(response, "Todas las sedes")
        self.assertNotContains(response, "Perfil de rol")
        self.assertNotContains(response, "Rol base")

    def test_create_user_and_permissions(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "create_user",
                "email": "nuevo@pariwana.test",
                "password": "StrongPass123",
                "first_name": "Nuevo",
                "last_name": "Usuario",
                "role": "supervisor",
                "can_access": "on",
                "can_schedule": "on",
                "area_ids": [str(self.area_1.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        created = User.objects.get(email="nuevo@pariwana.test")
        self.assertTrue(
            UserTenantRole.objects.filter(
                user=created,
                tenant=self.tenant,
                role=RoleChoices.SUPERVISOR,
            ).exists()
        )
        self.assertTrue(
            UserPropertyPermission.objects.filter(
                user=created,
                tenant=self.tenant,
                property=self.property,
                can_access=True,
                can_schedule=True,
            ).exists()
        )
        self.assertTrue(
            UserAreaPermission.objects.filter(
                user=created,
                tenant=self.tenant,
                property=self.property,
                area=self.area_1,
                can_view=True,
                can_schedule=True,
            ).exists()
        )

    def test_create_admin_with_multiple_selected_properties(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "create_user",
                "email": "multi-admin@pariwana.test",
                "password": "StrongPass123",
                "first_name": "Multi",
                "last_name": "Admin",
                "role": "admin",
                "property_ids": [str(self.property.id), str(self.property_2.id)],
                "can_access": "on",
                "can_manage_users": "on",
                "can_export_buk": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        created = User.objects.get(email="multi-admin@pariwana.test")
        tenant_role = UserTenantRole.objects.get(user=created, tenant=self.tenant)
        self.assertEqual(tenant_role.role, RoleChoices.ADMIN)
        self.assertFalse(tenant_role.all_properties_access)
        self.assertEqual(
            set(UserPropertyPermission.objects.filter(user=created, tenant=self.tenant).values_list("property_id", flat=True)),
            {self.property.id, self.property_2.id},
        )

    def test_create_operator_with_all_properties_access_reaches_future_property(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "create_user",
                "email": "all-operator@pariwana.test",
                "password": "StrongPass123",
                "first_name": "All",
                "last_name": "Operator",
                "role": "operator",
                "all_properties_access": "on",
                "property_ids": [str(self.property.id), str(self.property_2.id)],
                "can_access": "on",
                "can_schedule": "on",
                "can_export_buk": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        created = User.objects.get(email="all-operator@pariwana.test")
        tenant_role = UserTenantRole.objects.get(user=created, tenant=self.tenant)
        self.assertEqual(tenant_role.role, RoleChoices.OPERATOR)
        self.assertTrue(tenant_role.all_properties_access)
        self.assertTrue(tenant_role.property_permissions_template["can_schedule"])
        future_property = Property.objects.create(
            tenant=self.tenant,
            name="Pariwana Miraflores",
            slug="pariwana-miraflores",
        )
        self.assertIn(future_property.id, PermissionService.get_accessible_property_ids(created, self.tenant))
        self.assertTrue(PermissionService.user_can_property_action(created, self.tenant, future_property, "can_schedule"))

    def test_create_role_profile_and_apply_defaults_to_user(self):
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "create_role_profile",
                "name": "Supervisor con reportes",
                "code": "supervisor-reportes",
                "base_role": "supervisor",
                "can_access": "on",
                "can_schedule": "on",
                "can_view_reports": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        profile = RoleProfile.objects.get(tenant=self.tenant, code="supervisor-reportes")
        self.assertEqual(profile.base_role, RoleChoices.SUPERVISOR)
        self.assertTrue(profile.permissions["can_view_reports"])
        self.assertFalse(profile.permissions["can_manage_users"])

        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "create_user",
                "email": "reportes@pariwana.test",
                "password": "StrongPass123",
                "first_name": "Supervisor",
                "last_name": "Reportes",
                "role_profile_id": str(profile.id),
                "role": "operator",
                "apply_role_profile_defaults": "on",
                "area_ids": [str(self.area_1.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        created = User.objects.get(email="reportes@pariwana.test")
        tenant_role = UserTenantRole.objects.get(user=created, tenant=self.tenant)
        self.assertEqual(tenant_role.role, RoleChoices.SUPERVISOR)
        self.assertEqual(tenant_role.role_profile_id, profile.id)
        prop_perm = UserPropertyPermission.objects.get(
            user=created,
            tenant=self.tenant,
            property=self.property,
        )
        self.assertTrue(prop_perm.can_access)
        self.assertTrue(prop_perm.can_schedule)
        self.assertTrue(prop_perm.can_view_reports)
        self.assertFalse(prop_perm.can_manage_users)

    def test_update_user_permissions_and_areas(self):
        target = User.objects.create_user(email="target-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
            can_export_buk=False,
            can_manage_workers=False,
            can_manage_shifts=False,
        )
        UserAreaPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            area=self.area_1,
            can_view=True,
            can_schedule=True,
        )

        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "update_property_permissions",
                "user_id": str(target.id),
                "first_name": "Usuario",
                "last_name": "Actualizado",
                "role": "supervisor",
                "can_access": "on",
                "can_export_buk": "on",
                "area_ids": [str(self.area_2.id)],
            },
        )
        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertEqual(target.first_name, "Usuario")
        self.assertEqual(target.last_name, "Actualizado")
        role = UserTenantRole.objects.get(user=target, tenant=self.tenant)
        self.assertEqual(role.role, RoleChoices.SUPERVISOR)
        perm = UserPropertyPermission.objects.get(user=target, tenant=self.tenant, property=self.property)
        self.assertTrue(perm.can_access)
        self.assertFalse(perm.can_schedule)
        self.assertTrue(perm.can_export_buk)
        self.assertFalse(
            UserAreaPermission.objects.filter(
                user=target,
                tenant=self.tenant,
                property=self.property,
                area=self.area_1,
            ).exists()
        )
        self.assertTrue(
            UserAreaPermission.objects.filter(
                user=target,
                tenant=self.tenant,
                property=self.property,
                area=self.area_2,
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="update",
                entity_type="User",
                entity_id=str(target.id),
            ).exists()
        )

    def test_update_user_adds_multiple_properties_and_renders_permissions(self):
        target = User.objects.create_user(email="target-properties@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_schedule=True,
        )

        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "update_property_permissions",
                "user_id": str(target.id),
                "first_name": "Target",
                "last_name": "Properties",
                "role": "operator",
                "property_ids": [str(self.property.id), str(self.property_2.id)],
                "can_access": "on",
                "can_schedule": "on",
                "can_export_buk": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            set(UserPropertyPermission.objects.filter(user=target, tenant=self.tenant).values_list("property_id", flat=True)),
            {self.property.id, self.property_2.id},
        )
        role = UserTenantRole.objects.get(user=target, tenant=self.tenant)
        self.assertFalse(role.all_properties_access)
        response = self.client.get(reverse("webui-users-permissions"))
        self.assertContains(response, "Pariwana Cusco")
        self.assertContains(response, "Pariwana Lima")
        self.assertContains(response, "Horarios")
        self.assertContains(response, "BUK")

    def test_update_user_sets_all_properties_access(self):
        target = User.objects.create_user(email="target-all-properties@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_users=True,
        )

        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "update_property_permissions",
                "user_id": str(target.id),
                "first_name": "Target",
                "last_name": "All",
                "role": "admin",
                "all_properties_access": "on",
                "property_ids": [str(self.property.id), str(self.property_2.id)],
                "can_access": "on",
                "can_manage_users": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        role = UserTenantRole.objects.get(user=target, tenant=self.tenant)
        self.assertTrue(role.all_properties_access)
        self.assertTrue(role.property_permissions_template["can_manage_users"])
        response = self.client.get(reverse("webui-users-permissions"))
        self.assertContains(response, "Todas las sedes")

    def test_deactivate_user(self):
        target = User.objects.create_user(email="deactivate-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "deactivate_user",
                "user_id": str(target.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertFalse(target.is_active)

    def test_reactivate_user(self):
        target = User.objects.create_user(email="reactivate-ui@pariwana.test", password="StrongPass123", is_active=False)
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "reactivate_user",
                "user_id": str(target.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertTrue(target.is_active)

    def test_delete_inactive_user_permanently(self):
        target = User.objects.create_user(email="destroy-ui@pariwana.test", password="StrongPass123", is_active=False)
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "delete_user_permanently",
                "user_id": str(target.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(User.objects.filter(id=target.id).exists())

    def test_reset_user_password_records_audit(self):
        target = User.objects.create_user(email="reset-ui@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=target, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=target,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
        )
        self._activate_context()
        response = self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "reset_user_password",
                "user_id": str(target.id),
                "new_password": "NewStrongPass123",
                "confirm_password": "NewStrongPass123",
            },
        )
        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertTrue(target.check_password("NewStrongPass123"))
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="password_reset",
                entity_type="User",
                entity_id=str(target.id),
            ).exists()
        )

    def test_direct_post_cannot_update_permissions_for_user_from_another_tenant_by_id(self):
        foreign_tenant, foreign_property, target = self._create_foreign_tenant_user(
            email="foreign-update@pariwana.test"
        )
        original_role = UserTenantRole.objects.get(user=target, tenant=foreign_tenant)
        original_permission = UserPropertyPermission.objects.get(
            user=target,
            tenant=foreign_tenant,
            property=foreign_property,
        )
        self._activate_context()

        self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "update_property_permissions",
                "user_id": str(target.id),
                "first_name": "Alterado",
                "last_name": "Sin autorizacion",
                "role": "admin",
                "can_access": "on",
                "can_manage_users": "on",
                "property_ids": [str(self.property.id)],
                "area_ids": [str(self.area_1.id)],
            },
        )

        target.refresh_from_db()
        original_role.refresh_from_db()
        original_permission.refresh_from_db()
        self.assertEqual(target.first_name, "Sin cambios")
        self.assertEqual(target.last_name, "Externo")
        self.assertEqual(original_role.role, RoleChoices.OPERATOR)
        self.assertTrue(original_permission.can_access)
        self.assertFalse(UserTenantRole.objects.filter(user=target, tenant=self.tenant).exists())
        self.assertFalse(UserPropertyPermission.objects.filter(user=target, tenant=self.tenant).exists())
        self.assertFalse(UserAreaPermission.objects.filter(user=target, tenant=self.tenant).exists())

    def test_direct_post_cannot_reset_password_for_user_from_another_tenant_by_id(self):
        foreign_tenant, _, target = self._create_foreign_tenant_user(email="foreign-reset@pariwana.test")
        self._activate_context()

        self.client.post(
            reverse("webui-users-permissions"),
            {
                "action": "reset_user_password",
                "user_id": str(target.id),
                "new_password": "ChangedPass123",
                "confirm_password": "ChangedPass123",
            },
        )

        target.refresh_from_db()
        self.assertTrue(target.check_password("ForeignPass123"))
        self.assertFalse(target.check_password("ChangedPass123"))
        self.assertTrue(UserTenantRole.objects.filter(user=target, tenant=foreign_tenant).exists())
        self.assertFalse(
            AuditLog.objects.filter(
                tenant=self.tenant,
                entity_type="User",
                entity_id=str(target.id),
                action="password_reset",
            ).exists()
        )

    def test_direct_post_cannot_deactivate_user_from_another_tenant_by_id(self):
        foreign_tenant, _, target = self._create_foreign_tenant_user(email="foreign-deactivate@pariwana.test")
        self._activate_context()

        self.client.post(
            reverse("webui-users-permissions"),
            {"action": "deactivate_user", "user_id": str(target.id)},
        )

        target.refresh_from_db()
        self.assertTrue(target.is_active)
        self.assertTrue(UserTenantRole.objects.filter(user=target, tenant=foreign_tenant).exists())

    def test_direct_post_cannot_reactivate_user_from_another_tenant_by_id(self):
        foreign_tenant, _, target = self._create_foreign_tenant_user(
            email="foreign-reactivate@pariwana.test",
            is_active=False,
        )
        self._activate_context()

        self.client.post(
            reverse("webui-users-permissions"),
            {"action": "reactivate_user", "user_id": str(target.id)},
        )

        target.refresh_from_db()
        self.assertFalse(target.is_active)
        self.assertTrue(UserTenantRole.objects.filter(user=target, tenant=foreign_tenant).exists())

    def test_direct_post_cannot_delete_user_from_another_tenant_by_id(self):
        foreign_tenant, foreign_property, target = self._create_foreign_tenant_user(
            email="foreign-delete@pariwana.test",
            is_active=False,
        )
        target_id = target.id
        self._activate_context()

        self.client.post(
            reverse("webui-users-permissions"),
            {"action": "delete_user_permanently", "user_id": str(target_id)},
        )

        self.assertTrue(User.objects.filter(id=target_id, is_active=False).exists())
        self.assertTrue(UserTenantRole.objects.filter(user_id=target_id, tenant=foreign_tenant).exists())
        self.assertTrue(
            UserPropertyPermission.objects.filter(
                user_id=target_id,
                tenant=foreign_tenant,
                property=foreign_property,
            ).exists()
        )


class WebUiBackupTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time="06:00",
            end_time="14:45",
        )
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            document_number="56565656",
            first_name="Rocio",
            last_name="Solis",
            active=True,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date="2026-07-10",
            shift=self.shift,
        )
        self.user = User.objects.create_user(email="backup-admin@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_manage_workers=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="excel_import", is_enabled=True)

    def _activate_context(self):
        self.client.force_login(self.user)
        session = self.client.session
        session["ui_tenant_id"] = self.tenant.id
        session["ui_property_id"] = self.property.id
        session.save()

    def test_backup_download_returns_json(self):
        self._activate_context()
        response = self.client.get(reverse("webui-backup-download"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/json", response["Content-Type"])
        data = json.loads(response.content.decode("utf-8"))
        self.assertIn("meta", data)
        self.assertIn("workers", data)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="backup_export_json",
            ).exists()
        )

    def test_backup_restore_creates_records(self):
        self._activate_context()
        payload = {
            "areas": [{"name": "Bar", "type": "", "active": True}],
            "special_states": [{"name": "OFF", "buk_code": "OFF", "active": True}],
            "shifts": [
                {
                    "area_name": "Bar",
                    "name": "Bar-T",
                    "buk_code": "BAR-T",
                    "start_time": "14:00:00",
                    "end_time": "22:00:00",
                    "break_start": None,
                    "break_end": None,
                    "is_night_shift": False,
                    "active": True,
                }
            ],
            "workers": [
                {
                    "document_number": "90919191",
                    "first_name": "Luis",
                    "last_name": "Rojas",
                    "area_name": "Bar",
                    "active": True,
                    "start_date": None,
                    "end_date": None,
                    "buk_employee_code": None,
                    "metadata": {},
                }
            ],
            "assignments": [
                {
                    "document_number": "90919191",
                    "date": "2026-07-11",
                    "shift_name": "Bar-T",
                    "special_state_name": None,
                }
            ],
            "month_closures": [{"year": 2026, "month": 7, "status": "closed"}],
            "buk_export_config": {"sheet_name": "Reporte carga BUK"},
        }
        uploaded = SimpleUploadedFile(
            "backup.json",
            json.dumps(payload).encode("utf-8"),
            content_type="application/json",
        )
        response = self.client.post(reverse("webui-backup"), {"file": uploaded})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="90919191",
            ).exists()
        )
        self.assertTrue(
            ScheduleAssignment.objects.filter(
                tenant=self.tenant,
                property=self.property,
                worker__document_number="90919191",
                date="2026-07-11",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="backup_restore_json",
            ).exists()
        )

    def test_backup_restore_preview_shows_summary_without_applying(self):
        self._activate_context()
        payload = {
            "areas": [{"name": "Bar", "type": "", "active": True}],
            "workers": [
                {
                    "document_number": "93939393",
                    "first_name": "Lucho",
                    "last_name": "Paz",
                    "area_name": "Bar",
                    "active": True,
                    "start_date": None,
                    "end_date": None,
                    "buk_employee_code": None,
                    "metadata": {},
                }
            ],
        }
        uploaded = SimpleUploadedFile(
            "backup_preview.json",
            json.dumps(payload).encode("utf-8"),
            content_type="application/json",
        )
        response = self.client.post(
            reverse("webui-backup"),
            {"action": "preview_restore", "file": uploaded},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vista previa de restauracion")
        self.assertFalse(
            Worker.objects.filter(
                tenant=self.tenant,
                property=self.property,
                document_number="93939393",
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="backup_restore_preview_json",
            ).exists()
        )

    def test_backup_restore_sync_mode_deactivates_missing_records(self):
        extra_state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="VAC",
            buk_code="VAC",
            active=True,
        )
        extra_area = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        extra_shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=extra_area,
            name="Bar-T",
            buk_code="BAR-T",
            start_time="14:00",
            end_time="22:00",
            active=True,
        )
        extra_worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=extra_area,
            document_number="11112222",
            first_name="Extra",
            last_name="Persona",
            active=True,
        )

        self._activate_context()
        payload = {
            "areas": [{"name": "Recepcion", "type": "", "active": True}],
            "special_states": [{"name": "OFF", "buk_code": "OFF", "active": True}],
            "shifts": [
                {
                    "area_name": "Recepcion",
                    "name": "Manana",
                    "buk_code": "REC-M",
                    "start_time": "06:00:00",
                    "end_time": "14:45:00",
                    "break_start": None,
                    "break_end": None,
                    "is_night_shift": False,
                    "active": True,
                }
            ],
            "workers": [
                {
                    "document_number": "56565656",
                    "first_name": "Rocio",
                    "last_name": "Solis",
                    "area_name": "Recepcion",
                    "active": True,
                    "start_date": None,
                    "end_date": None,
                    "buk_employee_code": None,
                    "metadata": {},
                }
            ],
        }
        uploaded = SimpleUploadedFile(
            "backup_sync.json",
            json.dumps(payload).encode("utf-8"),
            content_type="application/json",
        )
        preview_response = self.client.post(
            reverse("webui-backup"),
            {"action": "preview_restore", "sync_mode": "1", "file": uploaded},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "Desactivar (solo sync)")

        apply_response = self.client.post(
            reverse("webui-backup"),
            {
                "action": "apply_restore",
                "sync_mode": "1",
                "source_name": "backup_sync.json",
                "payload_json": json.dumps(payload),
            },
        )
        self.assertEqual(apply_response.status_code, 302)
        extra_state.refresh_from_db()
        extra_shift.refresh_from_db()
        extra_worker.refresh_from_db()
        self.assertFalse(extra_state.active)
        self.assertFalse(extra_shift.active)
        self.assertFalse(extra_worker.active)


class WebUiSmokeCommandTests(TestCase):
    def test_smoke_test_webui_passes_after_bootstrap(self):
        call_command(
            "bootstrap_local_demo",
            password="StrongPass123",
            days=7,
            supervisor_areas="Recepción,Housekeeping",
        )
        call_command("smoke_test_webui")

    def test_smoke_test_webui_fails_when_demo_user_missing(self):
        call_command(
            "bootstrap_local_demo",
            password="StrongPass123",
            days=5,
            supervisor_areas="Recepción,Housekeeping",
        )
        User.objects.filter(email="operador.demo@pariwana.local").delete()
        with self.assertRaises(CommandError):
            call_command("smoke_test_webui")

    def test_qa_check_local_passes_after_bootstrap(self):
        call_command(
            "bootstrap_local_demo",
            password="StrongPass123",
            days=5,
            supervisor_areas="Recepción,Housekeeping",
        )
        call_command("qa_check_local")
