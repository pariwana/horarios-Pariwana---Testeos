from dataclasses import dataclass
from datetime import date, datetime, timedelta
import hashlib
from io import BytesIO

from django.db.models import Count, Q
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from apps.buk_exports.models import BukExportConfig, BukExportLog, BukTemplateCompareLog
from apps.scheduling.models import ScheduleAssignment
from apps.workers.models import Shift, SpecialState, Worker


@dataclass
class ValidationIssue:
    date: str
    property_name: str
    area_name: str
    worker_name: str
    problem: str
    severity: str
    suggested_action: str


@dataclass
class TemplateCompatibilityIssue:
    type: str
    code: str
    detail: str


class BukValidationService:
    @staticmethod
    def validate_assignments(*, tenant, property_obj, date_from, date_to, area_ids=None, worker_ids=None):
        issues = []

        workers = Worker.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        ).select_related("area")
        if area_ids:
            workers = workers.filter(area_id__in=area_ids)
        if worker_ids:
            workers = workers.filter(id__in=worker_ids)

        assignments = ScheduleAssignment.objects.filter(
            tenant=tenant,
            property=property_obj,
            date__gte=date_from,
            date__lte=date_to,
        ).select_related("worker", "worker__area", "shift", "special_state")
        if area_ids:
            assignments = assignments.filter(worker__area_id__in=area_ids)
        if worker_ids:
            assignments = assignments.filter(worker_id__in=worker_ids)

        for worker in workers.filter(Q(document_number__isnull=True) | Q(document_number__exact="")):
            issues.append(
                ValidationIssue(
                    date="-",
                    property_name=property_obj.name,
                    area_name=worker.area.name if worker.area_id else "",
                    worker_name=f"{worker.first_name} {worker.last_name}",
                    problem="Trabajador activo sin documento.",
                    severity="error",
                    suggested_action="Completar documento del trabajador.",
                )
            )
        for worker in workers.filter(area_id__isnull=True):
            issues.append(
                ValidationIssue(
                    date="-",
                    property_name=property_obj.name,
                    area_name="",
                    worker_name=f"{worker.first_name} {worker.last_name}",
                    problem="Trabajador activo sin area.",
                    severity="error",
                    suggested_action="Asignar area al trabajador antes de exportar.",
                )
            )
        for worker in workers:
            if worker.start_date and worker.start_date > date_to:
                issues.append(
                    ValidationIssue(
                        date="-",
                        property_name=property_obj.name,
                        area_name=worker.area.name if worker.area_id else "",
                        worker_name=f"{worker.first_name} {worker.last_name}",
                        problem="Trabajador activo fuera de vigencia: inicio posterior al rango.",
                        severity="warning",
                        suggested_action="Revisar estado activo o ajustar fecha de inicio.",
                    )
                )
            if worker.end_date and worker.end_date < date_from:
                issues.append(
                    ValidationIssue(
                        date="-",
                        property_name=property_obj.name,
                        area_name=worker.area.name if worker.area_id else "",
                        worker_name=f"{worker.first_name} {worker.last_name}",
                        problem="Trabajador activo fuera de vigencia: cese anterior al rango.",
                        severity="warning",
                        suggested_action="Revisar estado activo o ajustar fecha de cese.",
                    )
                )

        assigned_worker_ids = set(assignments.values_list("worker_id", flat=True).distinct())
        for worker in workers.exclude(id__in=assigned_worker_ids):
            issues.append(
                ValidationIssue(
                    date="-",
                    property_name=property_obj.name,
                    area_name=worker.area.name if worker.area_id else "",
                    worker_name=f"{worker.first_name} {worker.last_name}",
                    problem="Trabajador activo sin horario en el rango.",
                    severity="warning",
                    suggested_action="Asignar turno o estado especial antes de exportar.",
                )
            )

        for assignment in assignments:
            if assignment.worker.start_date and assignment.date < assignment.worker.start_date:
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem="Asignacion fuera de vigencia: fecha antes del inicio del trabajador.",
                        severity="error",
                        suggested_action="Corregir fecha o fecha de inicio del trabajador.",
                    )
                )
            if assignment.worker.end_date and assignment.date > assignment.worker.end_date:
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem="Asignacion fuera de vigencia: fecha posterior al cese del trabajador.",
                        severity="error",
                        suggested_action="Corregir fecha o fecha de cese del trabajador.",
                    )
                )
            if assignment.special_state_id and not (assignment.special_state.buk_code or "").strip():
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem=f"Estado especial sin codigo BUK: {assignment.special_state.name}.",
                        severity="error",
                        suggested_action="Completar codigo BUK del estado especial.",
                    )
                )
            if assignment.shift_id and not (assignment.shift.buk_code or "").strip():
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem=f"Turno asignado sin codigo BUK: {assignment.shift.name}.",
                        severity="error",
                        suggested_action="Completar codigo BUK del turno asignado.",
                    )
                )
            if assignment.shift_id and not assignment.shift.active:
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem=f"Asignacion usa turno inactivo: {assignment.shift.name}.",
                        severity="warning",
                        suggested_action="Reactivar turno o reasignar uno activo.",
                    )
                )
            if assignment.special_state_id and not assignment.special_state.active:
                issues.append(
                    ValidationIssue(
                        date=assignment.date.isoformat(),
                        property_name=property_obj.name,
                        area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                        worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                        problem=f"Asignacion usa estado especial inactivo: {assignment.special_state.name}.",
                        severity="warning",
                        suggested_action="Reactivar estado especial o corregir asignacion.",
                    )
                )
            if assignment.shift_id:
                expected_night = assignment.shift.end_time <= assignment.shift.start_time
                if bool(assignment.shift.is_night_shift) != bool(expected_night):
                    issues.append(
                        ValidationIssue(
                            date=assignment.date.isoformat(),
                            property_name=property_obj.name,
                            area_name=assignment.worker.area.name if assignment.worker.area_id else "",
                            worker_name=f"{assignment.worker.first_name} {assignment.worker.last_name}",
                            problem=f"Turno con bandera nocturna inconsistente: {assignment.shift.name}.",
                            severity="warning",
                            suggested_action="Revisar definicion del turno (is_night_shift).",
                        )
                    )

        scope_area_ids = set(workers.values_list("area_id", flat=True))
        shifts = Shift.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        )
        if area_ids:
            shifts = shifts.filter(area_id__in=area_ids)
        elif worker_ids:
            shifts = shifts.filter(area_id__in=scope_area_ids)
        for shift in shifts.filter(Q(buk_code__isnull=True) | Q(buk_code__exact="")):
            issues.append(
                ValidationIssue(
                    date="-",
                    property_name=property_obj.name,
                    area_name=shift.area.name if shift.area_id else "",
                    worker_name="-",
                    problem=f"Turno activo sin codigo BUK: {shift.name}.",
                    severity="error",
                    suggested_action="Completar codigo BUK del turno.",
                )
            )

        code_registry = {}
        for shift in shifts.exclude(Q(buk_code__isnull=True) | Q(buk_code__exact="")):
            code = (shift.buk_code or "").strip()
            if not code:
                continue
            code_registry.setdefault(code, []).append(f"Shift:{shift.id}")

        states = SpecialState.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        )
        for state in states.exclude(Q(buk_code__isnull=True) | Q(buk_code__exact="")):
            code = (state.buk_code or "").strip()
            if not code:
                continue
            code_registry.setdefault(code, []).append(f"SpecialState:{state.id}")

        for code, entries in code_registry.items():
            if len(entries) > 1:
                issues.append(
                    ValidationIssue(
                        date="-",
                        property_name=property_obj.name,
                        area_name="-",
                        worker_name="-",
                        problem=f"Codigo BUK duplicado: {code}.",
                        severity="error",
                        suggested_action="Unificar o corregir codigos BUK duplicados.",
                    )
                )

        duplicates = (
            ScheduleAssignment.objects.filter(
                tenant=tenant,
                property=property_obj,
                date__gte=date_from,
                date__lte=date_to,
            )
            .values("worker_id", "date")
            .annotate(total=Count("id"))
            .filter(total__gt=1)
        )
        duplicate_worker_ids = [item["worker_id"] for item in duplicates]
        worker_index = {
            worker.id: worker
            for worker in Worker.objects.filter(id__in=duplicate_worker_ids).select_related("area")
        }
        for item in duplicates:
            worker = worker_index.get(item["worker_id"])
            issues.append(
                ValidationIssue(
                    date=item["date"].isoformat(),
                    property_name=property_obj.name,
                    area_name=(worker.area.name if worker and worker.area_id else ""),
                    worker_name=(f"{worker.first_name} {worker.last_name}" if worker else "-"),
                    problem="Asignaciones duplicadas para trabajador y fecha.",
                    severity="error",
                    suggested_action="Eliminar duplicados y dejar una sola asignacion por dia.",
                )
            )
        return issues


class BukExportService:
    @staticmethod
    def _cleanup_mojibake_text(value):
        text = str(value or "")
        replacements = {
            "Ã": "Á",
            "Ã‰": "É",
            "Ã": "Í",
            "Ã“": "Ó",
            "Ãš": "Ú",
            "ÃƒÂ": "Á",
            "Ãƒâ€°": "É",
            "ÃƒÂ": "Í",
            "Ãƒâ€œ": "Ó",
            "ÃƒÅ¡": "Ú",
            "Ã¡": "á",
            "Ã©": "é",
            "Ã­": "í",
            "Ã³": "ó",
            "Ãº": "ú",
            "ÃƒÂ¡": "á",
            "ÃƒÂ©": "é",
            "ÃƒÂ­": "í",
            "ÃƒÂ³": "ó",
            "ÃƒÂº": "ú",
            "Ã±": "ñ",
            "Ã‘": "Ñ",
            "ÃƒÂ±": "ñ",
            "Ãƒâ€˜": "Ñ",
        }
        for source, target in replacements.items():
            text = text.replace(source, target)
        return text

    @staticmethod
    def _normalize_compare_text(value):
        text = BukExportService._cleanup_mojibake_text(value).strip().lower()
        replacements = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
            "ñ": "n",
        }
        for source, target in replacements.items():
            text = text.replace(source, target)
        return text

    @staticmethod
    def _coerce_header_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or "").strip()
        if not text:
            return None
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _detect_date_format(values):
        cleaned = [str(item or "").strip() for item in values if str(item or "").strip()]
        if not cleaned:
            return "unknown"
        patterns = {
            "%d-%m-%Y": all(BukExportService._coerce_header_date(v) is not None and "-" in v for v in cleaned),
            "%d/%m/%Y": all(BukExportService._coerce_header_date(v) is not None and "/" in v for v in cleaned),
        }
        for fmt, ok in patterns.items():
            if ok:
                return fmt
        if all(BukExportService._coerce_header_date(v) is not None for v in cleaned):
            return "date-parsed"
        return "mixed"

    @staticmethod
    def _extract_template_signature(worksheet):
        max_scan_row = min(max(worksheet.max_row, 1), 30)
        max_scan_col = min(max(worksheet.max_column, 1), 120)
        best_row = None
        best_date_cols = []
        for row_idx in range(1, max_scan_row + 1):
            date_cols = []
            for col_idx in range(1, max_scan_col + 1):
                value = worksheet.cell(row_idx, col_idx).value
                if BukExportService._coerce_header_date(value) is not None:
                    date_cols.append(col_idx)
            if len(date_cols) > len(best_date_cols):
                best_row = row_idx
                best_date_cols = date_cols
        if not best_date_cols:
            return {
                "header_row": None,
                "first_data_row": None,
                "first_date_col": None,
                "fixed_labels": [],
                "date_values": [],
                "date_format": "unknown",
                "freeze_panes": str(worksheet.freeze_panes or ""),
            }

        first_date_col = min(best_date_cols)
        fixed_labels = []
        for col_idx in range(1, first_date_col):
            value = worksheet.cell(best_row, col_idx).value
            text = BukExportService._cleanup_mojibake_text(value).strip()
            if text:
                fixed_labels.append(text)
        date_values = [worksheet.cell(best_row, col_idx).value for col_idx in best_date_cols]
        first_data_row = None
        for row_idx in range(best_row + 1, max_scan_row + 1):
            if str(worksheet.cell(row_idx, 1).value or "").strip():
                first_data_row = row_idx
                break
        if first_data_row is None:
            first_data_row = best_row + 1

        return {
            "header_row": best_row,
            "first_data_row": first_data_row,
            "first_date_col": first_date_col,
            "fixed_labels": fixed_labels,
            "date_values": date_values,
            "date_format": BukExportService._detect_date_format(date_values),
            "freeze_panes": str(worksheet.freeze_panes or ""),
        }

    @staticmethod
    def compare_template_compatibility(*, reference_file_bytes, candidate_file_bytes, sheet_name="Reporte carga BUK"):
        issues = []
        reference_wb = load_workbook(BytesIO(reference_file_bytes), data_only=True)
        candidate_wb = load_workbook(BytesIO(candidate_file_bytes), data_only=True)

        if sheet_name not in reference_wb.sheetnames:
            issues.append(
                TemplateCompatibilityIssue(
                    type="error",
                    code="reference_sheet_missing",
                    detail=f"No existe la hoja '{sheet_name}' en el archivo de referencia.",
                )
            )
            return {
                "is_compatible": False,
                "errors": [item.__dict__ for item in issues],
                "warnings": [],
                "reference": {},
                "candidate": {},
            }

        if sheet_name not in candidate_wb.sheetnames:
            issues.append(
                TemplateCompatibilityIssue(
                    type="error",
                    code="candidate_sheet_missing",
                    detail=f"No existe la hoja '{sheet_name}' en el archivo generado.",
                )
            )
            return {
                "is_compatible": False,
                "errors": [item.__dict__ for item in issues],
                "warnings": [],
                "reference": {},
                "candidate": {},
            }

        reference = BukExportService._extract_template_signature(reference_wb[sheet_name])
        candidate = BukExportService._extract_template_signature(candidate_wb[sheet_name])

        if reference["header_row"] is None or candidate["header_row"] is None:
            issues.append(
                TemplateCompatibilityIssue(
                    type="error",
                    code="header_not_detected",
                    detail="No se pudo detectar fila de encabezado con fechas en uno de los archivos.",
                )
            )
        else:
            if reference["header_row"] != candidate["header_row"]:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="error",
                        code="header_row_mismatch",
                        detail=(
                            "Fila de encabezado distinta: "
                            f"referencia={reference['header_row']} generado={candidate['header_row']}."
                        ),
                    )
                )
            if reference["first_date_col"] != candidate["first_date_col"]:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="error",
                        code="first_date_col_mismatch",
                        detail=(
                            "Columna inicial de fechas distinta: "
                            f"referencia={reference['first_date_col']} generado={candidate['first_date_col']}."
                        ),
                    )
                )

            ref_labels = [BukExportService._normalize_compare_text(item) for item in reference["fixed_labels"]]
            out_labels = [BukExportService._normalize_compare_text(item) for item in candidate["fixed_labels"]]
            if ref_labels != out_labels:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="error",
                        code="fixed_labels_mismatch",
                        detail=(
                            f"Columnas fijas distintas. referencia={reference['fixed_labels']} "
                            f"generado={candidate['fixed_labels']}."
                        ),
                    )
                )

            if reference["date_format"] != "unknown" and candidate["date_format"] != reference["date_format"]:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="warning",
                        code="date_format_mismatch",
                        detail=(
                            "Formato de fecha distinto en encabezado: "
                            f"referencia={reference['date_format']} generado={candidate['date_format']}."
                        ),
                    )
                )

            if candidate["first_data_row"] != reference["first_data_row"]:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="warning",
                        code="first_data_row_mismatch",
                        detail=(
                            "Fila inicial de datos distinta: "
                            f"referencia={reference['first_data_row']} generado={candidate['first_data_row']}."
                        ),
                    )
                )

            if candidate["freeze_panes"] != reference["freeze_panes"]:
                issues.append(
                    TemplateCompatibilityIssue(
                        type="warning",
                        code="freeze_panes_mismatch",
                        detail=(
                            "Congelado de panel distinto: "
                            f"referencia='{reference['freeze_panes']}' generado='{candidate['freeze_panes']}'."
                        ),
                    )
                )

        errors = [item.__dict__ for item in issues if item.type == "error"]
        warnings = [item.__dict__ for item in issues if item.type == "warning"]
        return {
            "is_compatible": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "reference": reference,
            "candidate": candidate,
        }

    @staticmethod
    def _safe_label(label, fallback):
        text = BukExportService._cleanup_mojibake_text(label).strip()
        if not text:
            return fallback
        if text in {"Ãrea", "Ã¡rea"}:
            return "Area"
        return text

    @staticmethod
    def _safe_label_v2(label, fallback):
        text = BukExportService._cleanup_mojibake_text(label).strip()
        if not text:
            return fallback
        lowered = text.lower()
        if lowered in {"area", "área"}:
            return "Área"
        return text

    @staticmethod
    def _get_export_config(*, tenant, property_obj):
        config, _ = BukExportConfig.objects.get_or_create(
            tenant=tenant,
            property=property_obj,
        )
        return config

    @staticmethod
    def _fixed_columns(config):
        columns = [("document", BukExportService._safe_label_v2(config.document_column_name, "RUT"))]
        if config.include_worker_name:
            columns.append(("name", BukExportService._safe_label_v2(config.name_column_name, "Nombre")))
        if config.include_area:
            columns.append(("area", BukExportService._safe_label_v2(config.area_column_name, "Área")))
        return columns

    @staticmethod
    def _fmt_date(current_date, fmt):
        if not fmt:
            return current_date.strftime("%d-%m-%Y")
        try:
            return current_date.strftime(fmt)
        except ValueError:
            return current_date.strftime("%d-%m-%Y")

    @staticmethod
    def _date_range(date_from, date_to):
        dates = []
        cursor = date_from
        while cursor <= date_to:
            dates.append(cursor)
            cursor = cursor + timedelta(days=1)
        return dates

    @staticmethod
    def _build_day_code_index(*, tenant, property_obj, date_from, date_to, area_ids=None, worker_ids=None):
        assignments = ScheduleAssignment.objects.filter(
            tenant=tenant,
            property=property_obj,
            date__gte=date_from,
            date__lte=date_to,
        ).select_related("worker", "shift", "special_state")
        if area_ids:
            assignments = assignments.filter(worker__area_id__in=area_ids)
        if worker_ids:
            assignments = assignments.filter(worker_id__in=worker_ids)
        index = {}
        for assignment in assignments:
            if assignment.special_state_id:
                code = assignment.special_state.buk_code or ""
            elif assignment.shift_id:
                code = assignment.shift.buk_code
            else:
                code = ""
            index[(assignment.worker_id, assignment.date)] = code
        return index

    @staticmethod
    def build_preview_rows(*, tenant, property_obj, date_from, date_to, area_ids=None, worker_ids=None):
        workers = Worker.objects.filter(
            tenant=tenant,
            property=property_obj,
            active=True,
        ).select_related("area").order_by("last_name", "first_name")
        if area_ids:
            workers = workers.filter(area_id__in=area_ids)
        if worker_ids:
            workers = workers.filter(id__in=worker_ids)
        day_codes = BukExportService._build_day_code_index(
            tenant=tenant,
            property_obj=property_obj,
            date_from=date_from,
            date_to=date_to,
            area_ids=area_ids,
            worker_ids=worker_ids,
        )
        dates = BukExportService._date_range(date_from, date_to)
        row_map = {}
        for worker in workers:
            row_map[worker.id] = {
                "document": worker.document_number,
                "name": f"{worker.first_name} {worker.last_name}".strip(),
                "area": worker.area.name if worker.area_id else "",
                "days": {d.isoformat(): day_codes.get((worker.id, d), "") for d in dates},
            }
        return list(row_map.values())

    @staticmethod
    def generate_xlsx_bytes(*, tenant, property_obj, date_from, date_to, area_ids=None, worker_ids=None):
        config = BukExportService._get_export_config(tenant=tenant, property_obj=property_obj)
        preview_rows = BukExportService.build_preview_rows(
            tenant=tenant,
            property_obj=property_obj,
            date_from=date_from,
            date_to=date_to,
            area_ids=area_ids,
            worker_ids=worker_ids,
        )
        dates = BukExportService._date_range(date_from, date_to)
        fixed_columns = BukExportService._fixed_columns(config)

        header_row = max(1, int(config.header_row or 2))
        month_row = max(1, header_row - 1)
        first_data_row = max(header_row + 1, int(config.first_data_row or (header_row + 1)))
        title_label = str(config.other_settings.get("title_label", "Trabajadores")).strip() or "Trabajadores"
        date_col_start = len(fixed_columns) + 1

        workbook = Workbook()
        ws = workbook.active
        ws.title = config.sheet_name or "Reporte carga BUK"

        ws.cell(month_row, 1, title_label)
        for col_idx, (_, label) in enumerate(fixed_columns, start=1):
            ws.cell(header_row, col_idx, label)

        month_label = date_from.strftime("%m-%Y")
        for idx, current_date in enumerate(dates, start=date_col_start):
            ws.cell(month_row, idx, month_label)
            ws.cell(header_row, idx, BukExportService._fmt_date(current_date, config.date_format))

        for row_idx, item in enumerate(preview_rows, start=first_data_row):
            for col_idx, (field_key, _) in enumerate(fixed_columns, start=1):
                ws.cell(row_idx, col_idx, item.get(field_key, ""))
            for col_idx, current_date in enumerate(dates, start=date_col_start):
                ws.cell(row_idx, col_idx, item["days"].get(current_date.isoformat(), ""))

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        max_col = len(fixed_columns) + len(dates)

        for row_number in {month_row, header_row}:
            for row in ws.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=max_col):
                for cell in row:
                    cell.font = bold
                    cell.alignment = center
                    cell.border = border

        last_data_row = first_data_row + len(preview_rows) - 1
        if last_data_row >= first_data_row:
            for row in ws.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = border

        default_widths = {"document": 12, "name": 28, "area": 18}
        for col_idx, (field_key, _) in enumerate(fixed_columns, start=1):
            ws.column_dimensions[ws.cell(header_row, col_idx).column_letter].width = default_widths.get(field_key, 18)
        for col_idx in range(date_col_start, date_col_start + len(dates)):
            ws.column_dimensions[ws.cell(header_row, col_idx).column_letter].width = 12

        freeze_column_letter = ws.cell(header_row, date_col_start).column_letter
        ws.freeze_panes = f"{freeze_column_letter}{first_data_row}"

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def generate_csv_text(*, tenant, property_obj, date_from, date_to, area_ids=None, worker_ids=None):
        config = BukExportService._get_export_config(tenant=tenant, property_obj=property_obj)
        preview_rows = BukExportService.build_preview_rows(
            tenant=tenant,
            property_obj=property_obj,
            date_from=date_from,
            date_to=date_to,
            area_ids=area_ids,
            worker_ids=worker_ids,
        )
        dates = BukExportService._date_range(date_from, date_to)
        fixed_columns = BukExportService._fixed_columns(config)
        headers = [label for _, label in fixed_columns] + [
            BukExportService._fmt_date(d, config.date_format) for d in dates
        ]
        lines = [",".join(headers)]
        for row in preview_rows:
            day_values = [row["days"].get(d.isoformat(), "") for d in dates]
            values = [str(row.get(field_key, "") or "") for field_key, _ in fixed_columns] + day_values
            escaped = [f"\"{v.replace('\"', '\"\"')}\"" if "," in v or "\"" in v else v for v in values]
            lines.append(",".join(escaped))
        return "\n".join(lines)

    @staticmethod
    def log_export(
        *,
        tenant,
        property_obj,
        date_from,
        date_to,
        generated_by,
        file_name,
        validation_issues,
        export_with_observations=False,
    ):
        error_count = sum(1 for issue in validation_issues if issue.severity == "error")
        warning_count = sum(1 for issue in validation_issues if issue.severity == "warning")
        if error_count == 0:
            validation_status = "ok"
        elif export_with_observations:
            validation_status = "observations"
        else:
            validation_status = "error"
        return BukExportLog.objects.create(
            tenant=tenant,
            property=property_obj,
            date_from=date_from,
            date_to=date_to,
            generated_by=generated_by,
            file_name=file_name,
            validation_status=validation_status,
            errors_count=error_count,
            warnings_count=warning_count,
        )

    @staticmethod
    def log_template_compare(
        *,
        tenant,
        property_obj,
        compared_by,
        date_from,
        date_to,
        sheet_name,
        reference_file_name,
        reference_file_bytes,
        result,
    ):
        reference_bytes = reference_file_bytes or b""
        reference_hash = hashlib.sha256(reference_bytes).hexdigest() if reference_bytes else ""
        return BukTemplateCompareLog.objects.create(
            tenant=tenant,
            property=property_obj,
            compared_by=compared_by,
            date_from=date_from,
            date_to=date_to,
            sheet_name=sheet_name or "Reporte carga BUK",
            reference_file_name=reference_file_name or "",
            reference_file_sha256=reference_hash,
            reference_file_size_bytes=len(reference_bytes),
            is_compatible=bool(result.get("is_compatible", False)),
            errors_count=len(result.get("errors", [])),
            warnings_count=len(result.get("warnings", [])),
            result_payload=result,
        )
