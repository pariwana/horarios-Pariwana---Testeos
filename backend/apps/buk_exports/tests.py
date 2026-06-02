from datetime import date, time
from io import BytesIO
from pathlib import Path
import tempfile

from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditLog
from apps.buk_exports.models import BukExportConfig, BukExportLog, BukTemplateCompareLog
from apps.buk_exports.services import BukExportService, BukValidationService
from apps.modules.models import ModuleActivation
from apps.tenants.models import Property, Tenant, TenantSupportAccessSession
from apps.users.models import RoleChoices, User, UserPropertyPermission, UserTenantRole
from apps.workers.models import Area, Shift, SpecialState, Worker
from apps.scheduling.models import ScheduleAssignment


class BukExportApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(email="op@pariwana.test", password="StrongPass123")
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepción")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Mañana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="12345678",
            first_name="Ana",
            last_name="Quispe",
            area=self.area,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date=date(2026, 4, 2),
            shift=self.shift,
        )
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_export", is_enabled=False)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_validator", is_enabled=True)
        self.client.force_authenticate(user=self.user)

    def test_export_denies_when_module_disabled(self):
        response = self.client.post(
            "/api/buk/export/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-02",
                "date_to": "2026-04-02",
                "format": "xlsx",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 403)


class BukExportSupportContextTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.super_admin = User.objects.create_user(
            email="super-buk@pariwana.test",
            password="StrongPass123",
            is_super_admin=True,
            is_staff=True,
        )
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="88888888",
            first_name="Luis",
            last_name="Rojas",
            area=self.area,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date=date(2026, 4, 5),
            shift=self.shift,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        self.support_session = TenantSupportAccessSession.objects.create(
            tenant=self.tenant,
            property=self.property,
            started_by=self.super_admin,
            reason="support buk preview",
        )
        self.client.force_authenticate(user=self.super_admin)

    def test_preview_works_with_support_session_without_tenant_id(self):
        response = self.client.post(
            "/api/buk/preview/",
            {
                "property_id": self.property.id,
                "date_from": "2026-04-05",
                "date_to": "2026-04-05",
            },
            format="json",
            HTTP_X_SUPPORT_SESSION_ID=str(self.support_session.id),
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        self.assertIn("rows", response.json())


class BukExportWithObservationsApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="77777777",
            first_name="Ana",
            last_name="SinDoc",
            area=self.area,
        )
        self.worker.document_number = ""
        self.worker.save(update_fields=["document_number", "updated_at"])
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 4, 7),
            shift=self.shift,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_export", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_validator", is_enabled=True)

    def test_operator_cannot_export_with_observations(self):
        operator = User.objects.create_user(email="op-obs@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=operator, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=operator,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        self.client.force_authenticate(user=operator)
        response = self.client.post(
            "/api/buk/export/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-07",
                "date_to": "2026-04-07",
                "format": "xlsx",
                "export_with_observations": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_can_export_with_observations(self):
        admin = User.objects.create_user(email="admin-obs@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=admin, tenant=self.tenant, role=RoleChoices.ADMIN)
        UserPropertyPermission.objects.create(
            user=admin,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        self.client.force_authenticate(user=admin)
        response = self.client.post(
            "/api/buk/export/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-07",
                "date_to": "2026-04-07",
                "format": "xlsx",
                "export_with_observations": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        log = BukExportLog.objects.latest("id")
        self.assertEqual(log.validation_status, "observations")


class BukExportAreaFilterApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area_1 = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.area_2 = Area.objects.create(tenant=self.tenant, property=self.property, name="Bar")
        self.shift_1 = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area_1,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )
        self.shift_2 = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area_2,
            name="Tarde",
            buk_code="BAR-T",
            start_time=time(14, 0),
            end_time=time(22, 0),
        )
        worker_1 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="10000001",
            first_name="Mario",
            last_name="Soto",
            area=self.area_1,
        )
        worker_2 = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="10000002",
            first_name="Lucia",
            last_name="Perez",
            area=self.area_2,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker_1,
            date=date(2026, 4, 8),
            shift=self.shift_1,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker_2,
            date=date(2026, 4, 8),
            shift=self.shift_2,
        )
        self.user = User.objects.create_user(email="op-area@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_validator", is_enabled=True)
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_export", is_enabled=True)
        self.client.force_authenticate(user=self.user)

    def test_preview_filters_rows_by_area_ids(self):
        response = self.client.post(
            "/api/buk/preview/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-08",
                "date_to": "2026-04-08",
                "area_ids": [self.area_1.id],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        rows = response.json()["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Mario Soto")

    def test_preview_filters_rows_by_worker_ids(self):
        mario = Worker.objects.get(tenant=self.tenant, property=self.property, document_number="10000001")
        response = self.client.post(
            "/api/buk/preview/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-08",
                "date_to": "2026-04-08",
                "worker_ids": [mario.id],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        rows = response.json()["rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Mario Soto")


class BukTemplateCompareApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )
        worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="30000001",
            first_name="Mario",
            last_name="Soto",
            area=self.area,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=worker,
            date=date(2026, 4, 12),
            shift=self.shift,
        )
        self.user = User.objects.create_user(email="op-compare@pariwana.test", password="StrongPass123")
        UserTenantRole.objects.create(user=self.user, tenant=self.tenant, role=RoleChoices.OPERATOR)
        UserPropertyPermission.objects.create(
            user=self.user,
            tenant=self.tenant,
            property=self.property,
            can_access=True,
            can_export_buk=True,
        )
        ModuleActivation.objects.create(tenant=self.tenant, module_key="buk_preview", is_enabled=True)
        self.client.force_authenticate(user=self.user)

    def test_compare_template_api_returns_compatible_result(self):
        reference_bytes = BukExportService.generate_xlsx_bytes(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 4, 12),
            date_to=date(2026, 4, 12),
        )
        uploaded = SimpleUploadedFile(
            "base.xlsx",
            reference_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/api/buk/compare-template/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-12",
                "date_to": "2026-04-12",
                "reference_file": uploaded,
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        payload = response.json()
        self.assertTrue(payload["is_compatible"])
        self.assertEqual(payload["errors"], [])
        self.assertIn("compare_log_id", payload)
        log = BukTemplateCompareLog.objects.get(id=payload["compare_log_id"])
        self.assertEqual(len(log.reference_file_sha256), 64)
        self.assertGreater(log.reference_file_size_bytes, 0)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="buk_compare_template",
                entity_type="BukTemplateCompareLog",
                entity_id=str(log.id),
            ).exists()
        )

    def test_compare_template_api_can_download_json_report(self):
        reference_bytes = BukExportService.generate_xlsx_bytes(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 4, 12),
            date_to=date(2026, 4, 12),
        )
        uploaded = SimpleUploadedFile(
            "base.xlsx",
            reference_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/api/buk/compare-template/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-12",
                "date_to": "2026-04-12",
                "download_report": "1",
                "reference_file": uploaded,
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        self.assertIn("application/json", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertEqual(BukTemplateCompareLog.objects.count(), 1)
        log = BukTemplateCompareLog.objects.first()
        self.assertEqual(len(log.reference_file_sha256), 64)
        self.assertGreater(log.reference_file_size_bytes, 0)

    def test_compare_template_api_denies_when_module_disabled(self):
        ModuleActivation.objects.filter(tenant=self.tenant, module_key="buk_preview").update(is_enabled=False)
        reference_bytes = BukExportService.generate_xlsx_bytes(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 4, 12),
            date_to=date(2026, 4, 12),
        )
        uploaded = SimpleUploadedFile(
            "base.xlsx",
            reference_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/api/buk/compare-template/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "date_from": "2026-04-12",
                "date_to": "2026-04-12",
                "reference_file": uploaded,
            },
            format="multipart",
        )
        self.assertEqual(response.status_code, 403)

    def test_compare_template_logs_api_lists_results_with_filters(self):
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok.xlsx",
            reference_file_sha256="a" * 64,
            reference_file_size_bytes=120,
            is_compatible=True,
            errors_count=0,
            warnings_count=1,
            result_payload={"is_compatible": True},
        )
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 4, 3),
            date_to=date(2026, 4, 4),
            sheet_name="Reporte carga BUK",
            reference_file_name="bad.xlsx",
            reference_file_sha256="b" * 64,
            reference_file_size_bytes=240,
            is_compatible=False,
            errors_count=2,
            warnings_count=0,
            result_payload={"is_compatible": False},
        )
        response = self.client.get(
            "/api/buk/compare-template-logs/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "is_compatible": "0",
                "limit": "5",
            },
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        payload = response.json()
        self.assertEqual(len(payload["results"]), 1)
        self.assertEqual(payload["results"][0]["reference_file_name"], "bad.xlsx")
        self.assertFalse(payload["results"][0]["is_compatible"])

    def test_compare_template_logs_api_downloads_json_by_id(self):
        item = BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok.xlsx",
            reference_file_sha256="a" * 64,
            reference_file_size_bytes=120,
            is_compatible=True,
            errors_count=0,
            warnings_count=1,
            result_payload={"is_compatible": True},
        )
        response = self.client.get(
            f"/api/buk/compare-template-logs/{item.id}/download/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
            },
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        self.assertIn("application/json", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="buk_compare_template_log_download_json",
                entity_type="BukTemplateCompareLog",
                entity_id=str(item.id),
            ).exists()
        )

    def test_compare_template_logs_api_supports_date_filter_and_pagination(self):
        for idx in range(1, 6):
            BukTemplateCompareLog.objects.create(
                tenant=self.tenant,
                property=self.property,
                compared_by=self.user,
                date_from=date(2026, 4, idx),
                date_to=date(2026, 4, idx),
                sheet_name="Reporte carga BUK",
                reference_file_name=f"log-{idx}.xlsx",
                reference_file_sha256=str(idx) * 64,
                reference_file_size_bytes=100 + idx,
                is_compatible=idx % 2 == 0,
                errors_count=0 if idx % 2 == 0 else 1,
                warnings_count=0,
                result_payload={"is_compatible": idx % 2 == 0},
            )
        response = self.client.get(
            "/api/buk/compare-template-logs/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "compared_from": "2026-01-01",
                "page_size": "2",
                "page": "1",
            },
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        payload = response.json()
        self.assertEqual(payload["pagination"]["page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 2)
        self.assertEqual(payload["pagination"]["total"], 5)
        self.assertEqual(payload["pagination"]["total_pages"], 3)
        self.assertEqual(len(payload["results"]), 2)

        response = self.client.get(
            "/api/buk/compare-template-logs/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "compared_from": "2099-01-01",
            },
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        payload = response.json()
        self.assertEqual(payload["pagination"]["total"], 0)
        self.assertEqual(len(payload["results"]), 0)

    def test_compare_template_logs_api_can_export_csv(self):
        BukTemplateCompareLog.objects.create(
            tenant=self.tenant,
            property=self.property,
            compared_by=self.user,
            date_from=date(2026, 4, 1),
            date_to=date(2026, 4, 2),
            sheet_name="Reporte carga BUK",
            reference_file_name="ok.csvtest.xlsx",
            reference_file_sha256="a" * 64,
            reference_file_size_bytes=120,
            is_compatible=True,
            errors_count=0,
            warnings_count=0,
            result_payload={"is_compatible": True},
        )
        response = self.client.get(
            "/api/buk/compare-template-logs/export-csv/",
            {
                "tenant_id": self.tenant.id,
                "property_id": self.property.id,
                "is_compatible": "1",
            },
        )
        self.assertEqual(response.status_code, 200, msg=response.content)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn("attachment; filename=", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        self.assertIn("reference_file_name", body)
        self.assertIn("ok.csvtest.xlsx", body)
        self.assertTrue(
            AuditLog.objects.filter(
                tenant=self.tenant,
                property=self.property,
                user=self.user,
                action="buk_compare_template_logs_export_csv",
                entity_type="BukTemplateCompareLog",
            ).exists()
        )


class BukExportServiceRegressionTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="22223333",
            first_name="Rosa",
            last_name="Paz",
            area=self.area,
        )
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 45),
        )

    def test_xlsx_uses_config_structure_and_special_state_code(self):
        state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Descanso",
            buk_code="DES",
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 4, 10),
            special_state=state,
        )
        BukExportConfig.objects.create(
            tenant=self.tenant,
            property=self.property,
            sheet_name="Reporte carga BUK",
            date_format="%d/%m/%Y",
            document_column_name="Documento",
            name_column_name="Trabajador",
            area_column_name="Area",
            header_row=4,
            first_data_row=6,
            other_settings={"title_label": "Carga BUK"},
        )

        xlsx_bytes = BukExportService.generate_xlsx_bytes(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 4, 10),
            date_to=date(2026, 4, 10),
        )
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["Reporte carga BUK"]

        self.assertEqual(ws.cell(3, 1).value, "Carga BUK")
        self.assertEqual(ws.cell(4, 1).value, "Documento")
        self.assertEqual(ws.cell(4, 2).value, "Trabajador")
        self.assertEqual(ws.cell(4, 3).value, "Área")
        self.assertEqual(ws.cell(4, 4).value, "10/04/2026")
        self.assertEqual(ws.cell(6, 1).value, "22223333")
        self.assertEqual(ws.cell(6, 2).value, "Rosa Paz")
        self.assertEqual(ws.cell(6, 3).value, "Recepcion")
        self.assertEqual(ws.cell(6, 4).value, "DES")
        self.assertEqual(str(ws.freeze_panes), "D6")

    def test_csv_respects_optional_fixed_columns(self):
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 4, 11),
            shift=self.shift,
        )
        BukExportConfig.objects.create(
            tenant=self.tenant,
            property=self.property,
            include_worker_name=False,
            include_area=False,
            document_column_name="RUT",
            date_format="%d-%m-%Y",
        )

        text = BukExportService.generate_csv_text(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 4, 11),
            date_to=date(2026, 4, 11),
        )
        lines = text.splitlines()
        self.assertEqual(lines[0], "RUT,11-04-2026")
        self.assertEqual(lines[1], "22223333,REC-M")


class BukTemplateCompatibilityTests(TestCase):
    @staticmethod
    def _build_template_bytes(*, fixed_labels=None, header_row=2, first_data_row=3, date_text="01-04-2026"):
        fixed_labels = fixed_labels or ["RUT", "Nombre", "Area"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte carga BUK"
        ws.cell(header_row - 1, 1, "Trabajadores")
        for index, label in enumerate(fixed_labels, start=1):
            ws.cell(header_row, index, label)
        date_col = len(fixed_labels) + 1
        ws.cell(header_row - 1, date_col, "04-2026")
        ws.cell(header_row, date_col, date_text)
        ws.cell(first_data_row, 1, "12345678")
        ws.cell(first_data_row, 2, "Ana Quispe")
        ws.cell(first_data_row, 3, "Recepcion")
        ws.freeze_panes = f"{ws.cell(header_row, date_col).column_letter}{first_data_row}"
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_template_comparator_detects_compatible_structure(self):
        reference = self._build_template_bytes()
        candidate = self._build_template_bytes()
        result = BukExportService.compare_template_compatibility(
            reference_file_bytes=reference,
            candidate_file_bytes=candidate,
        )
        self.assertTrue(result["is_compatible"])
        self.assertEqual(result["errors"], [])

    def test_template_comparator_flags_header_row_difference(self):
        reference = self._build_template_bytes(header_row=2, first_data_row=3)
        candidate = self._build_template_bytes(header_row=4, first_data_row=5)
        result = BukExportService.compare_template_compatibility(
            reference_file_bytes=reference,
            candidate_file_bytes=candidate,
        )
        self.assertFalse(result["is_compatible"])
        error_codes = {item["code"] for item in result["errors"]}
        self.assertIn("header_row_mismatch", error_codes)

    def test_template_comparator_flags_fixed_label_difference(self):
        reference = self._build_template_bytes(fixed_labels=["RUT", "Nombre", "Area"])
        candidate = self._build_template_bytes(fixed_labels=["Documento", "Nombre", "Area"])
        result = BukExportService.compare_template_compatibility(
            reference_file_bytes=reference,
            candidate_file_bytes=candidate,
        )
        self.assertFalse(result["is_compatible"])
        error_codes = {item["code"] for item in result["errors"]}
        self.assertIn("fixed_labels_mismatch", error_codes)

    def test_template_comparator_normalizes_mojibake_fixed_labels(self):
        reference = self._build_template_bytes(fixed_labels=["RUT", "Nombre", "Ãrea"])
        candidate = self._build_template_bytes(fixed_labels=["RUT", "Nombre", "Área"])
        result = BukExportService.compare_template_compatibility(
            reference_file_bytes=reference,
            candidate_file_bytes=candidate,
        )
        self.assertTrue(result["is_compatible"])
        self.assertEqual(result["errors"], [])


class BukValidationServiceRulesTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="55667788",
            first_name="Dina",
            last_name="Lopez",
            area=self.area,
            active=True,
        )
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Tarde",
            buk_code="REC-T",
            start_time=time(14, 0),
            end_time=time(22, 0),
            is_night_shift=False,
        )

    def test_validator_warns_active_worker_without_assignment(self):
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 1),
            date_to=date(2026, 5, 1),
        )
        self.assertTrue(any(issue.problem == "Trabajador activo sin horario en el rango." for issue in issues))
        self.assertTrue(any(issue.severity == "warning" for issue in issues))

    def test_validator_flags_special_state_without_buk_code_when_used(self):
        state = SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Licencia",
            buk_code="",
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 5, 2),
            special_state=state,
        )
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 2),
            date_to=date(2026, 5, 2),
        )
        self.assertTrue(any("Estado especial sin codigo BUK" in issue.problem for issue in issues))
        self.assertTrue(any(issue.severity == "error" for issue in issues))

    def test_validator_flags_buk_code_duplicate_between_shift_and_state(self):
        SpecialState.objects.create(
            tenant=self.tenant,
            property=self.property,
            name="Descanso",
            buk_code="REC-T",
        )
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 3),
            date_to=date(2026, 5, 3),
        )
        self.assertTrue(any("Codigo BUK duplicado: REC-T." == issue.problem for issue in issues))

    def test_validator_flags_assignment_outside_worker_start_date(self):
        self.worker.start_date = date(2026, 5, 15)
        self.worker.save(update_fields=["start_date", "updated_at"])
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 5, 10),
            shift=self.shift,
        )
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 10),
            date_to=date(2026, 5, 10),
        )
        self.assertTrue(any("fecha antes del inicio del trabajador" in issue.problem for issue in issues))

    def test_validator_flags_worker_active_outside_range_validity(self):
        self.worker.start_date = date(2026, 6, 1)
        self.worker.end_date = date(2026, 6, 30)
        self.worker.save(update_fields=["start_date", "end_date", "updated_at"])
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 10),
            date_to=date(2026, 5, 10),
        )
        self.assertTrue(any("inicio posterior al rango" in issue.problem for issue in issues))

    def test_validator_flags_assignment_with_inactive_shift(self):
        self.shift.active = False
        self.shift.save(update_fields=["active", "updated_at"])
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 5, 11),
            shift=self.shift,
        )
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 11),
            date_to=date(2026, 5, 11),
        )
        self.assertTrue(any("turno inactivo" in issue.problem for issue in issues))

    def test_validator_flags_assignment_shift_without_buk_code(self):
        self.shift.buk_code = ""
        self.shift.save(update_fields=["buk_code", "updated_at"])
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 5, 12),
            shift=self.shift,
        )
        issues = BukValidationService.validate_assignments(
            tenant=self.tenant,
            property_obj=self.property,
            date_from=date(2026, 5, 12),
            date_to=date(2026, 5, 12),
        )
        self.assertTrue(any("Turno asignado sin codigo BUK" in issue.problem for issue in issues))


class BukTemplateCompatibilityCommandTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pariwana Hostels", slug="pariwana-hostels")
        self.property = Property.objects.create(tenant=self.tenant, name="Pariwana Cusco", slug="pariwana-cusco")
        self.area = Area.objects.create(tenant=self.tenant, property=self.property, name="Recepcion")
        self.shift = Shift.objects.create(
            tenant=self.tenant,
            property=self.property,
            area=self.area,
            name="Manana",
            buk_code="REC-M",
            start_time=time(6, 0),
            end_time=time(14, 0),
        )
        self.worker = Worker.objects.create(
            tenant=self.tenant,
            property=self.property,
            document_number="11111111",
            first_name="Ana",
            last_name="Quispe",
            area=self.area,
        )
        ScheduleAssignment.objects.create(
            tenant=self.tenant,
            property=self.property,
            worker=self.worker,
            date=date(2026, 4, 2),
            shift=self.shift,
        )

    def test_command_succeeds_with_compatible_reference(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            reference_path = Path(tmpdir) / "reference.xlsx"
            output_path = Path(tmpdir) / "result.json"
            reference_bytes = BukExportService.generate_xlsx_bytes(
                tenant=self.tenant,
                property_obj=self.property,
                date_from=date(2026, 4, 2),
                date_to=date(2026, 4, 2),
            )
            reference_path.write_bytes(reference_bytes)

            call_command(
                "check_buk_template_compatibility",
                tenant_slug="pariwana-hostels",
                property_slug="pariwana-cusco",
                date_from="2026-04-02",
                date_to="2026-04-02",
                reference_file=str(reference_path),
                output_json=str(output_path),
            )
            payload = output_path.read_text(encoding="utf-8")
            self.assertIn('"is_compatible": true', payload)

    def test_command_fails_when_reference_sheet_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            reference_path = Path(tmpdir) / "reference_missing_sheet.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Otra hoja"
            ws.cell(1, 1, "x")
            wb.save(reference_path)

            with self.assertRaises(CommandError):
                call_command(
                    "check_buk_template_compatibility",
                    tenant_slug="pariwana-hostels",
                    property_slug="pariwana-cusco",
                    date_from="2026-04-02",
                    date_to="2026-04-02",
                    reference_file=str(reference_path),
                )
